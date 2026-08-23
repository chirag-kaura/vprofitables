"""
notifier.py — Email + WhatsApp Signal Delivery for GANN-ASTRO v3.7
Place in: core/notifier.py  (or root — app.py imports from here)

Supports:
  1. Email via Gmail SMTP (app password)
  2. WhatsApp via Twilio WhatsApp Sandbox (free tier available)
  3. WhatsApp via CallMeBot (free, no account needed — easiest)

Configuration (edit NOTIFICATION_CONFIG below or pass via settings):
  EMAIL_TO        : recipient email address
  EMAIL_FROM      : your Gmail address
  EMAIL_PASS      : Gmail App Password (not your login password)
                    Get at: myaccount.google.com/apppasswords
  WHATSAPP_METHOD : "callmebot" | "twilio" | "none"
  CALLMEBOT_PHONE : your WhatsApp number with country code e.g. +919876543210
  CALLMEBOT_KEY   : get free at api.callmebot.com/whatsapp.php
  TWILIO_SID      : Twilio Account SID (if using Twilio)
  TWILIO_TOKEN    : Twilio Auth Token
  TWILIO_FROM     : Twilio WhatsApp number e.g. whatsapp:+14155238886
  TWILIO_TO       : your WhatsApp number e.g. whatsapp:+919876543210
"""

import os, json, smtplib, urllib.request, urllib.parse
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from datetime import datetime
from typing import Dict, Optional

from core.paths import CONFIG_PATH
_CFG_PATH = CONFIG_PATH

def _load_cfg() -> Dict:
    # ── Defaults — no personal credentials in source code ────────────────────
    # Set GANN_EMAIL_FROM / GANN_EMAIL_PASS / GANN_EMAIL_TO env vars in
    # production (Oracle Cloud) or configure via the app's Settings → Notifications UI.
    defaults = {
        "EMAIL_ENABLED":    False,
        "EMAIL_TO":         os.environ.get("GANN_EMAIL_TO", ""),
        "EMAIL_FROM":       os.environ.get("GANN_EMAIL_FROM", ""),
        "EMAIL_PASS":       os.environ.get("GANN_EMAIL_PASS", ""),
        "WHATSAPP_METHOD":  "none",   # none | callmebot | twilio
        "CALLMEBOT_PHONE":  "",
        "CALLMEBOT_KEY":    "",
        "TWILIO_SID":       "",
        "TWILIO_TOKEN":     "",
        "TWILIO_FROM":      "whatsapp:+14155238886",
        "TWILIO_TO":        "",
    }
    try:
        if os.path.exists(_CFG_PATH):
            with open(_CFG_PATH, encoding="utf-8") as f:
                saved = json.load(f)
            defaults.update(saved)
    except Exception:
        pass
    # Environment variable overrides
    for key in defaults:
        ev = os.environ.get(f"GANN_{key}")
        if ev:
            defaults[key] = ev
    return defaults


def save_cfg(updates: Dict) -> bool:
    """Save notification config to gann_settings.json"""
    try:
        cfg = _load_cfg()
        cfg.update(updates)
        with open(_CFG_PATH, "w", encoding="utf-8") as f:
            json.dump(cfg, f, indent=2)
        return True
    except Exception as e:
        print(f"  [NOTIFY] Config save error: {e}")
        return False


# ─────────────────────────────────────────────────────────────────────────────
# SIGNAL FORMATTER
# ─────────────────────────────────────────────────────────────────────────────

def _format_signal_text(signal: Dict) -> tuple:
    """
    Format a forward signal into (subject, html_body, whatsapp_text).
    signal dict keys (from forward_signal endpoint):
      symbol, name, sector, analysis_date, inv_type,
      action, entry, stop_loss, target1, target2,
      rr_ratio, confidence, regime, hold_days,
      buy_date, sell_date, buy_time,
      reasons (list of strings),
      engine_scores (dict: technical/gann/natal/simons/fundamental/sentiment),
      bulk_signal, news_sentiment, wyckoff_phase, macro_note
    """
    sym         = signal.get("symbol", "")
    name        = signal.get("name", sym)
    sector      = signal.get("sector", "")
    dt          = signal.get("analysis_date", datetime.now().strftime("%Y-%m-%d"))
    inv_type    = signal.get("inv_type", "swing").replace("_", " ").title()
    action      = signal.get("action", "BUY")
    entry       = signal.get("entry", 0)
    sl          = signal.get("stop_loss", 0)
    t1          = signal.get("target1", 0)
    t2          = signal.get("target2", 0)
    rr          = signal.get("rr_ratio", 0)
    conf        = signal.get("confidence", 0)
    regime      = signal.get("regime", "")
    hold        = signal.get("hold_days", 0)
    buy_date    = signal.get("buy_date", "")
    sell_date   = signal.get("sell_date", "")
    buy_time    = signal.get("buy_time", "09:15 IST")
    reasons     = signal.get("reasons", [])
    scores      = signal.get("engine_scores", {})
    bulk        = signal.get("bulk_signal", "NEUTRAL")
    news        = signal.get("news_sentiment", "N/A")
    wyckoff     = signal.get("wyckoff_phase", "")
    macro       = signal.get("macro_note", "")
    ruling_pl   = signal.get("ruling_planet", "")

    risk_amt    = round(entry - sl, 2) if entry and sl else 0
    reward1_amt = round(t1 - entry, 2) if t1 and entry else 0
    reward2_amt = round(t2 - entry, 2) if t2 and entry else 0

    action_col  = "#00c87a" if action == "BUY" else "#e05050"
    conf_col    = "#00c87a" if conf >= 70 else "#e6b800" if conf >= 50 else "#e05050"

    # ── WhatsApp text (plain, concise) ───────────────────────────────────────
    wa = f"""🏦 *GANN-ASTRO SIGNAL* | {dt}
━━━━━━━━━━━━━━━━━━━━
{'🟢' if action=='BUY' else '🔴'} *{action}: {sym}* ({name})
📊 Sector: {sector} | Type: {inv_type}

💰 *ENTRY:* ₹{entry:,.2f}
🛡 *STOP LOSS:* ₹{sl:,.2f}  (Risk: ₹{risk_amt:,.2f})
🎯 *TARGET 1:* ₹{t1:,.2f}  (Reward: ₹{reward1_amt:,.2f})
🎯 *TARGET 2:* ₹{t2:,.2f}  (Reward: ₹{reward2_amt:,.2f})
📐 *R:R Ratio:* {rr}:1
⚡ *Confidence:* {conf}%

📅 Entry window: {buy_date} {buy_time}
📅 Exit by: {sell_date}
⏱ Hold: ~{hold} days

📈 *ENGINE SCORES:*
  Technical: {scores.get('technical','—')}  Gann: {scores.get('gann','—')}
  Simons: {scores.get('simons','—')}  Natal: {scores.get('natal','—')}
  Fundamental: {scores.get('fundamental','—')}  Sentiment: {scores.get('sentiment','—')}

🌊 Wyckoff: {wyckoff or 'N/A'}
📰 News: {news}  🏦 Bulk deals: {bulk}
🌌 Regime: {regime}  Ruling planet: {ruling_pl}
{('📌 Macro: ' + macro) if macro else ''}

Top reasons:
{chr(10).join('• ' + r for r in reasons[:4])}

⚠ Not financial advice. Trade at your own risk.
━━━━━━━━━━━━━━━━━━━━"""

    # ── Email subject ─────────────────────────────────────────────────────────
    subject = f"GANN-ASTRO {'🟢 BUY' if action=='BUY' else '🔴 SELL'} Signal: {sym} ₹{entry:,.0f} | R:R {rr}:1 | Conf {conf}%"

    # ── Email HTML body ───────────────────────────────────────────────────────
    reasons_html = "".join(f"<li style='margin:4px 0;color:#b0c8d8;font-size:13px;'>{r}</li>" for r in reasons[:8])
    scores_html  = "".join(
        f"<td style='padding:6px 10px;text-align:center;border:1px solid #1a3040;'>"
        f"<div style='font-size:10px;color:#4a6678;letter-spacing:1px;'>{k.upper()}</div>"
        f"<div style='font-size:14px;font-weight:bold;color:#00d4ff;'>{v}</div></td>"
        for k, v in scores.items() if v
    )

    html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"></head>
<body style="margin:0;padding:20px;background:#060f16;font-family:'Courier New',monospace;">
<div style="max-width:650px;margin:0 auto;background:#0d1e2d;border:1px solid #1a3040;border-radius:8px;overflow:hidden;">

  <!-- Header -->
  <div style="background:#0a1520;padding:16px 24px;border-bottom:1px solid #1a3040;">
    <div style="display:flex;justify-content:space-between;align-items:center;">
      <span style="font-size:18px;font-weight:bold;color:#e6b800;letter-spacing:2px;">GANN·ASTRO</span>
      <span style="font-size:11px;color:#4a6678;letter-spacing:1px;">SIMONS QUANT v3.7</span>
    </div>
    <div style="font-size:11px;color:#4a6678;margin-top:4px;">{dt} · {inv_type.upper()} SIGNAL</div>
  </div>

  <!-- Signal headline -->
  <div style="padding:20px 24px;border-bottom:1px solid #1a3040;background:{'rgba(0,200,122,0.06)' if action=='BUY' else 'rgba(224,80,80,0.06)'};">
    <div style="font-size:24px;font-weight:bold;color:{action_col};">
      {'🟢' if action=='BUY' else '🔴'} {action}: {sym}
    </div>
    <div style="font-size:14px;color:#b0c8d8;margin-top:4px;">{name} · {sector}</div>
  </div>

  <!-- Trade levels -->
  <div style="padding:20px 24px;display:grid;gap:2px;border-bottom:1px solid #1a3040;">
    <table style="width:100%;border-collapse:collapse;">
      <tr>
        <td style="padding:10px;background:#0a1520;border:1px solid #1a3040;width:50%;">
          <div style="font-size:10px;color:#4a6678;letter-spacing:1px;">ENTRY PRICE</div>
          <div style="font-size:22px;font-weight:bold;color:#00d4ff;">₹{entry:,.2f}</div>
        </td>
        <td style="padding:10px;background:#0a1520;border:1px solid #1a3040;width:50%;">
          <div style="font-size:10px;color:#4a6678;letter-spacing:1px;">CONFIDENCE</div>
          <div style="font-size:22px;font-weight:bold;color:{conf_col};">{conf}%</div>
        </td>
      </tr>
      <tr>
        <td style="padding:10px;background:#0a1520;border:1px solid #1a3040;">
          <div style="font-size:10px;color:#4a6678;letter-spacing:1px;">STOP LOSS</div>
          <div style="font-size:18px;font-weight:bold;color:#e05050;">₹{sl:,.2f}</div>
          <div style="font-size:11px;color:#4a6678;">Risk: ₹{risk_amt:,.2f} per share</div>
        </td>
        <td style="padding:10px;background:#0a1520;border:1px solid #1a3040;">
          <div style="font-size:10px;color:#4a6678;letter-spacing:1px;">R:R RATIO</div>
          <div style="font-size:18px;font-weight:bold;color:#e6b800;">{rr}:1</div>
        </td>
      </tr>
      <tr>
        <td style="padding:10px;background:#0a1520;border:1px solid #1a3040;">
          <div style="font-size:10px;color:#4a6678;letter-spacing:1px;">TARGET 1</div>
          <div style="font-size:18px;font-weight:bold;color:#00c87a;">₹{t1:,.2f}</div>
          <div style="font-size:11px;color:#4a6678;">Reward: ₹{reward1_amt:,.2f}</div>
        </td>
        <td style="padding:10px;background:#0a1520;border:1px solid #1a3040;">
          <div style="font-size:10px;color:#4a6678;letter-spacing:1px;">TARGET 2</div>
          <div style="font-size:18px;font-weight:bold;color:#8fd8a0;">₹{t2:,.2f}</div>
          <div style="font-size:11px;color:#4a6678;">Reward: ₹{reward2_amt:,.2f}</div>
        </td>
      </tr>
    </table>
  </div>

  <!-- Timing -->
  <div style="padding:16px 24px;background:#0a1520;border-bottom:1px solid #1a3040;">
    <div style="display:flex;gap:24px;flex-wrap:wrap;">
      <div><div style="font-size:10px;color:#4a6678;letter-spacing:1px;">ENTRY DATE</div>
           <div style="font-size:13px;color:#00c87a;font-weight:bold;">{buy_date} {buy_time}</div></div>
      <div><div style="font-size:10px;color:#4a6678;letter-spacing:1px;">EXIT BY</div>
           <div style="font-size:13px;color:#e05050;font-weight:bold;">{sell_date}</div></div>
      <div><div style="font-size:10px;color:#4a6678;letter-spacing:1px;">HOLD PERIOD</div>
           <div style="font-size:13px;color:#00d4ff;font-weight:bold;">~{hold} days</div></div>
      <div><div style="font-size:10px;color:#4a6678;letter-spacing:1px;">REGIME</div>
           <div style="font-size:13px;color:#e6b800;font-weight:bold;">{regime}</div></div>
    </div>
  </div>

  <!-- Engine scores -->
  <div style="padding:16px 24px;border-bottom:1px solid #1a3040;">
    <div style="font-size:10px;color:#4a6678;letter-spacing:2px;margin-bottom:10px;">ENGINE SCORES</div>
    <table style="width:100%;border-collapse:collapse;">{scores_html}</table>
  </div>

  <!-- Intelligence signals -->
  <div style="padding:16px 24px;border-bottom:1px solid #1a3040;">
    <div style="display:flex;gap:16px;flex-wrap:wrap;font-size:12px;">
      <div><span style="color:#4a6678;">Wyckoff: </span><span style="color:#cc88ff;">{wyckoff or 'N/A'}</span></div>
      <div><span style="color:#4a6678;">News: </span><span style="color:#00d4ff;">{news}</span></div>
      <div><span style="color:#4a6678;">Bulk deals: </span><span style="color:#e6b800;">{bulk}</span></div>
      <div><span style="color:#4a6678;">Ruling planet: </span><span style="color:#cc88ff;">{ruling_pl}</span></div>
    </div>
    {f'<div style="margin-top:8px;font-size:12px;color:#4a6678;">Macro: {macro}</div>' if macro else ''}
  </div>

  <!-- Reasons -->
  <div style="padding:16px 24px;border-bottom:1px solid #1a3040;">
    <div style="font-size:10px;color:#4a6678;letter-spacing:2px;margin-bottom:8px;">BUY REASONS</div>
    <ul style="margin:0;padding-left:18px;">{reasons_html}</ul>
  </div>

  <!-- Footer -->
  <div style="padding:12px 24px;text-align:center;font-size:10px;color:#2a4050;">
    ⚠ For educational purposes only. Not financial advice. All trades carry risk.<br>
    Generated by GANN-ASTRO v3.7 · {datetime.now().strftime('%Y-%m-%d %H:%M IST')}
  </div>

</div>
</body></html>"""

    return subject, html, wa


# ─────────────────────────────────────────────────────────────────────────────
# SEND FUNCTIONS
# ─────────────────────────────────────────────────────────────────────────────

def send_email(signal: Dict) -> Dict:
    """Send signal via Gmail SMTP."""
    cfg = _load_cfg()
    if not cfg.get("EMAIL_ENABLED") or not cfg.get("EMAIL_TO"):
        return {"ok": False, "error": "Email not configured or disabled"}

    subject, html, _ = _format_signal_text(signal)

    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject
        msg["From"]    = cfg["EMAIL_FROM"]
        msg["To"]      = cfg["EMAIL_TO"]
        msg.attach(MIMEText(html, "html", "utf-8"))

        with smtplib.SMTP_SSL("smtp.gmail.com", 465, timeout=15) as server:
            server.login(cfg["EMAIL_FROM"], cfg["EMAIL_PASS"])
            server.sendmail(cfg["EMAIL_FROM"], cfg["EMAIL_TO"], msg.as_string())

        print(f"  [NOTIFY] Email sent → {cfg['EMAIL_TO']}", flush=True)
        return {"ok": True, "method": "email", "to": cfg["EMAIL_TO"]}

    except smtplib.SMTPAuthenticationError:
        return {"ok": False, "error": "Gmail auth failed — check App Password at myaccount.google.com/apppasswords"}
    except Exception as e:
        return {"ok": False, "error": f"Email error: {e}"}


def send_whatsapp(signal: Dict) -> Dict:
    """Send signal via WhatsApp (CallMeBot free API or Twilio)."""
    cfg = _load_cfg()
    method = cfg.get("WHATSAPP_METHOD", "none")

    if method == "none":
        return {"ok": False, "error": "WhatsApp not configured"}

    _, _, wa_text = _format_signal_text(signal)

    if method == "callmebot":
        phone = cfg.get("CALLMEBOT_PHONE", "").strip()
        key   = cfg.get("CALLMEBOT_KEY", "").strip()
        if not phone or not key:
            return {"ok": False, "error": "CALLMEBOT_PHONE and CALLMEBOT_KEY required"}
        try:
            encoded = urllib.parse.quote(wa_text)
            url = f"https://api.callmebot.com/whatsapp.php?phone={phone}&text={encoded}&apikey={key}"
            req = urllib.request.Request(url, headers={"User-Agent": "GANN-ASTRO/3.7"})
            with urllib.request.urlopen(req, timeout=15) as resp:
                body = resp.read().decode("utf-8", errors="replace")
            print(f"  [NOTIFY] WhatsApp (CallMeBot) sent → {phone}", flush=True)
            return {"ok": True, "method": "callmebot", "to": phone, "response": body[:100]}
        except Exception as e:
            return {"ok": False, "error": f"CallMeBot error: {e}"}

    elif method == "twilio":
        try:
            from twilio.rest import Client as TwilioClient
        except ImportError:
            return {"ok": False, "error": "pip install twilio"}
        try:
            sid   = cfg.get("TWILIO_SID", "")
            token = cfg.get("TWILIO_TOKEN", "")
            frm   = cfg.get("TWILIO_FROM", "")
            to    = cfg.get("TWILIO_TO", "")
            if not all([sid, token, frm, to]):
                return {"ok": False, "error": "Twilio SID/Token/From/To all required"}
            client = TwilioClient(sid, token)
            msg = client.messages.create(body=wa_text[:1500], from_=frm, to=to)
            print(f"  [NOTIFY] WhatsApp (Twilio) sent → {to} SID={msg.sid}", flush=True)
            return {"ok": True, "method": "twilio", "to": to, "sid": msg.sid}
        except Exception as e:
            return {"ok": False, "error": f"Twilio error: {e}"}

    return {"ok": False, "error": f"Unknown method: {method}"}


def send_signal(signal: Dict) -> Dict:
    """Send via all configured channels."""
    results = {}
    cfg = _load_cfg()

    if cfg.get("EMAIL_ENABLED") and cfg.get("EMAIL_TO"):
        results["email"] = send_email(signal)

    if cfg.get("WHATSAPP_METHOD", "none") != "none":
        results["whatsapp"] = send_whatsapp(signal)

    if not results:
        return {"ok": False, "error": "No notification channels configured. Set up email or WhatsApp in Settings."}

    any_ok = any(v.get("ok") for v in results.values())
    return {"ok": any_ok, "channels": results}


def test_notification() -> Dict:
    """Send a test notification to verify configuration."""
    test_signal = {
        "symbol": "TEST", "name": "Test Signal", "sector": "IT",
        "analysis_date": datetime.now().strftime("%Y-%m-%d"),
        "inv_type": "swing", "action": "BUY",
        "entry": 1000.00, "stop_loss": 980.00, "target1": 1035.00, "target2": 1070.00,
        "rr_ratio": 1.75, "confidence": 72, "regime": "WEAK_BULL",
        "hold_days": 5, "buy_date": datetime.now().strftime("%Y-%m-%d"),
        "sell_date": "", "buy_time": "09:15 IST",
        "reasons": [
            "✅ This is a TEST notification from GANN-ASTRO v3.7",
            "Price above SMA20 + SMA50 (momentum healthy)",
            "Gann Sq9 support confirmed at ₹980",
            "If you received this — your notifications are working!",
        ],
        "engine_scores": {"technical": "18/25", "gann": "14/20", "simons": "12/20",
                          "natal": "10/20", "fundamental": "15/25", "sentiment": "+0.15"},
        "bulk_signal": "NEUTRAL", "news_sentiment": "+0.12",
        "wyckoff_phase": "PHASE_B_LATE", "macro_note": "Test only — ignore",
        "ruling_planet": "Mercury",
    }
    return send_signal(test_signal)


# ─────────────────────────────────────────────────────────────────────────────
# BATCH SIGNAL REPORT — Excel + single combined email for all-symbols scan
# ─────────────────────────────────────────────────────────────────────────────

def _build_excel_report(signals: list, scan_date: str, inv_type: str) -> bytes:
    """
    Build an Excel workbook with one row per signal.
    Columns mirror the backtest Excel report style.
    Returns raw bytes of the .xlsx file.
    """
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return b""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Forward Signals"

    # ── Colour palette ──────────────────────────────────────────────────────
    C_HDR   = "FF0D1B2A"   # dark navy
    C_GOLD  = "FFFFCC00"
    C_GREEN = "FF00FF88"
    C_RED   = "FFFF3355"
    C_CYAN  = "FF00D4FF"
    C_PURP  = "FFBB88FF"
    C_DIM   = "FF607080"
    C_BG1   = "FF0A1520"
    C_BG2   = "FF0D1B2A"
    C_WHITE = "FFE8F4FF"

    def hdr_font(col=C_GOLD):  return Font(bold=True, color=col, name="Courier New", size=9)
    def val_font(col=C_WHITE): return Font(color=col, name="Courier New", size=9)
    def fill(col):             return PatternFill("solid", fgColor=col)
    def center():              return Alignment(horizontal="center", vertical="center", wrap_text=False)
    def left():                return Alignment(horizontal="left",   vertical="center")
    thin = Side(style="thin", color="FF1A2A3A")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Title row ───────────────────────────────────────────────────────────
    ws.merge_cells("A1:R1")
    title_cell = ws["A1"]
    title_cell.value = f"GANN-ASTRO  ·  FORWARD SIGNAL REPORT  ·  {scan_date}  ·  {inv_type.upper().replace('_',' ')}"
    title_cell.font      = Font(bold=True, color=C_GOLD, name="Courier New", size=12)
    title_cell.fill      = fill(C_HDR)
    title_cell.alignment = center()
    ws.row_dimensions[1].height = 22

    # ── Sub-header ──────────────────────────────────────────────────────────
    ws.merge_cells("A2:R2")
    sub = ws["A2"]
    sub.value     = f"Generated: {scan_date}  |  Signals: {len(signals)}  |  Min Confidence shown: all passing"
    sub.font      = val_font(C_DIM)
    sub.fill      = fill(C_BG1)
    sub.alignment = center()
    ws.row_dimensions[2].height = 14

    # ── Column headers ──────────────────────────────────────────────────────
    HEADERS = [
        ("Symbol",      9),  ("Name",       22), ("Sector",     14),
        ("Date",        10), ("Type",        8), ("Entry ₹",    10),
        ("SL ₹",        10), ("T1 ₹",       10), ("T2 ₹",       10),
        ("R:R",          6), ("Conf%",        7), ("Hold(d)",     8),
        ("Regime",      14), ("Wyckoff",     12), ("News",        9),
        ("Bulk",         8), ("Planet",       9), ("Reasons",    50),
    ]
    COL_KEYS = [
        "symbol","name","sector","analysis_date","inv_type",
        "entry","stop_loss","target1","target2",
        "rr_ratio","confidence","hold_days",
        "regime","wyckoff_phase","news_sentiment","bulk_signal","ruling_planet","reasons",
    ]

    for ci, (hdr, w) in enumerate(HEADERS, 1):
        cell = ws.cell(row=3, column=ci, value=hdr)
        cell.font      = hdr_font()
        cell.fill      = fill(C_HDR)
        cell.alignment = center()
        cell.border    = border
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 16

    # ── Data rows ───────────────────────────────────────────────────────────
    for ri, sig in enumerate(signals, 4):
        bg = C_BG2 if ri % 2 == 0 else C_BG1
        conf = sig.get("confidence", 0)
        rr   = sig.get("rr_ratio", 0)
        pnl_col = C_GREEN if conf >= 70 else C_GOLD if conf >= 55 else C_WHITE

        reasons = sig.get("reasons", [])
        if isinstance(reasons, list):
            reasons_str = " | ".join(str(r)[:80] for r in reasons[:5])
        else:
            reasons_str = str(reasons)[:200]

        row_vals = {
            "symbol":        sig.get("symbol",""),
            "name":          sig.get("name",""),
            "sector":        sig.get("sector",""),
            "analysis_date": sig.get("analysis_date",""),
            "inv_type":      (sig.get("inv_type","")).replace("_"," ").upper(),
            "entry":         round(float(sig.get("entry") or 0), 2),
            "stop_loss":     round(float(sig.get("stop_loss") or 0), 2),
            "target1":       round(float(sig.get("target1") or 0), 2),
            "target2":       round(float(sig.get("target2") or 0), 2),
            "rr_ratio":      round(float(sig.get("rr_ratio") or 0), 2),
            "confidence":    conf,
            "hold_days":     sig.get("hold_days", 0),
            "regime":        sig.get("regime",""),
            "wyckoff_phase": sig.get("wyckoff_phase","N/A"),
            "news_sentiment":sig.get("news_sentiment","N/A"),
            "bulk_signal":   sig.get("bulk_signal","NEUTRAL"),
            "ruling_planet": sig.get("ruling_planet",""),
            "reasons":       reasons_str,
        }

        for ci, key in enumerate(COL_KEYS, 1):
            val  = row_vals.get(key, "")
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill   = fill(bg)
            cell.border = border

            # Per-column colour logic
            if key == "symbol":
                cell.font = Font(bold=True, color=C_GOLD, name="Courier New", size=9)
                cell.alignment = center()
            elif key == "entry":
                cell.font = val_font(C_CYAN)
                cell.alignment = center()
                cell.number_format = "#,##0.00"
            elif key in ("stop_loss",):
                cell.font = val_font(C_RED)
                cell.alignment = center()
                cell.number_format = "#,##0.00"
            elif key in ("target1","target2"):
                cell.font = val_font(C_GREEN)
                cell.alignment = center()
                cell.number_format = "#,##0.00"
            elif key == "rr_ratio":
                col = C_GREEN if rr >= 2 else C_GOLD if rr >= 1.5 else C_WHITE
                cell.font = val_font(col)
                cell.alignment = center()
                cell.number_format = "0.00"
            elif key == "confidence":
                cell.font = val_font(pnl_col)
                cell.alignment = center()
                cell.number_format = "0"
            elif key == "regime":
                col = C_GREEN if "BULL" in str(val) else C_RED if "BEAR" in str(val) else C_WHITE
                cell.font = val_font(col)
                cell.alignment = left()
            elif key == "bulk_signal":
                col = C_GREEN if val=="BUY" else C_RED if val=="SELL" else C_DIM
                cell.font = val_font(col)
                cell.alignment = center()
            elif key == "reasons":
                cell.font = val_font(C_DIM)
                cell.alignment = left()
            else:
                cell.font = val_font()
                cell.alignment = left()

        ws.row_dimensions[ri].height = 14

    # ── Freeze top 3 rows ────────────────────────────────────────────────────
    ws.freeze_panes = "A4"

    # ── Summary sheet ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "GANN-ASTRO Forward Signal Summary"
    ws2["A1"].font = Font(bold=True, color=C_GOLD, name="Courier New", size=11)
    ws2["A1"].fill = fill(C_HDR)
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 20

    summary_rows = [
        ("Scan Date",         scan_date),
        ("Investment Type",   inv_type.replace("_"," ").upper()),
        ("Total Signals",     len(signals)),
        ("Avg Confidence",    f"{round(sum(s.get('confidence',0) for s in signals)/max(len(signals),1), 1)}%"),
        ("High Conf (≥70%)",  sum(1 for s in signals if s.get('confidence',0) >= 70)),
        ("Med Conf (60-70%)", sum(1 for s in signals if 60 <= s.get('confidence',0) < 70)),
        ("Avg R:R",           f"{round(sum(s.get('rr_ratio',0) for s in signals)/max(len(signals),1), 2)}:1"),
    ]
    for ri2, (lbl, val) in enumerate(summary_rows, 3):
        ws2.cell(ri2, 1, lbl).font = hdr_font()
        ws2.cell(ri2, 1).fill = fill(C_BG2)
        ws2.cell(ri2, 2, val).font = val_font(C_CYAN)
        ws2.cell(ri2, 2).fill = fill(C_BG1)
        ws2.row_dimensions[ri2].height = 16

    # ── Serialize to bytes ────────────────────────────────────────────────────
    import io
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def send_signal_batch(signals: list, scan_date: str, inv_type: str) -> Dict:
    """
    Send ONE combined email with an Excel report attached covering all signals.
    Called after the all-symbols forward scan completes.
    """
    cfg = _load_cfg()
    results = {}

    # ── Build Excel attachment ─────────────────────────────────────────────
    xlsx_bytes = _build_excel_report(signals, scan_date, inv_type)

    # ── Email ─────────────────────────────────────────────────────────────
    if cfg.get("EMAIL_ENABLED") and cfg.get("EMAIL_TO"):
        try:
            from email.mime.base import MIMEBase
            from email import encoders as _enc

            subject = (f"GANN-ASTRO 📡 Forward Signals: {len(signals)} opportunities "
                       f"| {inv_type.upper().replace('_',' ')} | {scan_date}")

            # Build summary HTML
            rows_html = ""
            for s in sorted(signals, key=lambda x: x.get("confidence",0), reverse=True):
                conf = s.get("confidence", 0)
                rr   = s.get("rr_ratio", 0)
                c_col = "#00FF88" if conf >= 70 else "#FFCC00"
                rr_col = "#00FF88" if rr >= 2 else "#FFCC00"
                rows_html += f"""
                <tr style="border-bottom:1px solid #1a2a3a;">
                  <td style="padding:7px 10px;color:#FFCC00;font-weight:bold;">{s.get('symbol','')}</td>
                  <td style="padding:7px 10px;color:#b0c8d8;">{s.get('name','')}</td>
                  <td style="padding:7px 10px;color:#00D4FF;">₹{s.get('entry',0):,.2f}</td>
                  <td style="padding:7px 10px;color:#FF3355;">₹{s.get('stop_loss',0):,.2f}</td>
                  <td style="padding:7px 10px;color:#00FF88;">₹{s.get('target1',0):,.2f}</td>
                  <td style="padding:7px 10px;color:#00FF88;">₹{s.get('target2',0):,.2f}</td>
                  <td style="padding:7px 10px;color:{rr_col};">{rr:.2f}:1</td>
                  <td style="padding:7px 10px;color:{c_col};">{conf}%</td>
                  <td style="padding:7px 10px;color:#888;">{s.get('regime','')}</td>
                </tr>"""

            html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"></head>
<body style="margin:0;padding:20px;background:#060f16;font-family:'Courier New',monospace;">
  <div style="max-width:900px;margin:0 auto;">
    <div style="background:#0D1B2A;border:1px solid #FFCC00;padding:20px 24px;margin-bottom:16px;">
      <div style="font-size:22px;color:#FFCC00;font-weight:bold;letter-spacing:3px;">
        📡 GANN-ASTRO FORWARD SIGNALS
      </div>
      <div style="font-size:11px;color:#607080;margin-top:6px;letter-spacing:2px;">
        SCAN DATE: {scan_date} &nbsp;·&nbsp; TYPE: {inv_type.upper().replace('_',' ')} &nbsp;·&nbsp; SIGNALS: {len(signals)}
      </div>
    </div>

    <table style="width:100%;border-collapse:collapse;background:#0A1520;font-size:12px;">
      <thead>
        <tr style="background:#0D1B2A;">
          <th style="padding:8px 10px;color:#FFCC00;text-align:left;">SYMBOL</th>
          <th style="padding:8px 10px;color:#FFCC00;text-align:left;">NAME</th>
          <th style="padding:8px 10px;color:#FFCC00;">ENTRY</th>
          <th style="padding:8px 10px;color:#FFCC00;">SL</th>
          <th style="padding:8px 10px;color:#FFCC00;">T1</th>
          <th style="padding:8px 10px;color:#FFCC00;">T2</th>
          <th style="padding:8px 10px;color:#FFCC00;">R:R</th>
          <th style="padding:8px 10px;color:#FFCC00;">CONF</th>
          <th style="padding:8px 10px;color:#FFCC00;">REGIME</th>
        </tr>
      </thead>
      <tbody>{rows_html}</tbody>
    </table>

    <div style="margin-top:14px;padding:12px 16px;background:#0D1B2A;border:1px solid #1a3a5a;
         font-size:11px;color:#607080;letter-spacing:1px;">
      Full analysis with Gann, Wyckoff, Natal and Sentiment scores attached as Excel report.
      Forward Testing Report auto-updates daily in the GANN-ASTRO dashboard.
    </div>
  </div>
</body></html>"""

            msg = MIMEMultipart("mixed")
            msg["Subject"] = subject
            msg["From"]    = cfg["EMAIL_FROM"]
            msg["To"]      = cfg["EMAIL_TO"]

            msg.attach(MIMEText(html, "html", "utf-8"))

            if xlsx_bytes:
                fname = f"ForwardSignals_{scan_date.replace('-','')}.xlsx"
                part  = MIMEBase("application", "octet-stream")
                part.set_payload(xlsx_bytes)
                _enc.encode_base64(part)
                part.add_header("Content-Disposition", f'attachment; filename="{fname}"')
                msg.attach(part)

            with smtplib.SMTP_SSL("smtp.gmail.com", 465, timeout=15) as server:
                server.login(cfg["EMAIL_FROM"], cfg["EMAIL_PASS"])
                server.sendmail(cfg["EMAIL_FROM"], cfg["EMAIL_TO"], msg.as_string())

            print(f"  [NOTIFY] Batch email sent → {cfg['EMAIL_TO']} ({len(signals)} signals + Excel)", flush=True)
            results["email"] = {"ok": True, "signals": len(signals), "to": cfg["EMAIL_TO"]}

        except Exception as _ee:
            results["email"] = {"ok": False, "error": str(_ee)}

    # ── WhatsApp: send short summary text only ────────────────────────────
    if cfg.get("WHATSAPP_METHOD", "none") != "none":
        top3 = sorted(signals, key=lambda x: x.get("confidence",0), reverse=True)[:3]
        wa_text = f"📡 GANN-ASTRO {scan_date} — {len(signals)} SIGNALS ({inv_type.upper()})\n\n"
        for s in top3:
            wa_text += (f"▸ {s['symbol']}: Entry ₹{s.get('entry',0):.0f} "
                        f"SL ₹{s.get('stop_loss',0):.0f} T1 ₹{s.get('target1',0):.0f} "
                        f"R:R {s.get('rr_ratio',0):.1f}:1 Conf {s.get('confidence',0)}%\n")
        wa_text += f"\n[Full Excel report sent via email]"
        _dummy = {"symbol":"SCAN", "entry":0, "stop_loss":0, "target1":0, "target2":0,
                  "rr_ratio":0, "confidence":0, "regime":"", "action":"BUY",
                  "hold_days":0, "buy_date":"", "sell_date":"", "reasons":[], "engine_scores":{}}
        # Temporarily override format for WhatsApp
        try:
            _, _, _ = _format_signal_text(_dummy)  # noqa — just to check method available
            if cfg.get("WHATSAPP_METHOD") == "callmebot":
                phone = cfg.get("CALLMEBOT_PHONE","").strip()
                key   = cfg.get("CALLMEBOT_KEY","").strip()
                if phone and key:
                    import urllib.parse as _up, urllib.request as _ur
                    encoded = _up.quote(wa_text[:1500])
                    url = f"https://api.callmebot.com/whatsapp.php?phone={phone}&text={encoded}&apikey={key}"
                    with _ur.urlopen(_ur.Request(url, headers={"User-Agent":"GANN-ASTRO/3.8"}), timeout=15) as r:
                        results["whatsapp"] = {"ok": True, "method": "callmebot"}
        except Exception as _we:
            results["whatsapp"] = {"ok": False, "error": str(_we)}

    if not results:
        return {"ok": False, "error": "No notification channels configured"}

    any_ok = any(v.get("ok") for v in results.values())
    return {"ok": any_ok, "channels": results, "signals_sent": len(signals)}
