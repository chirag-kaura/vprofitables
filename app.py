"""
app.py — Vprofitables + Simons Quant Trading System v3.9
Screens: Dashboard | Scanner | Chart+Analysis | Simons Lab | Natal | Tools | Scheduler
"""
import json, math, http.server, urllib.parse, threading, webbrowser, sys, os, time, socketserver
import sqlite3                          # used throughout route() — imported once here
from datetime import date, timedelta, datetime
from typing import Dict, Any

# ── Frozen/Bundled executable detection (PyInstaller) ────────────────────────
# When packaged as a Windows .exe, __file__ is unreliable.
# sys.executable points to the .exe; sys._MEIPASS is the temp extraction dir.
_IS_FROZEN = getattr(sys, 'frozen', False)
if _IS_FROZEN:
    # Running as compiled .exe — all assets are in same folder as the executable
    BASE_DIR = os.path.dirname(sys.executable)
    # Also add the PyInstaller temp extraction dir for internal imports
    _MEIPASS = getattr(sys, '_MEIPASS', BASE_DIR)
    sys.path.insert(0, _MEIPASS)
else:
    # Running as normal Python script
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# ── Load environment variables from .env file ──
def _load_env():
    env_path = os.path.join(BASE_DIR, ".env")
    if os.path.exists(env_path):
        with open(env_path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                if "=" in line:
                    k, v = line.split("=", 1)
                    k = k.strip()
                    v = v.strip().strip('"').strip("'")
                    if k:
                        os.environ[k] = v
_load_env()

# numpy is optional but used in many serialisers — import once at top level
try:
    import numpy as _np_top
except ImportError:
    _np_top = None

if sys.platform == "win32":
    import io
    if sys.stdout is not None:
        if hasattr(sys.stdout, "reconfigure"):
            sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        elif hasattr(sys.stdout, "buffer"):
            sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    if sys.stderr is not None:
        if hasattr(sys.stderr, "reconfigure"):
            sys.stderr.reconfigure(encoding="utf-8", errors="replace")
        elif hasattr(sys.stderr, "buffer"):
            sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

sys.path.insert(0, BASE_DIR)

print("  [BOOT] Loading core modules...", flush=True)
try:
    print("  [BOOT]   core/ephemeris.py ...", end=" ", flush=True)
    from core.ephemeris  import get_all_planets
    print("OK", flush=True)
except Exception as _e: print(f"FAILED: {_e}", flush=True); raise

try:
    print("  [BOOT]   core/aspects.py ...", end=" ", flush=True)
    from core.aspects    import detect_aspects, detect_stations, detect_retrogrades
    print("OK", flush=True)
except Exception as _e: print(f"FAILED: {_e}", flush=True); raise

try:
    print("  [BOOT]   core/gann_math.py ...", end=" ", flush=True)
    from core.gann_math  import sq9_levels, gann_angles, time_cycles_from_pivot, confluence_score, planetary_price_map, sq9_from_atl, sq9_bounce_confirmed
    print("OK", flush=True)
except Exception as _e: print(f"FAILED: {_e}", flush=True); raise

try:
    print("  [BOOT]   core/scheduler.py ...", end=" ", flush=True)
    from core.scheduler  import (get_cached_prices, get_pivots_for_symbol, save_user_pivot,
                                  seed_static_pivots, detect_auto_pivots,
                                  get_scheduler, get_cached_quant, get_scheduler_log,
                                  init_db, DB_PATH, cache_quant)
    print("OK", flush=True)
except Exception as _e: print(f"FAILED: {_e}", flush=True); raise

try:
    print("  [BOOT]   core/signal_engine.py ...", end=" ", flush=True)
    from core.signal_engine import analyze_instrument, daily_astro_prefilter, get_planet_dashboard
    print("OK", flush=True)
except Exception as _e: print(f"FAILED: {_e}", flush=True); raise

try:
    print("  [BOOT]   core/quant_engine.py ...", end=" ", flush=True)
    from core.quant_engine  import full_quant_analysis, find_support_resistance
    print("OK", flush=True)
except Exception as _e: print(f"FAILED: {_e}", flush=True); raise

# ── v3.9 Unified Logic + Deep Signal Engine ──────────────────────────────────
try:
    from core.unified_logic import (
        INVESTMENT_TYPES, compute_score, passes_gate,
        compute_levels, compute_exit_plan, build_reasons, compute_acc_score,
        compute_conjunction_score, compute_levels_short,
        BAD_BUY_GOOD_SHORT, BEST_BUY_SYMBOLS, PLANET_TRADE_DIRECTION,
    )
    from core.deep_signal_engine import predict_reversal, get_model_status, train_models
    print("  [BOOT] v3.9 engines (unified_logic + deep_signal_engine) ... OK", flush=True)
except Exception as _e:
    print(f"  [BOOT] v3.9 engines WARN: {_e}", flush=True)
    # Stubs so app doesn't crash
    INVESTMENT_TYPES = {}
    def compute_score(*a, **k): return {"total": 50.0}
    def passes_gate(*a, **k): return True, "stub"
    def compute_levels(*a, **k): return {}
    def compute_exit_plan(*a, **k): return {"sell_date":"","hold_days":5,"sell_condition":"","trail_rule":"","exit_source":"stub"}
    def build_reasons(*a, **k): return [], []
    def predict_reversal(*a, **k): return {"direction":"NEUTRAL","confidence":0.5,"reversal_price":0,"reversal_date":"","days_to_reversal":5,"model_trained":False,"direction_prob":0.5,"expected_move_pct":0,"signal_alignment":0.5,"reversal_prob":0.5,"model_version":"stub","features_used":0}
    def get_model_status(): return {"trained": False}
    def train_models(**k): return {"error": "engine not loaded"}
    def compute_acc_score(*a, **k): return 2  # stub: allow backtest to run

# Wyckoff engine removed in v3.9 — replaced by unified Gann+Technical stack
# Stubs kept for any legacy references that may remain in DB/cache
wyckoff_phase    = lambda *a, **k: {"phase":"N/A","entry_signal":False,"confidence":0,"reason":"removed","regime":{"regime":"N/A","tradeable":True,"reason":"N/A","trend_strength":0}}
hedge_fund_levels= lambda *a, **k: {"entry":a[0][-1] if a and a[0] else 100,"entry_src":"Gann Sq9","sl":a[0][-1]*0.96 if a and a[0] else 96,"sl_src":"4% SL","t1":a[0][-1]*1.06 if a and a[0] else 106,"t1_src":"6% T1","t2":a[0][-1]*1.12 if a and a[0] else 112,"t2_src":"12% T2"}
mfe_mae          = lambda entry,exit_p,highs,lows: {"mfe_pct":round((max(highs)-entry)/entry*100,2) if highs else 0,"mae_pct":round((min(lows)-entry)/entry*100,2) if lows else 0,"mfe_abs":round(max(highs)-entry,2) if highs else 0,"mae_abs":round(min(lows)-entry,2) if lows else 0,"captured_pct":0}

try:
    print("  [BOOT]   core/notifier.py ...", end=" ", flush=True)
    from core.notifier import save_cfg as _notifier_save_cfg, _load_cfg as _notifier_load_cfg, \
                             test_notification as _notifier_test, send_signal as _notifier_send_signal, \
                             send_signal_batch as _notifier_send_signal_batch
    print("OK", flush=True)
except Exception as _e:
    print(f"WARN (non-fatal): {_e}", flush=True)
    # Notifier is optional — define no-op stubs so route() always has something to call
    def _notifier_save_cfg(cfg): return False
    def _notifier_load_cfg(): return {}
    def _notifier_test(): return {"ok": False, "error": "notifier not available"}
    def _notifier_send_signal(payload): return {"ok": False, "error": "notifier not available"}
    def _notifier_send_signal_batch(signals, date, inv_type): return {"ok": False, "error": "notifier not available"}

try:
    print("  [BOOT]   core/fundamental_engine.py ...", end=" ", flush=True)
    import core.fundamental_engine as _fundamental_engine
    print("OK", flush=True)
except Exception as _e:
    print(f"WARN (non-fatal): {_e}", flush=True)
    _fundamental_engine = None

try:
    print("  [BOOT]   core/sentiment_db.py ...", end=" ", flush=True)
    import core.sentiment_db as _sentiment_db
    print("OK", flush=True)
except Exception as _e:
    print(f"WARN (non-fatal): {_e}", flush=True)
    _sentiment_db = None

try:
    print("  [BOOT]   core/report_engine.py ...", end=" ", flush=True)
    import core.report_engine as _report_engine
    print("OK", flush=True)
except Exception as _e:
    print(f"WARN (non-fatal): {_e}", flush=True)
    _report_engine = None

try:
    print("  [BOOT]   core/fetch_institutional.py ...", end=" ", flush=True)
    import core.fetch_institutional as _fetch_institutional
    print("OK", flush=True)
except Exception as _e:
    print(f"WARN (non-fatal): {_e}", flush=True)
    _fetch_institutional = None

try:
    print("  [BOOT]   core/market_feedback.py ...", end=" ", flush=True)
    import core.market_feedback as _market_feedback
    print("OK", flush=True)
except Exception as _e:
    print(f"WARN (non-fatal): {_e}", flush=True)
    _market_feedback = None

try:
    print("  [BOOT]   core/market_brain_local.py ...", end=" ", flush=True)
    import core.market_brain_local as _market_brain_local
    print("OK", flush=True)
except Exception as _e:
    print(f"WARN (non-fatal): {_e}", flush=True)
    _market_brain_local = None

try:
    print("  [BOOT]   data/instruments.py ...", end=" ", flush=True)
    from data.instruments   import ALL_INSTRUMENTS, list_all_symbols, get_natal, get_instrument, get_transit_to_natal_aspects
    print("OK", flush=True)
except Exception as _e: print(f"FAILED: {_e}", flush=True); raise

print("  [BOOT] Core modules OK", flush=True)

PORT = int(os.environ.get("GANN_PORT", 8080))
print(f"  [BOOT] Port: {PORT}", flush=True)
print("  [BOOT] Initialising database...", end=" ", flush=True)
try:
    init_db()
    from core.personalization_db import init_personalization_db
    init_personalization_db()
    print("OK", flush=True)
except Exception as _idbe:
    print(f"FAILED: {_idbe}", flush=True)
    raise
# Initialise sentiment tables (news_sentiment + sentiment_labels)
try:
    from core.sentiment_db import init_sentiment_tables
    init_sentiment_tables()
    print("  [DB   ] Sentiment tables verified", flush=True)
except Exception as _se:
    print(f"  [WARN ] Sentiment DB init skipped: {_se}", flush=True)
# Initialise institutional data tables
try:
    if _fetch_institutional is not None:
        _fetch_institutional.init_institutional_tables()
        print("  [DB   ] Institutional tables verified", flush=True)
except Exception as _ie:
    print(f"  [WARN ] Institutional DB init skipped: {_ie}", flush=True)
# Initialise market feedback columns + run labelling in background
try:
    if _market_feedback is not None:
        _market_feedback.init_market_feedback_columns()
        print("  [DB   ] Market feedback columns verified", flush=True)
        # Apply market labels in background (non-blocking) on startup
        def _bg_label():
            try:
                n = _market_feedback.apply_market_labels(min_age_days=6, batch_size=200)
                if n: print(f"  [MKTF ] Auto-labelled {n} headlines on startup", flush=True)
            except Exception as _e:
                print(f"  [WARN ] Market label bg task: {_e}", flush=True)
        threading.Thread(target=_bg_label, daemon=True).start()
except Exception as _mfe:
    print(f"  [WARN ] Market feedback init skipped: {_mfe}", flush=True)
# Clear stale quant cache entries that lack open prices (one-time migration)
# (Runs after _db() is defined below — wrapped in try to be safe)
def _clear_stale_cache() -> None:
    try:
        _conn_init = _db()
        _n = _conn_init.execute('DELETE FROM quant_cache WHERE data NOT LIKE ?', ('%opens%',)).rowcount
        _conn_init.commit()
        _conn_init.close()
        if _n: print(f"[INIT] Cleared {_n} stale quant cache entries (missing opens)", flush=True)
    except Exception:
        pass

# ── ML Auto-Train on every startup (background, non-blocking) ──────────────
# Trains RandomForest + GradientBoost on your price history.
# Runs in background — server is ready immediately, training completes in ~60-90s.
# Model is only retrained if: no model exists OR price data is newer than last training.
def _auto_train_ml():
    try:
        print("  [ML  ] Auto-training Deep Signal Engine in background...", flush=True)
        # Check if we have enough data and if retraining is needed
        import sqlite3 as _sq_ml, os as _os_ml, pickle as _pk_ml
        _db_ml = _os_ml.path.join(_os_ml.path.dirname(_os_ml.path.abspath(__file__)), "market_data_v2.db")
        _model_dir = _os_ml.path.join(_os_ml.path.dirname(_os_ml.path.abspath(__file__)), "core", "models")
        _meta_path = _os_ml.path.join(_model_dir, "model_meta.pkl")
        # Check if model is fresh (trained within last 7 days) — skip if so
        _needs_train = True
        if _os_ml.path.exists(_meta_path):
            try:
                with open(_meta_path, "rb") as _mf: _meta = _pk_ml.load(_mf)
                from datetime import datetime as _dtml, timedelta as _tdd
                _trained = _dtml.fromisoformat(_meta.get("trained_at", "2000-01-01"))
                if (datetime.now() - _trained).days < 7:
                    print(f"  [ML  ] Model is recent ({_trained.date()}) — skipping retraining", flush=True)
                    _needs_train = False
            except Exception: pass
        if _needs_train:
            if _os_ml.path.exists(_db_ml):
                _conn_ml = sqlite3.connect(_db_ml, timeout=5)
                _bar_count = _conn_ml.execute("SELECT COUNT(*) FROM daily_prices WHERE close IS NOT NULL").fetchone()[0]
                _conn_ml.close()
                if _bar_count >= 500:
                    meta = train_models(lookback_years=3, forward_days=10, verbose=True)
                    if "error" in meta:
                        print(f"  [ML  ] Training skipped: {meta['error']}", flush=True)
                    else:
                        print(f"  [ML  ] Training complete — accuracy: {meta.get('dir_accuracy',0):.1%}  MAE: {meta.get('timing_mae',0):.1f}d", flush=True)
                else:
                    print(f"  [ML  ] Only {_bar_count} price bars — need 500+ for training (run download_history.py)", flush=True)
    except Exception as _ml_err:
        print(f"  [ML  ] Auto-train error (non-fatal): {_ml_err}", flush=True)

threading.Thread(target=_auto_train_ml, daemon=True, name="ML_AUTO_TRAIN").start()

def jdump(data):
    def s(o):
        if isinstance(o, (date, datetime)): return o.isoformat()
        if _np_top is not None:
            if isinstance(o, _np_top.integer):  return int(o)
            if isinstance(o, _np_top.floating):  return None if _np_top.isnan(o) else float(o)
            if isinstance(o, _np_top.bool_):     return bool(o)
            if isinstance(o, _np_top.ndarray):   return o.tolist()
        if hasattr(o, 'item'):   return o.item()
        if hasattr(o, 'tolist'): return o.tolist()
        raise TypeError(type(o))
    return json.dumps(data, default=s, indent=2).encode()


def _db(timeout: int = 5) -> sqlite3.Connection:
    """
    Open a SQLite connection to market_data_v2.db with the standard performance
    PRAGMAs.  Use this throughout route() instead of repeating
    ``import sqlite3 as _sq3; _db()`` inline.
    Caller is responsible for closing the connection.
    """
    conn = sqlite3.connect(DB_PATH, timeout=timeout)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")
    conn.execute("PRAGMA cache_size=-8000")   # 8 MB per connection
    conn.execute("PRAGMA temp_store=MEMORY")
    conn.execute("PRAGMA busy_timeout=30000")
    return conn

def _get_user_portfolio_id(conn, user_id):
    row = conn.execute("SELECT id FROM portfolios WHERE user_id=? AND name='Primary'", (user_id,)).fetchone()
    if row:
        return row[0]
    import uuid
    from datetime import datetime
    p_id = str(uuid.uuid4())
    conn.execute("INSERT INTO portfolios (id, user_id, name, created_at) VALUES (?, ?, 'Primary', ?)", (p_id, user_id, datetime.utcnow().isoformat()))
    conn.commit()
    return p_id

def _get_overview_data() -> dict:
    """Fetches dynamic data from market_data_v2.db for the Market Overview page."""
    import json
    
    # ── Self-contained technical indicator calculators ──
    from core.indicators import calculate_rsi, calculate_ema

    def calculate_macd(closes: list) -> float:
        if len(closes) < 34:
            return 0.0
        ema12 = [closes[0]]
        ema26 = [closes[0]]
        mult12 = 2.0 / 13.0
        mult26 = 2.0 / 27.0
        for p in closes[1:]:
            ema12.append((p - ema12[-1]) * mult12 + ema12[-1])
            ema26.append((p - ema26[-1]) * mult26 + ema26[-1])
        macd_line = [e12 - e26 for e12, e26 in zip(ema12, ema26)]
        signal = macd_line[0]
        mult9 = 2.0 / 10.0
        for m in macd_line[1:]:
            signal = (m - signal) * mult9 + signal
        return macd_line[-1] - signal

    conn = _db()
    res = {
        "indices": [],
        "most_bought": [],
        "gainers": [],
        "losers": [],
        "sectors": {},
        "performers": {"1W": [], "1M": [], "1Y": [], "5Y": []},
        "screeners": {"rsi": [], "macd": [], "sma": [], "ema": [], "pivot": []},
        "signals": {"candles": [], "patterns": [], "priceaction": []},
        "research": [],
        "pocket": [],
        "commodities": [],
        "news": []
    }
    
    try:
        dt_row = conn.execute("SELECT MAX(trade_date) FROM daily_prices").fetchone()
        if not dt_row or not dt_row[0]: return res
        max_dt = dt_row[0]
        
        # 1. Gainers
        rows = conn.execute("SELECT p.symbol, p.close, p.volume, p.change_pct FROM daily_prices p WHERE p.trade_date = ? AND p.change_pct IS NOT NULL ORDER BY p.change_pct DESC LIMIT 10", (max_dt,)).fetchall()
        res["gainers"] = [{"symbol": r[0], "name": "Equity", "ltp": round(r[1], 2), "chg": round((r[1]*r[3])/(100+r[3]), 2) if r[3] else 0, "chgPct": round(r[3], 2)} for r in rows]
        
        # 2. Losers
        rows = conn.execute("SELECT p.symbol, p.close, p.volume, p.change_pct FROM daily_prices p WHERE p.trade_date = ? AND p.change_pct IS NOT NULL ORDER BY p.change_pct ASC LIMIT 10", (max_dt,)).fetchall()
        res["losers"] = [{"symbol": r[0], "name": "Equity", "ltp": round(r[1], 2), "chg": round((r[1]*r[3])/(100+r[3]), 2) if r[3] else 0, "chgPct": round(r[3], 2)} for r in rows]
        
        # 3. Most Bought
        rows = conn.execute("SELECT p.symbol, p.close, p.volume, p.change_pct FROM daily_prices p WHERE p.trade_date = ? ORDER BY p.volume DESC LIMIT 10", (max_dt,)).fetchall()
        res["most_bought"] = [{"symbol": r[0], "name": "Equity", "ltp": round(r[1], 2), "chg": round((r[1]*(r[3] or 0))/(100+(r[3] or 0)), 2) if r[3] else 0, "chgPct": round(r[3] or 0, 2)} for r in rows]
        
        # 4. Pocket Friendly
        rows = conn.execute("SELECT p.symbol, p.close, p.volume, p.change_pct FROM daily_prices p WHERE p.trade_date = ? AND p.close < 200 ORDER BY p.volume DESC LIMIT 15", (max_dt,)).fetchall()
        res["pocket"] = [{"symbol": r[0], "name": "Equity", "ltp": round(r[1], 2), "chg": round((r[1]*(r[3] or 0))/(100+(r[3] or 0)), 2) if r[3] else 0, "chgPct": round(r[3] or 0, 2)} for r in rows]
        
        # 5. Performers (1W, 1M, 1Y, 5Y)
        try:
            dates = conn.execute("SELECT DISTINCT trade_date FROM daily_prices ORDER BY trade_date DESC LIMIT 260").fetchall()
            if len(dates) > 0:
                d_1w = dates[min(5, len(dates)-1)][0]
                d_1m = dates[min(20, len(dates)-1)][0]
                d_1y = dates[min(250, len(dates)-1)][0]
                d_5y = dates[-1][0]
                
                for tf, dt_old in [("1W", d_1w), ("1M", d_1m), ("1Y", d_1y), ("5Y", d_5y)]:
                    p_rows = conn.execute("SELECT p1.symbol, p1.close, p2.close FROM daily_prices p1 JOIN daily_prices p2 ON p1.symbol = p2.symbol WHERE p1.trade_date = ? AND p2.trade_date = ?", (max_dt, dt_old)).fetchall()
                    res["performers"][tf] = [
                        {"symbol": r[0], "name": "Equity", "ltp": round(r[1], 2), "chg": round(r[1]-r[2], 2), "returnPct": round((r[1]-r[2])/r[2]*100, 2)}
                        for r in p_rows if r[2]
                    ]
                    res["performers"][tf].sort(key=lambda x: x["returnPct"], reverse=True)
                    res["performers"][tf] = res["performers"][tf][:10]
        except Exception as pe:
            print(f"Error calculating performers: {pe}", flush=True)

        # 6. Real Indices (NIFTY50, BANKNIFTY, etc.)
        index_symbols = ["NIFTY50", "BANKNIFTY", "NIFTYIT", "NIFTYPHARMA", "NIFTYAUTO"]
        indices_data = []
        for sym in index_symbols:
            idx_row = conn.execute("""
                SELECT open, high, low, close, volume, change_pct 
                FROM daily_prices 
                WHERE symbol=? AND close IS NOT NULL 
                ORDER BY trade_date DESC LIMIT 1
            """, (sym,)).fetchone()
            
            if idx_row:
                open_p, high_p, low_p, close_p, vol_p, chg_pct = idx_row
                hist_rows = conn.execute("""
                    SELECT trade_date, close 
                    FROM daily_prices 
                    WHERE symbol=? AND close IS NOT NULL 
                    ORDER BY trade_date DESC LIMIT 30
                """, (sym,)).fetchall()
                history = [{"d": r[0], "c": float(r[1])} for r in reversed(hist_rows)]
                
                prev_close = close_p
                if len(history) >= 2:
                    prev_close = history[-2]["c"]
                
                chg = close_p - prev_close
                chg_pct_calc = (chg / prev_close * 100) if prev_close else 0.0
                
                indices_data.append({
                    "symbol": sym,
                    "ltp": round(close_p, 2),
                    "chg": round(chg, 2),
                    "chgPct": round(chg_pct_calc, 2),
                    "high": round(high_p or close_p, 2),
                    "low": round(low_p or close_p, 2),
                    "open": round(open_p or close_p, 2),
                    "prev_close": round(prev_close, 2),
                    "history": history
                })
        if indices_data:
            res["indices"] = indices_data

        # 7. Commodities
        commodity_symbols = ["GOLD", "SILVER", "COPPER", "CRUDEOIL", "NATURALGAS"]
        commodity_data = []
        for sym in commodity_symbols:
            com_row = conn.execute("""
                SELECT open, high, low, close, volume, change_pct 
                FROM daily_prices 
                WHERE symbol=? AND close IS NOT NULL 
                ORDER BY trade_date DESC LIMIT 1
            """, (sym,)).fetchone()
            
            if com_row:
                open_p, high_p, low_p, close_p, vol_p, chg_pct = com_row
                hist_rows = conn.execute("""
                    SELECT trade_date, close 
                    FROM daily_prices 
                    WHERE symbol=? AND close IS NOT NULL 
                    ORDER BY trade_date DESC LIMIT 30
                """, (sym,)).fetchall()
                history = [{"d": r[0], "c": float(r[1])} for r in reversed(hist_rows)]
                
                prev_close = close_p
                if len(history) >= 2:
                    prev_close = history[-2]["c"]
                
                chg = close_p - prev_close
                chg_pct_calc = (chg / prev_close * 100) if prev_close else 0.0
                
                commodity_data.append({
                    "symbol": sym,
                    "ltp": round(close_p, 2),
                    "chg": round(chg, 2),
                    "chgPct": round(chg_pct_calc, 2),
                    "high": round(high_p or close_p, 2),
                    "low": round(low_p or close_p, 2),
                    "open": round(open_p or close_p, 2),
                    "prev_close": round(prev_close, 2),
                    "history": history
                })
        if commodity_data:
            res["commodities"] = commodity_data

        # 8. News Sentiment Feed
        news_rows = conn.execute("SELECT symbol, title, snippet, source, published_at, raw_score, label FROM news_sentiment ORDER BY published_at DESC LIMIT 15").fetchall()
        res["news"] = [{"symbol": r[0], "title": r[1], "snippet": r[2], "source": r[3], "published_at": r[4], "score": r[5], "label": r[6]} for r in news_rows]

        # 9. Screeners & Pattern Signals via active scan
        active_symbols = ["NIFTY50", "BANKNIFTY", "NIFTYIT", "NIFTYPHARMA", "NIFTYAUTO"]
        equity_symbols = [r[0] for r in conn.execute("SELECT DISTINCT symbol FROM daily_prices WHERE symbol NOT LIKE '%NIFTY%' AND symbol NOT IN ('GOLD','SILVER','COPPER','CRUDEOIL','NATURALGAS','TEST_SYM') LIMIT 25").fetchall()]
        symbols_to_scan = active_symbols + equity_symbols
        
        from core.pattern_engine import detect as detect_patterns
        
        for sym in symbols_to_scan:
            rows_p = conn.execute("""
                SELECT close, high, low, volume, change_pct 
                FROM daily_prices 
                WHERE symbol=? AND close IS NOT NULL AND trade_date <= ? 
                ORDER BY trade_date DESC LIMIT 100
            """, (sym, max_dt)).fetchall()
            
            if len(rows_p) < 35:
                continue
                
            closes = [r[0] for r in reversed(rows_p)]
            highs = [r[1] or r[0] for r in reversed(rows_p)]
            lows = [r[2] or r[0] for r in reversed(rows_p)]
            volumes = [r[3] or 0.0 for r in reversed(rows_p)]
            
            ltp = closes[-1]
            chgPct = rows_p[0][4] or 0.0
            chg = (ltp * chgPct) / (100 + chgPct) if chgPct else 0.0
            
            item = {"symbol": sym, "name": "Equity", "ltp": round(ltp, 2), "chg": round(chg, 2), "chgPct": round(chgPct, 2)}
            
            # Technical Screeners
            rsi = calculate_rsi(closes)
            if rsi > 60 or rsi < 40:
                res["screeners"]["rsi"].append(item)
            macd_val = calculate_macd(closes)
            if macd_val != 0:
                res["screeners"]["macd"].append(item)
            if len(closes) >= 20:
                sma20 = sum(closes[-20:]) / 20.0
                if ltp > sma20:
                    res["screeners"]["sma"].append(item)
                ema20 = calculate_ema(closes, 20)
                if ltp > ema20:
                    res["screeners"]["ema"].append(item)
            if len(rows_p) >= 2:
                prev_high = rows_p[1][1] or rows_p[1][0]
                prev_low = rows_p[1][2] or rows_p[1][0]
                prev_close = rows_p[1][0]
                pp = (prev_high + prev_low + prev_close) / 3.0
                if ltp > pp:
                    res["screeners"]["pivot"].append(item)
                    
            # Pattern engine signals
            try:
                pat_res = detect_patterns(closes, highs, lows, volumes)
                if pat_res.pattern != "NONE":
                    signal_type = "Bullish"
                    if "BEAR" in pat_res.pattern or pat_res.pattern == "UTAD":
                        signal_type = "Bearish"
                    res["signals"]["patterns"].append({
                        "symbol": sym, "name": "Equity", "ltp": round(ltp, 2), "chg": round(chg, 2), "chgPct": round(chgPct, 2),
                        "signal": signal_type, "pattern": pat_res.pattern, "tf": "1d"
                    })
                if pat_res.inst_absorption:
                    res["signals"]["priceaction"].append({
                        "symbol": sym, "name": "Equity", "ltp": round(ltp, 2), "chg": round(chg, 2), "chgPct": round(chgPct, 2),
                        "signal": "Bullish", "pattern": "Inst. Absorption", "tf": "1d"
                    })
                elif pat_res.volume_exhaustion:
                    res["signals"]["priceaction"].append({
                        "symbol": sym, "name": "Equity", "ltp": round(ltp, 2), "chg": round(chg, 2), "chgPct": round(chgPct, 2),
                        "signal": "Bullish", "pattern": "Volume Exhaustion", "tf": "1d"
                    })
            except Exception: pass

        # 10. Volume Anomalies
        arows = conn.execute("SELECT v.symbol, v.signal, v.candle_type, v.price_change_pct, p.close FROM volume_anomalies v LEFT JOIN daily_prices p ON v.symbol = p.symbol AND v.trade_date = p.trade_date WHERE v.trade_date = ? LIMIT 15", (max_dt,)).fetchall()
        for sym, sig, can, pct, ltp in arows:
            if pct is None: pct = 0.0
            res["signals"]["candles"].append({
                "symbol": sym, "name": "Equity", "ltp": ltp or 0.0, "chg": round((ltp or 0.0)*pct/100,2), "chgPct": round(pct, 2),
                "signal": "Bullish" if pct > 0 else "Bearish", "pattern": can or "Unknown", "tf": "1d"
            })
            
        # 11. Research Recommendations
        frows = conn.execute("SELECT symbol, entry, stop_loss, target1, updated_at, inv_type FROM forward_signals WHERE status = 'OPEN' ORDER BY created_at DESC LIMIT 10").fetchall()
        for r in frows:
            sym, entry, sl, t1, upd, typ = r
            if not entry: continue
            res["research"].append({
                "symbol": sym, "name": "Equity", "type": typ, "ltp": entry, "chgPct": 0.0,
                "rec_price": entry, "rec_price_high": entry*1.01, "sl": sl, "target": t1, "potential": round(((t1 - entry) / entry) * 100, 2) if entry else 0, "updated": upd or max_dt
            })
            
        # 12. Sectors
        hard_sectors = {"NIFTY IT": ["TCS", "INFY", "WIPRO", "HCLTECH", "TECHM"], "NIFTY BANK": ["HDFCBANK", "ICICIBANK", "SBIN", "AXISBANK", "KOTAKBANK"]}
        for sec, syms in hard_sectors.items():
            sym_list = "','".join(syms)
            srows = conn.execute(f"SELECT p.symbol, p.close, p.change_pct FROM daily_prices p WHERE p.trade_date = ? AND p.symbol IN ('{sym_list}')", (max_dt,)).fetchall()
            if srows:
                res["sectors"][sec] = [{"symbol": r[0], "name": "Equity", "ltp": round(r[1],2), "chg": round(r[1]*r[2]/100,2) if r[2] else 0.0, "chgPct": round(r[2],2) if r[2] else 0.0} for r in srows]

    except Exception as e:
        print(f"Error generating overview: {e}", flush=True)
    finally:
        conn.close()
    return res


# ══════════════════════════════════════════════════════════════════════════
# HTML / CSS / JS — complete single-page application
# ══════════════════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════════════════
# Page modules — HTML templates + JavaScript for each page
# Backend endpoint handlers remain in this file (ep == "..." blocks below)
# ══════════════════════════════════════════════════════════════════════════
import sys as _sys, os as _os
_sys.path.insert(0, _os.path.dirname(_os.path.abspath(__file__)))

print("  [LOAD] Importing page modules...", flush=True)
_pages = ["page_auth", "page_onboarding", "page_overview","page_dashboard","page_scanner","page_chart","page_simons",
          "page_gann","page_natal","page_tools","page_advisor",
          "page_sentiment","page_fundamentals","page_trading","page_research",
          "page_risk","page_watchlist","page_analytics", "page_admin"]
for _pname in _pages:
    try:
        print(f"  [LOAD]   pages/{_pname}.py ...", end=" ", flush=True)
        import importlib as _iml
        _mod = _iml.import_module(f"pages.{_pname}")
        import sys as _sysp
        _sysp.modules[_pname] = _mod
        print("OK", flush=True)
    except Exception as _pe:
        import traceback as _tb
        print(f"FAILED: {_pe}", flush=True)
        _tb.print_exc()
        raise SystemExit(f"  [FATAL] Cannot load pages/{_pname}.py — fix above error and restart.")

from pages import (
    page_auth, page_onboarding,
    page_overview, page_dashboard, page_scanner, page_chart,
    page_simons, page_gann, page_natal, page_tools,
    page_advisor, page_sentiment, page_fundamentals,
    page_trading, page_research, page_risk,
    page_watchlist, page_analytics, page_admin,
)
print("  [LOAD] All page modules loaded OK", flush=True)


_clear_stale_cache()  # safe: _db() now defined

def _build_html():
    """
    Assemble the complete single-page application from page modules.
    Called once at startup — result cached in HTML global.

    To modify a page:
      - HTML layout  → edit pages/page_<name>.py  (HTML variable)
      - JS behaviour → edit pages/page_<name>.py  (JS variable)
      - API backend  → edit app.py               (if ep == "..." block)
    """
    import os as _o
    _d = _o.path.dirname(_o.path.abspath(__file__))

    # Shared HTML/JS fragments stored as plain files (no quote-escaping issues)
    def _read(path):
        with open(_o.path.join(_d, path), encoding="utf-8") as _fh:
            return _fh.read()

    html_head  = _read("pages/_html_head.html")   # <html>...<sidebar>...<main open>
    html_tools = page_tools.HTML                   # sq9/cycles/confluence/instruments pages
    shared_js  = _read("pages/_shared.js")         # nav(), api(), initSymbols(), etc
    tools_js   = page_tools.JS                     # sq9/cycles/confluence JS

    return (
        "<!DOCTYPE html>\n" +
        html_head + "\n" +

        # ── Page HTML blocks ──
        page_overview.HTML +
        page_dashboard.HTML +
        page_scanner.HTML +
        page_chart.HTML +
        page_simons.HTML +
        page_gann.HTML +
        page_natal.HTML +
        html_tools +
        page_advisor.HTML +
        page_sentiment.HTML +
        page_fundamentals.HTML +
        page_trading.HTML +
        page_research.HTML +
        page_risk.HTML +
        page_watchlist.HTML +
        page_analytics.HTML +
        page_admin.HTML +

        # ── Close main div and close app-layout, then inject full-screen overlays ──
        "\n</div><!-- /main -->\n</div><!-- /app-layout -->\n\n" +
        page_auth.HTML +
        page_onboarding.HTML +

        # ── Open script tags ──
        "\n<script>\n" +

        # ── JavaScript (shared first, then per-page) ──
        shared_js +
        page_auth.JS +
        page_onboarding.JS +
        page_overview.JS +
        page_dashboard.JS +
        page_scanner.JS +
        page_chart.JS +
        page_simons.JS +
        page_gann.JS +
        page_natal.JS +
        tools_js +
        page_advisor.JS +
        page_sentiment.JS +
        page_fundamentals.JS +
        page_trading.JS +
        page_research.JS +
        page_risk.JS +
        page_watchlist.JS +
        page_analytics.JS +
        page_admin.JS +

        "\n</script>\n</body>\n</html>\n"
    )


# Assemble SPA once at startup
print("  [LOAD] Building HTML (assembling all pages)...", flush=True)
try:
    HTML = _build_html()
    print(f"  [LOAD] HTML built OK ({len(HTML):,} chars)", flush=True)
except Exception as _bhe:
    import traceback as _bhtb
    print(f"  [FATAL] _build_html() crashed: {_bhe}", flush=True)
    _bhtb.print_exc()
    raise



# ── Live Price Caching & Google Finance Scraper ──────────────────────────────
_live_price_cache = {}  # { symbol: {"price": float, "close": float, "change_pct": float, "ts": float, "date": str} }

def get_live_price_google(symbol: str) -> dict:
    """
    Fetch real-time stock or index price from Google Finance with a 5-second in-memory cache.
    """
    import time
    from datetime import date
    global _live_price_cache

    now = time.time()
    if symbol in _live_price_cache:
        cached = _live_price_cache[symbol]
        if now - cached["ts"] < 5.0:  # 5 seconds TTL
            return cached

    ticker = symbol
    is_index = symbol in ["NIFTY50", "BANKNIFTY", "NIFTY", "BANK"]
    if symbol == "NIFTY50" or symbol == "NIFTY":
        ticker = "NIFTY_50"
        is_index = True
    elif symbol == "BANKNIFTY" or symbol == "BANK":
        ticker = "NIFTY_BANK"
        is_index = True

    exchange = "INDEXNSE" if is_index else "NSE"
    url = f"https://www.google.com/finance/quote/{ticker}:{exchange}"
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124.0.0.0 Safari/537.36'
    }
    try:
        import requests, re
        r = requests.get(url, headers=headers, timeout=5)
        if r.status_code == 200:
            # Main price matches class="N6SYTe" then jsname="Pdsbrc" containing the price
            match = re.search(r'class="N6SYTe"[^>]*>\s*<span jsname="Pdsbrc"[^>]*>\s*<span>[^0-9]*([\d,.]+)\s*</span>', r.text)
            if not match:
                match = re.search(r'<div[^>]*class="[^"]*N6SYTe[^"]*"[^>]*>.*?<span>[^0-9]*([\d,.]+)\s*</span>', r.text, re.DOTALL)

            if match:
                price = float(match.group(1).replace(",", ""))

                # Try to get percentage change (class YMlKec and jsname vY9t3b or similar)
                pct_change = 0.0
                match_pct = re.search(r'jsname="vY9t3b"[^>]*>\s*<span[^>]*>([^%]+)%', r.text)
                if match_pct:
                    try:
                        pct_change = float(match_pct.group(1).replace(",", "").replace("+", "").replace("-", "-").strip())
                    except ValueError:
                        pass

                res = {
                    "price": price,
                    "close": price,
                    "change_pct": pct_change,
                    "ts": now,
                    "date": date.today().isoformat(),
                    "source": "GOOGLE_FINANCE"
                }
                _live_price_cache[symbol] = res
                return res
    except Exception:
        pass
    return None


class Handler(http.server.BaseHTTPRequestHandler):
    def log_message(self, fmt, *args): pass
    def log_request(self, code="-", size="-"): pass

    def send_json(self, data, status=200):
        def ser(o):
            if isinstance(o, (date, datetime)): return o.isoformat()
            if _np_top is not None:
                if isinstance(o, _np_top.integer):  return int(o)
                if isinstance(o, _np_top.floating):  return None if _np_top.isnan(o) else float(o)
                if isinstance(o, _np_top.bool_):     return bool(o)
                if isinstance(o, _np_top.ndarray):   return o.tolist()
            if hasattr(o, 'item'):   return o.item()
            if hasattr(o, 'tolist'): return o.tolist()
            raise TypeError(type(o))
        body = json.dumps(data, default=ser, indent=2).encode()
        self.send_response(status)
        self.send_header("Content-Type", "application/json")
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store, no-cache, must-revalidate, max-age=0")
        self.end_headers()
        self.wfile.write(body)

    def send_html(self, html):
        try:
            body = html.encode()
            print(f"  [DEBUG] send_html: encoding OK, {len(body):,} bytes", flush=True)
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_header("Content-Length", str(len(body)))
            self.send_header("Cache-Control", "no-store, no-cache, must-revalidate, max-age=0")
            self.send_header("Pragma", "no-cache")
            self.send_header("Expires", "0")
            self.end_headers()
            print(f"  [DEBUG] send_html: headers sent, writing body...", flush=True)
            self.wfile.write(body)
            print(f"  [DEBUG] send_html: body written OK", flush=True)
        except Exception as _she:
            import traceback as _shtb
            print(f"  [DEBUG] send_html CRASHED: {_she}", flush=True)
            _shtb.print_exc()

    def do_POST(self):
        import time as _t
        _t0 = _t.time()
        parsed = urllib.parse.urlparse(self.path)
        content_length = int(self.headers.get('Content-Length', 0))
        post_data = self.rfile.read(content_length).decode('utf-8')
        params = {}
        if post_data:
            try:
                params = json.loads(post_data)
            except Exception:
                params = dict(urllib.parse.parse_qsl(post_data))
        # Merge with query parameters
        q_params = dict(urllib.parse.parse_qsl(parsed.query))
        params.update(q_params)

        if not parsed.path.startswith("/api/"):
            self.send_response(404); self.end_headers(); return
        ep = parsed.path[5:]
        hint = params.get("symbol", params.get("sym", ""))
        lbl = f" [{hint}]" if hint else ""
        print(f"  [REQ POST] /api/{ep}{lbl}", flush=True)
        try:
            result = self.route(ep, params)
            ms = (_t.time() - _t0)*1000
            tag = "SLOW" if ms>2000 else ("OK  " if ms<500 else "WAIT")
            print(f"  [{tag}] /api/{ep}{lbl}  {ms:.0f}ms", flush=True)
            if result is None:
                pass
            else:
                try:
                    from core.api_contracts import validate_response
                    validate_response(ep, result)
                except Exception as _ce:
                    print(f"  [CONTRACT ERROR] /api/{ep}: {_ce}", flush=True)
                self.send_json(result)
        except Exception as e:
            import traceback
            print(f"  [ERR  ] /api/{ep}{lbl}  {(_t.time()-_t0)*1000:.0f}ms  {e}", flush=True)
            traceback.print_exc()
            status_code = 401 if "Authentication required" in str(e) else 500
            self.send_json({"error": str(e)}, status_code)

    def do_GET(self):
        import time as _t
        _t0 = _t.time()
        parsed = urllib.parse.urlparse(self.path)
        params = dict(urllib.parse.parse_qsl(parsed.query))
        print(f"  [DEBUG] do_GET: {self.path}", flush=True)
        if parsed.path in ("/", "/index.html"):
            print("  [PAGE ] UI served", flush=True)
            print(f"  [DEBUG] HTML length: {len(HTML):,} chars", flush=True)
            try:
                _tstr = __import__("datetime").date.today().isoformat()
                _html_out = HTML.replace("{today_str}", _tstr)
                self.send_html(_html_out)
            except Exception as _pge:
                import traceback as _pgtb
                print(f"  [DEBUG] PAGE SERVE CRASHED: {_pge}", flush=True)
                _pgtb.print_exc()
            return
        if parsed.path == "/favicon.svg":
            _svg = b'''<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 32 32"><rect width="32" height="32" rx="6" fill="#060f16"/><polygon points="16,4 18.4,12.2 27,12.2 20.3,17.4 22.7,25.6 16,20.4 9.3,25.6 11.7,17.4 5,12.2 13.6,12.2" fill="#00d4ff"/></svg>'''
            self.send_response(200)
            self.send_header("Content-Type","image/svg+xml")
            self.send_header("Cache-Control","public,max-age=86400")
            self.send_header("Content-Length",str(len(_svg)))
            self.end_headers(); self.wfile.write(_svg); return
        if parsed.path == "/favicon.ico":
            self.send_response(204); self.end_headers(); return
        if not parsed.path.startswith("/api/"):
            self.send_response(404); self.end_headers(); return
        ep = parsed.path[5:]
        hint = params.get("symbol", params.get("sym", ""))
        lbl = f" [{hint}]" if hint else ""
        print(f"  [REQ  ] /api/{ep}{lbl}", flush=True)
        try:
            result = self.route(ep, params)
            ms = (_t.time() - _t0)*1000
            tag = "SLOW" if ms>2000 else ("OK  " if ms<500 else "WAIT")
            print(f"  [{tag}] /api/{ep}{lbl}  {ms:.0f}ms", flush=True)
            # Special case: raw xlsx binary response
            if isinstance(result, dict) and "_xlsx_bytes" in result:
                raw   = result["_xlsx_bytes"]
                fname = result.get("_filename", "report.xlsx")
                self.send_response(200)
                self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition", f'attachment; filename="{fname}"')
                self.send_header("Content-Length", str(len(raw)))
                self.send_header("Cache-Control", "no-store")
                self.end_headers()
                self.wfile.write(raw)
            elif result is None:
                pass  # endpoint already sent headers (e.g. portfolio_csv)
            else:
                try:
                    from core.api_contracts import validate_response
                    validate_response(ep, result)
                except Exception as _ce:
                    print(f"  [CONTRACT ERROR] /api/{ep}: {_ce}", flush=True)
                self.send_json(result)
        except Exception as e:
            import traceback
            print(f"  [ERR  ] /api/{ep}{lbl}  {(_t.time()-_t0)*1000:.0f}ms  {e}", flush=True)
            traceback.print_exc()
            self.send_json({"error": str(e)}, 500)


    def route(self, ep, p):
        import math
        from datetime import date, datetime
        today = date.today()

        # ── Parse JWT Authentication Token ──
        token = self.headers.get("Authorization")
        if not token and "token" in p:
            token = "Bearer " + p["token"]
        
        from core.auth import decode_access_token
        self.current_user = decode_access_token(token) if token else None
        
        # Public endpoints that DO NOT require authentication
        public_endpoints = {
            "all_symbols", "ticker", "auth/signup", "auth/login", "auth/google", "config", "favicon.ico", "favicon.svg",
            "price", "price_history", "instrument_info", "pivots_for_symbol", "ml_status",
            "strategies/intraday", "strategies/swing", "strategies/short_term", "strategies/long_term"
        }
        
        if ep not in public_endpoints and not self.current_user:
            raise Exception("Authentication required. Stale or invalid session.")

        # ── config ───────────────────────────────────────────────────
        if ep == "config":
            return {
                "google_client_id": os.environ.get("GOOGLE_CLIENT_ID", "")
            }

        # ── auth/signup ──────────────────────────────────────────────
        if ep == "auth/signup":
            import uuid
            from core.auth import hash_password, create_access_token
            email = p.get("email", "").strip().lower()
            password = p.get("password", "")
            google_sub = p.get("google_sub", "")
            display_name = p.get("display_name", email.split("@")[0])
            auth_method = p.get("auth_method", "password")

            if not email:
                raise Exception("Email is required")

            from core.personalization_db import get_db_connection
            conn = get_db_connection()
            try:
                exists = conn.execute("SELECT id FROM users WHERE email=?", (email,)).fetchone()
                if exists:
                    raise Exception("A user with this email already exists")

                user_id = str(uuid.uuid4())
                pwd_hash = hash_password(password) if password else None
                created_at = datetime.utcnow().isoformat()
                
                admin_email = os.environ.get("ADMIN_EMAIL", "chiragkaura2003@gmail.com").strip().lower()
                role = "ADMIN" if email == admin_email else "USER"
                
                conn.execute("""
                    INSERT INTO users (id, email, password_hash, google_sub, display_name, auth_method, role, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """, (user_id, email, pwd_hash, google_sub or None, display_name, auth_method, role, created_at))
                conn.commit()
                
                token = create_access_token(user_id, email, role)
                return {"token": token, "email": email, "role": role}
            finally:
                conn.close()

        # ── auth/login ───────────────────────────────────────────────
        if ep == "auth/login":
            from core.auth import verify_password, create_access_token
            import uuid
            email = p.get("email", "").strip().lower()
            password = p.get("password", "")
            google_sub = p.get("google_sub", "")
            display_name = p.get("display_name", email.split("@")[0] if email else "")

            from core.personalization_db import get_db_connection
            conn = get_db_connection()
            try:
                if google_sub:
                    # Google OAuth simulation login/signup
                    user = conn.execute("SELECT id, email, role FROM users WHERE google_sub=?", (google_sub,)).fetchone()
                    if not user:
                        # Fallback: check by email
                        user = conn.execute("SELECT id, email, role FROM users WHERE email=?", (email,)).fetchone()
                        if user:
                            # Link google_sub
                            conn.execute("UPDATE users SET google_sub=?, auth_method='both' WHERE id=?", (google_sub, user["id"]))
                            conn.commit()
                            user = conn.execute("SELECT id, email, role FROM users WHERE id=?", (user["id"],)).fetchone()
                        else:
                            # Create new Google user
                            user_id = str(uuid.uuid4())
                            created_at = datetime.utcnow().isoformat()
                            admin_email = os.environ.get("ADMIN_EMAIL", "chiragkaura2003@gmail.com").strip().lower()
                            role = "ADMIN" if email == admin_email else "USER"
                            conn.execute("""
                                INSERT INTO users (id, email, password_hash, google_sub, display_name, auth_method, role, created_at)
                                VALUES (?, ?, ?, ?, ?, 'google', ?, ?)
                            """, (user_id, email, None, google_sub, display_name, role, created_at))
                            conn.commit()
                            user = {"id": user_id, "email": email, "role": role}
                    
                    token = create_access_token(user["id"], user["email"], user["role"])
                    return {"token": token, "email": user["email"], "role": user["role"]}
                
                else:
                    # Password login
                    if not email or not password:
                        raise Exception("Email and password are required")
                    
                    user = conn.execute("SELECT id, email, password_hash, role FROM users WHERE email=?", (email,)).fetchone()
                    if not user or not user["password_hash"]:
                        raise Exception("Invalid email or password")
                    
                    if not verify_password(password, user["password_hash"]):
                        raise Exception("Invalid email or password")
                    
                    token = create_access_token(user["id"], user["email"], user["role"])
                    return {"token": token, "email": user["email"], "role": user["role"]}
            finally:
                conn.close()

        # ── auth/google (Real Google Verification) ────────────────────
        if ep == "auth/google":
            import requests as _req
            import uuid
            from core.auth import create_access_token
            google_token = p.get("id_token")
            if not google_token:
                raise Exception("Missing Google ID token")
            
            # Verify the ID token via Google TokenInfo API
            res = _req.get(f"https://oauth2.googleapis.com/tokeninfo?id_token={google_token}", timeout=10)
            if res.status_code != 200:
                raise Exception("Invalid Google ID token (verification failed)")
            
            id_info = res.json()
            client_id = os.environ.get("GOOGLE_CLIENT_ID")
            if client_id and client_id != "your-google-client-id-here.apps.googleusercontent.com":
                aud = id_info.get("aud")
                if aud != client_id:
                    raise Exception("Google token audience mismatch")
            
            email = id_info.get("email")
            google_sub = id_info.get("sub")
            name = id_info.get("name", email.split("@")[0] if email else "Google User")
            
            if not email or not google_sub:
                raise Exception("Invalid Google ID token claims")
            
            email = email.strip().lower()
            
            from core.personalization_db import get_db_connection
            conn = get_db_connection()
            try:
                user = conn.execute("SELECT id, email, role FROM users WHERE google_sub=?", (google_sub,)).fetchone()
                if not user:
                    user = conn.execute("SELECT id, email, role FROM users WHERE email=?", (email,)).fetchone()
                    if user:
                        conn.execute("UPDATE users SET google_sub=?, auth_method='both' WHERE id=?", (google_sub, user["id"]))
                        conn.commit()
                        user = conn.execute("SELECT id, email, role FROM users WHERE id=?", (user["id"],)).fetchone()
                    else:
                        user_id = str(uuid.uuid4())
                        created_at = datetime.utcnow().isoformat()
                        admin_email = os.environ.get("ADMIN_EMAIL", "chiragkaura2003@gmail.com").strip().lower()
                        role = "ADMIN" if email == admin_email else "USER"
                        conn.execute("""
                            INSERT INTO users (id, email, password_hash, google_sub, display_name, auth_method, role, created_at)
                            VALUES (?, ?, ?, ?, ?, 'google', ?, ?)
                        """, (user_id, email, None, google_sub, name, role, created_at))
                        conn.commit()
                        user = {"id": user_id, "email": email, "role": role}
                
                token = create_access_token(user["id"], user["email"], user["role"])
                return {"token": token, "email": user["email"], "role": user["role"]}
            finally:
                conn.close()

        # ── onboarding/check ─────────────────────────────────────────
        if ep == "onboarding/check":
            user_id = self.current_user["user_id"]
            from core.personalization_db import get_db_connection
            conn = get_db_connection()
            try:
                profile = conn.execute("SELECT id FROM risk_profiles WHERE user_id=?", (user_id,)).fetchone()
                return {"completed": profile is not None}
            finally:
                conn.close()

        # ── onboarding/submit ────────────────────────────────────────
        if ep == "onboarding/submit":
            import uuid
            user_id = self.current_user["user_id"]
            
            # Extract basic info
            age_range = p.get("age_range", "")
            occupation = p.get("occupation", "")
            location = p.get("location", "")
            
            # Experience
            experience_level = p.get("experience_level", "Intermediate")
            experience_duration = p.get("experience_duration", "")
            
            # Goals
            primary_goals = p.get("primary_goals", [])
            primary_goal = primary_goals[0] if primary_goals else "growth"
            
            # Risk
            risk_comfort = p.get("risk_comfort", "Moderate")
            risk_scenario_answer = p.get("risk_scenario_answer", "hold")
            drawdown_reaction = risk_scenario_answer
            
            # Preferences
            preferred_markets = p.get("preferred_markets", [])
            trading_styles = p.get("trading_styles", [])
            preferred_sectors = p.get("preferred_sectors", [])
            investment_horizon = p.get("investment_horizon", "Months")
            
            # Capital & Sizing
            starting_capital = float(p.get("starting_capital", 100000))
            max_position_pct = float(p.get("max_position_pct", 10.0))
            max_sector_pct = float(p.get("max_sector_pct", 30.0))
            onboarding_version = int(p.get("onboarding_version", 1))

            # Automatically compute horizon_weights from preferred investment_horizon
            horizon = investment_horizon.lower()
            if "intraday" in horizon:
                hw = {"intraday": 70, "swing": 20, "short": 10, "long": 0}
            elif "day" in horizon or "week" in horizon:
                hw = {"swing": 50, "short": 40, "long": 10}
            elif "month" in horizon:
                hw = {"swing": 20, "short": 50, "long": 30}
            else: # Years
                hw = {"swing": 0, "short": 20, "long": 80}
            
            import json as _json
            hw_str = _json.dumps(hw)
            
            # Excluded sectors = any sector not in preferred_sectors
            all_known_sectors = {"IT", "Banking", "Pharma", "Auto", "Energy", "FMCG", "Defence", "Infrastructure", "Telecom"}
            excluded = list(all_known_sectors - set(preferred_sectors))
            es_str = _json.dumps(excluded)

            from core.personalization_db import get_db_connection
            conn = get_db_connection()
            try:
                profile_id = str(uuid.uuid4())
                created_at = datetime.utcnow().isoformat()
                
                conn.execute("""
                    INSERT OR REPLACE INTO risk_profiles (
                        id, user_id, primary_goal, horizon_weights, drawdown_reaction,
                        experience_level, starting_capital, excluded_sectors,
                        max_position_pct, max_sector_pct, onboarding_version,
                        age_range, occupation, location, experience_duration,
                        primary_goals, risk_comfort, risk_scenario_answer,
                        preferred_markets, trading_styles, preferred_sectors,
                        investment_horizon, created_at, updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (profile_id, user_id, primary_goal, hw_str, drawdown_reaction,
                      experience_level, starting_capital, es_str,
                      max_position_pct, max_sector_pct, onboarding_version,
                      age_range, occupation, location, experience_duration,
                      _json.dumps(primary_goals), risk_comfort, risk_scenario_answer,
                      _json.dumps(preferred_markets), _json.dumps(trading_styles), _json.dumps(preferred_sectors),
                      investment_horizon, created_at, created_at))

                # 2. Create Default 'Primary' Portfolio for the user if it doesn't exist
                portfolio_id = str(uuid.uuid4())
                conn.execute("""
                    INSERT OR IGNORE INTO portfolios (id, user_id, name, created_at)
                    VALUES (?, ?, 'Primary', ?)
                """, (portfolio_id, user_id, created_at))
                
                # Update risk_settings for backwards compatibility
                conn.execute("""
                    INSERT OR REPLACE INTO risk_settings (user_id, capital, max_position_pct, max_sector_pct, kill_switch, updated_at)
                    VALUES (?, ?, ?, ?, 0, ?)
                """, (user_id, starting_capital, max_position_pct, max_sector_pct, created_at))
                
                conn.commit()
                return {"ok": True}
            finally:
                conn.close()

        # ── recommendations ──────────────────────────────────────────
        if ep == "recommendations":
            import uuid
            user_id = self.current_user["user_id"]
            
            from core.personalization_db import get_db_connection
            conn = get_db_connection()
            try:
                # 1. Fetch risk profile
                profile = conn.execute("SELECT * FROM risk_profiles WHERE user_id=?", (user_id,)).fetchone()
                if not profile:
                    raise Exception("Profile not found. Please complete onboarding first.")
                        # 2. Get user advisor candidates
                # Map primary_goal to advisor type and risk preference
                goal_types = {"growth": "swing", "income": "long", "preservation": "position", "learning": "swing"}
                adv_type = goal_types.get(profile["primary_goal"], "swing")
                
                reaction_risks = {"sell_all": "low", "sell_some": "balanced", "hold": "balanced", "buy_more": "high"}
                adv_risk = reaction_risks.get(profile["drawdown_reaction"], "balanced")

                advisor_params = {
                    "amount": total_capital,
                    "type": adv_type,
                    "risk": adv_risk,
                    "date": today.isoformat()
                }
                
                advisor_res = self.route("advisor", advisor_params)
                candidates = advisor_res.get("signals", [])
                
                # Call recommendation engine pipeline helper
                from core.recommendation_engine import generate_personalized_recommendations
                recs, blocked, block_reason = generate_personalized_recommendations(user_id, candidates, conn)
                
                if blocked:
                    return {"ok": True, "recommendations": [], "blocked": True, "reason": block_reason}
                
                # Clear past un-accepted recommendations to avoid clutter
                conn.execute("DELETE FROM recommendations WHERE user_id=? AND passed=1", (user_id,))
                
                passed_recommendations = []
                for r in recs:
                    rec_id = str(uuid.uuid4())
                    payload = json.dumps(r)
                    
                    reasoning_text = json.dumps({
                        "contributors": r.get("reasons", ["Technical breakout signal"]),
                        "score": r.get("score", 75),
                        "confidence": r.get("confidence", 80),
                        "engine_contributions": " · ".join(r.get("explanations", []))
                    })
                    
                    checks_log = json.dumps({
                        "sector_exclusion": {"passed": True},
                        "position_limit": {"passed": True, "limit_pct": r["sizing"]["max_position_pct"]},
                        "sector_limit": {"passed": True, "limit_pct": r["sizing"]["max_sector_pct"]},
                        "kill_switch": {"passed": True}
                    })
                    
                    conn.execute("""
                        INSERT INTO recommendations (
                            id, user_id, portfolio_id, rec_type, payload, reasoning, constraints_checked, passed, created_at
                        ) VALUES (?, ?, ?, 'new_capital', ?, ?, ?, ?, ?)
                    """, (rec_id, user_id, pf_id, payload, reasoning_text, checks_log, 1, created_at))
                    
                    r_copy = dict(r)
                    r_copy["recommendation_id"] = rec_id
                    passed_recommendations.append(r_copy)
                
                conn.commit()
                return {"ok": True, "recommendations": passed_recommendations}
            finally:
                conn.close()

        # ── recommendations/feedback ─────────────────────────────────
        if ep == "recommendations/feedback":
            import uuid
            user_id = self.current_user["user_id"]
            rec_id = p.get("recommendation_id")
            action = p.get("action") # accepted, rejected, modified
            
            if not rec_id or not action:
                raise Exception("recommendation_id and action parameters are required")

            from core.personalization_db import get_db_connection
            conn = get_db_connection()
            try:
                # Verify recommendation exists
                rec = conn.execute("SELECT id FROM recommendations WHERE id=? AND user_id=?", (rec_id, user_id)).fetchone()
                if not rec:
                    raise Exception("Recommendation not found")

                feedback_id = str(uuid.uuid4())
                created_at = datetime.utcnow().isoformat()
                
                conn.execute("""
                    INSERT INTO user_feedback (id, recommendation_id, user_id, action, created_at)
                    VALUES (?, ?, ?, ?, ?)
                """, (feedback_id, rec_id, user_id, action, created_at))
                
                # If accepted, mark the recommendation applied
                if action == "accepted":
                    conn.execute("UPDATE recommendations SET applied=1 WHERE id=?", (rec_id,))
                
                conn.commit()
                return {"ok": True, "message": f"Feedback action '{action}' recorded successfully."}
            finally:
                conn.close()

        # ── admin/users ──────────────────────────────────────────────
        if ep == "admin/users":
            if not self.current_user or self.current_user.get("role") != "ADMIN":
                raise Exception("Access denied: Administrator privileges required.")
            
            from core.personalization_db import get_db_connection
            conn = get_db_connection()
            try:
                users = conn.execute("SELECT id, email, display_name, auth_method, role, created_at FROM users").fetchall()
                user_list = [dict(u) for u in users]
                
                # Fetch system stats too
                inst_cnt = conn.execute("SELECT COUNT(*) FROM instruments").fetchone()[0]
                prices_cnt = conn.execute("SELECT COUNT(*) FROM daily_prices").fetchone()[0]
                recs_cnt = conn.execute("SELECT COUNT(*) FROM recommendations").fetchone()[0]
                
                return {
                    "users": user_list,
                    "stats": {
                        "instruments": inst_cnt,
                        "prices": prices_cnt,
                        "recommendations": recs_cnt,
                        "engine_status": "ONLINE",
                        "database_path": "market_data_v2.db"
                    }
                }
            finally:
                conn.close()

        # ── admin/update_role ────────────────────────────────────────
        if ep == "admin/update_role":
            if not self.current_user or self.current_user.get("role") != "ADMIN":
                raise Exception("Access denied: Administrator privileges required.")
            
            target_user_id = p.get("user_id")
            new_role = p.get("role")
            if not target_user_id or new_role not in ["USER", "ADMIN"]:
                raise Exception("Invalid request parameters.")
                
            from core.personalization_db import get_db_connection
            conn = get_db_connection()
            try:
                # Prevent self-demotion
                if target_user_id == self.current_user["user_id"]:
                    raise Exception("You cannot demote yourself.")
                    
                conn.execute("UPDATE users SET role=? WHERE id=?", (new_role, target_user_id))
                conn.commit()
                return {"ok": True}
            finally:
                conn.close()

        # ── all_symbols ──────────────────────────────────────────────
        if ep == "all_symbols":
            # Use ALL_INSTRUMENTS directly — local imports inside route() cause
            # UnboundLocalError on globals like ALL_INSTRUMENTS in same scope
            _idx = [{"symbol":s,"name":i.name} for s,i in ALL_INSTRUMENTS.items() if i.instrument_type=="INDEX"]
            _eq  = [{"symbol":s,"name":i.name} for s,i in ALL_INSTRUMENTS.items() if i.instrument_type=="EQUITY"]
            _cm  = [{"symbol":s,"name":i.name} for s,i in ALL_INSTRUMENTS.items() if i.instrument_type=="COMMODITY"]
            return {"indices": _idx, "equities": _eq, "commodities": _cm}

        # ── ticker ───────────────────────────────────────────────────────────
        if ep == "ticker":
            try:
                _conn_t = _db()
                _max = _conn_t.execute("SELECT MAX(trade_date) FROM daily_prices").fetchone()[0]
                if not _max:
                    _conn_t.close()
                    return {"prices": [], "date": None}
                # Fetch up to 60 symbols to keep ticker fast and relevant
                _rows = _conn_t.execute("SELECT symbol, close, change_pct FROM daily_prices WHERE trade_date=? AND close IS NOT NULL LIMIT 60", (_max,)).fetchall()
                _conn_t.close()
                prices = [{"sym": r[0], "price": float(r[1]), "chg": float(r[2] or 0.0)} for r in _rows]
                return {"prices": prices, "date": _max}
            except Exception as e:
                return {"error": str(e)}


        # ── nakshatra_today ──────────────────────────────────────────
        if ep == "nakshatra_today":
            from core.nakshatra_engine import get_current_nakshatra
            analysis_date = today
            if p.get("date"):
                try:
                    analysis_date = date.fromisoformat(p.get("date"))
                except ValueError:
                    pass
            nak = get_current_nakshatra(analysis_date)
            notes = {
                "VOLATILE": "High volatility expected. Ideal for options and quick swing trades. Avoid heavy initial positions.",
                "EXIT_ONLY": "Profit booking day. Avoid new long entries. Focus on risk management.",
                "ACCUMULATE": "Favorable accumulation day. Auspicious for long-term additions in blue-chips.",
                "DIRECTIONAL": "Decisive directional trends likely. Follow strong breakout patterns.",
                "BULLISH": "Bullish sentiment favored. Ride the consumer and momentum trends.",
                "STABLE": "Stable, steady market. Good day for foundation-building and low-beta investments.",
                "CAUTION": "Caution warranted. Hidden traps and sharp reversals possible. Do extra research.",
                "SPECULATIVE": "High speculative retail action. Options buying and tight stop losses are recommended."
            }
            nak["market_note"] = notes.get(nak["behavior"], "Neutral daily trends. Follow classic technical setups.")
            return nak

        # ── price ─────────────────────────────────────────────────────
        # Prices are fetched from yfinance at startup and stored in SQLite.
        # This endpoint simply reads from that cache — instant, no network call.
        if ep == "price":
            sym=p.get("symbol","").upper(); req_date=p.get("date",""); inst=ALL_INSTRUMENTS.get(sym)
            if not inst: return {"error":f"Unknown: {sym}"}
            if req_date and req_date!=today.isoformat():
                try:
                    _pc=_db()
                    _row=_pc.execute("SELECT trade_date,open,high,low,close,volume,change_pct FROM daily_prices WHERE symbol=? AND trade_date<=? AND close IS NOT NULL ORDER BY trade_date DESC LIMIT 1",(sym,req_date)).fetchone()
                    _pc.close()
                    if _row:
                        _td,_o,_h,_l,_c,_v,_chg=_row
                        return {"symbol":sym,"price":round(float(_c),2),"close":round(float(_c),2),"change_pct":round(float(_chg or 0),2),"high":round(float(_h or 0),2),"low":round(float(_l or 0),2),"open":round(float(_o or 0),2),"date":_td,"source":"EOD_CACHE","backtest":True}
                except Exception: pass
            
            # If live mode (today or no date specified), try to fetch from Google Finance
            if not req_date or req_date == today.isoformat():
                live = get_live_price_google(sym)
                if live:
                    cache=get_cached_prices(); row=cache.get(sym) or {}
                    return {
                        "symbol": sym,
                        "price": round(float(live["price"]), 2),
                        "close": round(float(live["price"]), 2),
                        "change_pct": round(float(live.get("change_pct") or 0.0), 2),
                        "high": round(float(row.get("high") or live["price"]), 2),
                        "low": round(float(row.get("low") or live["price"]), 2),
                        "open": round(float(row.get("open") or live["price"]), 2),
                        "date": live.get("date", ""),
                        "source": "GOOGLE_FINANCE",
                        "backtest": False
                    }

            cache=get_cached_prices(); row=cache.get(sym)
            if row and row.get("close"):
                return {"symbol":sym,"price":round(float(row["close"]),2),"close":round(float(row["close"]),2),"change_pct":round(float(row.get("change_pct") or 0),2),"high":round(float(row.get("high") or 0),2),"low":round(float(row.get("low") or 0),2),"open":round(float(row.get("open") or 0),2),"date":row.get("date",""),"source":"EOD_CACHE","backtest":False}
            return {"symbol":sym,"price":None,"close":None,"change_pct":None,"high":None,"low":None,"open":None,"date":"","source":"FALLBACK","backtest":False}

        # ── strategies/intraday ──────────────────────────────────────
        if ep == "strategies/intraday":
            sym = p.get("symbol", "").upper()
            target_dt = p.get("date")
            if not sym: raise Exception("Symbol is required")
            from core.strategy_engine import get_intraday_strategy_signal
            return get_intraday_strategy_signal(sym, target_dt)

        # ── strategies/swing ─────────────────────────────────────────
        if ep == "strategies/swing":
            sym = p.get("symbol", "").upper()
            target_dt = p.get("date")
            if not sym: raise Exception("Symbol is required")
            from core.strategy_engine import get_swing_strategy_signal
            return get_swing_strategy_signal(sym, target_dt)

        # ── strategies/short_term ────────────────────────────────────
        if ep == "strategies/short_term":
            sym = p.get("symbol", "").upper()
            target_dt = p.get("date")
            if not sym: raise Exception("Symbol is required")
            from core.strategy_engine import get_short_term_strategy_signal
            return get_short_term_strategy_signal(sym, target_dt)

        # ── strategies/long_term ─────────────────────────────────────
        if ep == "strategies/long_term":
            sym = p.get("symbol", "").upper()
            target_dt = p.get("date")
            if not sym: raise Exception("Symbol is required")
            from core.strategy_engine import get_long_term_strategy_signal
            return get_long_term_strategy_signal(sym, target_dt)

        # ── pivots_for_symbol ────────────────────────────────────────
        # Returns all pivot levels for a symbol (ATL + auto-detected + user-saved)
        if ep == "pivots_for_symbol":
            sym = p.get("symbol","").upper()
            # Run auto-detection first (no-op if <30 rows)
            try: detect_auto_pivots(sym)
            except Exception as _dap_err:
                print(f"  [WARN ] detect_auto_pivots({sym}): {_dap_err}", flush=True)
            # ── ATL reconciliation: instruments.py hardcoded vs DB auto-detected ──
            # If instruments.py ATL < DB-detected ATL → instruments.py wins (it's the true historical low)
            # Update DB with STATIC_VERIFIED source so it shows correctly in UI
            try:
                inst = ALL_INSTRUMENTS.get(sym)
                if inst and getattr(inst, 'all_time_low', None):
                    static_atl   = float(inst.all_time_low)
                    static_date  = str(getattr(inst, 'atl_date', None) or getattr(inst, 'inception_date', ''))
                    from datetime import datetime as _dt
                    _conn = _db()
                    _cur  = _conn.cursor()
                    # Get current ATL in DB for this symbol
                    row = _cur.execute(
                        "SELECT pivot_price, source FROM pivot_levels WHERE symbol=? AND label='ATL'",
                        (sym,)).fetchone()
                    if row:
                        db_atl, db_src = float(row[0]), row[1]
                        if static_atl < db_atl and db_src != 'USER':
                            # Instruments.py has a lower (more extreme) ATL — update DB
                            desc = (f"All-Time Low {static_atl:,.2f} — {static_date} "
                                    f"(instruments.py {static_atl:,.2f} < DB auto {db_atl:,.2f}, manually verified)")
                            _cur.execute("""UPDATE pivot_levels
                                SET pivot_price=?, pivot_date=?, source='STATIC_VERIFIED',
                                    description=?, updated_at=?
                                WHERE symbol=? AND label='ATL' AND source!='USER'""",
                                (static_atl, static_date, desc, _dt.now().isoformat(), sym))
                            _conn.commit()
                            print(f"  [ATL] {sym}: updated {db_atl} → {static_atl} ({static_date}) STATIC_VERIFIED")
                        elif static_atl > db_atl and db_src not in ('USER', 'STATIC', 'STATIC_VERIFIED'):
                            # DB has a lower ATL from history — update description to note it's DB-sourced
                            desc = (f"All-Time Low {db_atl:,.2f} — (DB historical low; "
                                    f"instruments.py has {static_atl:,.2f} but DB is lower)")
                            _cur.execute("""UPDATE pivot_levels
                                SET source='AUTO_VERIFIED', description=?, updated_at=?
                                WHERE symbol=? AND label='ATL' AND source NOT IN ('USER','STATIC','STATIC_VERIFIED')""",
                                (desc, _dt.now().isoformat(), sym))
                            _conn.commit()
                    else:
                        # No ATL in DB at all — seed from instruments.py
                        desc = f"All-Time Low {static_atl:,.2f} — {static_date} (instruments.py, manually researched)"
                        _cur.execute("""INSERT INTO pivot_levels(symbol,label,pivot_price,pivot_date,source,description,updated_at)
                            VALUES(?,?,?,?,?,?,?)""",
                            (sym, 'ATL', static_atl, static_date, 'STATIC_VERIFIED', desc, _dt.now().isoformat()))
                        _conn.commit()
                    _conn.close()
            except Exception as _atl_e:
                print(f"  [ATL reconcile warn] {sym}: {_atl_e}")
            pivots = get_pivots_for_symbol(sym)
            return {"symbol": sym, "pivots": pivots}

        # ── sentiment ────────────────────────────────────────────────
        # Hybrid Sentiment Model: volatility + pivots + candlestick psychology
        if ep == "sentiment":
            import math as _ms
            sym    = p.get("symbol","NIFTY50").upper()
            period = int(p.get("period", 60))
            inst   = ALL_INSTRUMENTS.get(sym)
            if not inst:
                return {"error": f"Unknown symbol: {sym}"}

            # ── Fetch price history from DB ──
            _conn_s = _db()
            rows = _conn_s.execute(
                "SELECT trade_date,open,high,low,close,volume FROM daily_prices "
                "WHERE symbol=? AND close IS NOT NULL ORDER BY trade_date DESC LIMIT ?",
                (sym, max(period+50, 300))).fetchall()
            _conn_s.close()
            if len(rows) < 10:
                return {"error": "Insufficient price history. Run scheduler first."}

            rows = list(reversed(rows))  # oldest first
            dates  = [r[0] for r in rows]
            opens  = [float(r[1]) for r in rows]
            highs  = [float(r[2]) for r in rows]
            lows   = [float(r[3]) for r in rows]
            closes = [float(r[4]) for r in rows]
            vols   = [float(r[5] or 0) for r in rows]
            N = len(closes)
            window = min(period, N)
            w_cl   = closes[-window:]
            w_hi   = highs[-window:]
            w_lo   = lows[-window:]
            w_op   = opens[-window:]
            w_dt   = dates[-window:]
            w_vl   = vols[-window:]

            # ── 1. VOLATILITY INDEX ──
            rets = [(w_cl[i]-w_cl[i-1])/w_cl[i-1] for i in range(1,len(w_cl))]
            vol_std   = (_ms.sqrt(sum(r**2 for r in rets)/len(rets)) * _ms.sqrt(252) * 100)
            vol_21    = (_ms.sqrt(sum(r**2 for r in rets[-20:])/20) * _ms.sqrt(252) * 100) if len(rets)>=20 else vol_std
            vol_ratio = vol_21 / max(vol_std, 0.001)
            # Fear: high recent vol vs avg = panic. Greed: low recent vol = complacency
            if vol_ratio > 1.4:   vol_sentiment = -0.7; vol_label = "FEAR — Spike in volatility"
            elif vol_ratio > 1.15: vol_sentiment = -0.3; vol_label = "CAUTION — Rising volatility"
            elif vol_ratio < 0.7:  vol_sentiment =  0.4; vol_label = "GREED — Complacency (low vol)"
            elif vol_ratio < 0.85: vol_sentiment =  0.2; vol_label = "CALM — Below-average volatility"
            else:                  vol_sentiment =  0.1; vol_label = "NEUTRAL — Normal volatility"
            vol_score = round(vol_21, 1)

            # ── 2. PIVOT INTERACTIONS (price vs key S/R) ──
            cmp   = w_cl[-1]
            high_window = max(w_hi)
            low_window  = min(w_lo)
            mid_pivot   = (high_window + low_window) / 2
            pct_range   = (cmp - low_window) / max(high_window - low_window, 0.001)
            # Where is price in the range? >70% = near top = overbought = greed; <30% = near bottom = fear
            if pct_range > 0.80:   piv_sentiment =  0.8; piv_label = f"EXTREME GREED — Price at {pct_range*100:.0f}% of range (near highs)"
            elif pct_range > 0.65: piv_sentiment =  0.5; piv_label = f"BULLISH — Price at {pct_range*100:.0f}% of range"
            elif pct_range > 0.45: piv_sentiment =  0.1; piv_label = f"NEUTRAL — Price mid-range ({pct_range*100:.0f}%)"
            elif pct_range > 0.25: piv_sentiment = -0.4; piv_label = f"BEARISH — Price at {pct_range*100:.0f}% of range"
            else:                  piv_sentiment = -0.8; piv_label = f"EXTREME FEAR — Price at {pct_range*100:.0f}% of range (near lows)"
            # ATH proximity
            ath_pct = (cmp - inst.all_time_high) / inst.all_time_high * 100
            atl_pct = (cmp - inst.all_time_low)  / inst.all_time_low  * 100
            piv_details = [
                f"Price: ₹{cmp:,.0f}  Range: ₹{low_window:,.0f}–₹{high_window:,.0f}",
                f"Position in window: {pct_range*100:.1f}% from low",
                f"From ATH: {ath_pct:.1f}%  From ATL: +{atl_pct:.1f}%",
            ]

            # ── 3. CANDLESTICK PSYCHOLOGY ──
            candle_signals = []
            candle_sentiment_sum = 0
            candle_count = 0
            # Check last 10 candles for patterns
            for ci in range(max(0,len(w_cl)-10), len(w_cl)):
                o,h,l2,c = w_op[ci],w_hi[ci],w_lo[ci],w_cl[ci]
                body    = abs(c - o)
                wick_up = h - max(o, c)
                wick_dn = min(o, c) - l2
                rng     = h - l2
                if rng < 0.001: continue
                body_pct   = body / rng
                wick_up_pct= wick_up / rng
                wick_dn_pct= wick_dn / rng
                dt_label   = w_dt[ci]

                # Hammer (bullish reversal): small body, long lower wick, at bottom
                if wick_dn_pct > 0.55 and body_pct < 0.35 and wick_up_pct < 0.15:
                    candle_signals.append({"date":dt_label,"pattern":"🔨 Hammer","emotion":"Recovery Optimism","score":0.6,"color":"#26a69a"})
                    candle_sentiment_sum += 0.6; candle_count += 1
                # Shooting star (bearish reversal): small body, long upper wick
                elif wick_up_pct > 0.55 and body_pct < 0.35 and wick_dn_pct < 0.15:
                    candle_signals.append({"date":dt_label,"pattern":"💫 Shooting Star","emotion":"Euphoria Peak / Rejection","score":-0.6,"color":"#ef5350"})
                    candle_sentiment_sum -= 0.6; candle_count += 1
                # Doji (indecision)
                elif body_pct < 0.1:
                    candle_signals.append({"date":dt_label,"pattern":"✚ Doji","emotion":"Market Indecision / Fear of Missing","score":0.0,"color":"#ffcc00"})
                    candle_count += 1
                # Marubozu bull (strong conviction buy)
                elif body_pct > 0.85 and c > o:
                    candle_signals.append({"date":dt_label,"pattern":"▲ Bull Marubozu","emotion":"Strong Buying Conviction","score":0.8,"color":"#26a69a"})
                    candle_sentiment_sum += 0.8; candle_count += 1
                # Marubozu bear (panic selling)
                elif body_pct > 0.85 and c < o:
                    candle_signals.append({"date":dt_label,"pattern":"▼ Bear Marubozu","emotion":"Panic Selling / Capitulation","score":-0.8,"color":"#ef5350"})
                    candle_sentiment_sum -= 0.8; candle_count += 1
                # Spinning top (uncertainty)
                elif body_pct < 0.30 and wick_up_pct > 0.25 and wick_dn_pct > 0.25:
                    candle_signals.append({"date":dt_label,"pattern":"⊕ Spinning Top","emotion":"Uncertainty — Bulls vs Bears","score":-0.1,"color":"#7aa8c0"})
                    candle_count += 1

            candle_sentiment = (candle_sentiment_sum / max(candle_count, 1)) if candle_count > 0 else 0.0
            if not candle_signals:
                candle_signals.append({"date":w_dt[-1],"pattern":"— No strong patterns","emotion":"Neutral price action","score":0,"color":"var(--dim)"})

            # ── 4. TREND EMOTION (momentum) ──
            sma20 = sum(w_cl[-20:])/20 if len(w_cl)>=20 else w_cl[-1]
            sma50 = sum(w_cl[-50:])/50 if len(w_cl)>=50 else w_cl[-1]
            sma_trend = (cmp - sma20) / sma20 * 100
            rsi_window = w_cl[-15:]
            gains = [max(0,rsi_window[i]-rsi_window[i-1]) for i in range(1,len(rsi_window))]
            losses= [max(0,rsi_window[i-1]-rsi_window[i]) for i in range(1,len(rsi_window))]
            avg_g = sum(gains)/max(len(gains),1); avg_l = sum(losses)/max(len(losses),1)
            rsi_val = 100 - 100/(1+avg_g/max(avg_l,0.001)) if avg_l > 0 else 100

            if rsi_val > 75:   trend_sentiment =  0.7; trend_label = f"GREED — RSI {rsi_val:.0f} (Overbought)"
            elif rsi_val > 60: trend_sentiment =  0.4; trend_label = f"BULLISH — RSI {rsi_val:.0f} (Strong)"
            elif rsi_val > 40: trend_sentiment =  0.0; trend_label = f"NEUTRAL — RSI {rsi_val:.0f}"
            elif rsi_val > 25: trend_sentiment = -0.4; trend_label = f"BEARISH — RSI {rsi_val:.0f} (Weak)"
            else:              trend_sentiment = -0.7; trend_label = f"FEAR — RSI {rsi_val:.0f} (Oversold)"
            trend_details = [
                f"RSI(14): {rsi_val:.1f}",
                f"Price vs SMA20: {sma_trend:+.2f}%",
                f"SMA20: ₹{sma20:,.0f}  SMA50: ₹{sma50:,.0f}",
                f"{'Above' if cmp>sma20 else 'Below'} SMA20 · {'Above' if cmp>sma50 else 'Below'} SMA50",
            ]

            # ── 5. VOLUME EMOTION ──
            avg_vol = sum(w_vl[-20:])/20 if len(w_vl)>=20 else (sum(w_vl)/len(w_vl) if w_vl else 1)
            recent_vol = w_vl[-1] if w_vl else 0
            vol_ratio2 = recent_vol / max(avg_vol, 1)
            price_dir  = (w_cl[-1] - w_cl[-2]) / w_cl[-2] if len(w_cl)>=2 else 0
            if vol_ratio2 > 2.0 and price_dir > 0:   vol_em = 0.7;  vol_em_label = "STRONG BUYING — High volume bull candle"
            elif vol_ratio2 > 2.0 and price_dir < 0: vol_em = -0.7; vol_em_label = "PANIC SELL — High volume bear candle"
            elif vol_ratio2 > 1.4 and price_dir > 0: vol_em = 0.4;  vol_em_label = "Buying interest increasing"
            elif vol_ratio2 > 1.4 and price_dir < 0: vol_em = -0.4; vol_em_label = "Selling pressure rising"
            elif vol_ratio2 < 0.5:                    vol_em = -0.1; vol_em_label = "Low volume — lack of conviction"
            else:                                     vol_em = 0.1;  vol_em_label = "Normal volume"

            # ── COMPOSITE SCORE (weighted average) ──
            # Weights: Pivot(25%) + Trend/RSI(25%) + Candle(20%) + Volatility(15%) + Volume(15%)
            composite = (piv_sentiment*0.25 + trend_sentiment*0.25 +
                         candle_sentiment*0.20 + vol_sentiment*0.15 + vol_em*0.15)
            composite = max(-1.0, min(1.0, composite))

            # Emotion state label
            if   composite >  0.60: emotion = "EXTREME GREED";  emo_col = "#ff4444"
            elif composite >  0.30: emotion = "GREED";           emo_col = "#ff8800"
            elif composite >  0.10: emotion = "OPTIMISM";        emo_col = "#26a69a"
            elif composite > -0.10: emotion = "NEUTRAL";         emo_col = "#7aa8c0"
            elif composite > -0.30: emotion = "CAUTION";         emo_col = "#ffcc00"
            elif composite > -0.60: emotion = "FEAR";            emo_col = "#ef5350"
            else:                   emotion = "EXTREME FEAR";    emo_col = "#cc00ff"

            # ── Rolling 14-day sentiment history ──
            history = []
            for hi in range(14, min(window, len(w_cl))):
                seg = w_cl[hi-14:hi]
                seg_h = w_hi[hi-14:hi]; seg_l = w_lo[hi-14:hi]
                seg_r = [(seg[j]-seg[j-1])/seg[j-1] for j in range(1,len(seg))]
                seg_std = _ms.sqrt(sum(r**2 for r in seg_r)/len(seg_r)) if seg_r else 0.01
                seg_rsi_g = [max(0,seg[j]-seg[j-1]) for j in range(1,len(seg))]
                seg_rsi_l = [max(0,seg[j-1]-seg[j]) for j in range(1,len(seg))]
                ag = sum(seg_rsi_g)/max(len(seg_rsi_g),1); al = sum(seg_rsi_l)/max(len(seg_rsi_l),1)
                seg_rsi = 100-100/(1+ag/max(al,0.001)) if al>0 else 50
                seg_pct = (seg[-1]-min(seg_l))/max(max(seg_h)-min(seg_l),0.001)
                seg_s = (seg_pct-0.5)*2*0.4 + (seg_rsi/100-0.5)*2*0.35 + (-seg_std*15)*0.25
                history.append({"date":w_dt[hi], "score":round(max(-1,min(1,seg_s)),3)})

            # ── Actionable signal ──
            if composite > 0.4:
                signal = "🟢 BULLISH SENTIMENT — Market is in an optimistic/greed phase for this symbol. Momentum players can ride the trend. Watch for overbought signals as a warning."
                signal_action = "Consider: Holding longs · Trailing stop advised · Avoid new shorts"
                signal_col = "#26a69a"
            elif composite > 0.1:
                signal = "🟡 MILDLY BULLISH — Cautious optimism. Good for accumulation if other signals align (Gann, Simons, Natal)."
                signal_action = "Consider: Partial entry · Accumulate on dips · Keep SL below support"
                signal_col = "#7FFFD4"
            elif composite > -0.1:
                signal = "⚪ NEUTRAL — No strong emotional extreme. Wait for a clearer sentiment signal before taking a position."
                signal_action = "Consider: Wait for confirmation · Check Gann time cycles · Monitor RSI"
                signal_col = "#7aa8c0"
            elif composite > -0.4:
                signal = "🟠 CAUTIOUS / FEAR — Market showing stress for this symbol. Possible accumulation zone if fundamentals are strong."
                signal_action = "Consider: Reduce position size · Watch for hammer/reversal candles · ATR-based SL"
                signal_col = "#ffcc00"
            else:
                signal = "🔴 EXTREME FEAR / PANIC — Deep oversold sentiment. Contrarian opportunity IF near strong S/R support. High risk."
                signal_action = "Consider: Wait for stabilisation · Watch volume for capitulation bottom · Gann pivot buy signal"
                signal_col = "#ef5350"

            return {
                "symbol":      sym,
                "name":        inst.name,
                "cmp":         round(cmp,2),
                "composite":   round(composite, 4),
                "emotion":     emotion,
                "emotion_color": emo_col,
                "components": {
                    "pivot":    {"score":round(piv_sentiment,3), "label":piv_label, "details":piv_details, "weight":25},
                    "trend":    {"score":round(trend_sentiment,3),"label":trend_label,"details":trend_details,"weight":25},
                    "candle":   {"score":round(candle_sentiment,3),"label":f"{len([c for c in candle_signals if c['score']!=0])} patterns detected","details":[],"weight":20},
                    "volatility":{"score":round(vol_sentiment,3),"label":vol_label,"details":[f"Annual vol: {vol_std:.1f}%",f"Recent 21d vol: {vol_21:.1f}%",f"Vol ratio: {vol_ratio:.2f}x"],"weight":15},
                    "volume":   {"score":round(vol_em,3),"label":vol_em_label,"details":[f"Today vol: {recent_vol/1e6:.2f}M",f"20d avg: {avg_vol/1e6:.2f}M",f"Ratio: {vol_ratio2:.2f}x"],"weight":15},
                },
                "candle_signals":  candle_signals[-6:],
                "pivot_details":   piv_details,
                "history":     history[-60:],
                "signal":      signal,
                "signal_action": signal_action,
                "signal_color":  signal_col,
                "rsi":         round(rsi_val,1),
                "vol_pct":     round(vol_21,1),
                "range_pct":   round(pct_range*100,1),
                "period":      window,
            }

        # ── external_sentiment ───────────────────────────────────────
        # Fetch news headlines + analyst data, run NLP scoring (free, no API key)
        if ep == "external_sentiment":
            sym  = p.get("symbol","").upper()
            force = p.get("force", "") == "1"
            inst = ALL_INSTRUMENTS.get(sym)
            if not inst:
                return {"error": f"Unknown symbol: {sym}"}
            try:
                from core.sentiment_external import get_external_sentiment
                result = get_external_sentiment(
                    symbol=sym,
                    yfinance_symbol=getattr(inst, "yfinance_symbol", "") or "",
                    company_name=getattr(inst, "name", sym),
                    instrument_type=getattr(inst, "instrument_type", "EQUITY"),
                    force_refresh=force,
                )
                return result
            except Exception as _ese:
                import traceback as _tb
                return {"error": str(_ese), "trace": _tb.format_exc()[-500:]}

        # ── rag_status — RAG + LLM system status ──────────────────────
        if ep == "rag_status":
            try:
                from core.rag_engine import get_rag_status
                from core.llm_extractor import get_extraction_status
                return {**get_rag_status(), **get_extraction_status()}
            except Exception as _e:
                return {"error": str(_e)}

        # ── rag_ingest — trigger ingest for one symbol or all ──────────
        if ep == "rag_ingest":
            sym  = p.get("symbol", "").upper()
            inst = ALL_INSTRUMENTS.get(sym) if sym else None
            try:
                from core.rag_engine import ingest_symbol, nightly_ingest, init_rag_tables
                init_rag_tables()
                if inst:
                    result = ingest_symbol(
                        symbol=sym,
                        yf_symbol=getattr(inst, "yfinance_symbol", "") or "",
                        company_name=getattr(inst, "name", sym),
                    )
                else:
                    import threading as _t
                    _t.Thread(target=nightly_ingest, daemon=True,
                              name="RAGNightlyIngest").start()
                    result = {"ok": True, "message": "Full nightly ingest started in background"}
                return {**result, "ok": True}
            except Exception as _e:
                import traceback as _tb
                return {"ok": False, "error": str(_e), "trace": _tb.format_exc()[-400:]}

        # ── llm_extract — get structured extraction for a symbol ───────
        if ep == "llm_extract":
            sym  = p.get("symbol", "").upper()
            fp   = p.get("fiscal_period", "")
            inst = ALL_INSTRUMENTS.get(sym)
            if not inst:
                return {"error": f"Unknown symbol: {sym}"}
            try:
                from core.llm_extractor import _get_cached, get_cached_extractions, extract
                from core.rag_engine import retrieve, RAG_AVAILABLE

                # ── 1. Always return cached data immediately (no wait) ─────────
                cached  = _get_cached(sym, fp)
                history = get_cached_extractions(sym)

                if cached:
                    # Fresh cache hit — return instantly, no extraction needed
                    return {"ok": True, "symbol": sym, "extraction": cached,
                            "history": history, "status": "cached"}

                # ── 2. Nothing cached — fire background extraction and return ──
                # The UI will poll or the user can re-click after a few seconds.
                def _bg_extract(sym=sym, fp=fp, inst=inst):
                    try:
                        inst_name = getattr(inst, "name", sym)
                        chunks    = retrieve(sym,
                                            f"{inst_name} earnings guidance results margins",
                                            k=6) if RAG_AVAILABLE else []
                        extract(sym, fiscal_period=fp, context_chunks=chunks)
                        print(f"  [LLM] Background extract done: {sym}", flush=True)
                    except Exception as _be:
                        print(f"  [LLM] Background extract error [{sym}]: {_be}", flush=True)

                import threading as _t
                _t.Thread(target=_bg_extract, daemon=True,
                          name=f"LLM-Extract-{sym}").start()

                # Return immediately with a "pending" payload so the UI can
                # show a spinner and poll once after ~5 seconds.
                return {
                    "ok":       True,
                    "symbol":   sym,
                    "extraction": None,
                    "history":  history,
                    "status":   "extracting",   # UI uses this to show spinner
                    "message":  f"Extracting {sym} in background — refresh in ~5s",
                }
            except Exception as _e:
                import traceback as _tb
                return {"ok": False, "error": str(_e), "trace": _tb.format_exc()[-400:]}

        # ── rag_qa — natural language Q&A over earnings docs ───────────
        if ep == "rag_qa":
            query  = p.get("query", "").strip()
            sym    = p.get("symbol", "").upper()
            if not query:
                return {"error": "query is required"}
            try:
                from core.rag_engine import retrieve, retrieve_any
                from core.llm_extractor import answer_question
                chunks = retrieve(sym, query, k=5) if sym else retrieve_any(query, k=5)
                result = answer_question(query, chunks, symbol=sym)
                return {"ok": True, **result}
            except Exception as _e:
                import traceback as _tb
                return {"ok": False, "error": str(_e), "trace": _tb.format_exc()[-400:]}

        # ── sentiment_db_stats ──────────────────────────────────────────
        if ep == "sentiment_db_stats":
            try:
                from core.sentiment_db import get_stats, get_symbol_sentiment_trend, init_sentiment_tables
                # Ensure tables exist (re-run migration if needed)
                init_sentiment_tables()
                sym   = p.get("symbol","")
                stats = get_stats()
                trend = get_symbol_sentiment_trend(sym, days=30) if sym else []
                return {"stats": stats, "trend": trend, "db_path": DB_PATH}
            except Exception as _sde:
                import traceback as _tb2
                return {"error": str(_sde), "trace": _tb2.format_exc()[-600:]}

        # ── sentiment_db_test ───────────────────────────────────────────────
        # Quick diagnostic: write one test row and read it back
        if ep == "sentiment_db_test":
            try:
                from core.sentiment_db import save_headlines, get_stats, init_sentiment_tables, DB_PATH as _sdb_path
                init_sentiment_tables()
                test_hl = [{
                    "title":          "TEST HEADLINE — DB connection working",
                    "snippet":        "Diagnostic test row",
                    "source":         "SYSTEM",
                    "url":            "",
                    "published":      __import__("datetime").datetime.now().strftime("%Y-%m-%d %H:%M"),
                    "score":          0.5,
                    "time_weight":    1.0,
                    "weighted_score": 0.5,
                    "label":          "BULLISH",
                }]
                n = save_headlines("TEST", test_hl, "EQUITY")
                stats = get_stats()
                return {
                    "ok":       True,
                    "inserted": n,
                    "db_path":  _sdb_path,
                    "total_rows": stats.get("total_headlines", 0),
                    "message":  f"Test row written to {_sdb_path}"
                }
            except Exception as _sdt:
                import traceback as _tb3
                return {"ok": False, "error": str(_sdt), "trace": _tb3.format_exc()[-800:]}

        # ── sentiment_fetch_all ─────────────────────────────────────────────
        # Trigger bulk news fetch for all symbols in background thread
        if ep == "sentiment_fetch_all":
            sym_filter = p.get("symbols","").upper().split(",") if p.get("symbols") else None
            try:
                import threading as _thr
                def _bulk_run():
                    try:
                        import sys as _s, os as _o
                        _s.path.insert(0, _o.path.dirname(_o.path.abspath(__file__)))
                        from bulk_news_fetch import bulk_fetch_all, INSTRUMENTS as _INST
                        _instruments = _INST
                        if sym_filter:
                            _wanted = set(sym_filter)
                            _instruments = [r for r in _INST if r[0] in _wanted]
                        bulk_fetch_all(delay_secs=1.5, max_per_symbol=20, verbose=True)
                    except Exception as _be:
                        print(f"  [BULK ] Error: {_be}", flush=True)
                _t = _thr.Thread(target=_bulk_run, daemon=True)
                _t.start()
                return {"ok": True, "message": "Bulk news fetch started in background. Check terminal for progress."}
            except Exception as _bfe:
                return {"error": str(_bfe)}

        # ── sentiment_label ──────────────────────────────────────────────────
        if ep == "sentiment_label":
            try:
                from core.sentiment_db import save_human_label
                sym   = p.get("symbol","").upper()
                title = p.get("title","")
                label = p.get("label","").upper()
                notes = p.get("notes","")
                if not sym or not title or not label:
                    return {"error": "symbol, title and label required"}
                ok = save_human_label(sym, title, label, notes)
                return {"ok": ok, "message": f"Label '{label}' saved for {sym}"}
            except Exception as _sle:
                return {"error": str(_sle)}

        # ── check_atl ────────────────────────────────────────────────
        # Debug: verify ATL stored in DB vs instruments.py value
        if ep == "check_atl":
            sym  = p.get("symbol","NIFTY50").upper()
            inst = ALL_INSTRUMENTS.get(sym)
            from datetime import datetime as _dtc
            _conn_c = _db()
            # Get current DB ATL
            db_row = _conn_c.execute(
                "SELECT pivot_price, pivot_date, source, description, updated_at "
                "FROM pivot_levels WHERE symbol=? AND label='ATL'", (sym,)).fetchone()
            # Get DB history min low
            db_min = _conn_c.execute(
                "SELECT MIN(low), MIN(close) FROM daily_prices WHERE symbol=?", (sym,)).fetchone()
            _conn_c.close()
            return {
                "symbol": sym,
                "instruments_atl": float(inst.all_time_low) if inst else None,
                "instruments_atl_date": str(getattr(inst,'atl_date',None) or '') if inst else None,
                "db_pivot_atl": {"price":db_row[0],"date":db_row[1],"source":db_row[2],
                                  "description":db_row[3],"updated_at":db_row[4]} if db_row else None,
                "db_history_min_low":   db_min[0] if db_min else None,
                "db_history_min_close": db_min[1] if db_min else None,
                "verdict": ("instruments.py ATL is lower — STATIC_VERIFIED" 
                            if inst and db_row and float(inst.all_time_low) < float(db_row[0])
                            else "DB history ATL is lower" 
                            if inst and db_row and float(inst.all_time_low) > float(db_row[0])
                            else "Equal or unknown")
            }

        # ── save_pivot ───────────────────────────────────────────────
        # Save a USER-defined custom pivot level
        if ep == "save_pivot":
            sym   = p.get("symbol","").upper()
            label = p.get("label","CUSTOM")
            price = float(p.get("price",0))
            dt    = p.get("date","")
            desc  = p.get("description","User-defined pivot")
            if not sym or not price or not dt:
                return {"error": "symbol, price, date required"}
            save_user_pivot(sym, label, price, dt, desc)
            return {"ok": True, "symbol": sym, "label": label, "price": price, "date": dt}

        # ── instrument_info ──────────────────────────────────────────
        # Returns ATL, ATH, inception_date, ruling_planet for a single instrument
        # Used by Gann Analysis to auto-fill pivot price and pivot date
        if ep == "instrument_info":
            sym  = p.get("symbol","").upper()
            inst = ALL_INSTRUMENTS.get(sym)
            if not inst:
                return {"error": f"Unknown: {sym}"}
                
            # Check for manual user pivot in DB
            from core.gann_math import detect_swing_pivots
            _conn_p = _db()
            manual = _conn_p.execute(
                "SELECT pivot_price, pivot_date, label FROM pivot_levels WHERE symbol=? AND source='USER' ORDER BY updated_at DESC LIMIT 1",
                (sym,)
            ).fetchone()
            
            p_price = None
            p_date = None
            p_source = "static_fallback"
            p_label = "ATL"
            
            if manual:
                p_price = float(manual[0])
                p_date = manual[1]
                p_source = "user_defined"
                p_label = manual[2]
            else:
                # Query history to run detect_swing_pivots
                rows = _conn_p.execute(
                    "SELECT trade_date, high, low, close FROM daily_prices WHERE symbol=? AND close IS NOT NULL ORDER BY trade_date ASC",
                    (sym,)
                ).fetchall()
                ohlc = [{"trade_date": r[0], "high": r[1], "low": r[2], "close": r[3]} for r in rows]
                pivs = detect_swing_pivots(ohlc, lookback=5)
                if pivs:
                    recent = pivs[0]
                    p_price = recent.price
                    p_date = recent.date
                    p_source = "auto_detected"
                    p_label = recent.label
                else:
                    p_price = float(inst.all_time_low)
                    p_date = str(inst.inception_date) if inst.inception_date else ""
                    p_source = "static_fallback"
                    p_label = "ATL"
            _conn_p.close()
            
            return {
                "symbol":         sym,
                "name":           inst.name,
                "atl":            inst.all_time_low,
                "ath":            inst.all_time_high,
                "inception_date": str(inst.inception_date) if inst.inception_date else "",
                "ruling_planet":  inst.ruling_planet,
                "exchange":       inst.exchange,
                # New Gann Hardening fields
                "pivot_price":    p_price,
                "pivot_date":     p_date,
                "pivot_source":   p_source,
                "pivot_label":    p_label
            }

        # ── instruments_full ─────────────────────────────────────────
        # Returns all 257 instruments + their cached prices in one call
        # Used by the Instruments DB page to avoid 257 individual /api/price calls
        if ep == "instruments_full":
            _ik = f"_if_{today}"
            if not hasattr(Handler, "_icache"): Handler._icache = {}
            _ic = Handler._icache.get(_ik)
            if _ic and (time.time()-_ic["ts"])<300:
                print("  [CACHE] instruments_full", flush=True)
                return _ic["data"]
            print(f"  [BUILD] instruments_full: {len(ALL_INSTRUMENTS)} records...", flush=True)
            cache = get_cached_prices()
            instruments_list = []
            _n=0
            for sym, inst in sorted(ALL_INSTRUMENTS.items()):
                _n+=1
                if _n%20==0: print(f"  [BUILD] {_n}/{len(ALL_INSTRUMENTS)}...", flush=True)
                row = cache.get(sym)
                price_data = {}
                if row and row.get("close"):
                    price_data = {
                        "price":      round(float(row["close"]), 2),
                        "change_pct": round(float(row.get("change_pct") or 0), 2),
                        "high":       round(float(row.get("high") or 0), 2),
                        "low":        round(float(row.get("low")  or 0), 2),
                        "date":       row.get("date",""),
                        "source":     "EOD_CACHE",
                    }
                elif inst.yfinance_symbol:
                    price_data = {"price": None, "source": "FALLBACK"}
                else:
                    price_data = {"price": None, "source": "NO_YF"}

                sig_score=0; sig_verdict=""; sig_stars=0; sig_action=""; aspects_str="—"

                instruments_list.append({
                    "symbol":          sym,
                    "name":            inst.name,
                    "exchange":        inst.exchange,
                    "instrument_type": inst.instrument_type,
                    "sector":          inst.sector or "—",
                    "ruling_planet":   inst.ruling_planet or "—",
                    "secondary_planet":inst.secondary_planet or "",
                    "yf_symbol":       inst.yfinance_symbol or "",
                    "atl":             inst.all_time_low,
                    "ath":             inst.all_time_high,
                    "ath_date":        str(inst.ath_date) if getattr(inst,"ath_date",None) else "UNKNOWN",
                    "inception_date":  str(inst.inception_date) if inst.inception_date else "—",
                    "signal_score":    sig_score,
                    "signal_verdict":  sig_verdict,
                    "signal_stars":    sig_stars,
                    "signal_action":   sig_action,
                    "aspects":         aspects_str,
                })

            # Build prices dict {symbol: price_data}
            # price=None means not in DB yet → UI leaves field empty
            prices = {}
            for sym, inst in ALL_INSTRUMENTS.items():
                row = cache.get(sym)
                if row and row.get("close"):
                    prices[sym] = {
                        "price":      round(float(row["close"]), 2),
                        "change_pct": round(float(row.get("change_pct") or 0), 2),
                        "high":       round(float(row.get("high") or 0), 2),
                        "low":        round(float(row.get("low")  or 0), 2),
                        "source":     "EOD_CACHE",
                        "date":       row.get("date",""),
                    }
                else:
                    prices[sym] = {
                        "price":      None,
                        "change_pct": None,
                        "source":     "FALLBACK" if inst.yfinance_symbol else "NO_YF",
                    }

            _r={"instruments":instruments_list,"prices":prices}
            Handler._icache[_ik]={"data":_r,"ts":time.time()}
            print(f"  [DONE ] instruments_full: {len(instruments_list)}", flush=True)
            return _r

        if ep == "overview_data":
            return _get_overview_data()

        # ── dashboard ────────────────────────────────────────────────
        if ep == "dashboard":
            dt = date.fromisoformat(p.get("date", today.isoformat()))
            _dk = f"_pd_{dt}"
            if not hasattr(Handler, "_pcache"): Handler._pcache = {}
            if _dk in Handler._pcache:
                print("  [CACHE] planet dashboard (mem)", flush=True)
                return Handler._pcache[_dk]
            # Try planet_cache DB first (written by BG thread at startup)
            try:
                _pc = _db()
                _row = _pc.execute(
                    "SELECT payload FROM planet_cache WHERE cache_date=?",
                    (dt.isoformat(),)
                ).fetchone()
                _pc.close()
                if _row:
                    _r = json.loads(_row[0])
                    Handler._pcache[_dk] = _r
                    print("  [CACHE] planet dashboard (DB)", flush=True)
                    return _r
            except Exception:
                pass
            print(f"  [VSOP ] Computing planets for {dt}...", flush=True)
            try:
                _r = get_planet_dashboard(dt)
                Handler._pcache[_dk] = _r
                if len(Handler._pcache)>7: del Handler._pcache[next(iter(Handler._pcache))]
                return _r
            except Exception as _e:
                return {"error":str(_e),"planets":{},"aspects":[],"stations":[],
                        "retrograde_planets":[],"retrograde_count":0}

        # ── scanner ──────────────────────────────────────────────────
        if ep == "scanner":
            dt = date.fromisoformat(p.get("date", today.isoformat()))
            _sk = f"_sc_{dt}"
            if not hasattr(Handler, "_scache"): Handler._scache = {}
            _sc = Handler._scache.get(_sk)
            if _sc and (time.time()-_sc["ts"])<60:
                print("  [CACHE] scanner", flush=True)
                return _sc["data"]
            print(f"  [SCAN ] Running scan {dt}...", flush=True)
            raw   = daily_astro_prefilter(dt)
            cache = get_cached_prices(dt)
            raw_list = raw if isinstance(raw, list) else raw.get("results") or []
            results = []
            for item in raw_list:
                sym   = item.get("symbol","")
                inst  = ALL_INSTRUMENTS.get(sym)
                prow  = cache.get(sym, {})
                price = float(prow.get("close") or 0) or (inst.all_time_high * 0.85 if inst else 0)
                results.append({
                    **item,
                    "exchange":      inst.exchange      if inst else item.get("exchange","NSE"),
                    "sector":        inst.sector        if inst else item.get("sector",""),
                    "ruling_planet": inst.ruling_planet if inst else item.get("ruling_planet",""),
                    "current_price": price,
                    "change_pct":    float(prow.get("change_pct") or 0),
                    "price_date":    prow.get("date",""),
                })
            _sd={"results":results}
            Handler._scache[_sk]={"data":_sd,"ts":time.time()}
            print(f"  [DONE ] Scanner: {len(results)} results", flush=True)
            return _sd

        # ── quant ─────────────────────────────────────────────────────
        if ep == "quant":
            sym=p.get("symbol","NIFTY50").upper(); inst=ALL_INSTRUMENTS.get(sym)
            force=p.get("force","")=="1"; req_date=p.get("date",""); backtest=req_date and req_date!=today.isoformat()
            sig_type = p.get("signal_type", "fourier")
            fwd_days = int(p.get("forward_days", "10"))
            if not inst: raise ValueError(f"Unknown symbol: {sym}")
            if backtest:
                try:
                    _pc=_db()
                    _row=_pc.execute("SELECT close FROM daily_prices WHERE symbol=? AND trade_date<=? AND close IS NOT NULL ORDER BY trade_date DESC LIMIT 1",(sym,req_date)).fetchone()
                    _pc.close(); price=float(_row[0]) if _row else (float(p.get("price") or 0) or inst.all_time_high*0.85)
                except Exception: price=float(p.get("price") or 0) or inst.all_time_high*0.85
            else: price=float(p.get("price") or 0) or inst.all_time_high*0.85
            
            # Bypass cache if custom backtesting parameters are requested
            if not force and not backtest and sig_type == "fourier" and fwd_days == 10:
                cached=get_cached_quant(sym,max_age_days=1)
                if cached and "backtest_swing" in cached:
                    cached["current_price"] = price
                    if abs(cached.get("support_resistance",{}).get("current_price",0) - price) / max(price,1) > 0.05:
                        chart = cached.get("chart",{})
                        cached["support_resistance"] = find_support_resistance(
                            chart.get("closes",[price]*50),
                            chart.get("highs",[price*1.01]*50),
                            chart.get("lows",[price*0.99]*50),
                            chart.get("volumes",[1000000]*50),
                            current_price=price,
                        )
                    try:
                        cached["db_pivots"] = get_pivots_for_symbol(sym)
                    except Exception:
                        pass
                    return cached

            # Run full analysis dynamically (reads from daily_prices DB first)
            result = full_quant_analysis(
                symbol=sym,
                yf_symbol=inst.yfinance_symbol,
                current_price=price,
                atl=inst.all_time_low,
                ath=inst.all_time_high,
                trend_up=True,
                as_of_date=req_date if backtest else None,
                signal_type=sig_type,
                forward_days=fwd_days,
            )
            result["backtest_date"] = req_date if backtest else None
            
            # Nakshatra Context
            analysis_dt = date.fromisoformat(req_date) if req_date else today
            t_planets = get_all_planets(analysis_dt)
            t_moon = t_planets["Moon"]
            result["transit_moon_nakshatra"] = t_moon.nakshatra
            result["transit_moon_nakshatra_lord"] = t_moon.nakshatra_lord
            result["transit_moon_nakshatra_sectors"] = __import__('core.ephemeris', fromlist=['get_nakshatra_info']).get_nakshatra_info(t_moon.longitude)["sectors"]
            result["instrument_sector"] = inst.sector
            
            # ── Extend chart with full history from DB (quant_engine limits to 504 bars) ──
            try:
                _conn_ext = _db()
                _end_d = req_date if backtest else today.isoformat()
                _rows_ext = _conn_ext.execute(
                    "SELECT trade_date,open,high,low,close,volume FROM daily_prices "
                    "WHERE symbol=? AND close IS NOT NULL AND trade_date<=? "
                    "ORDER BY trade_date ASC", (sym, _end_d)).fetchall()
                _conn_ext.close()
                if len(_rows_ext) > (len(result.get("chart",{}).get("closes",[]))):
                    _rows_ext = _rows_ext[-1200:]
                    result["chart"]["dates"]   = [r[0] for r in _rows_ext]
                    result["chart"]["opens"]   = [round(float(r[1]) if r[1] else float(r[4]),2) for r in _rows_ext]
                    result["chart"]["highs"]   = [round(float(r[2]) if r[2] else float(r[4]),2) for r in _rows_ext]
                    result["chart"]["lows"]    = [round(float(r[3]) if r[3] else float(r[4]),2) for r in _rows_ext]
                    result["chart"]["closes"]  = [round(float(r[4]),2) for r in _rows_ext]
                    result["chart"]["volumes"] = [int(r[5]) if r[5] else 0 for r in _rows_ext]
                    print(f"  [CHART] Extended to {len(_rows_ext)} bars from DB for {sym}", flush=True)
            except Exception as _ext_e:
                pass  # Non-critical — fall back to quant_engine chart data
            # Attach live pivot levels from DB
            try:
                result["db_pivots"] = get_pivots_for_symbol(sym)
            except Exception:
                pass
            # Attach DB coverage info
            try:
                conn2 = _db()
                row2  = conn2.execute(
                    "SELECT COUNT(*), MIN(trade_date), MAX(trade_date) FROM daily_prices WHERE symbol=?",
                    (sym,)).fetchone()
                conn2.close()
                result["db_coverage"] = {
                    "rows": row2[0], "from": row2[1], "to": row2[2]
                } if row2 and row2[0] else {"rows": 0}
            except Exception:
                result["db_coverage"] = {"rows": 0}
            # Attach Options GEX profile (v4.3)
            try:
                from core.options_engine import fetch_gex_profile
                result["gex_profile"] = fetch_gex_profile(inst.yfinance_symbol, price)
            except Exception:
                result["gex_profile"] = {"max_gamma_wall": None, "zero_gamma_level": None, "skew_ratio": 1.0, "status": "Error"}

            # Save dynamically calculated result to SQLite quant_cache for subsequent instant loads
            if not force and not backtest and sig_type == "fourier" and fwd_days == 10:
                try:
                    cache_quant(sym, result, today)
                    print(f"  [CACHE] Saved quant result for {sym} to DB", flush=True)
                except Exception as _ce:
                    print(f"  [CACHE] Failed to save quant result for {sym} to DB: {_ce}", flush=True)

            return result

        # ── gann_track_record ─────────────────────────────────────────
        if ep == "gann_track_record":
            try:
                sym = p.get("symbol", "").upper()
                conn = _db()
                cursor = conn.cursor()
                rows = cursor.execute("""
                    SELECT signal_subtype, direction, price_at_signal, outcome_price_10d
                    FROM signals
                    WHERE symbol=? AND outcome_price_10d IS NOT NULL AND signal_subtype IS NOT NULL
                """, (sym,)).fetchall()
                conn.close()
                
                subtypes = {}
                for subtype, direction, price, outcome in rows:
                    is_bull = (direction or "").upper() in ("BULLISH", "SUPPORT", "BUY")
                    hit = 1 if (is_bull and outcome > price) or (not is_bull and outcome < price) else 0
                    fwd_return = ((outcome - price) / price * 100) if is_bull else ((price - outcome) / price * 100)
                    subtypes.setdefault(subtype, []).append((hit, fwd_return))
                    
                track_record = {}
                for subtype, data in subtypes.items():
                    count = len(data)
                    hit_rate = sum(x[0] for x in data) / count * 100
                    avg_ret = sum(x[1] for x in data) / count
                    track_record[subtype] = {
                        "sample_size": count,
                        "hit_rate_pct": round(hit_rate, 1),
                        "avg_return_pct": round(avg_ret, 2)
                    }
                return {"ok": True, "symbol": sym, "track_record": track_record}
            except Exception as e:
                return {"ok": False, "error": str(e)}

        # ── analyze (full Vprofitables) ─────────────────────────────────
        if ep == "analyze":
            sym          = p.get("symbol","NIFTY50").upper()
            price        = float(p.get("price",24500))
            pivot_price  = float(p.get("pivot_price", price * 0.88))
            pivot_date   = date.fromisoformat(p.get("pivot_date",(today-timedelta(days=120)).isoformat()))
            analysis_date= date.fromisoformat(p.get("date", today.isoformat()))
            return analyze_instrument(
                symbol=sym, current_price=price,
                pivot_price=pivot_price, pivot_date=pivot_date,
                analysis_date=analysis_date,
                volume_spike   =bool(int(p.get("vol",0))),
                reversal_candle=bool(int(p.get("candle",0))),
                gap_opening    =bool(int(p.get("gap",0))),
            )

        # ── natal ─────────────────────────────────────────────────────
        if ep == "natal":
            sym  = p.get("symbol","NIFTY50").upper()
            dt   = date.fromisoformat(p.get("date", today.isoformat()))
            inst  = get_instrument(sym)
            natal = get_natal(sym)
            if not natal or not inst:
                raise ValueError(f"Unknown: {sym}")
            aspects = get_transit_to_natal_aspects(natal, dt)
            planets = {n: {"longitude":pl.longitude,"sign":pl.sign,
                           "sign_degree":pl.sign_degree,"retrograde":pl.retrograde}
                       for n,pl in natal.all_positions().items()}
            return {
                "symbol": sym, "name": inst.name, "exchange": inst.exchange,
                "sector": inst.sector,
                "inception_date": natal.inception_date.isoformat(),
                "inception_time": natal.inception_time_ist,
                "location": natal.location,
                "primary_ruler":   natal.primary_ruler,
                "secondary_ruler": natal.secondary_ruler,
                "tertiary_ruler":  natal.tertiary_ruler,
                "natal_planets":   planets,
                "transit_date":    dt.isoformat(),
                "transit_to_natal": aspects[:20],
                "ruler_activations":[a for a in aspects if a.get("is_ruler_activated")][:10],
                "bull_signals":    [a for a in aspects if a.get("nature")=="BULLISH"
                                    and a.get("orb",99) <= 3.0 and a.get("applying") is not False],
                "bear_signals":    [a for a in aspects if a.get("nature") in ("BEARISH","VOLATILE")
                                    and a.get("orb",99) <= 3.0 and a.get("applying") is not False],
                "transit_moon_nakshatra": get_all_planets(dt)["Moon"].nakshatra,
                "transit_moon_nakshatra_lord": get_all_planets(dt)["Moon"].nakshatra_lord,
                "transit_moon_nakshatra_sectors": __import__('core.ephemeris', fromlist=['get_nakshatra_info']).get_nakshatra_info(get_all_planets(dt)["Moon"].longitude)["sectors"],
            }

        # ── sq9 ───────────────────────────────────────────────────────
        if ep == "sq9":
            price = float(p.get("price",24500))
            atl   = float(p.get("atl",900))
            return {
                "price":    price,
                "levels":   [{"rotation":l.rotation,"above":l.above,"below":l.below,
                               "above_pct":l.above_pct,"below_pct":l.below_pct}
                              for l in sq9_levels(price)],
                "atl_data": sq9_from_atl(atl, price),
            }

        # ── cycles ────────────────────────────────────────────────────
        if ep == "cycles":
            pivot = date.fromisoformat(p.get("pivot",(today-timedelta(days=180)).isoformat()))
            cycs  = time_cycles_from_pivot(pivot, today)
            return {
                "pivot_date": pivot.isoformat(), "today": today.isoformat(),
                "cycles": [{"label":c.label,"target_date":c.target_date.isoformat(),
                             "days_remaining":c.days_remaining,"planet":c.planet_cycle}
                           for c in cycs],
            }

        # ── scheduler_status ─────────────────────────────────────────
        if ep == "scheduler_status":
            sched = get_scheduler()
            return {
                "status":        sched.status,
                "cached_prices": get_cached_prices(),
                "logs":          get_scheduler_log(10),
            }

        # ── scheduler_trigger ─────────────────────────────────────────
        if ep == "scheduler_trigger":
            sched = get_scheduler()
            return sched.trigger_now()

        # ── history_status ──────────────────────────────────────────
        if ep == "history_status":
            conn2 = _db()
            rows2 = conn2.execute("""
                SELECT symbol, COUNT(*) as cnt, MIN(trade_date), MAX(trade_date)
                FROM daily_prices GROUP BY symbol ORDER BY cnt DESC
            """).fetchall()
            total = conn2.execute("SELECT COUNT(*) FROM daily_prices").fetchone()[0]
            conn2.close()
            sym_data = [{"symbol":r[0],"rows":r[1],"from":r[2],"to":r[3]} for r in rows2]
            db_mb    = os.path.getsize(DB_PATH) / 1024 / 1024 if os.path.exists(DB_PATH) else 0
            return {
                "total_rows":    total,
                "symbols_count": len(sym_data),
                "db_size_mb":    round(db_mb, 2),
                "symbols":       sym_data[:50],   # top 50 by row count
                "download_hint": "Run: python download_history.py" if total < 10000 else "Data loaded",
            }

        # ── download_trigger ─────────────────────────────────────────
        # Trigger background download for a single symbol (non-blocking)
        if ep == "download_trigger":
            sym = p.get("symbol","").upper()
            inst = ALL_INSTRUMENTS.get(sym)
            if not inst or not inst.yfinance_symbol:
                return {"error": "Unknown symbol or no YF ticker"}
            import threading as _thr
            def _dl():
                try:
                    from download_history import download_symbol_history, save_ohlcv_rows, compute_zigzag_pivots, write_zigzag_to_db
                    rows = download_symbol_history(sym, inst.yfinance_symbol, inst.inception_date)
                    if rows:
                        save_ohlcv_rows(sym, rows)
                        pivots = compute_zigzag_pivots(sym)
                        if "_meta" in pivots:
                            write_zigzag_to_db(sym, pivots)
                    print(f"  [DL] {sym}: {len(rows)} rows + pivots refreshed")
                except Exception as e:
                    print(f"  [DL] {sym} error: {e}")
            _thr.Thread(target=_dl, daemon=True, name=f"DL_{sym}").start()
            return {"ok": True, "symbol": sym, "message": f"Download started for {sym}"}


        # ── advisor_plan — Personalized Portfolio Planner ─────────────────────
        if ep == "advisor_plan":
            try:
                inv_type   = p.get("type", "swing")
                risk_pref  = p.get("risk", "balanced")
                ratio_val  = float(p.get("ratio", 0.5))
                from core.portfolio_planner import generate_plan
                return generate_plan(inv_type=inv_type, risk_pref=risk_pref, ratio_astro_quant=ratio_val)
            except Exception as e:
                return {"ok": False, "error": str(e)}

        # ══════════════════════════════════════════════════════════════════
        # INVESTMENT AI ADVISOR
        # ══════════════════════════════════════════════════════════════════
        if ep == "advisor":
            import math as _math, sqlite3 as _sq3

            # ── Parse inputs ─────────────────────────────────────────────
            amount       = float(p.get("amount", 100000))
            inv_type     = p.get("type", "swing")        # swing/short/long
            risk_pref    = p.get("risk", "balanced")     # low/balanced/high
            start_date   = p.get("date", today.isoformat())
            n_stocks     = int(p.get("diversify", 3))    # 2-5
            sector_filter= p.get("sector","").lower()
            req_symbols  = p.get("symbols","").upper().split(",") if p.get("symbols") else []

            # ── Horizon map ──────────────────────────────────────────────
            # HORIZON = minimum hold estimate (days). Actual exit driven by price/cycle.
            HORIZON = {"intraday":1, "swing":5, "short":15, "long":90, "position":90}
            horizon_days = HORIZON.get(inv_type, 5)

            # ── Risk multipliers ─────────────────────────────────────────
            RISK_SL  = {"low":0.020, "balanced":0.035, "high":0.060}  # SL % below entry
            RISK_ALLOC={"low":0.25,  "balanced":0.35,  "high":0.50}   # max single alloc
            sl_pct   = RISK_SL.get(risk_pref, 0.035)
            max_alloc= RISK_ALLOC.get(risk_pref, 0.35)

            analysis_date = date.fromisoformat(start_date)
            from core.nakshatra_engine import get_current_nakshatra
            nak_today_global = get_current_nakshatra(analysis_date)

            # ── Step 1: Get planet dashboard ─────────────────────────────
            planet_data = get_planet_dashboard(analysis_date)
            active_aspects = planet_data.get("aspects", [])
            stations       = planet_data.get("stations", [])
            retrograde     = planet_data.get("retrograde_planets", [])

            # ── Step 2: Get prices for analysis date ─────────────────────
            prices = get_cached_prices(analysis_date)

            # ── Step 3: Scan all symbols ──────────────────────────────────
            scan_raw = daily_astro_prefilter(analysis_date)
            scan_list = scan_raw if isinstance(scan_raw, list) else scan_raw.get("results",[])

            # ── Market overview from indices + MCX (for context only) ────────
            # Compute overall market regime using index instruments
            mkt_bull = 0; mkt_bear = 0
            try:
                _all_aspects = detect_aspects(analysis_date)
                for _sym, _inst in ALL_INSTRUMENTS.items():
                    if _inst.instrument_type in ("INDEX", "COMMODITY"):
                        _rl_asp = [a for a in _all_aspects
                                   if a.planet_a == _inst.ruling_planet or a.planet_b == _inst.ruling_planet]
                        mkt_bull += sum(1 for a in _rl_asp if a.bullish_bearish == "BULLISH" and a.orb <= 4)
                        mkt_bear += sum(1 for a in _rl_asp if a.bullish_bearish == "BEARISH" and a.orb <= 4)
            except Exception: pass
            mkt_regime_bearish = mkt_bear > mkt_bull + 2

            # ── Compute global reversal forecast ─────────────────────────────
            # For each investment horizon scan forward to find when sky clears
            WAIT_SCAN = {"swing": 15, "short": 45, "long": 120, "hedge_fund": 20}
            reversal_forecast = {}  # {type: days_to_wait}
            for _itype, _window in WAIT_SCAN.items():
                for _fwd in range(1, _window + 1):
                    _fdate = analysis_date + timedelta(days=_fwd)
                    if _fdate.weekday() >= 5: continue
                    _fpd   = get_planet_dashboard(_fdate)
                    _fbull = sum(1 for a in _fpd.get("aspects",[]) if a.get("direction")=="BULLISH")
                    _fbear = sum(1 for a in _fpd.get("aspects",[]) if a.get("direction")=="BEARISH")
                    _fmal  = len([r for r in _fpd.get("retrograde_planets",[])
                                  if r in ["Saturn","Mars","Rahu","Ketu"]])
                    if _fbull > _fbear + 1 and _fmal <= 1:
                        reversal_forecast[_itype] = _fwd
                        break
                if _itype not in reversal_forecast:
                    reversal_forecast[_itype] = _window  # no improvement found in window

            # Filter by sector if requested — EQUITY ONLY for recommendations
            auto_sectors = []
            if sector_filter == "auto":
                from core.nakshatra_engine import get_current_nakshatra
                nak_today = get_current_nakshatra(analysis_date)
                auto_sectors = [s.lower() for s in nak_today.get("fav_sectors", [])]

            def sym_ok(sym):
                inst = ALL_INSTRUMENTS.get(sym)
                if not inst: return False
                if inst.instrument_type != "EQUITY": return False   # ← Equity only
                if sector_filter:
                    if sector_filter == "auto":
                        inst_sec = (inst.sector or "").lower()
                        if not any(s in inst_sec for s in auto_sectors):
                            return False
                    else:
                        if sector_filter not in (inst.sector or "").lower(): return False
                if req_symbols and sym not in req_symbols: return False
                return True

            # ── Step 4: Score every symbol ────────────────────────────────
            candidates = []
            scored_syms = set()
            score_count = 0
            max_scan_limit = 40

            for item in scan_list:
                sym = item.get("symbol","")
                if not sym or sym in scored_syms: continue
                if not sym_ok(sym): continue
                if not req_symbols and score_count >= max_scan_limit:
                    break
                scored_syms.add(sym)
                score_count += 1
                inst = ALL_INSTRUMENTS.get(sym)
                if not inst: continue

                price_history  = []   # reset per-symbol — must be before any use below
                reversal_dates = []   # reset per-symbol — used in sell date logic
                # Initialise sentiment vars early — used in scoring before news fetch block
                news_score   = None
                news_label   = "NEUTRAL"
                bulk_signal  = "NEUTRAL"
                bulk_net_val = 0.0
                prow  = prices.get(sym, {})
                price = float(prow.get("close") or 0)
                if price <= 0:
                    # get_cached_prices already does a 7-day fallback, but if still
                    # missing, query the DB directly for this symbol's latest close
                    try:
                        _pb = _db()
                        _pr = _pb.execute(
                            "SELECT close FROM daily_prices WHERE symbol=? "
                            "AND close IS NOT NULL ORDER BY trade_date DESC LIMIT 1",
                            (sym,)
                        ).fetchone()
                        _pb.close()
                        price = float(_pr[0]) if _pr and _pr[0] else inst.all_time_high * 0.5
                    except Exception:
                        price = inst.all_time_high * 0.5

                if inv_type == "intraday":
                    try:
                        import yfinance as yf
                        from core.intraday_engine import analyze_intraday
                        from core.nakshatra_engine import compute_nakshatra_alignment
                        from core.gann_math import get_intraday_reversal_times
                        
                        yf_sym = inst.yfinance_symbol
                        df_15m = yf.Ticker(yf_sym).history(period="1d", interval="15m")
                        if not df_15m.empty:
                            # Bypass the 15-minute yfinance delay by overlaying the real-time Google Finance price
                            live_p = get_live_price_google(sym)
                            if live_p and live_p.get("price"):
                                lp = float(live_p["price"])
                                df_15m.loc[df_15m.index[-1], "Close"] = lp
                                if lp > df_15m.loc[df_15m.index[-1], "High"]:
                                    df_15m.loc[df_15m.index[-1], "High"] = lp
                                if lp < df_15m.loc[df_15m.index[-1], "Low"]:
                                    df_15m.loc[df_15m.index[-1], "Low"] = lp
                            res_int = analyze_intraday(sym, analysis_date, df_15m)
                            if res_int.get("ok"):
                                entry = res_int["entry"]
                                sl = res_int["stop_loss"]
                                t1 = res_int["target1"]
                                t2 = res_int["target2"]
                                entry_source = res_int["reasons"][0] if res_int["reasons"] else "Gann Intraday"
                                t1_source = "VWAP Attractor"
                                t2_source = "VWAP Dev Band"
                                sl_source = "Gann Offset"
                                total = res_int["confidence"]
                                buy_reasons = res_int["reasons"]
                                sell_reasons = [
                                    f"T1 ₹{t1:.2f} (VWAP Target) — trade target",
                                    f"T2 ₹{t2:.2f} (VWAP Band) — maximum extension target",
                                    f"Auto-exit: All intraday trades auto-liquidate at 15:15 IST",
                                    f"SL ₹{sl:.2f} — hard stop loss below Gann support"
                                ]
                                hold_days = 1
                                buy_time_str = "Hora / Sidereal window"
                                sell_time_str = "15:15 IST (Auto Close)"
                                buy_date_str = analysis_date.isoformat()
                                buy_price = entry
                                buy_condition = "Astro-Quant Confluence Triggered"
                                sell_date_str = analysis_date.isoformat()
                                sell_price = t1
                                sell_price_2 = t2
                                sell_condition = "T1 hit or same-session close"
                                sell_condition2 = "T2 hit or same-session close"
                                reversal_dates = get_intraday_reversal_times(sym, analysis_date, ruling_planet=inst.ruling_planet)
                                
                                nak_align = compute_nakshatra_alignment(
                                    symbol=sym,
                                    analysis_date=analysis_date,
                                    inv_type=inv_type,
                                    ruling_planet=inst.ruling_planet,
                                    sector=inst.sector
                                )
                                
                                try:
                                    _pc = _db()
                                    _rows = _pc.execute("""
                                        SELECT trade_date, close FROM daily_prices
                                        WHERE symbol=? AND trade_date<=? AND close IS NOT NULL
                                        ORDER BY trade_date DESC LIMIT 30""",
                                        (sym, start_date)).fetchall()
                                    _pc.close()
                                    price_history = [{"date": r[0], "close": round(float(r[1]), 2)} for r in reversed(_rows)]
                                except Exception:
                                    price_history = []
                                    
                                skyBull_count = sum(1 for a in active_aspects if a.get("direction") == "BULLISH")
                                skyBear_count = sum(1 for a in active_aspects if a.get("direction") == "BEARISH")
                                
                                upside_t1 = round((t1 - entry) / entry * 100, 2) if entry > 0 else 0.0
                                upside_t2 = round((t2 - entry) / entry * 100, 2) if entry > 0 else 0.0
                                risk_amt = round((entry - sl) / entry * 100, 2) if entry > 0 else 2.0
                                rr_ratio = round((t1 - entry) / max(entry - sl, 0.01), 2) if entry > sl else 1.5

                                candidates.append({
                                    "symbol":        sym,
                                    "nakshatra_info": {
                                        "name":          nak_align["nakshatra"],
                                        "number":        nak_align["number"],
                                        "ruler":         nak_align["ruler"],
                                        "guna":          nak_align["guna"],
                                        "behavior":      nak_align["behavior"],
                                        "trade_style":   nak_align["trade_style"],
                                        "favored_today": nak_align["favored_today"],
                                        "nak_score":     nak_align["nak_score"],
                                        "pada":          nak_align["pada"],
                                        "caution":       nak_align["caution"],
                                        "rahu_kaal":     nak_align["rahu_kaal"],
                                        "abhijit_muhurat": nak_align["abhijit_muhurat"],
                                    },
                                    "name":          inst.name,
                                    "exchange":      inst.exchange,
                                    "sector":        inst.sector,
                                    "ruling_planet": inst.ruling_planet,
                                    "price":         round(price, 2),
                                    "entry":         round(entry, 2),
                                    "stop_loss":     round(sl, 2),
                                    "target1":       round(t1, 2),
                                    "target2":       round(t2, 2),
                                    "entry_source":  entry_source,
                                    "t1_source":     t1_source,
                                    "t2_source":     t2_source,
                                    "sl_source":     sl_source,
                                    "inv_type":      inv_type,
                                    "risk_pref":     risk_pref,
                                    "upside_t1_pct": upside_t1,
                                    "upside_t2_pct": upside_t2,
                                    "fourier_buy_price":  None,
                                    "fourier_sell_price": None,
                                    "fourier_buy_date":   None,
                                    "fourier_sell_date":  None,
                                    "regime_metrics":     {},
                                    "gex_profile":        {"max_gamma_wall": None, "zero_gamma_level": None, "skew_ratio": 1.0},
                                    "dominant_cycle":     {},
                                    "fourier_r2":         0.0,
                                    "hold_days":     1,
                                    "buy_date":        buy_date_str,
                                    "buy_price":       buy_price,
                                    "buy_time":        buy_time_str,
                                    "buy_condition":   buy_condition,
                                    "sell_date":       sell_date_str,
                                    "sell_price":      sell_price,
                                    "sell_price_2":    sell_price_2,
                                    "sell_time":       sell_time_str,
                                    "sell_condition":  sell_condition,
                                    "sell_condition2": sell_condition2,
                                    "reversal_dates":  reversal_dates,
                                    "sky_bull":        skyBull_count,
                                    "sky_bear":        skyBear_count,
                                    "confidence":      total,
                                    "confluence_score": round(total, 1),
                                    "passed_gate":    True,
                                    "gate_reason":    "Astro-Quant Intraday Alignment",
                                    "price_history":  price_history,
                                    "buy_reasons":    buy_reasons,
                                    "sell_reasons":   sell_reasons,
                                    "risk_pct":        risk_amt,
                                    "rr_ratio":        rr_ratio,
                                    "regime":          "INTRADAY",
                                    "gann_score":      int(total * 0.25),
                                    "quant_score":     int(total * 0.25),
                                    "natal_score":     int(total * 0.20),
                                    "planet_score":    int(total * 0.15),
                                    "fund_score":      0,
                                    "tech_100":        75.0,
                                    "simons_100":      50.0,
                                    "fund_grade":      "N/A",
                                    "fund_verdict":    "N/A",
                                    "fund_ratios":     {},
                                    "fund_signals":    [],
                                    "planet_text":     ["Intraday planetary hour active"],
                                    "bull_signals":    0,
                                    "bear_signals":    0,
                                    "natal_aspects":   [],
                                    "supports":        [round(entry*0.99,2), round(entry*0.98,2)],
                                    "resistances":     [round(entry*1.01,2), round(entry*1.02,2)],
                                    "news_score":      0.0,
                                    "news_label":      "NEUTRAL",
                                    "news_headline":   "Intraday Trade Scan",
                                    "bulk_signal":     "NEUTRAL",
                                    "bulk_net_val_cr": 0.0,
                                    "bulk_deals_30d":  [],
                                    "inst_acc_score":  0,
                                    "inst_acc_signals": [],
                                    "tech_momentum":   "NEUTRAL",
                                    "tech_score":      5,
                                    "ml_direction":    "NEUTRAL",
                                    "ml_confidence":   0.0,
                                    "ml_reversal_price": price,
                                    "ml_reversal_date":  "",
                                    "ml_days_to_rev":  1,
                                    "ml_model_trained": False,
                                    "ml_expected_move": 0.0,
                                    "ml_signal_alignment": 0.0,
                                    "ml_reversal_map": [],
                                    "ml_swing": {
                                        "direction":      "NEUTRAL",
                                        "confidence":     0.0,
                                        "direction_prob": 0.5,
                                        "reversal_price": price,
                                        "reversal_date":  "",
                                        "days_to_rev":    1,
                                        "expected_move":  0.0,
                                        "signal_alignment": 0.0,
                                        "model_trained":  False,
                                        "reversal_map":   [],
                                        "horizon":        "intraday",
                                    },
                                    "ml_short": {
                                        "direction":      "NEUTRAL",
                                        "confidence":     0.0,
                                        "direction_prob": 0.5,
                                        "reversal_price": price,
                                        "reversal_date":  "",
                                        "days_to_rev":    1,
                                        "expected_move":  0.0,
                                        "signal_alignment": 0.0,
                                        "model_trained":  False,
                                        "reversal_map":   [],
                                        "horizon":        "intraday",
                                    },
                                    "ml_long": {
                                        "direction":      "NEUTRAL",
                                        "confidence":     0.0,
                                        "direction_prob": 0.5,
                                        "reversal_price": price,
                                        "reversal_date":  "",
                                        "days_to_rev":    1,
                                        "expected_move":  0.0,
                                        "signal_alignment": 0.0,
                                        "model_trained":  False,
                                        "reversal_map":   [],
                                        "horizon":        "intraday",
                                    },
                                })
                                continue
                    except Exception as _ie:
                        print(f"  [INTRADAY] Failed to process {sym}: {_ie}", flush=True)

                # ── Gann confluence (0-25) ────────────────────────────────
                gann_score = 0
                sq9_data   = {}
                angle_data = []
                cycle_data = []
                try:
                    pivot_price = price * 0.88
                    pivot_date  = analysis_date - timedelta(days=90)
                    gann_res = analyze_instrument(
                        symbol=sym, current_price=price,
                        pivot_price=pivot_price, pivot_date=pivot_date,
                        analysis_date=analysis_date,
                    )
                    conf = gann_res.get("confluence",{})
                    raw_score = conf.get("score",0)
                    gann_score = min(25, int(raw_score / 25 * 25))
                    sq9_data   = gann_res.get("sq9",{})
                    angle_data = gann_res.get("angles",[])
                    cycle_data = gann_res.get("time_cycles",[])
                    sq9_levels_list = sq9_data.get("levels",[]) if isinstance(sq9_data,dict) else []
                except Exception: pass

                # ── Quant/regime (0-25) ───────────────────────────────────
                # ── Quant/regime + Simons Fourier (0-25) ────────────────
                quant_score   = 0
                regime_str    = "UNKNOWN"
                sr_data       = {}
                chart_data    = {}
                fourier_data  = {}
                # Simons: per-symbol Fourier trough = buy price, peak = sell price
                fourier_buy_price  = None   # lowest forecast price in buy horizon
                fourier_sell_price = None   # highest forecast price in sell horizon
                fourier_buy_date   = None   # date of trough in buy horizon
                fourier_sell_date  = None   # date of peak in sell horizon
                quant_sr_support   = None   # strongest quant S/R support below price
                quant_sr_resist    = None   # strongest quant S/R resistance above price
                regime_metrics     = {}
                regime_data        = {}   # full regime dict — used by long cycle gates
                try:
                    qres = full_quant_analysis(
                        symbol=sym, yf_symbol=inst.yfinance_symbol,
                        current_price=price, atl=inst.all_time_low, ath=inst.all_time_high,
                        as_of_date=start_date if start_date != today.isoformat() else None
                    )
                    reg = qres.get("regime",{})
                    regime_str    = reg.get("regime","UNKNOWN")
                    regime_metrics= reg.get("metrics",{})
                    regime_data   = reg   # full regime dict for long cycle gate
                    regime_map = {"STRONG_BULL":25,"WEAK_BULL":18,"SIDEWAYS":10,
                                  "WEAK_BEAR":5,"STRONG_BEAR":0,"HIGH_VOLATILITY":8}
                    quant_score = regime_map.get(regime_str, 10)
                    sr_data     = qres.get("support_resistance",{})
                    chart_data  = qres.get("chart",{})
                    fourier_data= qres.get("fourier",{})

                    # ── Simons Fourier: find trough (buy) and peak (sell) ──
                    # forecast_60d = list of (date_str, price) for next 60 days
                    fc60 = fourier_data.get("forecast_60d", [])
                    if qres.get("data_source") == "synthetic":
                        fc60 = []
                    if fc60:
                        # Horizon windows per investment type
                        BUY_WINDOW  = {"swing": 15, "short": 30, "long": 60, "hedge_fund": 60}
                        SELL_WINDOW = {"swing": 15, "short": 45, "long": 60, "hedge_fund": 60}
                        bw = BUY_WINDOW.get(inv_type, 10)
                        sw = SELL_WINDOW.get(inv_type, 10)
                        # Find trough in buy window (lowest Fourier forecast)
                        buy_slice  = [(d,p) for d,p in fc60[:bw] if p > 0]
                        sell_slice = [(d,p) for d,p in fc60[:sw] if p > 0]
                        if buy_slice:
                            trough = min(buy_slice, key=lambda x: x[1])
                            # Only use if trough is below current price (genuine dip)
                            if trough[1] < price * 0.995:
                                fourier_buy_price = round(trough[1], 2)
                                fourier_buy_date  = trough[0]
                        if sell_slice:
                            peak = max(sell_slice, key=lambda x: x[1])
                            # Only use if peak is above current price (genuine upside)
                            if peak[1] > price * 1.005:
                                fourier_sell_price = round(peak[1], 2)
                                fourier_sell_date  = peak[0]

                    # ── Quant S/R: best support/resistance for this symbol ──
                    supports_list  = sr_data.get("supports",[])
                    resists_list   = sr_data.get("resistances",[])
                    # Best support: strongest level BELOW current price
                    below_sup = [s for s in supports_list if s.get("price",0) < price * 0.998]
                    if below_sup:
                        quant_sr_support = max(below_sup, key=lambda x: x.get("strength_score",0))
                    # Best resistance: strongest level ABOVE current price
                    above_res = [r for r in resists_list if r.get("price",0) > price * 1.002]
                    if above_res:
                        quant_sr_resist = min(above_res, key=lambda x: x.get("price",0))
                except Exception: pass

                # ── ML Reversal Prediction (v3.9) ─────────────────────────────
                # ── Investment-type-aware ML predictions (v3.9.1) ───────────────
                # Each type uses a different Fourier forecast window and look-back
                # so the model reasons about the correct time horizon for each tab.
                #
                #   swing: 5–15 day window  → short Fourier forecast (15d)
                #   short: 15–45 day window → medium Fourier forecast (30d)
                #   long:  3–18 month window→ full Fourier forecast (60d)

                _ml_default = {
                    "direction": "NEUTRAL", "confidence": 0.0, "direction_prob": 0.5,
                    "reversal_price": price, "reversal_date": "", "days_to_reversal": 5,
                    "model_trained": False, "signal_alignment": 0.5, "expected_move_pct": 0.0,
                    "reversal_prices": [], "reversal_days": [],
                }
                ml_result       = dict(_ml_default)
                ml_result_swing = dict(_ml_default)
                ml_result_short = dict(_ml_default)
                ml_result_long  = dict(_ml_default)

                try:
                    _ml_closes  = chart_data.get("closes",  [])
                    _ml_highs   = chart_data.get("highs",   [])
                    _ml_lows    = chart_data.get("lows",    [])
                    _ml_volumes = chart_data.get("volumes", [])
                    _ml_fc60    = fourier_data.get("forecast_60d", [])
                    _ml_r2      = fourier_data.get("r_squared", 0.0)
                    _ml_bulk    = (1.0 if bulk_signal == "BUY" else
                                  (-1.0 if bulk_signal == "SELL" else 0.0))
                    _ml_news    = news_score or 0.0

                    def _fourier_phase_for_window(fc_data, window_days):
                        """Extract Fourier phase and days-to-trough within a look-ahead window."""
                        phase = 0.5; dtrou = 999
                        if fc_data:
                            _pts = [p2 for _, p2 in fc_data[:window_days] if p2 > 0]
                            if _pts:
                                _tv = min(_pts)
                                _ti = _pts.index(_tv)
                                phase = _ti / max(len(_pts), 1)
                                dtrou = _ti + 1
                        return phase, dtrou

                    def _reversal_map(closes, highs, lows, price, days_ahead):
                        """Scan next days_ahead bars of Fourier forecast for multiple reversal points.
                        Returns list of {day, price, type} dicts for mini-chart overlay."""
                        if not _ml_fc60:
                            return []
                        pts = [(d, p) for d, p in _ml_fc60[:days_ahead] if p > 0]
                        if len(pts) < 4:
                            return []
                        prices_only = [p for _, p in pts]
                        reversals = []
                        # Find local extrema in forecast
                        for k in range(1, len(prices_only) - 1):
                            p_prev, p_cur, p_next = prices_only[k-1], prices_only[k], prices_only[k+1]
                            if p_cur == min(prices_only[k-1:k+2]):   # local trough
                                reversals.append({"day": pts[k][0], "price": round(p_cur, 2), "type": "LOW"})
                            elif p_cur == max(prices_only[k-1:k+2]): # local peak
                                reversals.append({"day": pts[k][0], "price": round(p_cur, 2), "type": "HIGH"})
                        return reversals[:6]  # cap at 6 reversal points

                    if len(_ml_closes) >= 60:
                        # ── SWING (5-15 days): use 15-day Fourier window ──
                        _sw_phase, _sw_dtrou = _fourier_phase_for_window(_ml_fc60, 15)
                        ml_result_swing = predict_reversal(
                            closes=_ml_closes[-60:],  # last 60 bars is enough for swing
                            highs=_ml_highs[-60:], lows=_ml_lows[-60:],
                            volumes=_ml_volumes[-60:],
                            current_price=price, analysis_date=analysis_date,
                            fourier_phase=_sw_phase, days_to_trough=_sw_dtrou,
                            fourier_r2=_ml_r2, news_score=_ml_news, bulk_signal=_ml_bulk,
                            symbol=sym,
                        )
                        ml_result_swing["reversal_map"] = _reversal_map(
                            _ml_closes, _ml_highs, _ml_lows, price, days_ahead=15)

                        # ── SHORT (15-45 days): use 45-day Fourier window ──
                        _sh_phase, _sh_dtrou = _fourier_phase_for_window(_ml_fc60, 45)
                        ml_result_short = predict_reversal(
                            closes=_ml_closes[-120:],  # 120 bars for short-term context
                            highs=_ml_highs[-120:], lows=_ml_lows[-120:],
                            volumes=_ml_volumes[-120:],
                            current_price=price, analysis_date=analysis_date,
                            fourier_phase=_sh_phase, days_to_trough=_sh_dtrou,
                            fourier_r2=_ml_r2, news_score=_ml_news, bulk_signal=_ml_bulk,
                            symbol=sym,
                        )
                        ml_result_short["reversal_map"] = _reversal_map(
                            _ml_closes, _ml_highs, _ml_lows, price, days_ahead=45)

                        # ── LONG (3-18 months): use full 60-day Fourier window ──
                        _lg_phase, _lg_dtrou = _fourier_phase_for_window(_ml_fc60, 60)
                        ml_result_long = predict_reversal(
                            closes=_ml_closes,  # full history for long-term
                            highs=_ml_highs, lows=_ml_lows, volumes=_ml_volumes,
                            current_price=price, analysis_date=analysis_date,
                            fourier_phase=_lg_phase, days_to_trough=_lg_dtrou,
                            fourier_r2=_ml_r2, news_score=_ml_news, bulk_signal=_ml_bulk,
                            symbol=sym,
                        )
                        ml_result_long["reversal_map"] = _reversal_map(
                            _ml_closes, _ml_highs, _ml_lows, price, days_ahead=60)

                        # Active result for this inv_type tab
                        ml_result = (ml_result_swing if inv_type in ("swing", "intraday") else
                                     ml_result_short if inv_type == "short" else
                                     ml_result_long)

                except Exception: pass

                # ── Natal signal (0-25) ───────────────────────────────────
                natal_score  = 0
                bull_signals = []
                bear_signals = []
                natal_aspects= []
                try:
                    natal = get_natal(sym)
                    if natal:
                        aspects = get_transit_to_natal_aspects(natal, analysis_date)
                        bull_signals = [a for a in aspects if a.get("nature")=="BULLISH"
                                        and a.get("orb",99)<=3.0 and a.get("applying") is not False]
                        bear_signals = [a for a in aspects if a.get("nature") in ("BEARISH","VOLATILE")
                                        and a.get("orb",99)<=3.0 and a.get("applying") is not False]
                        natal_aspects= aspects[:8]
                        bull_count = len(bull_signals)
                        bear_count = len(bear_signals)
                        if bull_count > bear_count:
                            natal_score = min(25, 10 + (bull_count - bear_count)*3)
                        elif bear_count > bull_count:
                            natal_score = max(0, 12 - (bear_count - bull_count)*3)
                        else:
                            natal_score = 12
                except Exception: pass

                # ── Planetary signal from ruling aspects (0-25) ───────────
                planet_score = 0
                _days_to_trough = 999
                _ruling_aspect_applying = False
                try:
                    _dom_cycles = fourier_data.get("dominant_cycles", [])
                    _days_to_trough = _dom_cycles[0].get("days_to_next_trough", 999) if _dom_cycles else 999
                except Exception: pass
                try:
                    _is_commodity = (inst.instrument_type == "COMMODITY")
                    _ruling_aspects = [a for a in detect_aspects(analysis_date, heliocentric=_is_commodity)
                                       if a.planet_a==inst.ruling_planet or a.planet_b==inst.ruling_planet]
                    _ruling_aspect_applying = any(getattr(a, "applying", False) for a in _ruling_aspects)
                    _bull_ruling = sum(1 for a in _ruling_aspects if a.bullish_bearish=="BULLISH" and a.orb<=5)
                    _bear_ruling = sum(1 for a in _ruling_aspects if a.bullish_bearish=="BEARISH" and a.orb<=5)
                    if _bull_ruling > _bear_ruling:
                        planet_score = min(25, 8 + _bull_ruling * 4)
                    elif _bull_ruling == _bear_ruling and _bull_ruling > 0:
                        planet_score = 8
                    for _s in stations:
                        if _s.get("planet") == inst.ruling_planet:
                            planet_score = min(25, planet_score + 8)
                except Exception: pass

                # ── Fundamental score (0-25) — Simons meets Graham ──────
                fund_score   = 12   # neutral default if fetch fails / not installed
                fund_grade   = "B"
                fund_verdict = "HOLD"
                fund_ratios  = {}
                fund_signals = []
                try:
                    if _fundamental_engine is not None:
                        _fadv = _fundamental_engine.fundamental_advisor_score(sym, inst.yfinance_symbol)
                        fund_score   = _fadv.get("fundamental_score", 12)
                        fund_grade   = _fadv.get("grade", "B")
                        fund_verdict = _fadv.get("verdict", "HOLD")
                        fund_ratios  = _fadv.get("key_ratios", {})
                        fund_signals = _fadv.get("signals", [])
                except Exception: pass   # never crash advisor due to fundamentals

                # ── Station/retrograde adjustments (type-weighted) ──────────
                adj = 0
                ruling = [inst.ruling_planet, inst.secondary_planet]
                # Stations matter most for swing (timing), least for long (noise)
                _station_wt  = {"intraday": 5, "swing": 5, "short": 3, "long": 1}.get(inv_type, 3)
                _retro_wt    = {"intraday": 8, "swing": 8, "short": 5, "long": 2}.get(inv_type, 5)
                for s in stations:
                    if s.get("planet") in ruling: adj += _station_wt
                for r in retrograde:
                    if r in ["Saturn","Mars","Rahu","Ketu"] and r in ruling: adj -= _retro_wt
                # Long-term: Jupiter/Saturn retro is actually a BUYING opportunity
                if inv_type == "long":
                    for r in retrograde:
                        if r in ["Jupiter","Saturn"] and r in ruling:
                            adj += 3  # mean-reversion opportunity for long-term buyer

                # ── Pre-compute Sq9 levels needed by scoring block ────────────
                # Also capture ML signals for use in scoring
                _ml_dir_prob  = ml_result.get("direction_prob", 0.5)
                _ml_conf      = ml_result.get("confidence", 0.0)
                _ml_sq9_prox  = min([abs(price-s)/price
                                     for s in [round(max(0.01,math.sqrt(price)-d)**2,2) for d in [0.25,0.5,1.0]]], default=0.5)
                import math as _m_pre
                _sqp_pre = _m_pre.sqrt(price) if price > 0 else 1.0
                sq9_sup1 = round(max(0.01, _sqp_pre - 0.5)**2, 2)
                sq9_sup2 = round(max(0.01, _sqp_pre - 1.0)**2, 2)
                sq9_res1 = round((_sqp_pre + 0.5)**2, 2)
                sq9_res2 = round((_sqp_pre + 1.0)**2, 2)

                # ══════════════════════════════════════════════════════════════
                # TYPE-DIFFERENTIATED SCORING — each mode has its own weights
                # SWING:  Gann35 + Technical30 + Natal25 + Sentiment10 + Fund0
                # SHORT:  Simons/Fourier30 + Technical25 + Gann20 + Natal15 + Fund5 + Sent5
                # LONG:   Fundamental40 + Gann25 + Simons/Fourier20 + Natal15 + Technical0
                # ══════════════════════════════════════════════════════════════

                # ── Raw engine values (all on 0-100 scale internally) ─────────────
                # gann_score:   0-25  → scale to 0-100 for weight math
                # quant_score:  0-25 (regime map)
                # natal_score:  0-25
                # planet_score: 0-25
                # fund_score:   0-25

                gann_100   = min(100, gann_score   * 4)    # 0-25 → 0-100
                quant_100  = min(100, quant_score  * 4)    # 0-25 → 0-100
                natal_100  = min(100, natal_score  * 4)    # 0-25 → 0-100
                planet_100 = min(100, planet_score * 4)    # 0-25 → 0-100
                fund_100   = min(100, fund_score   * 4)    # 0-25 → 0-100

                # ── Technical sub-score (RSI + SMA) — uses chart_data from quant engine ──
                tech_100 = 40  # neutral default
                try:
                    _c_t = chart_data.get("closes", [])  # already fetched by quant engine
                    if len(_c_t) >= 14:
                        _sma20_t = sum(_c_t[-20:])/min(20, len(_c_t))
                        _sma50_t = sum(_c_t[-50:])/min(50, len(_c_t))
                        # RSI(14)
                        _g_t, _l_t = 0.0, 0.0
                        for _ri in range(1, 15):
                            if _ri >= len(_c_t): break
                            _ch_t = _c_t[-_ri] - _c_t[-_ri-1]
                            if _ch_t > 0: _g_t += _ch_t
                            else: _l_t -= _ch_t
                        _ag_t, _al_t = _g_t/14, _l_t/14
                        rsi_t = round(100 - 100/(1 + _ag_t/max(_al_t, 0.001)), 1)
                        _cur_t = _c_t[-1]
                        tech_100 = 0
                        if _cur_t > _sma20_t: tech_100 += 30
                        if _cur_t > _sma50_t: tech_100 += 20
                        if 40 <= rsi_t <= 65:   tech_100 += 30
                        elif 30 < rsi_t < 40:   tech_100 += 10
                        elif 65 < rsi_t <= 75:  tech_100 += 15
                        # BB position bonus
                        _bb_m_t = _sma20_t
                        _bb_s_t = (sum((_x-_bb_m_t)**2 for _x in _c_t[-20:])/min(20,len(_c_t)))**0.5
                        _bb_pct_t = (_cur_t - (_bb_m_t - 2*_bb_s_t)) / max(4*_bb_s_t, 0.001)
                        if 0.25 <= _bb_pct_t <= 0.65: tech_100 += 20
                except Exception: tech_100 = 40  # neutral fallback

                # ── Simons Fourier sub-score — is price near a Fourier trough? ───
                simons_100 = quant_100  # base: regime score
                try:
                    fc60_s = fourier_data.get("forecast_60d", [])
                    if fc60_s and price > 0:
                        # Trough proximity score: how close is current price to forecast trough?
                        buy_w_map = {"swing":5, "short":20, "long":60}
                        _bw = buy_w_map.get(inv_type, 15)
                        _buy_sl = [(d_,p_) for d_,p_ in fc60_s[:_bw] if p_>0]
                        if _buy_sl:
                            _trough_p = min(_buy_sl, key=lambda x: x[1])[1]
                            _dist_pct  = abs(price - _trough_p) / price
                            # Closer to trough = higher score
                            if _dist_pct <= 0.02:   simons_100 = min(100, simons_100 + 40)  # within 2%
                            elif _dist_pct <= 0.05: simons_100 = min(100, simons_100 + 20)  # within 5%
                            elif _dist_pct <= 0.10: simons_100 = min(100, simons_100 + 10)
                        r2 = fourier_data.get("r_squared", 0)
                        simons_100 = min(100, simons_100 + int(r2*20))  # higher R² = more reliable cycle
                except Exception: pass

                # ── Sentiment sub-score ───────────────────────────────────────────
                sent_100 = 50  # neutral default
                try:
                    if news_score is not None:
                        if news_score >= 0.25:    sent_100 = 85
                        elif news_score >= 0.08:  sent_100 = 65
                        elif news_score <= -0.25: sent_100 = 15
                        elif news_score <= -0.08: sent_100 = 35
                    if bulk_signal == "BUY":      sent_100 = min(100, sent_100 + 15)
                    elif bulk_signal == "SELL":   sent_100 = max(0,   sent_100 - 15)
                    # ── LLM guidance boost (CACHE ONLY — never triggers live LLM) ──
                    # Live extraction only happens via explicit /api/llm_extract call
                    # or the nightly ingest job. This keeps advisor scans fast.
                    try:
                        from core.llm_extractor import _get_cached
                        _cached = _get_cached(sym)
                        if _cached:
                            _guidance = _cached.get("guidance_direction", "none")
                            _tone     = float(_cached.get("mgmt_tone", 0.0) or 0.0)
                            if _guidance == "raised":
                                sent_100 = min(100, sent_100 + 10)
                            elif _guidance == "lowered":
                                sent_100 = max(0,   sent_100 - 15)
                            # Blend mgmt_tone into sent_100 (max ±8 nudge)
                            sent_100 = max(0, min(100, sent_100 + int(_tone * 8)))
                    except Exception:
                        pass
                except Exception: pass

                # ── TYPE-SPECIFIC GATE — use chart_data.closes (already fetched) ──
                # Soft filters: only reject clearly unsuitable symbols.
                # If chart_data is empty the gates are skipped (fail-open).
                _skip_sym = False
                _gc = chart_data.get("closes", [])  # closes from quant engine

                def _safe_ratio(v, default=10.0):
                    """Parse fund_ratios value — handles '48.4%','2.3x','₹...',float,None."""
                    if v is None or v in ("—","","-"): return default
                    try: return float(v)
                    except (TypeError, ValueError):
                        s = str(v).replace(",","").replace("₹","").replace("x","").replace("%","").strip()
                        try: return float(s)
                        except: return default

                # Single-stock mode: never filter out — always show the report
                _is_single = bool(req_symbols and len(req_symbols) == 1)

                # ── Options Gamma (GEX) calculations (v4.3) ──
                gex_profile = {"max_gamma_wall": None, "zero_gamma_level": None, "skew_ratio": 1.0}
                if _is_single:
                    try:
                        from core.options_engine import fetch_gex_profile
                        gex_profile = fetch_gex_profile(inst.yfinance_symbol, price)
                    except Exception:
                        pass

                # Pre-compute S/R levels + RSI/SMA so all inv_type branches can use them
                import math as _m_early
                _sqp_early   = _m_early.sqrt(price)
                sq9_sup1 = round(max(0.01, _sqp_early - 0.5)**2, 2)
                sq9_sup2 = round(max(0.01, _sqp_early - 1.0)**2, 2)
                sq9_res1 = round((_sqp_early + 0.5)**2, 2)
                sq9_res2 = round((_sqp_early + 1.0)**2, 2)
                sq9_res3 = round((_sqp_early + 1.5)**2, 2)
                sq9_res4 = round((_sqp_early + 2.0)**2, 2)
                all_sup = sorted([r for r in sr_data.get("supports",[])    if r.get("price",0) < price],
                                  key=lambda x: x.get("price",0), reverse=True)
                all_res = sorted([r for r in sr_data.get("resistances",[]) if r.get("price",0) > price],
                                  key=lambda x: x.get("price",0))
                sqp = _sqp_early  # alias used by long/swing trade sections below

                # Pre-compute rsi, sma20, sma50, quant_regime for passes_gate calls
                _cl_pre = chart_data.get("closes", []) if chart_data else []
                sma20 = sum(_cl_pre[-20:]) / min(20, len(_cl_pre)) if len(_cl_pre) >= 2 else price
                sma50 = sum(_cl_pre[-50:]) / min(50, len(_cl_pre)) if len(_cl_pre) >= 2 else price
                rsi   = 50.0
                if len(_cl_pre) >= 15:
                    _rg = [max(0, _cl_pre[-k] - _cl_pre[-k-1]) for k in range(1, 15)]
                    _rl = [max(0, _cl_pre[-k-1] - _cl_pre[-k]) for k in range(1, 15)]
                    _ag = sum(_rg) / 14; _al = sum(_rl) / 14
                    rsi = round(100 - 100 / (1 + _ag / max(_al, 0.001)), 1)
                quant_regime = regime_str  # set above in quant block, default "UNKNOWN"

                if inv_type in ("swing", "intraday") and not _is_single:
                    # ── 10-YEAR PROVEN QUALITY FILTER ─────────────────────────
                    # Source: 2,486 real trades 2016-2026. MARUTI=76.5% WR vs RELIANCE=33%.
                    # Symbol quality > regime > RSI for swing BUY decisions.

                    # Gate 1: Symbol blacklist (10yr data)
                    if sym in BAD_BUY_GOOD_SHORT:
                        _skip_sym = True  # poor BUY symbols — consider SHORT instead

                    # Gate 2: Best 7 symbol whitelist — 63.8% WR filter
                    # If restricting to top symbols only (quality mode), enforce whitelist
                    _swing_quality_mode = True  # set False to allow all non-blacklisted symbols
                    if _swing_quality_mode and sym not in BEST_BUY_SYMBOLS and not _is_single:
                        # Allow non-whitelist symbols at lower signal threshold
                        # (they still fire but need higher conjunction score)
                        pass  # handled in conjunction scoring below

                    # Gate 3: RSI extreme filter (still valid — hard extremes)
                    if len(_gc) >= 15:
                        _sma20_g = sum(_gc[-20:]) / min(20, len(_gc))
                        _g_g, _l_g = 0.0, 0.0
                        for _ri in range(1, 15):
                            if _ri >= len(_gc): break
                            _ch_g = _gc[-_ri] - _gc[-_ri-1]
                            if _ch_g > 0: _g_g += _ch_g
                            else: _l_g -= _ch_g
                        _rsi_g = round(100 - 100/(1 + _g_g/14/max(_l_g/14, 0.001)), 1)
                        if _rsi_g > 85: _skip_sym = True   # extreme overbought only
                        # Note: RSI < 28 removed — oversold = reversal opportunity in 10yr data
                    if news_score is not None and news_score <= -0.40: _skip_sym = True  # only hard bearish news blocked

                elif inv_type == "short" and not _is_single:
                    # SHORT: full Advisory Report gate via passes_gate
                    # Rules 2,4,5,6: Regime×RSI, News, Sq9, Vol, MaxConcurrent
                    _fwd_atr14 = 0.0
                    _fwd_gc = chart_data.get("closes", []) if chart_data else []
                    _fwd_gh = chart_data.get("highs",  []) if chart_data else []
                    _fwd_gl = chart_data.get("lows",   []) if chart_data else []
                    _fwd_gv = chart_data.get("volumes",[]) if chart_data else []
                    if len(_fwd_gh) >= 15:
                        _fwd_atr14 = sum(
                            max(_fwd_gh[-k]-_fwd_gl[-k],
                                abs(_fwd_gh[-k]-_fwd_gc[-k-1]),
                                abs(_fwd_gl[-k]-_fwd_gc[-k-1]))
                            for k in range(1,15)
                        ) / 14
                    _fwd_avg_vol = (sum(_fwd_gv[-11:-1])/10) if len(_fwd_gv) >= 11 else 1
                    _fwd_vol_spike = round((_fwd_gv[-1] if _fwd_gv else _fwd_avg_vol) / max(_fwd_avg_vol,1), 2)
                    _fwd_sq9p = min(abs(price - sq9_sup1)/max(price,1), abs(price - sq9_sup2)/max(price,1))
                    _fwd_sup_px = all_sup[0]["price"] if all_sup else price * 0.97
                    _fwd_fractal_touches = sum(
                        1 for lo in _fwd_gl if abs(lo - _fwd_sup_px)/max(_fwd_sup_px,1) < 0.015
                    )
                    _fwd_news = float(news_score) if news_score not in (None, "N/A") else 0.0
                    _fwd_open_pos = 0
                    try:
                        _fwd_conn = _db()
                        _fwd_open_pos = _fwd_conn.execute(
                            "SELECT COUNT(*) FROM forward_signals WHERE status='OPEN' AND inv_type='short'"
                        ).fetchone()[0]
                        _fwd_conn.close()
                    except Exception: pass
                    _fwd_gate_ok, _fwd_gate_reason = passes_gate(
                        inv_type="short", is_single_stock=False,
                        rsi=rsi, price=price, sma20=sma20, sma50=sma50,
                        regime=quant_regime,
                        vol_spike=_fwd_vol_spike,
                        sq9_proximity=_fwd_sq9p,
                        fractal_touches=_fwd_fractal_touches,
                        news_score=_fwd_news,
                        open_positions=_fwd_open_pos,
                        pe=_safe_ratio((fund_ratios or {}).get("pe_ttm"), 50.0),
                        roe=_safe_ratio((fund_ratios or {}).get("roe"),   10.0),
                    )
                    if not _fwd_gate_ok:
                        _skip_sym = True

                elif not _is_single:  # long — skip gate for single stock
                    # ── LONG: Cycle-phase gate (Cycle Strategy Rules 2, 3, 4, 6, 7) ──

                    # Compute 52-week high/low from price history
                    _52wk_high = 0.0; _52wk_low = 0.0
                    try:
                        _52conn = _db()
                        _date_str_52 = str(analysis_date)
                        _52row = _52conn.execute(
                            "SELECT MAX(high), MIN(low) FROM daily_prices WHERE symbol=? "
                            "AND trade_date >= date(?, '-365 days') AND date <= ?",
                            (inst.symbol, _date_str_52, _date_str_52)
                        ).fetchone()
                        _52conn.close()
                        if _52row and _52row[0]:
                            _52wk_high = float(_52row[0])
                            _52wk_low  = float(_52row[1])
                    except Exception: pass
                    if _52wk_high <= 0: _52wk_high = price * 1.20
                    if _52wk_low  <= 0: _52wk_low  = price * 0.70

                    # Compute Nifty50 ATH gap and RSI for macro cycle filter (Rule 7)
                    _nifty_ath_gap = 0.5  # default: assume not near ATH
                    _nifty_rsi = 50.0
                    try:
                        _nc = _db()
                        _nifty_price_row = _nc.execute(
                            "SELECT close FROM daily_prices WHERE symbol='NIFTY50' "
                            "ORDER BY trade_date DESC LIMIT 1"
                        ).fetchone()
                        _nifty_high_row = _nc.execute(
                            "SELECT MAX(high) FROM daily_prices WHERE symbol='NIFTY50'"
                        ).fetchone()
                        _nc.close()
                        if _nifty_price_row and _nifty_high_row and _nifty_high_row[0]:
                            _np = float(_nifty_price_row[0])
                            _nh = float(_nifty_high_row[0])
                            _nifty_ath_gap = (_nh - _np) / _nh if _nh > 0 else 0.5
                    except Exception: pass

                    # Rule 6: Check Sq9 bounce confirmation using recent closes/volumes
                    _sq9_bounce_ok = False
                    try:
                        from core.gann_math import sq9_bounce_confirmed as _sq9bc
                        _cl_g = chart_data.get("closes", []) if chart_data else []
                        _vl_g = chart_data.get("volumes", []) if chart_data else []
                        _sq9_sup_check = sq9_sup1  # nearest Sq9 support
                        if len(_cl_g) >= 5:
                            _sq9_bounce_ok = _sq9bc(
                                price=price,
                                recent_closes=_cl_g[-15:],
                                recent_volumes=_vl_g[-15:] if _vl_g else [],
                                sq9_level=_sq9_sup_check,
                            )
                    except Exception: pass

                    # Fundamental disqualifiers (still apply)
                    _roe_g = _safe_ratio((fund_ratios or {}).get("roe"),     10.0)
                    _de_g  = _safe_ratio((fund_ratios or {}).get("de_ratio"),  1.0)
                    _pe_g  = _safe_ratio((fund_ratios or {}).get("pe_ttm"),  30.0)
                    _rev_g = _safe_ratio((fund_ratios or {}).get("revenue_growth"), 5.0)

                    # Call passes_gate with full cycle data
                    _long_gate_ok, _long_gate_reason = passes_gate(
                        inv_type="long", is_single_stock=False,
                        rsi=rsi, price=price, sma20=sma20, sma50=sma50,
                        regime=quant_regime,
                        vol_spike=_fwd_vol_spike if '_fwd_vol_spike' in dir() else 1.0,
                        sq9_proximity=_ml_sq9_prox,
                        trend_strength=float(regime_data.get("trend_strength", 0)),
                        price_52wk_high=_52wk_high,
                        price_52wk_low=_52wk_low,
                        nifty_ath_gap=_nifty_ath_gap,
                        nifty_rsi=_nifty_rsi,
                        sq9_bounce_confirmed=_sq9_bounce_ok,
                        pe=_pe_g, roe=_roe_g, de=_de_g, rev_growth=_rev_g,
                    )
                    if not _long_gate_ok:
                        _skip_sym = True
                        _gate_reason = _long_gate_reason

                if _skip_sym: continue

                # ── Nakshatra Alignment Score ──────────────────────────────
                from core.nakshatra_engine import compute_nakshatra_alignment
                nak_align = compute_nakshatra_alignment(
                    symbol=sym,
                    analysis_date=analysis_date,
                    inv_type=inv_type,
                    ruling_planet=inst.ruling_planet,
                    sector=inst.sector
                )
                nak_score = nak_align["nak_score"]

                # ── APPLY TYPE-SPECIFIC WEIGHTS → final 0-100 score ──────────────
                if inv_type in ("swing", "intraday"):
                    # SWING: 10yr-corrected scoring
                    # ML is CONTRARIAN: low ml_direction_prob (DOWN) = BUY reversal signal
                    natal_combined = round((natal_100 + planet_100) / 2, 1)
                    # Vol spike: compute from chart data for compute_score
                    _sw_vol_spike = 1.0
                    if chart_data and len(chart_data.get("volumes", [])) >= 12:
                        _sw_vols = chart_data["volumes"]
                        _sw_avg  = sum(_sw_vols[-11:-1]) / 10
                        _sw_vol_spike = round(_sw_vols[-1] / max(_sw_avg, 1), 2)
                    _sc = compute_score("swing", gann_100=gann_100, technical_100=tech_100,
                                        natal_100=natal_combined, sentiment_100=sent_100,
                                        ml_direction_prob=_ml_dir_prob, ml_confidence=_ml_conf,
                                        sq9_proximity=_ml_sq9_prox, regime=quant_regime,
                                        vol_spike=_sw_vol_spike, nak_score=nak_score,
                                        data_source=qres.get("data_source", "real"),
                                        days_to_trough=_days_to_trough,
                                        ruling_aspect_applying=_ruling_aspect_applying,
                                        bulk_signal=(1.0 if bulk_signal == "BUY" else -1.0 if bulk_signal == "SELL" else 0.0))
                    total = _sc["total"]
                    _sq9_close = _ml_sq9_prox
                    # RR quality check (10yr: block RR > 1.5 for swing)
                    _sw_rr_ok = True
                    if all_res and all_sup:
                        _sw_entry_est = price
                        _sw_t1_est    = all_res[0]["price"] if all_res else price * 1.05
                        _sw_sl_est    = all_sup[0]["price"] if all_sup else price * 0.985
                        _sw_rr = (_sw_t1_est - _sw_entry_est) / max(_sw_entry_est - _sw_sl_est, 0.01)
                        if _sw_rr > 1.75:
                            total = max(0, total - 15)   # penalise wide-target swing trades
                            _sw_rr_ok = False

                elif inv_type == "short":
                    # SHORT: use unified_logic.compute_score (Simons35+Tech25+Gann20+Natal15+Sent5)
                    natal_combined = round((natal_100 + planet_100) / 2, 1)
                    _sc = compute_score("short", gann_100=gann_100, technical_100=tech_100,
                                        simons_100=simons_100, natal_100=natal_combined,
                                        fundamental_100=fund_100, sentiment_100=sent_100,
                                        ml_direction_prob=_ml_dir_prob, ml_confidence=_ml_conf,
                                        regime=quant_regime, nak_score=nak_score,
                                        data_source=qres.get("data_source", "real"),
                                        days_to_trough=_days_to_trough,
                                        ruling_aspect_applying=_ruling_aspect_applying,
                                        bulk_signal=(1.0 if bulk_signal == "BUY" else -1.0 if bulk_signal == "SELL" else 0.0))
                    total = _sc["total"]

                else:  # position / long — CYCLE WAVE scoring (Cycle Strategy Rules 1, 8)
                    # POSITION TRADE: Find where we are in the Accumulation→Distribution wave
                    # Outer planets only for natal (Jupiter/Saturn cycles = years-long)
                    outer_bull = sum(1 for a in bull_signals
                                     if a.get("transit_planet","") in ("Jupiter","Saturn","Rahu","Ketu")
                                     and a.get("orb",99) <= 5)
                    outer_bear = sum(1 for a in bear_signals
                                     if a.get("transit_planet","") in ("Jupiter","Saturn","Rahu","Ketu")
                                     and a.get("orb",99) <= 5)
                    outer_score = min(100, 50 + (outer_bull - outer_bear) * 15)

                    # Collect 52wk data for cycle scoring (may already be set in gate block)
                    _ls_52h = _52wk_high if '_52wk_high' in dir() and _52wk_high > 0 else price * 1.20
                    _ls_52l = _52wk_low  if '_52wk_low'  in dir() and _52wk_low  > 0 else price * 0.70
                    _ls_ts  = float(regime_data.get("trend_strength", 0))
                    _ls_vs  = _fwd_vol_spike if '_fwd_vol_spike' in dir() else 1.0
                    _ls_sq9b = _sq9_bounce_ok if '_sq9_bounce_ok' in dir() else False
                    # Pass sector label so compute_score can apply 8yr sector bonuses
                    _sector_label = inst.sector if hasattr(inst, 'sector') else ""

                    # ── Gann Wave Position: are we near the Accumulation Low? ────
                    # Use multi-year price data (sr_data supports/resistances from quant engine)
                    # Find the DEEPEST support (accumulation zone) and HIGHEST resistance (distribution zone)
                    all_sup_pos = sorted([r.get("price",0) for r in sr_data.get("supports",[]) if r.get("price",0) > 0])
                    all_res_pos = sorted([r.get("price",0) for r in sr_data.get("resistances",[]) if r.get("price",0) > 0])
                    wave_low    = min(all_sup_pos) if all_sup_pos else price * 0.60
                    wave_high   = max(all_res_pos) if all_res_pos else price * 2.00
                    wave_range  = max(wave_high - wave_low, price * 0.10)
                    wave_pos    = (price - wave_low) / wave_range   # 0=at base, 1=at top
                    # Best entry = wave_pos 0-35% (in accumulation or early markup)
                    # Gann wave score: 100 at base, 0 at top
                    gann_wave_score = max(0, min(100, int((1 - wave_pos) * 100)))
                    # Bonus: if price is at a multi-year Sq9 structural support
                    import math as _mpos
                    _sqp_pos = _mpos.sqrt(price)
                    _pos_sq9_sups = [round(max(0.01,_sqp_pos-d)**2,2) for d in [2.0,3.0,4.0,5.0,6.0]]
                    _near_wave_base = any(abs(price-s)/price < 0.05 for s in _pos_sq9_sups)
                    if _near_wave_base: gann_wave_score = min(100, gann_wave_score + 20)

                    # ── Simons dominant cycle — are we in a trough? ─────────────
                    # For position trade, we want the DOMINANT multi-year cycle at trough
                    dom_cycles_pos = fourier_data.get("dominant_cycles", [])
                    cycle_score_pos = 50  # neutral
                    if dom_cycles_pos:
                        dc_pos = dom_cycles_pos[0]  # longest dominant cycle
                        dp_pos = dc_pos.get("days_to_next_trough", 999)
                        # Already in trough or approaching: high score
                        if dp_pos <= 30:    cycle_score_pos = 95   # AT the trough NOW
                        elif dp_pos <= 90:  cycle_score_pos = 80   # approaching trough
                        elif dp_pos <= 180: cycle_score_pos = 65   # within 6 months
                        else:               cycle_score_pos = 35   # far from trough
                        r2_pos = fourier_data.get("r_squared", 0)
                        cycle_score_pos = min(100, cycle_score_pos + int(r2_pos * 15))

                    # Rule 8: Use compute_score with cycle-weighted bonuses/penalties
                    _long_sc = compute_score(
                        "long",
                        gann_100=gann_wave_score,
                        technical_100=tech_100,
                        simons_100=cycle_score_pos,
                        natal_100=outer_score,
                        fundamental_100=fund_100,
                        sentiment_100=50.0,
                        ml_direction_prob=_ml_dir_prob,
                        ml_confidence=_ml_conf,
                        sq9_proximity=_ml_sq9_prox,
                        trend_strength=_ls_ts,
                        price_52wk_high=_ls_52h,
                        price_52wk_low=_ls_52l,
                        price_cur=price,
                        vol_spike=_ls_vs,
                        sq9_bounce_confirmed=_ls_sq9b,
                        regime=quant_regime,
                        nak_score=nak_score,
                        data_source=qres.get("data_source", "real"),
                        days_to_trough=_days_to_trough,
                        ruling_aspect_applying=_ruling_aspect_applying,
                        bulk_signal=(1.0 if bulk_signal == "BUY" else -1.0 if bulk_signal == "SELL" else 0.0),
                    )
                    total = _long_sc["total"]
                    # Outer planet bonus on top (Jupiter/Saturn alignments = major wave turns)
                    if outer_bull > outer_bear:
                        total = min(100, total + (outer_bull - outer_bear) * 3)
                    # WEAK_BULL regime bonus (63-67% WR over 8yr — best entry regime)
                    if quant_regime == "WEAK_BULL":
                        total = min(100, total + 15)
                    # SIDEWAYS + RSI 65-75 combo = 88% WR over 8yr
                    if quant_regime == "SIDEWAYS" and 65 <= rsi <= 75:
                        total = min(100, total + 12)
                    # Accumulation entry (TS<-10): extra confidence from 8yr 63.9% WR
                    if _ls_ts < -10:
                        total = min(100, total + 10)
                    # Sector bonuses/penalties from 8yr backtest data
                    _sec_lbl = (inst.sector if hasattr(inst, 'sector') else "").lower()
                    if 'finance' in _sec_lbl:
                        total = min(100, total + 12)   # Finance 70.6% WR
                    elif 'metal' in _sec_lbl:
                        total = min(100, total + 10)   # Metals 61.3% WR
                    elif 'power' in _sec_lbl:
                        total = min(100, total + 10)   # Power 61.1% WR
                    elif 'fmcg' in _sec_lbl or 'consumer' in _sec_lbl:
                        total = min(100, total + 8)    # FMCG 56.2% WR
                    elif 'oil' in _sec_lbl or 'gas' in _sec_lbl:
                        total = min(100, total + 8)    # Oil & Gas 56.0% WR
                    elif 'insur' in _sec_lbl:
                        total = max(0, total - 12)     # Insurance 38.9% WR
                    elif 'cement' in _sec_lbl:
                        total = max(0, total - 12)     # Cement 38.5% WR
                    elif 'mining' in _sec_lbl or 'coal' in _sec_lbl:
                        total = max(0, total - 10)     # Mining 38.5% WR
                    elif 'bank' in _sec_lbl:
                        total = max(0, total - 5)      # Banking 46.2% WR vs top sectors
                    # Fundamental quality BONUS
                    _roe_b  = _safe_ratio((fund_ratios or {}).get("roe"),  10.0)
                    _rev_b  = _safe_ratio((fund_ratios or {}).get("revenue_growth"), 0.0)
                    _de_b   = _safe_ratio((fund_ratios or {}).get("de_ratio"),  1.0)
                    if _roe_b >= 20 and _rev_b >= 15: total = min(100, total + 8)  # high quality premium
                    if _de_b < 0.3: total = min(100, total + 4)  # debt-free bonus

                total = round(max(0, min(100, total)), 1)

                # Type-specific minimum threshold
                # Single stock: always show even if confidence is low
                THRESHOLDS = {"swing": 38, "short": 35, "long": 36} if not _is_single else {"swing": 0, "short": 0, "long": 0, "position": 0}
                if total < THRESHOLDS.get(inv_type, 40): continue

                # Preserve engine scores for display (scaled back to /25 format)
                gann_sc20   = round(gann_100   * 0.20, 1)   # display as /20
                quant_sc20  = round(simons_100 * 0.20, 1)   # display as /20
                natal_sc20  = round(natal_100  * 0.20, 1)   # display as /20
                planet_sc15 = round(planet_100 * 0.15, 1)   # display as /15

                # ── Entry / SL / Target — DIFFERENT logic per investment type ──
                # (sq9_sup/res, all_sup, all_res already computed above before gate checks)

                # ── SWING TRADE (5-15 days) ────────────────────────────────────
                # Buy: CMP or slight pullback to nearest support (max 1% below CMP)
                # SL:  tight — 1.5-2% below entry (risk: low/bal/high = 1.5/2/3%)
                # T1:  nearest resistance (1-3% above entry)
                # T2:  next resistance (3-6% above entry)
                if inv_type in ("swing", "intraday"):
                    # 10yr data: RR 1.0-1.5 = 54.4% WR. Keep T1 tight.
                    swing_sl_pct = {"low":0.015, "balanced":0.020, "high":0.03}.get(risk_pref, 0.020)
                    swing_t1_pct = {"low":0.030, "balanced":0.040, "high":0.05}.get(risk_pref, 0.040)
                    swing_t2_pct = {"low":0.060, "balanced":0.080, "high":0.10}.get(risk_pref, 0.080)
                    # FIX: Entry = nearest Sq9 support or S/R level below CMP (deterministic for backtesting)
                    # Never use raw CMP — always anchor to a computed price level
                    _swing_sq9_sup = sq9_sup1  # closest sq9 below price
                    if all_sup and all_sup[0].get("price", 0) > price * (1 - swing_sl_pct * 3):
                        # Use nearest S/R support as entry (swing buy just above it)
                        entry = round(all_sup[0]["price"] * 1.001, 2)
                        entry_source = "Sq9 S/R support (swing entry)"
                    elif _swing_sq9_sup > price * (1 - swing_sl_pct * 3):
                        entry = round(_swing_sq9_sup, 2)
                        entry_source = "Sq9 support level (swing entry)"
                    else:
                        # Fallback: price rounded to nearest 0.5 sq9 unit
                        import math as _msw
                        _sqp_sw = _msw.sqrt(price)
                        entry = round(max(0.01, _sqp_sw - 0.25) ** 2, 2)
                        entry_source = "Sq9 nearest price anchor (swing)"
                    # Clamp entry: must not be more than 2% above or below CMP
                    entry = round(max(price * 0.985, min(price * 1.005, entry)), 2)
                    # SL: nearest support below entry or fixed %
                    if all_sup and all_sup[0].get("price",0) > entry * (1 - swing_sl_pct * 2):
                        sl = round(all_sup[0]["price"] * 0.998, 2)
                        sl_source = "Nearest S/R support"
                    else:
                        sl = round(entry * (1 - swing_sl_pct), 2)
                        sl_source = f"Swing SL {swing_sl_pct*100:.1f}%"
                    
                    # T1: nearest resistance AT LEAST swing_t1_pct away
                    _min_t1 = entry * (1 + swing_t1_pct)
                    _valid_res_t1 = [r["price"] for r in all_res if r.get("price", 0) >= _min_t1]
                    if _valid_res_t1:
                        t1 = round(_valid_res_t1[0], 2)
                        t1_source = "Nearest S/R resistance"
                    else:
                        t1 = round(entry * (1 + swing_t1_pct), 2)
                        t1_source = f"Swing T1 {swing_t1_pct*100:.1f}%"
                        
                    # T2: 2nd resistance AT LEAST swing_t2_pct away
                    _min_t2 = entry * (1 + swing_t2_pct)
                    _valid_res_t2 = [r["price"] for r in all_res if r.get("price", 0) >= _min_t2]
                    if len(_valid_res_t2) >= 2:
                        t2 = round(_valid_res_t2[1], 2)
                        t2_source = "2nd S/R resistance"
                    elif len(_valid_res_t2) == 1:
                        t2 = round(_valid_res_t2[0], 2)
                        t2_source = "Next S/R resistance"
                    else:
                        t2 = round(entry * (1 + swing_t2_pct), 2)
                        t2_source = f"Swing T2 {swing_t2_pct*100:.1f}%"

                # ── SHORT TERM (15-45 days) ────────────────────────────────────
                # Buy: slight dip entry (wait for 1-2% pullback, use Fourier trough)
                # SL:  moderate — 3-5% below entry
                # T1:  2nd or 3rd resistance level (5-12% above entry)
                # T2:  4th+ resistance or ATH projection
                elif inv_type == "short":
                    # SHORT TERM — Advisory Report rules applied:
                    # Rule 1: ATR-based SL (1.5×ATR14 instead of fixed 4%)
                    # Rule 3: 60/40 partial exit at T1, trail on 3-day low
                    # Rules 2,4,5,6: already gated in passes_gate above

                    # ── Compute ATR14 for adaptive SL ─────────────────────────
                    _sh_atr14 = 0.0
                    if chart_data and len(chart_data.get("highs", [])) >= 15:
                        _sh_h = chart_data["highs"][-14:]
                        _sh_l = chart_data["lows"][-14:]
                        _sh_cprev = chart_data["closes"][-15:]
                        _sh_atr14 = sum(
                            max(_sh_h[k] - _sh_l[k],
                                abs(_sh_h[k] - _sh_cprev[k]),
                                abs(_sh_l[k]  - _sh_cprev[k]))
                            for k in range(14)
                        ) / 14

                    # ── Compute vol spike for advisory gate ───────────────────
                    _sh_vols = chart_data.get("volumes", []) if chart_data else []
                    _sh_avg_vol = (sum(_sh_vols[-11:-1]) / 10) if len(_sh_vols) >= 11 else 1
                    _sh_cur_vol = _sh_vols[-1] if _sh_vols else _sh_avg_vol
                    _sh_vol_spike = round(_sh_cur_vol / max(_sh_avg_vol, 1), 2)

                    # ── Sq9 proximity + fractal touch count ───────────────────
                    _sh_sq9_prox = min(abs(price - sq9_sup1) / max(price, 1),
                                       abs(price - sq9_sup2) / max(price, 1))
                    _sh_frac_lows = chart_data.get("lows", []) if chart_data else []
                    _sh_sup_px = all_sup[0]["price"] if all_sup else price * 0.97
                    _sh_fractal_touches = sum(
                        1 for lo in _sh_frac_lows
                        if abs(lo - _sh_sup_px) / max(_sh_sup_px, 1) < 0.015
                    )

                    # ── Count open short positions for max_concurrent gate ─────
                    _sh_open_pos = 0
                    try:
                        _sh_conn = _db()
                        _sh_open_pos = _sh_conn.execute(
                            "SELECT COUNT(*) FROM forward_signals WHERE status='OPEN' AND inv_type='short'"
                        ).fetchone()[0]
                        _sh_conn.close()
                    except Exception: pass

                    # ── Advisory gates (only in portfolio scan, not single stock) ─
                    if not _is_single:
                        if inv_type == "long":
                            # Single stock long: compute 52wk range + cycle data for gate
                            _ss_52h = 0.0; _ss_52l = 0.0; _ss_nifty_gap = 0.5; _ss_nifty_rsi = 50.0
                            _ss_sq9b = False; _ss_ts = 0.0
                            try:
                                _ss_conn = _db()
                                _ss_row = _ss_conn.execute(
                                    "SELECT MAX(high), MIN(low) FROM daily_prices WHERE symbol=? "
                                    "AND trade_date >= date(?, '-365 days') AND date <= ?",
                                    (inst.symbol, str(analysis_date), str(analysis_date))
                                ).fetchone()
                                _ss_np = _ss_conn.execute(
                                    "SELECT close FROM daily_prices WHERE symbol='NIFTY50' ORDER BY trade_date DESC LIMIT 1"
                                ).fetchone()
                                _ss_nh = _ss_conn.execute(
                                    "SELECT MAX(high) FROM daily_prices WHERE symbol='NIFTY50'"
                                ).fetchone()
                                _ss_conn.close()
                                if _ss_row and _ss_row[0]:
                                    _ss_52h = float(_ss_row[0]); _ss_52l = float(_ss_row[1])
                                if _ss_np and _ss_nh and _ss_nh[0]:
                                    _snp = float(_ss_np[0]); _snh = float(_ss_nh[0])
                                    _ss_nifty_gap = (_snh - _snp) / _snh if _snh > 0 else 0.5
                            except Exception: pass
                            _ss_ts = float(regime_data.get("trend_strength", 0))
                            try:
                                from core.gann_math import sq9_bounce_confirmed as _sq9bc2
                                _cl_ss = chart_data.get("closes", []) if chart_data else []
                                _vl_ss = chart_data.get("volumes", []) if chart_data else []
                                if len(_cl_ss) >= 5:
                                    _ss_sq9b = _sq9bc2(price=price, recent_closes=_cl_ss[-15:],
                                                       recent_volumes=_vl_ss[-15:] if _vl_ss else [],
                                                       sq9_level=sq9_sup1)
                            except Exception: pass
                            _sh_gate_ok, _sh_gate_reason = passes_gate(
                                inv_type="long", is_single_stock=False,
                                rsi=rsi, price=price, sma20=sma20, sma50=sma50,
                                regime=quant_regime,
                                vol_spike=_sh_vol_spike,
                                sq9_proximity=_sh_sq9_prox,
                                trend_strength=_ss_ts,
                                price_52wk_high=_ss_52h,
                                price_52wk_low=_ss_52l,
                                nifty_ath_gap=_ss_nifty_gap,
                                nifty_rsi=_ss_nifty_rsi,
                                sq9_bounce_confirmed=_ss_sq9b,
                                pe=_safe_ratio((fund_ratios or {}).get("pe_ttm"), 50.0),
                                roe=_safe_ratio((fund_ratios or {}).get("roe"), 10.0),
                                de=_safe_ratio((fund_ratios or {}).get("de_ratio"), 1.0),
                                rev_growth=_safe_ratio((fund_ratios or {}).get("revenue_growth"), 5.0),
                            )
                        else:
                            _sh_gate_ok, _sh_gate_reason = passes_gate(
                                inv_type="short", is_single_stock=False,
                                rsi=rsi, price=price, sma20=sma20, sma50=sma50,
                                regime=quant_regime,
                                vol_spike=_sh_vol_spike,
                                sq9_proximity=_sh_sq9_prox,
                                fractal_touches=_sh_fractal_touches,
                                news_score=float(news_score) if news_score not in (None, "N/A") else 0.0,
                                open_positions=_sh_open_pos,
                            )
                        if not _sh_gate_ok:
                            continue

                    # ── Levels: ATR-based SL via compute_levels ────────────────
                    # ml_rev_price: use reversal_price from ml_result only if ML predicts UP
                    _sh_ml_rev_price = ml_result.get("reversal_price") if ml_result.get("direction") == "UP" else None
                    _sh_ml_conf      = _ml_conf  # already extracted from ml_result above
                    _sh_lvl = compute_levels(
                        inv_type="short", risk_pref=risk_pref, price=price,
                        all_sup=all_sup, all_res=all_res,
                        fourier_buy_price=fourier_buy_price,
                        fourier_buy_date=str(fourier_buy_date) if fourier_buy_date else None,
                        fourier_sell_price=fourier_sell_price,
                        analysis_date=analysis_date,
                        ml_reversal_price=_sh_ml_rev_price,
                        ml_confidence=_sh_ml_conf,
                        atr14=_sh_atr14,
                    )
                    entry        = _sh_lvl["entry"]
                    entry_source = _sh_lvl["entry_src"]
                    sl           = _sh_lvl["sl"]
                    sl_source    = _sh_lvl["sl_src"]
                    t1           = _sh_lvl["t1"]
                    t1_source    = _sh_lvl["t1_src"]
                    t2           = _sh_lvl["t2"]
                    t2_source    = _sh_lvl["t2_src"]

                # ── LONG TERM (3-18 months) ───────────────────────────────────
                # Buy: meaningful dip / accumulation zone (5-15% below CMP)
                # SL:  wide — 8-15% below entry (position sizing handles risk)
                # T1:  major structural resistance (20-50% above entry)
                # T2:  long-term ATH projection or Sq9 extended levels
                elif inv_type in ("hedge_fund", "position"):
                    # ═══════════════════════════════════════════════════════
                    # GANN POSITION TRADE — ABSOLUTE WAVE (Accumulation → Distribution)
                    # Entry: THE Accumulation Zone Low (structural base of the wave)
                    # SL:    Below accumulation structure — wave thesis broken
                    # T1:    50% wave midpoint (Gann 45° retracement target)
                    # T2:    Distribution High = complete wave exit (max profit)
                    # Time:  Gann cycle peak — months to years, not days
                    # ═══════════════════════════════════════════════════════
                    import math as _mhf

                    # ── Step 1: Find Accumulation Zone Low (wave base) ──────────
                    # Use ALL historical supports to find the DEEPEST structural base
                    # This is not the nearest support — it's the ABSOLUTE LOW of the wave
                    _all_sup_sorted = sorted([r.get("price",0) for r in sr_data.get("supports",[])
                                              if r.get("price",0) > 0])
                    _all_res_sorted = sorted([r.get("price",0) for r in sr_data.get("resistances",[])
                                              if r.get("price",0) > 0], reverse=True)

                    # Accumulation zone: lowest 25% of historical support cluster
                    # If price is already near the base — entry at current zone
                    # If price is mid-wave — wait for pullback to nearest wave support
                    _wave_low  = min(_all_sup_sorted) if _all_sup_sorted else price * 0.55
                    _wave_high = max(_all_res_sorted) if _all_res_sorted else price * 2.50
                    _wave_rng  = max(_wave_high - _wave_low, price * 0.20)
                    _wave_pos_pct = (price - _wave_low) / _wave_rng   # 0=base 1=peak

                    # Entry logic: depends on WHERE in the wave we are now
                    if _wave_pos_pct <= 0.30:
                        # IN accumulation zone — buy here (best entry)
                        # Use nearest structural support as exact entry
                        _acc_sups = sorted([r.get("price",0) for r in all_sup
                                            if r.get("price",0) > price * 0.92], reverse=True)
                        if _acc_sups:
                            entry = round(_acc_sups[0] * 1.001, 2)
                            entry_source = f"Accumulation Zone Base ₹{_acc_sups[0]:,.2f} — wave LOW entry"
                        else:
                            entry = round(price * 0.99, 2)
                            entry_source = "At Accumulation Zone — enter at CMP"
                    elif _wave_pos_pct <= 0.55:
                        # EARLY MARKUP phase — still worth entering, wait for pullback
                        _markup_sup = sorted([r.get("price",0) for r in all_sup
                                              if r.get("price",0) > price * 0.88], reverse=True)
                        if _markup_sup:
                            entry = round(_markup_sup[0] * 1.001, 2)
                            entry_source = f"Early Markup pullback entry ₹{_markup_sup[0]:,.2f} (wave {_wave_pos_pct:.0%})"
                        else:
                            entry = round(price * 0.96, 2)  # wait for 4% pullback
                            entry_source = f"Markup phase — enter on 4% pullback (wave {_wave_pos_pct:.0%})"
                    else:
                        # MID or LATE wave — enter only at next major Gann Sq9 support
                        _sqp_hf = _mhf.sqrt(price)
                        _gann_sup = round(max(0.01, _sqp_hf - 2.0) ** 2, 2)  # 2-unit below
                        entry = round(max(price * 0.85, _gann_sup), 2)
                        entry_source = f"Late markup — wait for Gann Sq9 correction to ₹{entry:,.2f} (wave {_wave_pos_pct:.0%})"

                    entry = round(max(price * 0.82, min(price * 1.01, entry)), 2)

                    # ── Step 2: SL — below Accumulation Zone structure ──────────
                    # SL is NOT a percentage — it's below the STRUCTURAL BASE of the wave
                    # If this breaks, the entire wave thesis is invalidated
                    _acc_zone_sl = _wave_low * 0.96  # 4% below absolute wave low
                    _structural_sups = sorted([r.get("price",0) for r in all_sup
                                               if r.get("price",0) < entry * 0.93], reverse=True)
                    if _structural_sups:
                        sl = round(max(_acc_zone_sl, _structural_sups[0] * 0.97), 2)
                        sl_source = f"Below wave structure ₹{_structural_sups[0]:,.2f} — thesis invalidated below this"
                    else:
                        sl = round(max(_acc_zone_sl, entry * 0.88), 2)
                        sl_source = f"Below Accumulation Zone — wave thesis fails below ₹{sl:,.2f}"

                    # ── Step 3: T1 — Wave Midpoint (Gann 45° / 50% retracement) ─
                    # T1 = 50% of the entire Accumulation→Distribution wave
                    # This is where you reduce 50% of position and trail SL
                    _wave_midpoint = round(_wave_low + _wave_rng * 0.50, 2)
                    # Also compute Gann Sq9 levels for distribution zone
                    _sqp_wave_high = _mhf.sqrt(_wave_high)
                    _gann_dist_1   = round((_mhf.sqrt(entry) + 2.0) ** 2, 2)
                    _gann_dist_2   = round((_mhf.sqrt(entry) + 4.0) ** 2, 2)
                    # T1: max of wave midpoint and nearest major resistance above entry+20%
                    _maj_res_t1 = sorted([r.get("price",0) for r in all_res
                                          if r.get("price",0) > entry * 1.20])
                    if _maj_res_t1:
                        t1 = round(max(_wave_midpoint, _maj_res_t1[0]), 2)
                        t1_source = f"Wave midpoint / major S/R ₹{t1:,.2f} — take 50% off here"
                    else:
                        t1 = round(max(_wave_midpoint, _gann_dist_1), 2)
                        t1_source = f"Gann Sq9 wave midpoint ₹{t1:,.2f} — partial exit 50%"

                    # ── Step 4: T2 — Distribution High (complete wave exit) ───────
                    # T2 = THE ABSOLUTE TOP of this wave — maximum profit point
                    # Use the highest historical resistance OR Gann extended Sq9 projection
                    _maj_res_t2 = sorted([r.get("price",0) for r in all_res
                                          if r.get("price",0) > t1 * 1.15], reverse=True)
                    if _maj_res_t2:
                        t2 = round(_maj_res_t2[0], 2)
                        t2_source = f"Distribution Zone High ₹{t2:,.2f} — ABSOLUTE WAVE EXIT"
                    elif _wave_high > t1 * 1.10:
                        t2 = round(_wave_high, 2)
                        t2_source = f"Historical Distribution High ₹{t2:,.2f} — complete wave"
                    else:
                        t2 = round(max(_gann_dist_2, entry * 1.80), 2)
                        t2_source = f"Gann Sq9 wave projection ₹{t2:,.2f} — extended distribution"

                    # Store wave metadata for display
                    _wave_pos_label = ("ACCUMULATION ZONE" if _wave_pos_pct <= 0.30 else
                                       "EARLY MARKUP" if _wave_pos_pct <= 0.55 else
                                       "LATE MARKUP / DISTRIBUTION")

                else:  # long — cycle-aligned targets (Fix 1: T1=16% matched to winner MFE 22.3%)
                    # 8yr data: T1 not hit = 177 trades, 5.1% WR. T1 hit = 169 trades, 100% WR.
                    # Reduce T1 from 30% to 16% — reachable by real cycle winners
                    # Accumulation entries (TS<-10): 13% T1 for faster lock-in
                    _is_acc_adv = _ls_ts < -10 if '_ls_ts' in dir() else False
                    long_sl_pct  = {"low":0.07,  "balanced":0.09, "high":0.12}.get(risk_pref, 0.09)
                    long_t1_pct  = {"low":0.12,  "balanced":0.13 if _is_acc_adv else 0.16, "high":0.20}.get(risk_pref, 0.13 if _is_acc_adv else 0.16)
                    long_t2_pct  = {"low":0.28,  "balanced":0.35, "high":0.50}.get(risk_pref, 0.35)
                    # Entry: Fourier trough only if REALISTIC (within 12% of CMP, within 60 days)
                    _long_fourier_ok = (fourier_buy_price and
                                        fourier_buy_price > price * 0.88 and
                                        fourier_buy_price < price * 1.005 and
                                        fourier_buy_date and
                                        (date.fromisoformat(fourier_buy_date) - analysis_date).days <= 60)
                    if _long_fourier_ok:
                        entry = round(fourier_buy_price, 2)
                        entry_source = f"Simons Fourier cycle trough ₹{entry:,.2f} — accumulation entry"
                    elif all_sup and all_sup[0].get("price",0) > price * 0.90:
                        entry = round(all_sup[0]["price"] * 1.001, 2)
                        entry_source = f"Accumulate at S/R support ₹{all_sup[0]['price']:,.2f}"
                    elif all_sup and all_sup[0].get("price",0) > price * 0.88:
                        entry = round(all_sup[0]["price"] * 1.001, 2)
                        entry_source = f"Major support ₹{all_sup[0]['price']:,.2f} — long-term entry"
                    else:
                        # Use Gann Sq9 as long-term entry anchor — max 8% below CMP
                        _sqp_l = math.sqrt(price)
                        _sq9_long = round(max(0.01, _sqp_l - 1.0)**2, 2)
                        entry = round(max(price * 0.92, _sq9_long), 2)
                        entry_source = f"Gann Sq9 ₹{entry:,.2f} — long-term accumulation zone"
                    # SL: major structural support — wide to avoid noise
                    deep_sup = [r for r in all_sup if r.get("price",0) < entry * (1 - long_sl_pct * 0.5)]
                    if deep_sup:
                        sl = round(deep_sup[0]["price"] * 0.995, 2)
                        sl_source = "Major structural support (long-term SL)"
                    else:
                        sl = round(entry * (1 - long_sl_pct), 2)
                        sl_source = f"Long-term SL {long_sl_pct*100:.0f}% — trend invalidation"
                    # T1: cycle amplitude target — 16% balanced (Fix 1, 8yr data)
                    # Use nearest resistance above entry+10% or config % whichever gives
                    # a realistic target aligned with winner MFE distribution
                    _t1_floor = entry * (1 + long_t1_pct)        # config-based minimum
                    _t1_ceil  = entry * (1 + long_t1_pct * 1.5)  # cap: 1.5x config
                    res_major = [r for r in all_res if r.get("price",0) > entry * 1.10]
                    if res_major and _t1_floor <= res_major[0]["price"] <= _t1_ceil:
                        t1 = round(res_major[0]["price"], 2)
                        t1_source = f"Cycle S/R target ₹{t1:,.2f} — exit 50% here, trail remainder"
                    elif fourier_sell_price and _t1_floor <= fourier_sell_price <= _t1_ceil:
                        t1 = round(fourier_sell_price, 2)
                        t1_source = f"Fourier cycle peak ₹{t1:,.2f} — partial exit"
                    else:
                        t1 = round(_t1_floor, 2)
                        t1_source = f"Cycle T1 {long_t1_pct*100:.0f}% ₹{t1:,.2f} — exit 50%, trail rest"
                    # T2: distribution zone (full cycle target — trail to here, no time stop)
                    res_deep = [r for r in all_res if r.get("price",0) > t1 * 1.12]
                    if res_deep:
                        t2 = round(res_deep[-1]["price"], 2)
                        t2_source = f"Distribution zone ₹{t2:,.2f} — FULL CYCLE EXIT (trail)"
                    elif all_res and all_res[-1].get("price",0) > t1 * 1.05:
                        t2 = round(all_res[-1]["price"], 2)
                        t2_source = f"Structural high ₹{t2:,.2f} — trail to distribution"
                    else:
                        ath_proxy = round((sqp + 2.5)**2, 2)
                        t2 = max(round(entry * (1 + long_t2_pct), 2), ath_proxy)
                        t2_source = f"Cycle T2 {long_t2_pct*100:.0f}% ₹{t2:,.2f} — trail freely"

                # Validate: T1 must be > entry, T2 > T1, SL < entry
                if t1 <= entry: t1 = round(entry * (1 + sl_pct * 3), 2)
                if t2 <= t1:    t2 = round(t1 * (1 + sl_pct * 3), 2)
                if sl >= entry: sl = round(entry * (1 - sl_pct), 2)

                rr_ratio = round((t1 - entry) / max(entry - sl, 0.01), 2)
                risk_amt  = round((entry - sl) / entry * 100, 2)
                # Upside % for display
                upside_t1 = round((t1 - entry) / entry * 100, 1)
                upside_t2 = round((t2 - entry) / entry * 100, 1)

                # ── Gann Reversal: "Price OR Date — whichever comes first" ─
                import math as _math2
                skyBull_count = sum(1 for a in active_aspects if a.get("direction")=="BULLISH")
                skyBear_count = sum(1 for a in active_aspects if a.get("direction")=="BEARISH")

                # ── Dynamic hold_days ──────────────────────────────────────────────
                # Hold is NOT fixed. It is determined by:
                #  1. Gann time cycle expiry AFTER buy date (best signal)
                #  2. Simons Fourier peak date (cycle-based exit)
                #  3. Fallback: type-specific RANGE (not a fixed number)
                # MIN/MAX ranges encode the wave philosophy:
                # Swing: 5–15d (exit earlier if T1 hit, extend if trend continues)
                # Short: 15–45d (Elliott impulse wave 3 duration)
                # Long:  3–18m (90-540d, wave 1-5 full cycle)
                # Position: 3–18m (90-540d, Accumulation→Distribution complete wave)
                HOLD_RANGE = {
                    "intraday": (1,   1),
                    "swing":    (5,   15),
                    "short":    (15,  45),
                    "long":     (90,  540),
                    "position": (90,  540),
                }
                hold_min, hold_max = HOLD_RANGE.get(inv_type, (5, 30))
                hold_days = hold_min   # will be updated by sell date logic below

                # ── BUY trigger: find best FUTURE entry day (NOT always today) ──
                # If sky is bearish today skip it; scan forward for a bullish window.
                PLANET_HOURS = {
                    "Sun":"10:00-11:30 IST","Moon":"09:15-09:45 IST",
                    "Mercury":"11:30-13:00 IST","Venus":"13:00-14:30 IST",
                    "Mars":"09:15-10:30 IST","Jupiter":"10:30-12:00 IST",
                    "Saturn":"14:30-15:30 IST","Rahu":"12:00-13:00 IST","Ketu":"12:00-13:00 IST",
                }
                buy_time_str = PLANET_HOURS.get(inst.ruling_planet, "09:20 IST")

                today_bull_c = sum(1 for a in active_aspects if a.get("direction")=="BULLISH")
                today_bear_c = sum(1 for a in active_aspects if a.get("direction")=="BEARISH")
                today_mal_c  = len([r for r in retrograde if r in ["Saturn","Mars","Rahu","Ketu"]])
                sky_bad_today = (today_bear_c > today_bull_c) or (today_mal_c >= 2)

                scan_start = 1 if sky_bad_today else 0
                max_scan   = {"swing": 10, "short": 30, "long": 90, "position": 90}.get(inv_type, 15)
                buy_date_str   = None
                buy_price      = round(entry, 2)
                buy_condition  = ""
                best_buy_score = -999

                # Get this symbol's natal chart once (for per-symbol natal transit scoring)
                _natal_obj = None
                try: _natal_obj = get_natal(sym)
                except Exception: pass

                try:
                    for fwd in range(scan_start, max_scan + 1):
                        fwd_date = analysis_date + timedelta(days=fwd)
                        if fwd_date.weekday() >= 5: continue
                        fwd_pd   = get_planet_dashboard(fwd_date)
                        fwd_bull = sum(1 for a in fwd_pd.get("aspects",[]) if a.get("direction")=="BULLISH")
                        fwd_bear = sum(1 for a in fwd_pd.get("aspects",[]) if a.get("direction")=="BEARISH")
                        fwd_mal  = len([r for r in fwd_pd.get("retrograde_planets",[])
                                        if r in ["Saturn","Mars","Rahu","Ketu"]])
                        fwd_stat = len(fwd_pd.get("stations",[]))
                        ruling_station = any(s.get("planet") == inst.ruling_planet
                                             for s in fwd_pd.get("stations",[]))
                        cyc_due = any(
                            abs((analysis_date + timedelta(days=c.get("days_remaining",999)) - fwd_date).days) <= 1
                            for c in cycle_data if abs(c.get("days_remaining",999)) <= max_scan * 2)
                        if fwd_bear > fwd_bull + 2 and fwd_mal >= 2 and not cyc_due: continue
                        fwd_score = (fwd_bull - fwd_bear) * 3 + (12 if cyc_due else 0) + fwd_stat * 4 \
                                    + (10 if ruling_station else 0) - fwd_mal * 2
                        # ── Per-symbol natal transit score — unique reversal day per stock ──
                        # Each symbol has its own inception date (natal chart), so transits
                        # to its natal planets peak on different days for each symbol
                        if _natal_obj:
                            try:
                                _fwd_tr = get_transit_to_natal_aspects(_natal_obj, fwd_date)
                                _n_bull = sum(1 for a in _fwd_tr
                                              if a.get("nature")=="BULLISH" and a.get("orb",99)<=4.0)
                                _n_bear = sum(1 for a in _fwd_tr
                                              if a.get("nature")=="BEARISH" and a.get("orb",99)<=4.0)
                                _ruler  = sum(1 for a in _fwd_tr
                                              if a.get("is_ruler_activated") and a.get("orb",99)<=3.0)
                                fwd_score += (_n_bull - _n_bear) * 5 + _ruler * 8
                            except Exception: pass
                        if fwd_score > best_buy_score:
                            best_buy_score = fwd_score
                            buy_date_str   = fwd_date.isoformat()
                            if ruling_station:
                                buy_condition = f"{inst.ruling_planet} stations on {fwd_date.strftime('%d-%b-%Y')} — high-energy reversal entry"
                            elif cyc_due:
                                buy_condition = f"Gann cycle completes on {fwd_date.strftime('%d-%b-%Y')} — cycle-reversal entry"
                            elif _natal_obj and fwd_score > 10:
                                try:
                                    _ra = [a for a in get_transit_to_natal_aspects(_natal_obj, fwd_date)
                                           if a.get("is_ruler_activated") and a.get("orb",99)<=3]
                                    if _ra:
                                        _p = _ra[0].get("transit_planet","") or _ra[0].get("planets","")
                                        buy_condition = f"Natal: {_p} activates {sym} ruler on {fwd_date.strftime('%d-%b-%Y')} — symbol reversal"
                                    else:
                                        buy_condition = f"{sym} natal transits peak on {fwd_date.strftime('%d-%b-%Y')} — symbol-specific entry"
                                except Exception:
                                    buy_condition = f"Best window for {sym} on {fwd_date.strftime('%d-%b-%Y')}"
                            elif fwd_bull > fwd_bear + 1:
                                buy_condition = f"Sky turns bullish ({fwd_bull} BULL vs {fwd_bear} BEAR) on {fwd_date.strftime('%d-%b-%Y')}"
                            else:
                                buy_condition = f"Best available window for {sym} on {fwd_date.strftime('%d-%b-%Y')}"
                        # Don't break early — scan full window to find best day
                        # Only stop early if we found an exceptional score (station + cycle)
                        if best_buy_score >= 35 and fwd >= 2: break
                except Exception: pass

                if buy_date_str is None:
                    _wait = reversal_forecast.get(inv_type, horizon_days)
                    _fdate = analysis_date + timedelta(days=_wait)
                    while _fdate.weekday() >= 5: _fdate += timedelta(days=1)
                    buy_date_str  = _fdate.isoformat()
                    buy_condition = f"Next clear bullish window ~{_wait} trading days away ({_fdate.strftime('%d-%b-%Y')})"

                # ── Reversal dates: scan horizon + 10 extra days ──
                reversal_scan_window = horizon_days + 10
                reversal_dates = []
                try:
                    for fwd in range(1, reversal_scan_window + 1):
                        fwd_date = analysis_date + timedelta(days=fwd)
                        if fwd_date.weekday() >= 5: continue
                        fwd_pd   = get_planet_dashboard(fwd_date)
                        fwd_bull = sum(1 for a in fwd_pd.get("aspects",[]) if a.get("direction")=="BULLISH")
                        fwd_bear = sum(1 for a in fwd_pd.get("aspects",[]) if a.get("direction")=="BEARISH")
                        fwd_stat = len(fwd_pd.get("stations",[]))
                        cyc_due  = any(
                            abs((analysis_date + timedelta(days=c.get("days_remaining",999)) - fwd_date).days) <= 1
                            for c in cycle_data if abs(c.get("days_remaining",999)) <= reversal_scan_window)
                        rev_score = fwd_stat*8 + (10 if cyc_due else 0) + abs(fwd_bull - fwd_bear)*2
                        if rev_score >= 10:
                            reversal_dates.append({
                                "date":    fwd_date.isoformat(),
                                "score":   rev_score,
                                "bull":    fwd_bull,
                                "bear":    fwd_bear,
                                "cycle":   cyc_due,
                                "station": fwd_stat > 0,
                                "bias":    "BULLISH" if fwd_bull > fwd_bear else "BEARISH" if fwd_bear > fwd_bull else "VOLATILE",
                            })
                    reversal_dates.sort(key=lambda x: x["score"], reverse=True)
                    reversal_dates = reversal_dates[:5]
                except Exception: pass

                # ════════════════════════════════════════════════════════════════
                # DYNAMIC SELL DATE — "price OR date, whichever comes first"
                # Priority:
                #  1. Gann time cycle peak in [hold_min, hold_max] window
                #  2. Simons Fourier peak date
                #  3. Gann reversal date (station/conjunction)
                #  4. hold_min + extension if trend confirmation signals exist
                # Never a fixed hold — always driven by the market's own rhythm.
                # ════════════════════════════════════════════════════════════════
                buy_date_obj = date.fromisoformat(buy_date_str)
                days_to_buy  = (buy_date_obj - analysis_date).days
                sell_price       = round(t1, 2)
                sell_price_2     = round(t2, 2)
                sell_time_str    = "15:10 IST — 20 min before close"
                sell_condition2  = (f"T2 ₹{t2:,.2f} — trail SL to cost after T1 hit, "
                                    f"let remaining position run to T2 freely")
                sell_date_str    = None
                sell_condition   = ""
                best_sell_score  = -999

                # ── Pass 1: Gann cycles in the type-appropriate window ──────────
                for cyc in cycle_data:
                    dr = cyc.get("days_remaining", 999)
                    # Cycle must fire AFTER entry date
                    if dr <= days_to_buy: continue
                    actual_hold = dr - days_to_buy
                    # Must be within the valid hold range for this investment type
                    if actual_hold < hold_min: continue
                    if actual_hold > hold_max: continue
                    tgt_date = analysis_date + timedelta(days=dr)
                    while tgt_date.weekday() >= 5: tgt_date += timedelta(days=1)
                    # Score: prefer cycles that match ruling planet and land in mid-range
                    mid_range = (hold_min + hold_max) // 2
                    prox = max(0, 10 - abs(actual_hold - mid_range) // (hold_max // 10 + 1))
                    planet_match = cyc.get("planet","") == inst.ruling_planet
                    cyc_score = prox + (8 if planet_match else 0)
                    if cyc_score > best_sell_score:
                        best_sell_score = cyc_score
                        sell_date_str   = tgt_date.isoformat()
                        hold_days       = actual_hold
                        cyc_label       = cyc.get("label","").split("--")[0].strip()
                        sell_condition  = (
                            f"Exit T1=₹{t1:,.2f} OR {tgt_date.strftime('%d-%b-%Y')} "
                            f"({cyc_label}) — whichever comes first. "
                            f"If above T1 → trail SL, extend to T2."
                        )

                # ── Pass 2: Simons Fourier peak in window ───────────────────────
                if sell_date_str is None and fourier_sell_price and fourier_sell_date:
                    try:
                        _fp_date = date.fromisoformat(fourier_sell_date)
                        _fp_hold = (_fp_date - buy_date_obj).days
                        if hold_min <= _fp_hold <= hold_max:
                            sell_date_str  = _fp_date.isoformat()
                            hold_days      = _fp_hold
                            sell_condition = (
                                f"Exit T1=₹{t1:,.2f} OR Simons cycle peak {_fp_date.strftime('%d-%b-%Y')} "
                                f"(Fourier peak ₹{fourier_sell_price:,.2f}) — whichever first. "
                                f"Trail SL after T1 hit."
                            )
                    except Exception: pass

                # ── Pass 3: Gann reversal dates in window ───────────────────────
                if sell_date_str is None:
                    for rev in reversal_dates:
                        try:
                            _rv_date = date.fromisoformat(rev["date"])
                            _rv_hold = (_rv_date - buy_date_obj).days
                            if hold_min <= _rv_hold <= hold_max and rev.get("score",0) >= 12:
                                sell_date_str  = _rv_date.isoformat()
                                hold_days      = _rv_hold
                                _rv_bias = rev.get("bias","VOLATILE")
                                sell_condition = (
                                    f"Exit T1=₹{t1:,.2f} OR Gann reversal {_rv_date.strftime('%d-%b-%Y')} "
                                    f"({_rv_bias} — score {rev['score']}) — whichever first."
                                )
                                break
                        except Exception: continue

                # ── Pass 4: Dynamic fallback — anchored to T1 price proximity ───
                # Hold until the price is EXPECTED to reach T1 based on momentum,
                # not a calendar date. Estimate from % gain per day in trend.
                if sell_date_str is None:
                    # Estimate days to reach T1 based on recent daily momentum
                    try:
                        _closes_recent = chart_data.get("closes", [])[-20:]
                        if len(_closes_recent) >= 10:
                            # Average daily move over last 10 days
                            _daily_moves = [abs(_closes_recent[k]-_closes_recent[k-1])/_closes_recent[k-1]
                                            for k in range(1,len(_closes_recent))]
                            _avg_daily_pct = sum(_daily_moves)/len(_daily_moves) if _daily_moves else 0.005
                            _gain_needed   = max((t1 - entry) / entry, 0.01)
                            _est_days      = max(hold_min, min(hold_max, int(_gain_needed / max(_avg_daily_pct, 0.001))))
                        else:
                            _est_days = hold_min + (hold_max - hold_min) // 3
                    except Exception:
                        _est_days = hold_min + (hold_max - hold_min) // 3
                    # Snap to next Gann reversal date if within ±5 days of estimate
                    _sd_fallback = buy_date_obj + timedelta(days=_est_days)
                    for rev in reversal_dates:
                        try:
                            _rv2 = date.fromisoformat(rev["date"])
                            if abs((_rv2 - _sd_fallback).days) <= 5 and hold_min <= (_rv2-buy_date_obj).days <= hold_max:
                                _sd_fallback = _rv2
                                break
                        except Exception: pass
                    while _sd_fallback.weekday() >= 5: _sd_fallback += timedelta(days=1)
                    sell_date_str  = _sd_fallback.isoformat()
                    hold_days      = (_sd_fallback - buy_date_obj).days
                    sell_condition = (
                        f"Exit T1=₹{t1:,.2f} OR {_sd_fallback.strftime('%d-%b-%Y')} — whichever comes first. "
                        f"If T1 hit → trail SL to cost, continue to T2. "
                        f"If trend extends → re-assess on each reversal date."
                    )




                # ── Sky-bearish flag (for reasons display only) ──────────────────────
                # We no longer override entry price based on sky — entry is always anchored
                # to a REAL price level close to CMP, not to a distant historical support.
                _sky_bearish_now = (
                    sum(1 for a in active_aspects if a.get("direction") == "BEARISH") >
                    sum(1 for a in active_aspects if a.get("direction") == "BULLISH") + 1
                )
                _days_to_buy = (date.fromisoformat(buy_date_str) - analysis_date).days if buy_date_str else 0
                # Safety clamp: entry must be within 8% of CMP regardless of what was computed above
                _max_entry_gap = {"intraday": 0.005, "swing": 0.02, "short": 0.05, "long": 0.12, "position": 0.15}.get(inv_type, 0.05)
                if entry < price * (1 - _max_entry_gap):
                    entry = round(price * (1 - _max_entry_gap), 2)
                    entry_source = entry_source + f" [clamped to {_max_entry_gap*100:.1f}% below CMP]"
                if entry > price * 1.01:
                    entry = round(price * 1.005, 2)
                    entry_source = entry_source + " [clamped to CMP]"
                # Recompute SL/T1/T2 from clamped entry
                if sl >= entry:
                    sl = round(entry * (1 - sl_pct), 2)
                if t1 <= entry * 1.005:
                    # T1 must be meaningfully above entry
                    _min_t1 = {"intraday":0.005,"swing":0.025,"short":0.06,"long":0.20,"position":0.25}.get(inv_type,0.06)
                    if all_res:
                        _res_above = [r for r in all_res if r.get("price",0) > entry*(1+_min_t1*0.5)]
                        t1 = round(_res_above[0]["price"],2) if _res_above else round(entry*(1+_min_t1),2)
                    else:
                        t1 = round(entry * (1 + _min_t1), 2)
                if t2 <= t1 * 1.01:
                    _min_t2 = {"intraday":0.010,"swing":0.05,"short":0.12,"long":0.40,"position":0.60}.get(inv_type,0.12)
                    if all_res and len(all_res) >= 2:
                        _res_t2 = [r for r in all_res if r.get("price",0) > t1*1.02]
                        t2 = round(_res_t2[0]["price"],2) if _res_t2 else round(entry*(1+_min_t2),2)
                    else:
                        t2 = round(entry * (1 + _min_t2), 2)
                rr_ratio = round((t1 - entry) / max(entry - sl, 0.01), 2)
                risk_amt  = round((entry - sl) / entry * 100, 2)

                # ── Build type-differentiated reasons ────────────────────────
                buy_reasons  = []
                sell_reasons = []
                supports = sr_data.get("supports",[])
                resists  = sr_data.get("resistances",[])
                regime_label = regime_str.replace("_"," ")

                if inv_type in ("swing", "intraday"):
                    # SWING: lead with Gann, then Technical, then Natal timing
                    buy_reasons.append(f"Entry ({entry_source}): ₹{entry:,.2f}")
                    _sq9_dist = round(abs(price - sq9_sup1)/price*100,1)
                    buy_reasons.append(f"Gann Sq9 support ₹{sq9_sup1:,.2f} — {_sq9_dist}% below CMP (key level)")
                    buy_reasons.append(f"Technical: {regime_label} — RSI/SMA momentum confirms swing entry")
                    if bull_signals:
                        b = bull_signals[0]
                        buy_reasons.append(f"Natal: {b['transit_planet']} {b['aspect']} {b['natal_planet']} (orb {b['orb']:.2f}°) — timing signal")
                    for a in angle_data[:1]:
                        if not a.get("above_current",True):
                            buy_reasons.append(f"Gann 1×1 angle support ₹{a['price_at_date']:,.0f} — trend intact")
                    if news_score and news_score > 0.08:
                        buy_reasons.append(f"News: {news_label} ({news_score:+.3f}) — positive flow aids momentum")
                    sell_reasons.append(f"T1 — Nearest Sq9 resistance ₹{t1:,.2f} ({t1_source}) — exit here")
                    sell_reasons.append(f"SL — ₹{sl:,.2f} ({sl_source}) — tight swing stop, breach = exit")
                    sell_reasons.append(f"Time: exit within {hold_days} sessions regardless of outcome")
                    for b in bear_signals[:1]:
                        sell_reasons.append(f"Watch: {b['transit_planet']} {b['aspect']} {b['natal_planet']} — bearish aspect forming")

                elif inv_type == "short":
                    # SHORT: lead with Simons Fourier, then Technical, then Gann, then Natal
                    buy_reasons.append(f"Entry ({entry_source}): ₹{entry:,.2f}")
                    dom_cycles = fourier_data.get("dominant_cycles",[])
                    if dom_cycles:
                        dc = dom_cycles[0]
                        dp = dc.get("days_to_next_trough",0)
                        buy_reasons.append(f"Simons Fourier: dominant {dc.get('gann_label','')} cycle — trough due in {dp}d (buy the dip)")
                    if fourier_buy_price:
                        buy_reasons.append(f"Simons cycle trough price: ₹{fourier_buy_price:,.2f} — accumulate here or below")
                    buy_reasons.append(f"Technical: above SMA50, {regime_label} — medium trend intact for 1–4 week hold")
                    _sq9_zone = f"₹{sq9_sup1:,.2f}–₹{sq9_sup2:,.2f}"
                    buy_reasons.append(f"Gann Sq9 support zone {_sq9_zone} — structural floor")
                    if bull_signals:
                        b = bull_signals[0]
                        buy_reasons.append(f"Natal: {b['transit_planet']} {b['aspect']} {b['natal_planet']} (orb {b['orb']:.2f}°)")
                    if fund_grade and fund_grade not in ("D","F"):
                        buy_reasons.append(f"Fundamental: Grade {fund_grade} — {fund_verdict} (no valuation red flags)")
                    if bulk_signal == "BUY":
                        buy_reasons.append(f"Institutional: Net BUY ₹{bulk_net_val:.1f}Cr (30d) — smart money accumulating")
                    sell_reasons.append(f"T1 ₹{t1:,.2f} ({t1_source}) — EXIT 60% at T1 (lock partial profit)")
                    if fourier_sell_price:
                        sell_reasons.append(f"Simons Fourier peak ₹{fourier_sell_price:,.2f} — primary exit zone")
                    _atr_disp = round(_sh_atr14, 2) if _sh_atr14 > 0 else 0
                    sell_reasons.append(f"SL ₹{sl:,.2f} ({sl_source}) — ATR({_atr_disp})×1.5, adaptive to volatility")
                    sell_reasons.append(f"After T1: trail remaining 40% on 3-day low; T2 timeout 10d post-T1")
                    sell_reasons.append(f"T2 ₹{t2:,.2f} — remaining 40% runs free; SL locked at entry cost (no loss possible)")
                    sell_reasons.append(f"Dead-position cut: exit if MFE<2% and held >{max(5,hold_days//2)}d (v2 rule)")
                    sell_reasons.append(f"Hold estimate: {hold_days}d (exit at T1 OR sell date — whichever comes first)")
                    for b in bear_signals[:2]:
                        sell_reasons.append(f"Watch: {b['transit_planet']} {b['aspect']} {b['natal_planet']} — bearish aspect")
                    if news_score and isinstance(news_score, (int, float)) and news_score < -0.30:
                        sell_reasons.append(f"⚠ Bearish news ({news_score:+.3f}) in {quant_regime} — tighten SL")

                else:  # position / long — ABSOLUTE WAVE reasoning
                    buy_reasons.append(f"Entry ({entry_source}): ₹{entry:,.2f}")
                    # Wave position context
                    _wp_lbl = getattr(locals(),"_wave_pos_label",
                              "ACCUMULATION ZONE" if (price - min([r.get("price",price*0.6)
                              for r in sr_data.get("supports",[{"price":price*0.6}])]))/
                              max((max([r.get("price",price*2) for r in sr_data.get("resistances",[{"price":price*2}])])-
                                   min([r.get("price",price*0.6) for r in sr_data.get("supports",[{"price":price*0.6}])])),1)
                              <= 0.35 else "MARKUP PHASE")
                    buy_reasons.append(f"Wave Position: {_wp_lbl} — buying the structural base, not chasing price")
                    # Fundamental lead — you hold for years
                    _pe_d  = (fund_ratios or {}).get("pe_ttm","N/A")
                    _roe_d = (fund_ratios or {}).get("roe","N/A")
                    _de_d  = (fund_ratios or {}).get("de_ratio","N/A")
                    _rev_d = (fund_ratios or {}).get("revenue_growth","N/A")
                    buy_reasons.append(f"Fundamental: Grade {fund_grade} — {fund_verdict} | P/E {_pe_d} ROE {_roe_d} D/E {_de_d} Rev {_rev_d}")
                    for sig in (fund_signals or [])[:2]:
                        buy_reasons.append(f"Fund: {sig}")
                    # Simons dominant cycle — are we at a cycle trough?
                    dom_cycles = fourier_data.get("dominant_cycles",[])
                    if dom_cycles:
                        dc = dom_cycles[0]
                        dp = dc.get("days_to_next_trough",999)
                        dp_pk = dc.get("days_to_next_peak",999)
                        buy_reasons.append(f"Simons: {dc.get('gann_label','')} dominant cycle — trough in {dp}d, peak in {dp_pk}d — buy the trough")
                    # Gann wave levels
                    buy_reasons.append(f"Gann: SL below wave structure ₹{sl:,.2f} — thesis fails only if this breaks")
                    for a in angle_data[:1]:
                        if not a.get("above_current",True):
                            buy_reasons.append(f"Gann angle support ₹{a['price_at_date']:,.0f} — major trend line")
                    # Outer planet catalysts only
                    for b in bull_signals[:2]:
                        if b.get("transit_planet","") in ("Jupiter","Saturn","Rahu","Ketu"):
                            buy_reasons.append(f"Natal cycle: {b['transit_planet']} {b['aspect']} {b['natal_planet']} (orb {b['orb']:.2f}°) — long wave catalyst")
                    # Exit plan
                    sell_reasons.append(f"T1 ₹{t1:,.2f} ({t1_source}) — EXIT 50% of position here")
                    sell_reasons.append(f"T2 ₹{t2:,.2f} ({t2_source}) — EXIT remaining 50%, this is the WAVE PEAK")
                    sell_reasons.append(f"SL ₹{sl:,.2f} — HARD STOP below wave structure (no averaging down below this)")
                    sell_reasons.append(f"Hold: {hold_days} days estimated — review on Jupiter/Saturn transits")
                    sell_reasons.append(f"Trail: After T1 hit → move SL to entry cost → let T2 run freely")
                    for b in bear_signals[:2]:
                        if b.get("transit_planet","") in ("Jupiter","Saturn"):
                            sell_reasons.append(f"Watch: {b['transit_planet']} {b['aspect']} — long-cycle warning, be alert near T1")

                # ── Planetary influence text ──────────────────────────────
                planet_text = []
                planet_text.append(f"Primary Ruler: {inst.ruling_planet} · Secondary: {inst.secondary_planet}")
                # Simons regime signal advice
                reg_advice = qres.get("regime",{}).get("signal_advice","") if 'qres' in dir() else ""
                if reg_advice:
                    planet_text.append(f"📊 Simons: {reg_advice[:80]}")
                for b in bull_signals[:2]:
                    apl = "↗ Applying" if b.get("applying") else "↘ Separating"
                    planet_text.append(f"✦ {b['transit_planet']} {b['aspect']} Natal {b['natal_planet']} — orb {b['orb']:.2f}° {apl}")
                for asp in active_aspects[:2]:
                    if asp.get("direction") == "BULLISH":
                        planet_text.append(f"✦ Sky: {asp.get('planets','')} — bullish ({asp.get('orb',0):.2f}° orb)")
                for st in stations[:1]:
                    planet_text.append(f"⚡ Station Alert: {st['planet']} {st['direction']} on {st['date']} ({st['days_away']:+d}d)")

                # ── News sentiment score for this symbol ─────────────────
                news_score    = None
                news_label    = "NEUTRAL"
                news_headline = ""
                bulk_signal   = "NEUTRAL"
                bulk_net_val  = 0
                bulk_deals_30d = []
                try:
                    _ns_conn = _db()
                    _ns_rows = _ns_conn.execute("""
                        SELECT COALESCE(calibrated_score, raw_score) as eff_score, label, title, published_at
                        FROM news_sentiment
                        WHERE symbol=? AND published_at >= date(?, '-30 days')
                        ORDER BY published_at DESC LIMIT 10
                    """, (sym, start_date)).fetchall()
                    _ns_conn.close()
                    if _ns_rows:
                        _scores = [float(r[0] or 0) for r in _ns_rows]
                        news_score = round(sum(_scores) / len(_scores), 3)
                        if news_score >= 0.25:   news_label = "BULLISH"
                        elif news_score >= 0.08: news_label = "MILD POSITIVE"
                        elif news_score <= -0.25: news_label = "BEARISH"
                        elif news_score <= -0.08: news_label = "MILD NEGATIVE"
                        else:                     news_label = "NEUTRAL"
                        news_headline = _ns_rows[0][2][:80] if _ns_rows[0][2] else ""
                except Exception: pass

                # ── Bulk/Block deal signal for this symbol ───────────────
                try:
                    _bd_conn = _db()
                    _bd_rows = _bd_conn.execute("""
                        SELECT deal_date, deal_type, quantity, price, deal_kind, client_name
                        FROM bulk_block_deals
                        WHERE symbol=? AND deal_date >= date(?, '-30 days')
                        ORDER BY deal_date DESC
                    """, (sym, start_date)).fetchall()
                    _bd_conn.close()
                    if _bd_rows:
                        _buy_val = sum(float(r[2] or 0)*float(r[3] or 0) for r in _bd_rows if r[1]=="BUY")
                        _sel_val = sum(float(r[2] or 0)*float(r[3] or 0) for r in _bd_rows if r[1]=="SELL")
                        bulk_net_val = round((_buy_val - _sel_val) / 1e7, 2)  # in crores
                        if bulk_net_val > 5:    bulk_signal = "STRONG_BUY"
                        elif bulk_net_val > 1:  bulk_signal = "BUY"
                        elif bulk_net_val < -5: bulk_signal = "STRONG_SELL"
                        elif bulk_net_val < -1: bulk_signal = "SELL"
                        else:                   bulk_signal = "NEUTRAL"
                        bulk_deals_30d = [{"date":r[0],"type":r[1],"qty":r[2],
                                           "price":r[3],"kind":r[4],"client":r[5][:40] if r[5] else ""}
                                          for r in _bd_rows[:5]]
                        # Add to buy/sell reasons
                        if bulk_net_val > 1:
                            buy_reasons.append(f"Bulk/Block: Net institutional BUY ₹{bulk_net_val:.1f} Cr (30d) — accumulation signal")
                        elif bulk_net_val < -1:
                            sell_reasons.append(f"Bulk/Block: Net institutional SELL ₹{abs(bulk_net_val):.1f} Cr (30d) — distribution signal")
                except Exception: pass

                # ── Institutional accumulation score ──────────────────
                inst_adv_score = 0
                inst_adv_signals = []
                try:
                    # Use recent price data from sr_data / price_history
                    _ph = price_history  # already fetched above
                    if len(_ph) >= 12:
                        _pc = [h["close"] for h in _ph]
                        # Get highs/lows from DB
                        _ih_conn = _db()
                        _ih_rows = _ih_conn.execute(
                            "SELECT high, low, volume FROM daily_prices "
                            "WHERE symbol=? AND trade_date<=? AND close IS NOT NULL "
                            "ORDER BY trade_date DESC LIMIT 15", (sym, start_date)).fetchall()
                        _ih_conn.close()
                        if len(_ih_rows) >= 10:
                            _ph_list = list(reversed(_ih_rows))
                            _hh = [float(r[0] or _pc[-1]) for r in _ph_list]
                            _ll = [float(r[1] or _pc[-1]) for r in _ph_list]
                            _vv = [int(r[2] or 0) for r in _ph_list]
                            _cc = _pc[-len(_ph_list):]

                            # Higher lows
                            hl_cnt = sum(1 for k in range(1,len(_ll)) if _ll[k]>_ll[k-1])
                            if hl_cnt/max(len(_ll)-1,1) >= 0.65:
                                inst_adv_score += 30
                                inst_adv_signals.append(f"Higher lows {hl_cnt}/{len(_ll)-1} sessions")

                            # Range tightening
                            _rng = [_hh[k]-_ll[k] for k in range(len(_hh))]
                            if len(_rng) >= 4:
                                import math as _im
                                _re = _im.sqrt(sum(x**2 for x in _rng[:len(_rng)//2])/(len(_rng)//2))
                                _rl = _im.sqrt(sum(x**2 for x in _rng[len(_rng)//2:])/(len(_rng)-len(_rng)//2))
                                if _rl < _re * 0.70:
                                    inst_adv_score += 25
                                    inst_adv_signals.append(f"Range tightening to {_rl/_re*100:.0f}% (coil forming)")

                            # Volume dry-up
                            _ve = sum(_vv[:len(_vv)//2])/max(len(_vv)//2,1)
                            _vl = sum(_vv[len(_vv)//2:])/max(len(_vv)-len(_vv)//2,1)
                            _prng = (max(_cc)-min(_cc))/max(_cc) if _cc else 1
                            if _vl < _ve*0.70 and _prng < 0.04:
                                inst_adv_score += 25
                                inst_adv_signals.append(f"Volume dry-up {_vl/_ve*100:.0f}% with stable price")

                            # Delivery proxy: up-vol vs dn-vol ratio
                            _uv = [_vv[k] for k in range(1,len(_cc)) if len(_cc)>k and _cc[k]>_cc[k-1]]
                            _dv = [_vv[k] for k in range(1,len(_cc)) if len(_cc)>k and _cc[k]<_cc[k-1]]
                            _ur = (sum(_uv)/len(_uv)) / max(sum(_dv)/max(len(_dv),1),1) if _uv else 1
                            if _ur >= 1.5:
                                inst_adv_score += 20
                                inst_adv_signals.append(f"Delivery ratio {_ur:.2f}x (institutions on buy side)")

                            inst_adv_score = min(inst_adv_score, 100)
                except Exception: pass

                # Boost confidence score with institutional signal
                if inst_adv_score >= 70:
                    total = min(100, round(total + 5, 1))
                    buy_reasons.append(f"Institutional accumulation STRONG ({inst_adv_score}/100): {' | '.join(inst_adv_signals[:2])}")
                elif inst_adv_score >= 40:
                    total = min(100, round(total + 2, 1))
                    buy_reasons.append(f"Institutional accumulation MODERATE ({inst_adv_score}/100): {inst_adv_signals[0] if inst_adv_signals else ''}")

                # ── Technical momentum scoring (replaces Wyckoff in v3.9) ─────
                # Uses RSI + MACD-proxy + BB position from price_history
                tech_momentum_score = 0
                tech_momentum_signal = "NEUTRAL"
                try:
                    _ph2 = price_history[-60:] if price_history else []
                    if len(_ph2) >= 20:
                        _c2 = [h["close"] for h in _ph2]
                        # RSI(14)
                        _g2,_l2 = 0.0,0.0
                        for _ri in range(1, min(15, len(_c2))):
                            _ch = _c2[-_ri] - _c2[-_ri-1]
                            if _ch > 0: _g2 += _ch
                            else:       _l2 -= _ch
                        _ag2,_al2 = _g2/14, _l2/14
                        _rsi2 = round(100 - 100/(1+_ag2/max(_al2,0.001)), 1)
                        # SMA20 / SMA50 trend
                        _sma20_2 = sum(_c2[-20:])/20 if len(_c2)>=20 else _c2[-1]
                        _sma50_2 = sum(_c2[-50:])/50 if len(_c2)>=50 else _c2[-1]
                        _cur2    = _c2[-1]
                        # BB position (price within 20-bar Bollinger)
                        _bb_mid  = _sma20_2
                        _bb_std  = (sum((_x-_bb_mid)**2 for _x in _c2[-20:])/20)**0.5
                        _bb_pct  = (_cur2 - (_bb_mid - 2*_bb_std)) / max(4*_bb_std, 0.001)
                        # Score: RSI 40-65 = healthy momentum, above SMA20 & SMA50 = trend
                        if 40 <= _rsi2 <= 65:  tech_momentum_score += 4
                        elif 65 < _rsi2 <= 75: tech_momentum_score += 2  # mildly overbought
                        if _cur2 > _sma20_2:   tech_momentum_score += 3
                        if _cur2 > _sma50_2:   tech_momentum_score += 3  # strong uptrend
                        if 0.3 <= _bb_pct <= 0.7: tech_momentum_score += 2  # healthy BB zone
                        if _rsi2 < 30: tech_momentum_score -= 3  # oversold — avoid
                        if _cur2 < _sma50_2: tech_momentum_score -= 2  # below 50MA
                        tech_momentum_score = max(0, min(10, tech_momentum_score))
                        if tech_momentum_score >= 7: tech_momentum_signal = "BULLISH"
                        elif tech_momentum_score >= 5: tech_momentum_signal = "MODERATE"
                        elif tech_momentum_score <= 2: tech_momentum_signal = "WEAK"
                        if tech_momentum_score >= 5:
                            total = min(100, round(total + tech_momentum_score * 0.8, 1))
                            buy_reasons.append(
                                f"Tech momentum: RSI {_rsi2:.0f} | BB {_bb_pct:.0%} | "                                f"{'Above' if _cur2>_sma20_2 else 'Below'} SMA20 | Score {tech_momentum_score}/10")
                        elif tech_momentum_score <= 2:
                            total = max(0, round(total - 5, 1))
                            sell_reasons.append(f"Tech momentum weak: RSI {_rsi2:.0f} / score {tech_momentum_score}/10")
                except Exception as _tm_err:
                    pass

                # Add news to buy/sell reasons
                if news_score is not None:
                    if news_score >= 0.2:
                        buy_reasons.append(f"News sentiment: {news_label} ({news_score:+.3f}) — positive news flow supports entry")
                    elif news_score <= -0.2:
                        sell_reasons.append(f"News sentiment: {news_label} ({news_score:+.3f}) — negative news flow is a headwind")

                # Append Nakshatra reason to buy reasons if score is high
                if nak_score >= 5:
                    buy_reasons.append(
                        f"Nakshatra: Moon in {nak_align['nakshatra']} ({nak_align['ruler']}) — "
                        f"{inst.sector} sector cosmically favored today (+{nak_score} pts)"
                    )

                # ── Price history for chart ───────────────────────────────
                price_history = []
                try:
                    _pc = _db()
                    _rows = _pc.execute(
                        """SELECT trade_date, close FROM daily_prices
                           WHERE symbol=? AND trade_date<=? AND close IS NOT NULL
                           ORDER BY trade_date DESC LIMIT 90""",
                        (sym, start_date)).fetchall()
                    _pc.close()
                    price_history = [{"date":r[0],"close":round(float(r[1]),2)} for r in reversed(_rows)]
                except Exception: pass

                candidates.append({
                    "symbol":        sym,
                    "nakshatra_info": {
                        "name":          nak_align["nakshatra"],
                        "number":        nak_align["number"],
                        "ruler":         nak_align["ruler"],
                        "guna":          nak_align["guna"],
                        "behavior":      nak_align["behavior"],
                        "trade_style":   nak_align["trade_style"],
                        "favored_today": nak_align["favored_today"],
                        "nak_score":     nak_align["nak_score"],
                        "pada":          nak_align["pada"],
                        "caution":       nak_align["caution"],
                        "rahu_kaal":     nak_align["rahu_kaal"],
                        "abhijit_muhurat": nak_align["abhijit_muhurat"],
                    },
                    "name":          inst.name,
                    "exchange":      inst.exchange,
                    "sector":        inst.sector,
                    "ruling_planet": inst.ruling_planet,
                    "price":         round(price, 2),
                    "entry":         round(entry, 2),
                    "stop_loss":     round(sl, 2),
                    "target1":       round(t1, 2),
                    "target2":       round(t2, 2),
                    "entry_source":  entry_source,
                    "t1_source":     t1_source,
                    "t2_source":     t2_source,
                    "sl_source":     sl_source,
                    "inv_type":      inv_type,
                    "risk_pref":     risk_pref,
                    "upside_t1_pct": upside_t1,
                    "upside_t2_pct": upside_t2,
                    "fourier_buy_price":  fourier_buy_price,
                    "fourier_sell_price": fourier_sell_price,
                    "fourier_buy_date":   fourier_buy_date,
                    "fourier_sell_date":  fourier_sell_date,
                    "regime_metrics":     regime_metrics,
                    "gex_profile":        gex_profile,
                    "dominant_cycle":     (fourier_data.get("dominant_cycles") or [{}])[0],
                    "fourier_r2":         fourier_data.get("r_squared", 0),
                    "hold_days":     hold_days,
                    "buy_date":        buy_date_str,
                    "buy_price":       buy_price,
                    "buy_time":        buy_time_str,
                    "buy_condition":   buy_condition,
                    "sell_date":       sell_date_str,
                    "sell_price":      sell_price,
                    "sell_price_2":    sell_price_2,
                    "sell_time":       sell_time_str,
                    "sell_condition":  sell_condition,
                    "sell_condition2": sell_condition2,
                    "reversal_dates":  reversal_dates,
                    "sky_bull":        skyBull_count,
                    "sky_bear":        skyBear_count,
                    "confidence":    total,
                    "risk_pct":      risk_amt,
                    "rr_ratio":      rr_ratio,
                    "regime":        regime_str,
                    "gann_score":    gann_sc20,
                    "quant_score":   quant_sc20,
                    "natal_score":   natal_sc20,
                    "planet_score":  planet_sc15,
                    "fund_score":    fund_score,
                    "tech_100":      tech_100,
                    "simons_100":    simons_100,
                    "fund_grade":    fund_grade,
                    "fund_verdict":  fund_verdict,
                    "fund_ratios":   fund_ratios,
                    "fund_signals":  fund_signals,
                    "buy_reasons":   buy_reasons[:6],
                    "sell_reasons":  sell_reasons[:5],
                    "planet_text":   planet_text[:5],
                    "bull_signals":  len(bull_signals),
                    "bear_signals":  len(bear_signals),
                    "natal_aspects": natal_aspects[:5],
                    "supports":      [round(s.get("price",0),2) for s in supports[:3]],
                    "resistances":   [round(r.get("price",0),2) for r in resists[:3]],
                    "price_history":   price_history,
                    # ── Sentiment + Bulk/Block (NEW) ──
                    "news_score":      news_score,
                    "news_label":      news_label,
                    "news_headline":   news_headline,
                    "bulk_signal":       bulk_signal,
                    "bulk_net_val_cr":   bulk_net_val,
                    "bulk_deals_30d":    bulk_deals_30d,
                    "inst_acc_score":    inst_adv_score,
                    "inst_acc_signals":  inst_adv_signals,
                    "tech_momentum":     tech_momentum_signal,
                    "tech_score":        tech_momentum_score,
                    # v3.9.1 ML — investment-type-aware predictions (all 3 types)
                    "ml_direction":      ml_result.get("direction", "NEUTRAL"),
                    "ml_confidence":     ml_result.get("confidence", 0.0),
                    "ml_reversal_price": ml_result.get("reversal_price", price),
                    "ml_reversal_date":  ml_result.get("reversal_date", ""),
                    "ml_days_to_rev":    ml_result.get("days_to_reversal", 5),
                    "ml_model_trained":  ml_result.get("model_trained", False),
                    "ml_expected_move":  ml_result.get("expected_move_pct", 0.0),
                    "ml_signal_alignment": ml_result.get("signal_alignment", 0.0),
                    "ml_reversal_map":   ml_result.get("reversal_map", []),
                    # All 3 types available for the UI to show per-tab
                    "ml_swing": {
                        "direction":      ml_result_swing.get("direction", "NEUTRAL"),
                        "confidence":     ml_result_swing.get("confidence", 0.0),
                        "direction_prob": ml_result_swing.get("direction_prob", 0.5),
                        "reversal_price": ml_result_swing.get("reversal_price", price),
                        "reversal_date":  ml_result_swing.get("reversal_date", ""),
                        "days_to_rev":    ml_result_swing.get("days_to_reversal", 5),
                        "expected_move":  ml_result_swing.get("expected_move_pct", 0.0),
                        "signal_alignment": ml_result_swing.get("signal_alignment", 0.0),
                        "model_trained":  ml_result_swing.get("model_trained", False),
                        "reversal_map":   ml_result_swing.get("reversal_map", []),
                        "horizon":        "5–15 days",
                    },
                    "ml_short": {
                        "direction":      ml_result_short.get("direction", "NEUTRAL"),
                        "confidence":     ml_result_short.get("confidence", 0.0),
                        "direction_prob": ml_result_short.get("direction_prob", 0.5),
                        "reversal_price": ml_result_short.get("reversal_price", price),
                        "reversal_date":  ml_result_short.get("reversal_date", ""),
                        "days_to_rev":    ml_result_short.get("days_to_reversal", 5),
                        "expected_move":  ml_result_short.get("expected_move_pct", 0.0),
                        "signal_alignment": ml_result_short.get("signal_alignment", 0.0),
                        "model_trained":  ml_result_short.get("model_trained", False),
                        "reversal_map":   ml_result_short.get("reversal_map", []),
                        "horizon":        "15–45 days",
                    },
                    "ml_long": {
                        "direction":      ml_result_long.get("direction", "NEUTRAL"),
                        "confidence":     ml_result_long.get("confidence", 0.0),
                        "direction_prob": ml_result_long.get("direction_prob", 0.5),
                        "reversal_price": ml_result_long.get("reversal_price", price),
                        "reversal_date":  ml_result_long.get("reversal_date", ""),
                        "days_to_rev":    ml_result_long.get("days_to_reversal", 5),
                        "expected_move":  ml_result_long.get("expected_move_pct", 0.0),
                        "signal_alignment": ml_result_long.get("signal_alignment", 0.0),
                        "model_trained":  ml_result_long.get("model_trained", False),
                        "reversal_map":   ml_result_long.get("reversal_map", []),
                        "horizon":        "3–18 months",
                    },
                })

            # ── Step 5: Sort and select top N ─────────────────────────────
            candidates.sort(key=lambda x: x["confidence"], reverse=True)
            selected = candidates[:min(n_stocks, len(candidates))]

            if not selected:
                return {"error": "No suitable symbols found for current date and filters. Try a different date or risk preference."}

            # ── Step 6: Portfolio allocation (half-Kelly) ─────────────────
            total_conf = sum(c["confidence"] for c in selected)
            for c in selected:
                raw_alloc = (c["confidence"] / max(total_conf,1)) * amount
                raw_alloc = min(raw_alloc, amount * max_alloc)
                c["allocation"]    = round(raw_alloc, 0)
                c["allocation_pct"]= round(raw_alloc / amount * 100, 1)
                c["shares"]        = int(raw_alloc // c["entry"]) if c["entry"] > 0 else 0
                c["risk_amount"]   = round(c["shares"] * (c["entry"] - c["stop_loss"]), 2)
                c["reward_amount"] = round(c["shares"] * (c["target1"] - c["entry"]), 2)

            # Normalize allocations to sum to amount
            alloc_sum = sum(c["allocation"] for c in selected)
            scale = amount / max(alloc_sum, 1)
            for c in selected:
                c["allocation"]     = round(c["allocation"] * scale, 0)
                c["allocation_pct"] = round(c["allocation"] / amount * 100, 1)
                c["shares"]         = int(c["allocation"] // c["entry"]) if c["entry"] > 0 else 0
                c["risk_amount"]    = round(c["shares"] * (c["entry"] - c["stop_loss"]), 2)
                c["reward_amount"]  = round(c["shares"] * (c["target1"] - c["entry"]), 2)

            # ── Step 7: Portfolio-level stats ─────────────────────────────
            total_risk   = sum(c["risk_amount"] for c in selected)
            total_reward = sum(c["reward_amount"] for c in selected)
            avg_conf     = round(sum(c["confidence"] for c in selected) / len(selected), 1)
            portfolio_rr = round(total_reward / max(total_risk, 1), 2)

            # ── FIX: Sector index comparison for each recommendation ──────
            # Maps each stock's sector to its benchmark index for directional context
            SECTOR_INDEX_MAP = {
                "IT": "NIFTYIT", "Technology": "NIFTYIT",
                "Banking": "BANKNIFTY", "Finance": "BANKNIFTY",
                "Pharma": "NIFTYPHARMA", "Healthcare": "NIFTYPHARMA",
                "Auto": "NIFTYAUTO", "Automobile": "NIFTYAUTO",
                "Broad Market": "NIFTY50", "Oil & Gas": "NIFTY50",
                "FMCG": "NIFTY50", "Consumer": "NIFTY50",
                "Metals": "NIFTY50", "Cement": "NIFTY50",
                "Power": "NIFTY50", "Energy": "NIFTY50",
            }
            _idx_prices = {}
            try:
                _all_px = get_cached_prices(analysis_date)
                for _c in selected:
                    _sec = _c.get("sector", "")
                    _idx_sym = SECTOR_INDEX_MAP.get(_sec, "NIFTY50")
                    if _idx_sym not in _idx_prices:
                        _iprow = _all_px.get(_idx_sym, {})
                        _iprice = float(_iprow.get("close") or 0)
                        # Get index price 20 days ago for trend
                        _idx_trend = "SIDEWAYS"
                        try:
                            _ic = _db()
                            _irows = _ic.execute(
                                "SELECT close FROM daily_prices WHERE symbol=? AND trade_date<=? "
                                "AND close IS NOT NULL ORDER BY trade_date DESC LIMIT 21",
                                (_idx_sym, start_date)).fetchall()
                            _ic.close()
                            if len(_irows) >= 5:
                                _iprice_now = float(_irows[0][0])
                                _iprice_20d = float(_irows[min(20, len(_irows)-1)][0])
                                _ichg = (_iprice_now - _iprice_20d) / max(_iprice_20d, 1) * 100
                                if _ichg > 2:   _idx_trend = "BULLISH"
                                elif _ichg < -2: _idx_trend = "BEARISH"
                                else:            _idx_trend = "SIDEWAYS"
                                _idx_prices[_idx_sym] = {"price": round(_iprice_now, 2), "trend": _idx_trend, "chg20d": round(_ichg, 2)}
                        except Exception:
                            _idx_prices[_idx_sym] = {"price": 0, "trend": "UNKNOWN", "chg20d": 0}
                    _idx_info = _idx_prices.get(SECTOR_INDEX_MAP.get(_c.get("sector",""), "NIFTY50"), {})
                    _c["sector_index"]       = SECTOR_INDEX_MAP.get(_c.get("sector", ""), "NIFTY50")
                    _c["sector_index_trend"] = _idx_info.get("trend", "UNKNOWN")
                    _c["sector_index_chg"]   = _idx_info.get("chg20d", 0)
                    # Divergence flag: stock signal vs sector trend
                    _stk_bull = _c.get("confidence", 0) >= 55
                    _sec_bull = _idx_info.get("trend") == "BULLISH"
                    _sec_bear = _idx_info.get("trend") == "BEARISH"
                    if _stk_bull and _sec_bear:
                        _c["sector_divergence"] = "CAUTION: Stock bullish but sector index bearish"
                    elif not _stk_bull and _sec_bull:
                        _c["sector_divergence"] = "NOTE: Stock weak but sector index is rising"
                    else:
                        _c["sector_divergence"] = ""
            except Exception as _se_err:
                pass  # never crash advisor over sector comparison

            return {
                "ok":           True,
                "analysis_date":start_date,
                "inv_type":     inv_type,
                "risk_pref":    risk_pref,
                "amount":       amount,
                "nakshatra_today": nak_today_global,
                "interpreted":  {
                    "amount":   amount,
                    "type":     inv_type,
                    "risk":     risk_pref,
                    "sector":   sector_filter,
                    "symbols":  ",".join(req_symbols) if req_symbols else "",
                    "n_stocks": n_stocks,
                },
                "n_selected":   len(selected),
                "avg_confidence":avg_conf,
                "portfolio_rr": portfolio_rr,
                "total_risk":   round(total_risk, 2),
                "total_reward": round(total_reward, 2),
                "planet_dashboard": {
                    "aspects":   active_aspects[:6],
                    "stations":  stations[:3],
                    "retrograde":retrograde,
                },
                "reversal_forecast":  reversal_forecast,
                "market_overview": {
                    "mkt_bull": mkt_bull,
                    "mkt_bear": mkt_bear,
                    "market_bearish": mkt_regime_bearish,
                    "swing_wait":  reversal_forecast.get("swing", 0),
                    "short_wait":  reversal_forecast.get("short", 0),
                    "long_wait":   reversal_forecast.get("long", 0),
                },
                "recommendations": selected,
            }

        # ══════════════════════════════════════════════════════════════
        # FUNDAMENTAL ANALYSIS
        # ══════════════════════════════════════════════════════════════
        if ep == "shareholding":
            sym = p.get("symbol","").upper()
            try:
                _sc = _db()
                _rows = _sc.execute(
                    "SELECT quarter,fii_pct,dii_pct,promoter_pct,retail_pct,fii_change,dii_change "
                    "FROM shareholding WHERE symbol=? ORDER BY quarter DESC LIMIT 4",
                    (sym,)).fetchall()
                _sc.close()
                if _rows:
                    latest = _rows[0]
                    return {
                        "ok": True, "symbol": sym,
                        "quarter":      latest[0],
                        "fii_pct":      round(float(latest[1] or 0), 2),
                        "dii_pct":      round(float(latest[2] or 0), 2),
                        "promoter_pct": round(float(latest[3] or 0), 2),
                        "retail_pct":   round(float(latest[4] or 0), 2),
                        "fii_change":   round(float(latest[5] or 0), 2),
                        "dii_change":   round(float(latest[6] or 0), 2),
                        "history":      [{"quarter":r[0],"fii":r[1],"dii":r[2],
                                          "promoter":r[3],"retail":r[4]} for r in _rows],
                    }
                return {"ok": False, "symbol": sym, "error": "No shareholding data"}
            except Exception as _e:
                return {"ok": False, "error": str(_e)}

        if ep == "fundamentals":
            sym    = p.get("symbol","").upper()
            force  = p.get("refresh","0") == "1"
            if not sym:
                return {"error": "symbol required"}
            inst = ALL_INSTRUMENTS.get(sym)
            if not inst or inst.instrument_type != "EQUITY":
                return {"error": f"{sym} is not a tracked equity"}
            if _fundamental_engine is None:
                return {"error": "fundamental_engine not available"}

            result = _fundamental_engine.compare_vs_peers(sym, inst.yfinance_symbol)
            # Add fundamental signals
            tgt_data = _fundamental_engine.fetch_fundamentals(sym, inst.yfinance_symbol, force_refresh=force)
            adv      = _fundamental_engine.fundamental_advisor_score(sym, inst.yfinance_symbol)
            result["fundamental_signals"] = adv.get("signals", [])
            result["fetch_date"]          = tgt_data.get("fetch_date","")
            result["source"]              = tgt_data.get("source","")
            return result

        # ── master_report ─────────────────────────────────────────────
        if ep == "master_report":
            import importlib.util as _ilu, os as _os2, traceback as _tb

            sym   = p.get("symbol","").upper()
            req_date = p.get("date", today.isoformat())
            # Clamp future dates to today
            if req_date > today.isoformat():
                req_date = today.isoformat()
            as_of = date.fromisoformat(req_date)
            inst  = ALL_INSTRUMENTS.get(sym)
            if not inst:
                raise ValueError(f"Unknown symbol: {sym}")

            # ── Get current price ──────────────────────────────────
            try:
                _cr = _db()
                _pr = _cr.execute(
                    "SELECT close FROM daily_prices WHERE symbol=? AND trade_date<=? "
                    "AND close IS NOT NULL ORDER BY trade_date DESC LIMIT 1",
                    (sym, req_date)).fetchone()
                _cr.close()
                cur_price = float(_pr[0]) if _pr else float(p.get("price",0) or inst.all_time_high*0.85)
            except Exception:
                cur_price = float(p.get("price",0) or inst.all_time_high*0.85)

            print(f"  [REPORT] {sym} @ ₹{cur_price} on {req_date}", flush=True)

            # ── Quant data ─────────────────────────────────────────
            quant_d = {}
            try:
                backtest = req_date != today.isoformat()
                quant_d  = full_quant_analysis(
                    symbol=sym, yf_symbol=inst.yfinance_symbol,
                    current_price=cur_price, atl=inst.all_time_low,
                    ath=inst.all_time_high, trend_up=True,
                    as_of_date=req_date if backtest else None,
                )
                print(f"  [REPORT] quant OK", flush=True)
            except Exception as _e:
                print(f"  [REPORT] quant FAIL: {_e}", flush=True)
                quant_d = {"regime": {"metrics": {}, "regime": "UNKNOWN", "regime_advice": ""},
                           "chart": {"closes":[],"highs":[],"lows":[],"volumes":[],"dates":[]},
                           "support_resistance": {"supports":[],"resistances":[]}}

            # Ensure chart key exists
            if "chart" not in quant_d:
                quant_d["chart"] = {"closes":[],"highs":[],"lows":[],"volumes":[],"dates":[]}
            if "support_resistance" not in quant_d:
                quant_d["support_resistance"] = {"supports":[],"resistances":[]}

            # Extend chart from DB
            try:
                _conn2 = _db()
                _rows2 = _conn2.execute(
                    "SELECT trade_date,open,high,low,close,volume FROM daily_prices "
                    "WHERE symbol=? AND close IS NOT NULL AND trade_date<=? "
                    "ORDER BY trade_date ASC", (sym, req_date)).fetchall()
                _conn2.close()
                if _rows2:
                    quant_d["chart"]["dates"]   = [r[0] for r in _rows2]
                    quant_d["chart"]["opens"]   = [round(float(r[1] or r[4]),2) for r in _rows2]
                    quant_d["chart"]["highs"]   = [round(float(r[2] or r[4]),2) for r in _rows2]
                    quant_d["chart"]["lows"]    = [round(float(r[3] or r[4]),2) for r in _rows2]
                    quant_d["chart"]["closes"]  = [round(float(r[4]),2) for r in _rows2]
                    quant_d["chart"]["volumes"] = [int(r[5] or 0) for r in _rows2]
                    quant_d["support_resistance"] = find_support_resistance(
                        quant_d["chart"]["closes"],
                        quant_d["chart"]["highs"],
                        quant_d["chart"]["lows"],
                        quant_d["chart"]["volumes"],
                        current_price=cur_price,
                    )
                    print(f"  [REPORT] chart extended: {len(_rows2)} bars", flush=True)
            except Exception as _e:
                print(f"  [REPORT] chart extend FAIL: {_e}", flush=True)

            # ── Gann data ──────────────────────────────────────────
            gann_d = {}
            try:
                pivot_px = cur_price * 0.88
                pivot_dt = as_of - timedelta(days=90)
                gann_d   = analyze_instrument(
                    symbol=sym, current_price=cur_price,
                    pivot_price=pivot_px, pivot_date=pivot_dt,
                    analysis_date=as_of,
                )
                print(f"  [REPORT] gann OK  score={gann_d.get('confluence',{}).get('score','?')}", flush=True)
            except Exception as _e:
                print(f"  [REPORT] gann FAIL: {_e}", flush=True)
                _tb.print_exc()

            # ── Natal data ─────────────────────────────────────────
            natal_d  = None
            try:
                natal_obj = get_natal(sym)
                if natal_obj:
                    natal_aspects = get_transit_to_natal_aspects(natal_obj, as_of)
                    bull_sigs = [a for a in natal_aspects if a.get("nature")=="BULLISH" and a.get("orb",99)<=3]
                    bear_sigs = [a for a in natal_aspects if a.get("nature")=="BEARISH" and a.get("orb",99)<=3]
                    ruler_acts= [a for a in natal_aspects if a.get("is_ruler_activated")]
                    natal_d   = {
                        "bull_signals":      bull_sigs,
                        "bear_signals":      bear_sigs,
                        "ruler_activations": ruler_acts[:5],
                        "transit_to_natal":  natal_aspects[:15],
                        "primary_ruler":     natal_obj.primary_ruler,
                        "secondary_ruler":   natal_obj.secondary_ruler,
                    }
                    print(f"  [REPORT] natal OK  bull={len(bull_sigs)} bear={len(bear_sigs)}", flush=True)
            except Exception as _e:
                print(f"  [REPORT] natal FAIL: {_e}", flush=True)

            # ── Fundamental data ───────────────────────────────────
            fund_d = None
            if inst.instrument_type == "EQUITY":
                try:
                    if _fundamental_engine is not None:
                        fund_d   = _fundamental_engine.compare_vs_peers(sym, inst.yfinance_symbol)
                        _adv2    = _fundamental_engine.fundamental_advisor_score(sym, inst.yfinance_symbol)
                        fund_d["fundamental_signals"] = _adv2.get("signals", [])
                    print(f"  [REPORT] fund OK", flush=True)
                except Exception as _e:
                    print(f"  [REPORT] fund FAIL: {_e}", flush=True)

            # ── Sentiment data ─────────────────────────────────────
            sent_d = None
            try:
                _sc2  = _db()
                _sr2  = _sc2.execute(
                    "SELECT trade_date,open,high,low,close,volume FROM daily_prices "
                    "WHERE symbol=? AND close IS NOT NULL ORDER BY trade_date DESC LIMIT 90",
                    (sym,)).fetchall()
                _sc2.close()
                if len(_sr2) >= 10:
                    _sr2 = list(reversed(_sr2))
                    _cl2 = [float(r[4]) for r in _sr2]
                    _rets2  = [(_cl2[i]-_cl2[i-1])/_cl2[i-1] for i in range(1,len(_cl2))]
                    _vstd2  = (math.sqrt(sum(r**2 for r in _rets2)/len(_rets2))*math.sqrt(252)*100) if _rets2 else 20
                    _v212   = (math.sqrt(sum(r**2 for r in _rets2[-20:])/20)*math.sqrt(252)*100) if len(_rets2)>=20 else _vstd2
                    _vratio2= _v212 / max(_vstd2, 0.001)
                    if _vratio2 > 1.4:    _vlbl2="FEAR — Spike in volatility"
                    elif _vratio2 > 1.15: _vlbl2="CAUTION — Rising volatility"
                    elif _vratio2 < 0.7:  _vlbl2="GREED — Complacency (low vol)"
                    else:                 _vlbl2="CALM — Normal volatility"
                    # Fetch news headlines from DB
                    _news_items = []
                    try:
                        if _sentiment_db is not None:
                            _news_items = _sentiment_db.get_recent_headlines(sym, limit=5)
                    except Exception as _ne:
                        print(f"  [REPORT] news FAIL: {_ne}", flush=True)
                    # Weighted sentiment from news scores
                    _ns2 = 0.0
                    if _news_items:
                        _ns2 = sum(float(n.get("weighted_score",0) or 0) for n in _news_items) / len(_news_items)
                    if _ns2 >= 0.3:    _slbl2="BULLISH"
                    elif _ns2 >= 0.1:  _slbl2="MILDLY POSITIVE"
                    elif _ns2 <= -0.3: _slbl2="BEARISH"
                    elif _ns2 <= -0.1: _slbl2="MILDLY NEGATIVE"
                    else:              _slbl2="NEUTRAL"
                    sent_d = {
                        "overall_score":   round(_ns2, 3),
                        "sentiment_label": _slbl2,
                        "vol_label":       _vlbl2,
                        "news_items":      _news_items[:3],
                        "candle_signals":  [],
                    }
                    print(f"  [REPORT] sentiment OK  score={_ns2:.2f}", flush=True)
            except Exception as _e:
                print(f"  [REPORT] sentiment FAIL: {_e}", flush=True)
                _tb.print_exc()

            # ── Load report engine and generate ───────────────────
            try:
                if _report_engine is None:
                    raise ValueError("report_engine.py not loaded at boot — check core/ directory")
                _rmod = _report_engine
                print(f"  [REPORT] report_engine ready", flush=True)
            except Exception as _e:
                print(f"  [REPORT] report_engine LOAD FAIL: {_e}", flush=True)
                _tb.print_exc()
                raise ValueError(f"report_engine.py missing from core/: {_e}")

            _advisor_action_param = p.get("advisor_action", "")
            # Fetch 52wk range for cycle context in long-type reports
            _r_52h = 0.0; _r_52l = 0.0; _r_ts = 0.0; _r_wvp = 0.5
            try:
                _r_conn = _db()
                _r_row = _r_conn.execute(
                    "SELECT MAX(high), MIN(low) FROM daily_prices WHERE symbol=? "
                    "AND trade_date >= date(?, '-365 days') AND trade_date <= ?",
                    (sym, as_of.isoformat(), as_of.isoformat())
                ).fetchone()
                _r_conn.close()
                if _r_row and _r_row[0]:
                    _r_52h = float(_r_row[0]); _r_52l = float(_r_row[1])
                    _r_rng = _r_52h - _r_52l
                    if _r_rng > 0:
                        _r_wvp = (cur_price - _r_52l) / _r_rng
                if quant_d:
                    _r_ts = float(quant_d.get("regime", {}).get("trend_strength", 0))
            except Exception: pass

            _inv_type_param = p.get("inv_type", "short")
            report = _rmod.generate_master_report(
                sym            = sym,
                cur            = cur_price,
                dt             = as_of,
                quant_data     = quant_d,
                gann_data      = gann_d,
                natal_data     = natal_d,
                fund_data      = fund_d,
                sent_data      = sent_d,
                advisor_action = _advisor_action_param,
                inv_type       = _inv_type_param,
                wave_pos_pct   = _r_wvp,
                price_52wk_high= _r_52h,
                price_52wk_low = _r_52l,
                trend_strength = _r_ts,
            )
            print(f"  [REPORT] done  bias={report.get('trade_setup',{}).get('bias','?')}", flush=True)
            return report

        # ── institutional ─────────────────────────────────────────────
        if ep == "institutional":
            sym  = p.get("symbol","").upper()
            days = int(p.get("days", 365))
            if not sym:
                raise ValueError("symbol required")
            if _fetch_institutional is None:
                return {"error": "fetch_institutional not available"}
            # Auto-compute volume anomalies on first request (fast — local DB only)
            try:
                _fetch_institutional.compute_volume_anomalies(sym)
            except Exception:
                pass
            return _fetch_institutional.get_institutional_data(sym, days=days)

        # ── market_feedback ───────────────────────────────────────────
        if ep == "market_feedback":
            if _market_feedback is None:
                return {"error": "market_feedback not available"}
            action   = p.get("action", "report")   # report | label | train | pipeline
            sym      = p.get("symbol") or None
            if action == "label":
                n = _market_feedback.apply_market_labels(symbol=sym)
                return {"labelled": n, "status": "ok"}
            elif action == "train":
                result = _market_feedback.train_market_model(symbol=sym)
                return result or {"error": "Insufficient data for training"}
            elif action == "pipeline":
                return _market_feedback.run_pipeline(symbol=sym, retrain=True)
            else:   # report
                return _market_feedback.generate_accuracy_report(symbol=sym)

        # ── sentiment_trend ───────────────────────────────────────────
        if ep == "sentiment_trend":
            sym  = p.get("symbol","").upper()
            days = int(p.get("days", 90))
            if not sym: raise ValueError("symbol required")
            if _sentiment_db is None:
                return {"symbol": sym, "days": days, "trend": []}
            trend = _sentiment_db.get_symbol_sentiment_trend(sym, days=days)
            for r in trend:
                r["avg_effective"] = r.get("avg_weighted") or r.get("avg_raw") or 0
            return {"symbol": sym, "days": days, "trend": trend}

        # ── price_history ─────────────────────────────────────────────
        if ep == "price_history":
            sym  = p.get("symbol","").upper()
            days = int(p.get("days", 90))
            if not sym: raise ValueError("symbol required")
            _c  = _db()
            _rows = _c.execute("""
                SELECT trade_date, close FROM daily_prices
                WHERE symbol=? AND close IS NOT NULL
                ORDER BY trade_date DESC LIMIT ?
            """, (sym, days)).fetchall()
            _c.close()
            _rows = list(reversed(_rows))
            return {
                "symbol": sym,
                "dates":  [r[0] for r in _rows],
                "closes": [round(float(r[1]),2) for r in _rows],
            }

        # ── ephemeris_range ──────────────────────────────────────────
        if ep == "ephemeris_range":
            from core.ephemeris import build_ephemeris_range
            start_str = p.get("start_date")
            end_str = p.get("end_date")
            if not start_str or not end_str:
                raise ValueError("start_date and end_date are required")
            start_dt = date.fromisoformat(start_str)
            end_dt = date.fromisoformat(end_str)
            data = build_ephemeris_range(start_dt, end_dt)
            return {"range": data}

        # ══════════════════════════════════════════════════════════════
        # BACKTEST EXCEL EXPORT
        # ══════════════════════════════════════════════════════════════
        if ep == "backtest_export":
            import math as _mbt, io as _io_bt
            # Unified logic stack — no Wyckoff; using Gann+Technical signals

            sym_filter  = p.get("symbol", "").upper() or None
            start_str   = p.get("start_date", "2024-01-01")
            end_str     = p.get("end_date", today.isoformat())
            inv_type_bt = p.get("type", "swing")   # swing | short | long | hedge_fund

            # Phase 4 Walk-Forward Cutoff calculation (70% In-Sample, 30% Out-of-Sample)
            try:
                start_dt = date.fromisoformat(start_str)
                end_dt = date.fromisoformat(end_str)
            except Exception:
                start_dt = date(2024, 1, 1)
                end_dt = today
            total_days = (end_dt - start_dt).days
            cutoff_days = int(total_days * 0.70)
            cutoff_dt = start_dt + timedelta(days=cutoff_days)
            cutoff_str = cutoff_dt.isoformat()

            # ── Type-specific config ───────────────────────────────────────
            # TYPE_CFG aligned with unified_logic.INVESTMENT_TYPES (3 types only)
            # max_bars: max hold in bars matching hold_max from INVESTMENT_TYPES
            # step:     how many bars to skip between signal checks
            # min_rr:   minimum risk:reward to accept a trade
            # lookback: bars of history for S/R / wave calculation
            TYPE_CFG = {
                "intraday": {"max_bars": 1,  "step": 1,  "min_rr": 1.5, "lookback": 5},
                "swing":  {"max_bars": 15,  "step": 3,  "min_rr": 1.0, "lookback": 20},
                "short":  {"max_bars": 45,  "step": 8,  "min_rr": 1.0, "lookback": 40},
                "long":   {"max_bars": 540, "step": 15, "min_rr": 1.0, "lookback": 120},
                # legacy aliases (map to nearest equivalent)
                "position":   {"max_bars": 540, "step": 15, "min_rr": 1.5, "lookback": 120},
                "hedge_fund": {"max_bars": 540, "step": 15, "min_rr": 1.5, "lookback": 120},
            }
            # Normalise any legacy type names
            if inv_type_bt in ("position", "hedge_fund"): inv_type_bt = "long"
            cfg = TYPE_CFG.get(inv_type_bt, TYPE_CFG["swing"])

            _cbt = _db(timeout=10)
            _cbt.row_factory = sqlite3.Row

            syms_to_scan = [sym_filter] if sym_filter else [
                s for s, i in ALL_INSTRUMENTS.items() if i.instrument_type == "EQUITY"
            ]

            # ── Trade simulator — Advisory Report Rules 1,3 applied ────────
            # Rule 1: ATR-based SL already in levels (computed before _simulate)
            # Rule 3: SHORT TERM — 60% exit at T1, trail 40% on 3-day low,
            #         T2 timeout 10 days post-T1
            def _simulate(closes, highs, lows, volumes, levels, i, dates):
                date_to_idx = {d: idx for idx, d in enumerate(dates)}
                entry_i = min(i + 1, len(closes) - 1)
                # Limit order entry fill: get filled at support entry price or next day close, whichever is lower (v4.4)
                actual_entry = min(closes[entry_i], levels["entry"])

                # Scale levels to actual entry price
                scale   = actual_entry / max(levels["entry"], 0.01)
                sl_live = round(levels["sl"] * scale, 2)
                t1_live = round(levels["t1"] * scale, 2)
                t2_live = round(levels["t2"] * scale, 2)

                trailing_sl = sl_live
                t1_hit = t2_hit = trailing_active = False
                t1_hit_bar = None          # bar index when T1 was hit (for T2 timeout)
                partial_exit_done = False   # 60% already booked at T1
                partial_pnl = 0.0          # PnL from the 60% already exited
                exit_date = exit_price = exit_reason = None
                outcome = "OPEN"
                trade_highs = []; trade_lows = []
                max_mfe_seen = 0.0         # track peak favorable move (for dead-position exit)

                for j in range(entry_i + 1, min(entry_i + cfg["max_bars"] + 1, len(closes))):
                    bh = highs[j]; bl = lows[j]; bc = closes[j]
                    trade_highs.append(bh); trade_lows.append(bl)
                    hold = j - entry_i
                    # Track max favorable excursion (for dead-position detection)
                    cur_mfe = (bh - actual_entry) / actual_entry
                    if cur_mfe > max_mfe_seen:
                        max_mfe_seen = cur_mfe

                    # ── T1 Hit ────────────────────────────────────────────────
                    if not t1_hit and bh >= t1_live:
                        t1_hit = True
                        t1_hit_bar = j

                        if inv_type_bt == "short":
                            # Short Term Restructure: 50% exit at T1, trail 50%
                            partial_exit_done = True
                            partial_pnl = round((t1_live - actual_entry) * 0.50, 4)
                            # No progressive trailing SL (keeps SL at breakeven)
                            trailing_active = False
                            # SL must be at minimum at entry cost (can't lose money now)
                            trailing_sl = max(trailing_sl, actual_entry * 1.002)

                        elif inv_type_bt == "swing":
                            # Swing Restructure: 50% exit at T1, trail remaining 50%
                            partial_exit_done = True
                            partial_pnl = round((t1_live - actual_entry) * 0.50, 4)
                            trailing_active = True
                            # SL moved to entry cost to lock in break-even on the rest
                            trailing_sl = max(trailing_sl, actual_entry * 1.002)

                        elif inv_type_bt in ("long", "hedge_fund"):
                            # v4.6 Long restructure: 50% exit at T1, trail 50% with wide 3% stop
                            partial_exit_done = True
                            partial_pnl = round((t1_live - actual_entry) * 0.50, 4)
                            trailing_active = True
                            # SL moves to break-even (entry + 0.3%) — capital protected
                            trailing_sl = max(trailing_sl, actual_entry * 1.003)

                    # ── T2 Hit ────────────────────────────────────────────────
                    if t1_hit and not t2_hit and bh >= t2_live:
                        t2_hit = True
                        if inv_type_bt in ("short", "swing"):
                            # Remaining 50% exits at T2
                            t2_pnl = round((t2_live - actual_entry) * 0.50, 4)
                            blended_pnl = partial_pnl + t2_pnl
                            exit_price = round(actual_entry + blended_pnl, 2)
                            exit_reason = f"T2 Hit (50% at T1 + 50% at T2)"
                        else:
                            exit_price = t2_live
                            exit_reason = "T2 Hit"
                        exit_date = dates[j]
                        outcome = "WIN_T2"; break

                    # ── Dead Position Early Exit (Advisory Report v2, Rule Change 04) ──
                    # If trade held >50% of max hold, MFE never reached 2%, and price is
                    # still below entry → cut the dead position, redeploy capital
                    if (inv_type_bt == "short" and not t1_hit
                            and hold > cfg["max_bars"] * 0.50
                            and max_mfe_seen < 0.02
                            and bc < actual_entry):
                        exit_date   = dates[j]
                        exit_price  = round(bc, 2)
                        exit_reason = (f"Dead position cut: held {hold}d, MFE only "
                                       f"{max_mfe_seen*100:.1f}% — never moved favorably")
                        pnl_dead    = round(bc - actual_entry, 2)
                        outcome     = "LOSS"
                        break

                    # ── T2 Timeout post-T1 ────────────────────────────────────────────
                    # short/swing: 5/15 days post-T1
                    # long: 60 days post-T1 (wider window for multi-month moves)
                    _bt_timeout = 5 if inv_type_bt == "swing" else 60 if inv_type_bt in ("long", "hedge_fund") else 15
                    _bt_rem_pct = 0.50  # all types: 50% remaining after T1 partial exit
                    if t1_hit and not t2_hit and t1_hit_bar and (j - t1_hit_bar) >= _bt_timeout:
                        # Exit remaining portion at current close
                        rem_pnl = round((bc - actual_entry) * _bt_rem_pct, 4)
                        blended_pnl = partial_pnl + rem_pnl
                        exit_price = round(actual_entry + blended_pnl, 2)
                        exit_date  = dates[j]
                        exit_reason = f"T2 timeout ({_bt_timeout}d post-T1) — exit remaining {_bt_rem_pct*100:.0f}% at ₹{bc:.2f}"
                        outcome = "WIN" if exit_price > actual_entry else "LOSS"
                        break

                    # ── Trailing SL update ────────────────────────────────────
                    if trailing_active:
                        if inv_type_bt in ("short", "swing") and t1_hit:
                            if inv_type_bt == "swing":
                                new_trail = round(bc * 0.985, 2)  # tight 1.5% progressive trail for swing
                            else:
                                # 3-day low trail (Advisory Rule 3)
                                _trail_window = trade_lows[-3:] if len(trade_lows) >= 3 else trade_lows
                                new_trail = round(min(_trail_window) * 0.998, 2)
                        elif inv_type_bt in ("long", "hedge_fund") and t1_hit:
                            # v4.6: Wide 5% trailing stop for long-term holds
                            # 3% was too tight for multi-month positions — normal vol is 2-4%/week
                            new_trail = round(bc * 0.950, 2)
                        else:
                            new_trail = round(bc * 0.985, 2)
                        trailing_sl = max(trailing_sl, new_trail)
                        # Never let trailing SL go below entry cost (T1 already hit)
                        if t1_hit:
                            trailing_sl = max(trailing_sl, actual_entry * 1.001)
                    elif inv_type_bt == "hedge_fund" and hold > 8:
                        trailing_active = True
                        trailing_sl = max(trailing_sl, round(bc * 0.982, 2))

                    # ── SL Hit ────────────────────────────────────────────────
                    if bl <= trailing_sl:
                        exit_date = dates[j]; exit_price = trailing_sl
                        if (inv_type_bt in ("short", "swing", "long", "hedge_fund") and t1_hit and partial_exit_done):
                            # partial exit PnL calculation
                            rem_pnl = round((trailing_sl - actual_entry) * _bt_rem_pct, 4)
                            blended_pnl = partial_pnl + rem_pnl
                            exit_price  = round(actual_entry + blended_pnl, 2)
                            exit_reason = (f"Trail SL ₹{trailing_sl:.2f} ({_bt_rem_pct*100:.0f}% remaining — "
                                           f"{(1-_bt_rem_pct)*100:.0f}% locked at T1 ₹{t1_live:.2f})")
                            outcome = "WIN" if exit_price > actual_entry else "LOSS"
                        elif trailing_active:
                            exit_reason = f"Trailing SL ₹{trailing_sl:.2f} ({'T1 locked' if t1_hit else 'progressive trail'})"
                            outcome = "WIN" if t1_hit else "LOSS"
                        else:
                            exit_reason = f"SL Hit ₹{trailing_sl:.2f}"
                            outcome = "LOSS"
                        break

                if not exit_date:
                    last_j = min(entry_i + cfg["max_bars"], len(closes)-1)
                    exit_date  = dates[last_j]
                    bc_last    = closes[last_j]
                    hold_final = last_j - entry_i
                    if t1_hit and partial_exit_done:
                        rem_pnl = round((bc_last - actual_entry) * _bt_rem_pct, 4)
                        blended = partial_pnl + rem_pnl
                        exit_price = round(actual_entry + blended, 2)
                        exit_reason = f"Time exit {hold_final}d ({(1-_bt_rem_pct)*100:.0f}% at T1, {_bt_rem_pct*100:.0f}% at close)"
                    else:
                        exit_price = bc_last
                        exit_reason = f"Time exit {hold_final}d {'(T1 reached, trailed)' if t1_hit else '(no signal hit)'}"
                    outcome = "WIN" if exit_price > actual_entry else "LOSS"

                hold_days = (date_to_idx[exit_date] - entry_i) if exit_date in date_to_idx else cfg["max_bars"]
                pnl_abs   = round(exit_price - actual_entry, 2)
                pnl_pct   = round(pnl_abs / actual_entry * 100, 2)
                actual_rr = round(pnl_abs / max(actual_entry - sl_live, 0.01), 2)

                _mfe_result = mfe_mae(actual_entry, exit_price, trade_highs, trade_lows)

                return {
                    "actual_entry":    round(actual_entry, 2),
                    "sl_live":         round(sl_live, 2),
                    "t1_live":         round(t1_live, 2),
                    "t2_live":         round(t2_live, 2),
                    "trailing_sl":     round(trailing_sl, 2),
                    "trailing_active": trailing_active,
                    "t1_hit":          t1_hit,
                    "t2_hit":          t2_hit,
                    "partial_exit":    partial_exit_done,
                    "partial_pnl":     round(partial_pnl, 4),
                    "exit_price":      round(exit_price, 2),
                    "exit_date":       exit_date,
                    "exit_reason":     exit_reason,
                    "outcome":         outcome,
                    "hold_days":       hold_days,
                    "pnl_abs":         pnl_abs,
                    "pnl_pct":         pnl_pct,
                    "actual_rr":       actual_rr,
                    "mfe_pct":         _mfe_result["mfe_pct"],
                    "mae_pct":         _mfe_result["mae_pct"],
                    "mfe_abs":         _mfe_result["mfe_abs"],
                    "mae_abs":         _mfe_result["mae_abs"],
                    "captured_pct":    _mfe_result["captured_pct"],
                }

            # ── Dynamic level calculator using unified_logic ──────────────────
            # This is the SAME logic as the advisor — guaranteed consistency.
            def _levels(closes, highs, lows, volumes, i, inv_type, atr14=0.0):
                cur = closes[i]
                sqp = _mbt.sqrt(cur)

                # Build S/R lists from fractal analysis (lookback)
                lb = min(i, 120 if inv_type == "swing" else 200)
                h_sr = highs[max(0,i-lb):i+1]; lo_sr = lows[max(0,i-lb):i+1]
                sups_f = []; ress_f = []
                for k in range(5, len(lo_sr)-5):
                    if lo_sr[k] == min(lo_sr[k-5:k+6]): sups_f.append({"price": lo_sr[k]})
                    if h_sr[k]  == max(h_sr[k-5:k+6]):  ress_f.append({"price": h_sr[k]})
                # Sort nearest first
                all_sup_l = sorted([s for s in sups_f if s["price"] < cur*0.998], key=lambda x: x["price"], reverse=True)
                all_res_l = sorted([r for r in ress_f if r["price"] > cur*1.002], key=lambda x: x["price"])

                # Wave position for long type
                lb_wv = min(i, 250)
                wv_lo = min(lows[max(0,i-lb_wv):i+1]) if lb_wv > 10 else cur*0.70
                wv_hi = max(highs[max(0,i-lb_wv):i+1]) if lb_wv > 10 else cur*1.50
                wv_rng = max(wv_hi - wv_lo, cur*0.05)
                wv_pos = (cur - wv_lo) / wv_rng

                # Trend strength for cycle-phase SL/T1 adaptation (8yr: TS<-10 = 63.9% WR)
                _bt_ts = 0.0
                if lb_wv > 20:
                    _c20  = closes[max(0,i-20):i+1]
                    _sma10 = sum(_c20[-10:]) / 10 if len(_c20) >= 10 else cur
                    _sma20 = sum(_c20) / len(_c20) if _c20 else cur
                    _bt_ts = ((_sma10 - _sma20) / _sma20) * 100 if _sma20 > 0 else 0.0

                # ML reversal for level computation (use cached prediction if available)
                _ml_rev_price = None; _ml_conf_bt = 0.0
                try:
                    if len(closes) >= 60:
                        _ml_bt = predict_reversal(
                            closes=closes[max(0,i-100):i+1],
                            highs=highs[max(0,i-100):i+1],
                            lows=lows[max(0,i-100):i+1],
                            volumes=volumes[max(0,i-100):i+1],
                            current_price=cur,
                            symbol=_sym,
                        )
                        _ml_rev_price = _ml_bt.get("reversal_price")
                        _ml_conf_bt   = _ml_bt.get("confidence", 0.0)
                        _ml_dir_bt    = _ml_bt.get("direction", "NEUTRAL")
                        # Only use ML reversal price if direction matches expected (BUY)
                        if _ml_dir_bt != "UP": _ml_rev_price = None
                except Exception: pass

                # Use unified_logic.compute_levels (atr14 = ATR-based SL for short)
                # Use the trend_strength already computed in gate block for consistency
                _bt_ts_for_levels = _bt_ts_live if '_bt_ts_live' in dir() else _bt_ts
                lvl = compute_levels(
                    inv_type=inv_type, risk_pref="balanced",
                    price=cur, all_sup=all_sup_l, all_res=all_res_l,
                    ml_reversal_price=_ml_rev_price, ml_confidence=_ml_conf_bt,
                    wave_pos_pct=wv_pos, wave_low=wv_lo, wave_high=wv_hi,
                    atr14=atr14,
                    trend_strength=float(_bt_ts_for_levels),
                )
                return {
                    "entry":    lvl["entry"],     "entry_src": lvl["entry_src"],
                    "sl":       lvl["sl"],        "sl_src":    lvl["sl_src"],
                    "t1":       lvl["t1"],        "t1_src":    lvl["t1_src"],
                    "t2":       lvl["t2"],        "t2_src":    lvl["t2_src"],
                    "risk":     lvl["risk"],      "reward1":   lvl["reward1"],
                    "reward2":  lvl["reward2"],
                    "rr1":      lvl["rr_ratio"],  "rr2":       lvl["rr_ratio2"],
                    "spring_low": wv_lo, "range_high": wv_hi,
                }



            # ── Main scan ──────────────────────────────────────────────────
            all_trades = []
            for _sym in syms_to_scan:
                inst_bt = ALL_INSTRUMENTS.get(_sym)
                if not inst_bt: continue

                _px = _cbt.execute("""
                    SELECT trade_date,open,high,low,close,volume
                    FROM daily_prices WHERE symbol=? AND trade_date>=? AND trade_date<=?
                    AND close IS NOT NULL ORDER BY trade_date ASC
                """, (_sym, start_str, end_str)).fetchall()
                if len(_px) < 25: continue

                dates_bt  = [r[0] for r in _px]
                opens_bt  = [float(r[1] or r[4]) for r in _px]
                closes_bt = [float(r[4]) for r in _px]
                highs_bt  = [float(r[2]) for r in _px]
                lows_bt   = [float(r[3]) for r in _px]
                vols_bt   = [int(r[5] or 0) for r in _px]

                i = max(25, cfg["lookback"])
                while i < len(dates_bt) - 5:
                    cur_px = closes_bt[i]
                    analysis_dt = dates_bt[i]

                    # ── Unified signal logic (Gann+Technical, v3.9 — no Wyckoff) ─────
                    # Common indicators computed for all investment types
                    sma20 = sum(closes_bt[max(0,i-20):i+1]) / min(20, i+1)
                    sma50 = sum(closes_bt[max(0,i-50):i+1]) / min(50, i+1)
                    sma200= sum(closes_bt[max(0,i-200):i+1]) / min(200, i+1)

                    # RSI(14) — used by all modes
                    rtail = closes_bt[max(0,i-14):i+1]
                    rsi = 50.0
                    if len(rtail) >= 15:
                        g,l=0.0,0.0
                        for ri in range(1,len(rtail)):
                            ch=rtail[ri]-rtail[ri-1]
                            if ch>0:g+=ch
                            else:l-=ch
                        ag,al=g/14,l/14
                        rsi=round(100-100/(1+ag/al),1) if al>0 else 100.0

                    # Gann Sq9 proximity check — is price near a Sq9 level?
                    import math as _mbt2
                    _sqp_bt = _mbt2.sqrt(cur_px)
                    _sq9_sups_bt = [round(max(0.01,_sqp_bt-d)**2,2) for d in [0.25,0.5,1.0]]
                    _sq9_ress_bt = [round((_sqp_bt+d)**2,2) for d in [0.25,0.5,1.0]]
                    _near_sq9_sup = any(abs(cur_px-s)/cur_px < 0.015 for s in _sq9_sups_bt)
                    _near_sq9_res = any(abs(cur_px-r)/cur_px < 0.015 for r in _sq9_ress_bt)

                    # Bollinger Band position
                    _bb_m = sma20
                    _bb_s = (sum((closes_bt[max(0,i-19+k)]- _bb_m)**2 for k in range(min(20,i+1)))/min(20,i+1))**0.5
                    _bb_pct_bt = (cur_px-(_bb_m-2*_bb_s))/max(4*_bb_s,0.001)

                    # Tech signal: regime indicator (replaces Wyckoff regime)
                    _above_200  = cur_px > sma200
                    _above_50   = cur_px > sma50
                    _above_20   = cur_px > sma20
                    _tech_regime= ("BULL" if (_above_200 and _above_50) else
                                   "WEAK_BULL" if _above_50 else
                                   "BEAR" if not _above_200 else "SIDEWAYS")

                    # Avoid buying swing/short setups in bearish regimes (v4.5 Restructure)
                    if inv_type_bt in ("swing", "short") and _tech_regime in ("BEAR", "STRONG_BEAR"):
                        i += cfg["step"]; continue

                    # v4.6: Only block long entries in extreme STRONG_BEAR (allows mild BEAR — contrarian)
                    if inv_type_bt in ("long", "hedge_fund") and _tech_regime == "STRONG_BEAR":
                        i += cfg["step"]; continue

                    # ── Signal conditions — UNIFIED LOGIC (matches advisor exactly) ──
                    # Same gates as unified_logic.passes_gate() for consistency

                    # ── Signal gate: unified_logic.passes_gate() + ML augmentation ──
                    # Same gates as advisor — guaranteed consistency across all modules
                    # Compute ATR14 for adaptive SL
                    _bt_atr14 = 0.0
                    if i >= 14:
                        _bt_atr14 = sum(
                            max(highs_bt[i-k] - lows_bt[i-k],
                                abs(highs_bt[i-k] - closes_bt[i-k-1]),
                                abs(lows_bt[i-k]  - closes_bt[i-k-1]))
                            for k in range(1, 15)
                        ) / 14

                    # Vol spike
                    _bt_avg_vol = (sum(vols_bt[max(0,i-10):i]) / max(10, 1)) if i > 0 else 1
                    _bt_vol_spike = round(vols_bt[i] / max(_bt_avg_vol, 1), 2)

                    # Sq9 proximity
                    import math as _mbt2
                    _bt_sqp = _mbt2.sqrt(cur_px)
                    _bt_sq9s1 = round(max(0.01, _bt_sqp - 0.5)**2, 2)
                    _bt_sq9s2 = round(max(0.01, _bt_sqp - 1.0)**2, 2)
                    _bt_sq9_prox = min(abs(cur_px - _bt_sq9s1)/max(cur_px,1),
                                       abs(cur_px - _bt_sq9s2)/max(cur_px,1))

                    # Fractal touch count
                    _bt_lb_lows = lows_bt[max(0,i-120):i]
                    _bt_fractal_sup = min(_bt_lb_lows) if _bt_lb_lows else cur_px * 0.97
                    _bt_fractal_touches = sum(
                        1 for lo in _bt_lb_lows
                        if abs(lo - _bt_fractal_sup) / max(_bt_fractal_sup,1) < 0.015
                    )

                    # News score (already fetched in this loop? — use 0.0 as default for gate)
                    _bt_news_score = 0.0
                    try:
                        _bt_ns = _cbt.execute("""
                            SELECT AVG(COALESCE(calibrated_score, raw_score, 0))
                            FROM news_sentiment
                            WHERE symbol=? AND published_at >= date(?,'-15 days') AND published_at <= ?
                        """, (_sym, analysis_dt, analysis_dt)).fetchone()
                        _bt_news_score = float(_bt_ns[0]) if _bt_ns and _bt_ns[0] else 0.0
                    except Exception: pass

                    # For long-type backtest: compute cycle params from available history
                    _bt_52h = max(highs_bt[max(0,i-252):i+1]) if i > 20 else cur_px * 1.2
                    _bt_52l = min(lows_bt[max(0,i-252):i+1])  if i > 20 else cur_px * 0.7
                    # Trend strength from SMA10 vs SMA20 (fast approximation)
                    _bt_sma10 = sum(closes_bt[max(0,i-10):i+1]) / min(10, i+1)
                    _bt_sma20b = sum(closes_bt[max(0,i-20):i+1]) / min(20, i+1)
                    _bt_ts_live = ((_bt_sma10 - _bt_sma20b) / _bt_sma20b) * 100 if _bt_sma20b > 0 else 0.0

                    # ── 5-Condition Accumulation Score (required for long gate) ──
                    # compute_acc_score was removed from app.py in v3.9 diff but
                    # never moved anywhere — restored in unified_logic.py.
                    _bt_acc_score = 0
                    _bt_sq9_bounce = False
                    if inv_type_bt == "long":
                        try:
                            import math as _mbt_acc
                            _bt_sq9p = _mbt_acc.sqrt(cur_px)
                            _bt_sq9_sup = round(max(0.01, _bt_sq9p - 0.5) ** 2, 2)
                            _cl_acc = closes_bt[max(0, i-252):i+1]
                            _vl_acc = vols_bt[max(0, i-252):i+1]
                            _hi_acc = highs_bt[max(0, i-252):i+1]
                            _lo_acc = lows_bt[max(0, i-252):i+1]
                            _bt_acc_score = compute_acc_score(
                                closes=_cl_acc, highs=_hi_acc,
                                lows=_lo_acc, volumes=_vl_acc,
                                price=cur_px,
                                price_52wk_high=_bt_52h,
                                price_52wk_low=_bt_52l,
                                rsi=rsi,
                            )
                            # Sq9 bounce confirmation
                            from core.gann_math import sq9_bounce_confirmed as _sq9bc_bt
                            if len(_cl_acc) >= 5:
                                _bt_sq9_bounce = _sq9bc_bt(
                                    price=cur_px,
                                    recent_closes=_cl_acc[-15:],
                                    recent_volumes=_vl_acc[-15:] if _vl_acc else [],
                                    sq9_level=_bt_sq9_sup,
                                )
                        except Exception: pass

                    # Swing low % for long gate D
                    _bt_swing_low_pct = 0.0
                    if inv_type_bt == "long" and i >= 20:
                        _swl = min(lows_bt[max(0, i-20):i+1])
                        _bt_swing_low_pct = (cur_px - _swl) / cur_px * 100 if cur_px > 0 else 0.0

                    _bt_gate_ok, _bt_gate_reason = passes_gate(
                        inv_type=inv_type_bt,
                        is_single_stock=False,
                        rsi=rsi,
                        price=cur_px, sma20=sma20, sma50=sma50,
                        regime=_tech_regime,
                        vol_spike=_bt_vol_spike,
                        sq9_proximity=_bt_sq9_prox,
                        fractal_touches=_bt_fractal_touches,
                        news_score=_bt_news_score,
                        open_positions=0,
                        # Long-type cycle params (Fix 2 — consistency with advisor)
                        trend_strength=_bt_ts_live,
                        price_52wk_high=_bt_52h,
                        price_52wk_low=_bt_52l,
                        nifty_ath_gap=0.5,    # neutral: backtest doesn't have live NIFTY ATH
                        nifty_rsi=50.0,        # neutral default
                        sq9_bounce_confirmed=_bt_sq9_bounce,
                        acc_score=_bt_acc_score,
                        swing_low_pct=_bt_swing_low_pct,
                        symbol=_sym,
                        ruling_planet=inst_bt.ruling_planet,
                    )
                    if not _bt_gate_ok:
                        i += cfg["step"]; continue

                    # ── ML signal augmentation (v3.9) ───────────────────────────
                    # Predict direction from all available features at this bar
                    _ml_bt_sig = {"direction":"NEUTRAL","confidence":0.0,"direction_prob":0.5,
                                  "reversal_prob":0.5,"signal_alignment":0.5}
                    try:
                        if len(closes_bt) >= 60:
                            _ml_bt_sig = predict_reversal(
                                closes=closes_bt[max(0,i-100):i+1],
                                highs=highs_bt[max(0,i-100):i+1],
                                lows=lows_bt[max(0,i-100):i+1],
                                volumes=vols_bt[max(0,i-100):i+1],
                                current_price=cur_px,
                                symbol=_sym,
                            )
                    except Exception: pass

                    # If ML model is trained and has high confidence opposing our signal → skip
                    _ml_dir_bt = _ml_bt_sig.get("direction", "NEUTRAL")
                    _ml_conf_bt = _ml_bt_sig.get("confidence", 0.0)
                    if _ml_conf_bt > 0.75 and _ml_dir_bt == "DOWN":
                        i += cfg["step"]; continue   # ML strongly disagrees → skip this bar

                    # ── Calculate unified advisor confidence score inside backtest (v4.4) ──
                    gann_100 = 85.0 if _near_sq9_sup else 40.0
                    tech_100 = 50.0
                    if _tech_regime in ("BULL", "STRONG_BULL"):
                        tech_100 = 80.0
                    elif _tech_regime == "WEAK_BULL":
                        tech_100 = 65.0
                    elif _tech_regime == "BEAR":
                        tech_100 = 35.0
                    if 45 <= rsi <= 55:
                        tech_100 = min(100.0, tech_100 + 15.0)

                    simons_100 = 50.0
                    if _bb_pct_bt < 0.35:
                        simons_100 = 75.0

                    natal_100 = 50.0
                    try:
                        from datetime import date as _dt_bt3
                        _dt_obj = _dt_bt3.fromisoformat(analysis_dt) if isinstance(analysis_dt, str) else analysis_dt
                        _bt_pd = get_planet_dashboard(_dt_obj)
                        _bt_bull_c = sum(1 for a in _bt_pd.get("aspects", []) if a.get("direction") == "BULLISH")
                        _bt_bear_c = sum(1 for a in _bt_pd.get("aspects", []) if a.get("direction") == "BEARISH")
                        natal_100 = min(100.0, max(0.0, 50.0 + (_bt_bull_c - _bt_bear_c) * 12.0))
                    except Exception:
                        natal_100 = 60.0

                    fund_grade = getattr(inst_bt, "fundamental_grade", "C")
                    fund_100 = {"A": 90.0, "B": 80.0, "C": 60.0, "D": 40.0, "F": 20.0}.get(fund_grade, 60.0)

                    sent_100 = 50.0
                    if _bt_news_score > 0.08:
                        sent_100 = 80.0
                    elif _bt_news_score < -0.08:
                        sent_100 = 30.0

                    _bt_nak_score = 0.0
                    try:
                        from core.nakshatra_engine import compute_nakshatra_alignment
                        _dt_obj = _dt_bt3.fromisoformat(analysis_dt) if isinstance(analysis_dt, str) else analysis_dt
                        _bt_nak = compute_nakshatra_alignment(
                            symbol=_sym,
                            analysis_date=_dt_obj,
                            inv_type=inv_type_bt,
                            ruling_planet=inst_bt.ruling_planet,
                            sector=inst_bt.sector
                        )
                        _bt_nak_score = _bt_nak.get("nak_score", 0.0)
                    except Exception: pass

                    # Fetch bulk block deals signal early (v4.4)
                    _bd = _cbt.execute("""
                        SELECT deal_type, SUM(quantity*price)
                        FROM bulk_block_deals WHERE symbol=? AND deal_date >= date(?,'-20 days')
                        AND deal_date <= ? GROUP BY deal_type
                    """, (_sym, analysis_dt, analysis_dt)).fetchall()
                    buy_v  = sum(float(r[1] or 0) for r in _bd if r[0]=="BUY")
                    sell_v = sum(float(r[1] or 0) for r in _bd if r[0]=="SELL")
                    bd_net = round((buy_v-sell_v)/1e7, 2)
                    bd_sig = "BUY" if bd_net>1 else "SELL" if bd_net<-1 else "NEUTRAL"

                    _bt_sc = compute_score(
                        inv_type_bt,
                        gann_100=gann_100,
                        technical_100=tech_100,
                        simons_100=simons_100,
                        natal_100=natal_100,
                        fundamental_100=fund_100,
                        sentiment_100=sent_100,
                        ml_direction_prob=_ml_bt_sig.get("direction_prob", 0.5),
                        ml_confidence=_ml_conf_bt,
                        sq9_proximity=_bt_sq9_prox,
                        regime=_tech_regime,
                        nak_score=_bt_nak_score,
                        data_source="real",
                        days_to_trough=999,
                        ruling_aspect_applying=False,
                        bulk_signal=(1.0 if bd_sig == "BUY" else -1.0 if bd_sig == "SELL" else 0.0)
                    )

                    # FILTER BY HIGH CONVICTION THRESHOLD (v4.4/v4.6 restructure)
                    # swing/short: 50.0 (strong signal quality required)
                    # long: 45.0 (relaxed — quality pullback approach, gates do the filtering)
                    _bt_threshold = 50.0 if inv_type_bt in ("swing", "short") else 45.0
                    if _bt_sc["total"] < _bt_threshold:
                        i += cfg["step"]; continue

                    signal_ok = True

                    # Tech regime dict for reasons (replaces Wyckoff dict)
                    reg_d = {"regime":_tech_regime, "tradeable":_above_50,
                             "reason":f"Price {'above' if _above_50 else 'below'} SMA50",
                             "trend_strength": round(min(100,(cur_px/sma200-1)*100),1)}
                    # Breakout detection (replaces Wyckoff breakout)
                    _prev_high = max(highs_bt[max(0,i-5):i]) if i>0 else cur_px
                    _atr5 = (sum(abs(highs_bt[max(0,i-5+k)]-lows_bt[max(0,i-5+k)]) for k in range(min(5,i+1)))/max(5,1)) if i>0 else cur_px*0.02
                    _is_breakout = cur_px > _prev_high and vols_bt[i] > sum(vols_bt[max(0,i-10):i])/max(10,1)*1.5 if i>0 else False
                    brkout = {"score":80 if _is_breakout else 20, "vol_spike":vols_bt[i]/max(sum(vols_bt[max(0,i-10):i])/max(10,1),1),
                              "range_expansion":(highs_bt[i]-lows_bt[i])/_atr5,"close_strength":_bb_pct_bt,
                              "signals":[],"is_breakout":_is_breakout}

                    if not signal_ok:
                        i += cfg["step"]; continue

                    # ── Dynamic levels ────────────────────────────────────
                    lvl = _levels(closes_bt, highs_bt, lows_bt, vols_bt, i, inv_type_bt,
                                      atr14=_bt_atr14 if inv_type_bt == "short" else 0.0)
                    if lvl["rr1"] < cfg["min_rr"]:
                        i += cfg["step"]; continue

                    # ── Institutional + news data ──────────────────────────
                    try:
                        _ns = _cbt.execute("""
                            SELECT AVG(COALESCE(calibrated_score, raw_score, 0)) FROM news_sentiment
                            WHERE symbol=? AND published_at >= date(?,'-15 days') AND published_at <= ?
                        """, (_sym, analysis_dt, analysis_dt)).fetchone()
                        news_sc = round(float(_ns[0] or 0), 3) if _ns and _ns[0] else None
                    except Exception:
                        try:
                            _ns = _cbt.execute("""
                                SELECT AVG(raw_score) FROM news_sentiment
                                WHERE symbol=? AND published_at >= date(?,'-15 days') AND published_at <= ?
                            """, (_sym, analysis_dt, analysis_dt)).fetchone()
                            news_sc = round(float(_ns[0] or 0), 3) if _ns and _ns[0] else None
                        except Exception:
                            news_sc = None

                    # Bulk deals already fetched early in the loop (v4.4)

                    # ── Dynamic exit plan (unified_logic) ──────────────────
                    from datetime import date as _dt_bt
                    _buy_dt_bt  = _dt_bt.fromisoformat(dates_bt[min(i+1, len(dates_bt)-1)])
                    _anal_dt_bt = _dt_bt.fromisoformat(analysis_dt)
                    _exit_plan  = compute_exit_plan(
                        inv_type=inv_type_bt,
                        entry=lvl["entry"], t1=lvl["t1"], t2=lvl["t2"],
                        buy_date=_buy_dt_bt, analysis_date=_anal_dt_bt,
                        chart_closes=closes_bt[max(0, i-20):i+1],
                    )

                    # ── Simulate trade ─────────────────────────────────────
                    tr = _simulate(closes_bt, highs_bt, lows_bt, vols_bt, lvl, i, dates_bt)

                    # ── Build reasons ──────────────────────────────────────
                    buy_reas = []
                    # ML signal first if model is trained
                    if _ml_bt_sig.get("model_trained") and _ml_conf_bt > 0.60:
                        buy_reas.append(f"ML: {_ml_dir_bt} conf {_ml_conf_bt:.0%} — rev ₹{_ml_bt_sig.get('reversal_price',0):,.2f}")
                    buy_reas.append(f"Tech Regime: {reg_d['regime']} — {reg_d['reason'][:60]}")
                    buy_reas.append(f"RSI: {rsi:.1f} | SMA20: ₹{sma20:.2f} | SMA50: ₹{sma50:.2f}")
                    buy_reas.append(f"BB Position: {_bb_pct_bt:.0%} | Gann Sq9 near: {'Yes' if _near_sq9_sup else 'No'}")
                    buy_reas.append(f"Entry: {lvl['entry_src']}")
                    buy_reas.append(f"SL: {lvl['sl_src']} | T1: {lvl['t1_src']}")
                    if news_sc and news_sc > 0.1: buy_reas.append(f"News bullish {news_sc:+.3f}")
                    if bd_sig == "BUY": buy_reas.append(f"Bulk deal buy ₹{bd_net:.1f}Cr")
                    if brkout["is_breakout"]: buy_reas.append(f"Breakout: vol {brkout['vol_spike']:.2f}x + range {brkout['range_expansion']:.2f}x ATR")

                    sell_reas = []
                    sell_reas.append(tr["exit_reason"])
                    sell_reas.append(f"Risk ₹{lvl['risk']:.2f} | Reward1 ₹{lvl['reward1']:.2f} | R:R {lvl['rr1']:.2f}:1")
                    if tr["trailing_active"]: sell_reas.append(f"Trailing SL at ₹{tr['trailing_sl']:.2f}")
                    if news_sc and news_sc < -0.1: sell_reas.append(f"News headwind {news_sc:+.3f}")

                    all_trades.append({
                        "Symbol":              _sym,
                        "Name":                inst_bt.name,
                        "Sector":              inst_bt.sector,
                        "Investment Type":     ("Hedge Fund" if inv_type_bt=="hedge_fund"
                                                else inv_type_bt.capitalize()),
                        "Validation Regime":   "IN-SAMPLE" if dates_bt[min(i+1,len(dates_bt)-1)] < cutoff_str else "OUT-OF-SAMPLE",
                        "Tech Regime":         reg_d["regime"],
                        "RSI at Entry":         rsi,
                        "Analysis Date":        analysis_dt,
                        "Entry Date":          dates_bt[min(i+1,len(dates_bt)-1)],
                        "Entry Price (₹)":     tr["actual_entry"],
                        "Stop Loss (₹)":       tr["sl_live"],
                        "Target 1 (₹)":        tr["t1_live"],
                        "Target 2 (₹)":        tr["t2_live"],
                        "SL Source":           lvl["sl_src"],
                        "T1 Source":           lvl["t1_src"],
                        "Risk (₹)":            lvl["risk"],
                        "Reward 1 (₹)":        lvl["reward1"],
                        "Reward 2 (₹)":        lvl["reward2"],
                        "R:R (T1)":            lvl["rr1"],
                        "R:R (T2)":            lvl["rr2"],
                        "Exit Date":           tr["exit_date"],
                        "Exit Price (₹)":      tr["exit_price"],
                        "Exit Reason":         tr["exit_reason"],
                        "Actual Hold Days":    tr["hold_days"],
                        "Expected Hold Days":  _exit_plan.get("hold_days", tr["hold_days"]),
                        "Exit Source":         _exit_plan.get("exit_source", "simulated"),
                        "Trail Rule":          _exit_plan.get("trail_rule", "")[:120],
                        "Trailing SL (₹)":     tr["trailing_sl"],
                        "Trailing Active":     "Yes" if tr["trailing_active"] else "No",
                        "T1 Achieved":         "Yes" if tr["t1_hit"] else "No",
                        "T2 Achieved":         "Yes" if tr["t2_hit"] else "No",
                        "P&L (₹)":             tr["pnl_abs"],
                        "P&L (%)":             tr["pnl_pct"],
                        "Actual R:R":          tr["actual_rr"],
                        "Outcome":             tr["outcome"],
                        "MFE (%)":             tr["mfe_pct"],
                        "MAE (%)":             tr["mae_pct"],
                        "MFE (₹)":             tr["mfe_abs"],
                        "MAE (₹)":             tr["mae_abs"],
                        "Capture Efficiency %":tr["captured_pct"],
                        "BB Position %":       round(_bb_pct_bt*100, 1),
                        "Breakout":             "Yes" if brkout["is_breakout"] else "No",
                        "Vol Spike":            round(brkout["vol_spike"], 2),
                        "Range Expansion":      round(brkout["range_expansion"], 2),
                        "Market Regime":        reg_d["regime"],
                        "Trend Strength":       round(_bt_ts_live, 2),
                        "Gann Sq9 Near Sup":    "Yes" if _near_sq9_sup else "No",
                        "News Sentiment":      f"{news_sc:+.3f}" if news_sc is not None else "N/A",
                        "Bulk Deal Signal":    bd_sig,
                        "Bulk Net (Cr)":       bd_net,
                        "SMA20 at Entry":      round(sma20, 2),
                        "SMA50 at Entry":      round(sma50, 2),
                        "Buy Reasons":         " | ".join(buy_reas[:6]),
                        "Sell Reasons":        " | ".join(sell_reas[:4]),
                        "Ruling Planet":       inst_bt.ruling_planet,
                        "ML Direction":        _ml_bt_sig.get("direction","N/A"),
                        "ML Confidence":       f"{_ml_bt_sig.get('confidence',0):.0%}",
                        "ML Rev Price":        f"₹{_ml_bt_sig.get('reversal_price',0):,.2f}",
                        # v2 Rule 3 — partial 60/40 exit tracking
                        "Partial Exit":        "Yes" if tr.get("partial_exit") else "No",
                        "Partial PnL (60%)":   round(tr.get("partial_pnl", 0.0) / max(tr["actual_entry"], 0.01) * 100, 2) if tr.get("partial_exit") else 0.0,
                    })

                    i += max(tr["hold_days"], cfg["step"])

            _cbt.close()

            if not all_trades:
                return {"error": "No trades generated. Check date range, symbol, and price history in DB."}

            # ════════════════════════════════════════════════════
            # EXCEL REPORT
            # ════════════════════════════════════════════════════
            try:
                import openpyxl
                from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
                from openpyxl.utils import get_column_letter
            except ImportError:
                return {"error": "pip install openpyxl"}

            wb = openpyxl.Workbook()
            C_HDR="0d1e2d"; C_WIN="0a2a18"; C_LOSS="2a0a0a"; C_ALT="0f1e2e"
            C_GOLD="e6b800"; C_GRN="00c87a"; C_RED="e05050"; C_CYN="00d4ff"
            C_WHT="d8e8f0"; C_DIM="4a6678"; C_PURPLE="cc88ff"; C_ORG="ff9944"
            B=Side(style="thin",color="1a3040")
            bdr=Border(left=B,right=B,top=B,bottom=B)
            def _f(h): return PatternFill("solid",fgColor=h)
            def _fn(c=C_WHT,bold=False,sz=9): return Font(name="Consolas",color=c,bold=bold,size=sz)
            def _al(h="center",v="center",wrap=False):
                return Alignment(horizontal=h,vertical=v,wrap_text=wrap)

            # ── Stats ──
            def get_sub_stats(trades_list):
                if not trades_list:
                    return {
                        "count": 0, "wins": 0, "losses": 0, "t2_wins": 0,
                        "win_rate": 0.0, "avg_win": 0.0, "avg_loss": 0.0,
                        "expectancy": 0.0, "total_pnl": 0.0, "avg_rr": 0.0,
                        "avg_hold": 0.0, "t1_wins": 0, "avg_mfe": 0.0,
                        "avg_mae": 0.0, "avg_cap": 0.0,
                        "mc_mdd_95": 0.0, "deflated_sharpe": 0.0
                    }
                wins_sub = [t for t in trades_list if t["Outcome"] in ("WIN", "WIN_T2")]
                losses_sub = [t for t in trades_list if t["Outcome"] == "LOSS"]
                t2_wins_sub = [t for t in trades_list if t["Outcome"] == "WIN_T2"]
                t1_wins_sub = [t for t in trades_list if t["T1 Achieved"] == "Yes"]
                
                total_pnl_sub = sum(t["P&L (%)"] for t in trades_list)
                avg_win_sub = round(sum(t["P&L (%)"] for t in wins_sub) / len(wins_sub), 2) if wins_sub else 0.0
                avg_loss_sub = round(sum(t["P&L (%)"] for t in losses_sub) / len(losses_sub), 2) if losses_sub else 0.0
                win_rate_sub = round(len(wins_sub) / len(trades_list) * 100, 1)
                exp_sub = round((win_rate_sub / 100.0 * avg_win_sub) + ((1.0 - win_rate_sub / 100.0) * avg_loss_sub), 2)
                avg_rr_sub = round(sum(t["R:R (T1)"] for t in trades_list) / len(trades_list), 2)
                avg_hold_sub = round(sum(t["Actual Hold Days"] for t in trades_list) / len(trades_list), 1)
                avg_mfe_sub = round(sum(t.get("Max Favorable (%)", 0.0) for t in trades_list) / len(trades_list), 2)
                avg_mae_sub = round(sum(t.get("Max Adverse (%)", 0.0) for t in trades_list) / len(trades_list), 2)
                avg_cap_sub = round(sum(t.get("Capture Efficiency (%)", 0.0) for t in trades_list) / len(trades_list), 1)
                
                # Phase 5: Advanced Metrics
                from core.advanced_metrics import run_monte_carlo_drawdown, compute_deflated_sharpe
                returns_list = [t["P&L (%)"] for t in trades_list]
                mc_res = run_monte_carlo_drawdown(returns_list, iterations=1000)
                
                # Basic Sharpe calc for DSR
                ret_std = np.std(returns_list) if len(returns_list) > 1 else 1.0
                sharpe = (sum(returns_list) / len(returns_list)) / ret_std if ret_std > 0 else 0
                # Using 1000 iterations to simulate curve fitting discovery penalty
                dsr = compute_deflated_sharpe(sharpe, trials=1000, variance=ret_std)
                
                return {
                    "count": len(trades_list),
                    "wins": len(wins_sub),
                    "losses": len(losses_sub),
                    "t2_wins": len(t2_wins_sub),
                    "win_rate": win_rate_sub,
                    "avg_win": avg_win_sub,
                    "avg_loss": avg_loss_sub,
                    "expectancy": exp_sub,
                    "total_pnl": round(total_pnl_sub, 1),
                    "avg_rr": avg_rr_sub,
                    "avg_hold": avg_hold_sub,
                    "t1_wins": len(t1_wins_sub),
                    "avg_mfe": avg_mfe_sub,
                    "avg_mae": avg_mae_sub,
                    "avg_cap": avg_cap_sub,
                    "mc_mdd_95": mc_res["max_drawdown_95"],
                    "deflated_sharpe": dsr
                }

            is_trades = [t for t in all_trades if t["Validation Regime"] == "IN-SAMPLE"]
            oos_trades = [t for t in all_trades if t["Validation Regime"] == "OUT-OF-SAMPLE"]

            is_s = get_sub_stats(is_trades)
            oos_s = get_sub_stats(oos_trades)
            all_s = get_sub_stats(all_trades)

            best_t   = max(all_trades, key=lambda t:t["P&L (%)"])
            worst_t  = min(all_trades, key=lambda t:t["P&L (%)"])
            type_label = "Hedge Fund" if inv_type_bt=="hedge_fund" else inv_type_bt.capitalize()

            # ═══════ SHEET 1: SUMMARY ═══════
            ws1 = wb.active; ws1.title = "📊 Summary"
            ws1.sheet_properties.tabColor = "00c87a"
            ws1.merge_cells("A1:D1")
            c=ws1["A1"]
            c.value = f"Vprofitables UNIFIED BACKTEST — {sym_filter or 'ALL EQUITIES'} — {start_str} → {end_str} — {type_label.upper()} [Gann+Technical+Natal+Fundamental+Sentiment]"
            c.font=Font(name="Consolas",bold=True,color=C_GOLD,size=13)
            c.fill=_f(C_HDR); c.alignment=_al(); ws1.row_dimensions[1].height=26

            def _sec(ws, row, title):
                ws.merge_cells(f"A{row}:D{row}")
                c=ws.cell(row=row,column=1,value=title)
                c.font=_fn(C_CYN,True,9); c.fill=_f("081520"); c.alignment=_al()

            _sec(ws1, 2, "PERFORMANCE METRICS OVER VIEW")
            
            # Write 4-column headers at row 3
            ws1.cell(row=3, column=1, value="Metric").font = _fn(C_CYN, True, 10)
            ws1.cell(row=3, column=1).fill = _f("081520")
            ws1.cell(row=3, column=1).alignment = _al("left")
            ws1.cell(row=3, column=1).border = bdr

            ws1.cell(row=3, column=2, value="In-Sample (70% Optimization)").font = _fn(C_GOLD, True, 10)
            ws1.cell(row=3, column=2).fill = _f("081520")
            ws1.cell(row=3, column=2).alignment = _al()
            ws1.cell(row=3, column=2).border = bdr

            ws1.cell(row=3, column=3, value="Walk-Forward (30% Validation)").font = _fn(C_GOLD, True, 10)
            ws1.cell(row=3, column=3).fill = _f("081520")
            ws1.cell(row=3, column=3).alignment = _al()
            ws1.cell(row=3, column=3).border = bdr

            ws1.cell(row=3, column=4, value="Combined (All)").font = _fn(C_CYN, True, 10)
            ws1.cell(row=3, column=4).fill = _f("081520")
            ws1.cell(row=3, column=4).alignment = _al()
            ws1.cell(row=3, column=4).border = bdr

            ws1.row_dimensions[3].height = 24

            rows_data = [
                ("Total Trades",          lambda s: s["count"], C_CYN),
                ("Wins (incl. T2)",       lambda s: f"{s['wins']} ({s['t2_wins']} @ T2)", C_GRN),
                ("Losses",                lambda s: s["losses"], C_RED),
                ("Win Rate",              lambda s: f"{s['win_rate']}%", C_GRN),
                ("Avg Win %",             lambda s: f"+{s['avg_win']}%", C_GRN),
                ("Avg Loss %",            lambda s: f"{s['avg_loss']}%", C_RED),
                ("Expectancy/Trade",      lambda s: f"{s['expectancy']:+.2f}%", C_GRN),
                ("Total P&L",             lambda s: f"{s['total_pnl']:+.1f}%", C_GRN),
                ("Avg Dynamic R:R",       lambda s: f"{s['avg_rr']}:1", C_CYN),
                ("Avg Hold Days",         lambda s: f"{s['avg_hold']}d", C_WHT),
                ("T1 Achievement Rate",   lambda s: f"{s['t1_wins']}/{s['count']}" if s['count'] > 0 else "0/0", C_GRN),
                ("T2 Achievement Rate",   lambda s: f"{s['t2_wins']}/{s['count']}" if s['count'] > 0 else "0/0", C_GOLD),
                ("Avg MFE %",             lambda s: f"+{s['avg_mfe']}%", C_GRN),
                ("Avg MAE %",             lambda s: f"-{s['avg_mae']}%", C_RED),
                ("Avg Capture Efficiency",lambda s: f"{s['avg_cap']}%", C_GOLD),
                ("True Max Drawdown (95%)", lambda s: f"{s['mc_mdd_95']}%", C_RED),
                ("Deflated Sharpe (DSR)",   lambda s: f"{s['deflated_sharpe']}", C_CYN),
            ]

            last_row = 3
            for ri, (lbl, val_fn, col) in enumerate(rows_data, 4):
                lc = ws1.cell(row=ri, column=1, value=lbl)
                lc.font = _fn(C_DIM, True, 9); lc.fill = _f(C_HDR); lc.alignment = _al("left"); lc.border = bdr
                
                # In-Sample cell
                is_val = val_fn(is_s)
                vc1 = ws1.cell(row=ri, column=2, value=is_val)
                vc1.font = Font(name="Consolas", bold=True, color=col, size=10)
                vc1.fill = _f("0a1825"); vc1.alignment = _al("right"); vc1.border = bdr
                
                # Out-of-Sample cell
                oos_val = val_fn(oos_s)
                vc2 = ws1.cell(row=ri, column=3, value=oos_val)
                vc2.font = Font(name="Consolas", bold=True, color=col, size=10)
                vc2.fill = _f("0a1825"); vc2.alignment = _al("right"); vc2.border = bdr
                
                # Combined cell
                all_val = val_fn(all_s)
                vc3 = ws1.cell(row=ri, column=4, value=all_val)
                vc3.font = Font(name="Consolas", bold=True, color=col, size=10)
                vc3.fill = _f("0a1825"); vc3.alignment = _al("right"); vc3.border = bdr

                ws1.row_dimensions[ri].height = 20
                last_row = ri

            # Best/Worst/Strategy section below the table
            _sec(ws1, last_row + 2, "BEST / WORST TRADES & STRATEGY CONTEXT")
            ws1.row_dimensions[last_row + 2].height = 20
            
            strat_info = [
                ("Best Trade",             f"{best_t['Symbol']} {best_t['P&L (%)']:+.1f}%", C_GRN),
                ("Worst Trade",            f"{worst_t['Symbol']} {worst_t['P&L (%)']:+.1f}%", C_RED),
                ("Strategy",               f"Unified: Gann + Technical + Natal + Fundamental + Sentiment + {type_label}", C_CYN),
                ("Period",                 f"{start_str} → {end_str}", C_WHT),
                ("Walk-Forward Cutoff",    f"{cutoff_str} (70% In-Sample / 30% Out-of-Sample)", C_GOLD),
            ]
            for idx, (lbl, val, col) in enumerate(strat_info, last_row + 3):
                lc = ws1.cell(row=idx, column=1, value=lbl)
                lc.font = _fn(C_DIM, True, 9); lc.fill = _f(C_HDR); lc.alignment = _al("left"); lc.border = bdr
                
                # Merge columns 2 to 4 for values
                ws1.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=4)
                vc = ws1.cell(row=idx, column=2, value=val)
                vc.font = Font(name="Consolas", bold=True, color=col, size=10)
                vc.fill = _f("0a1825"); vc.alignment = _al("left"); vc.border = bdr
                
                # Apply borders to the merged cells manually (openpyxl merged cell border fix)
                for c_col in (2, 3, 4):
                    ws1.cell(row=idx, column=c_col).border = bdr
                
                ws1.row_dimensions[idx].height = 20
                last_row = idx

            # Tech regime summary (replaces Wyckoff phase breakdown)
            _sec(ws1, last_row + 2, "TECH REGIME BREAKDOWN")
            ws1.row_dimensions[last_row + 2].height = 20
            
            from collections import Counter
            phase_ct = Counter(t.get("Tech Regime","N/A") for t in all_trades)
            last_row = last_row + 2
            for ri, (ph, cnt) in enumerate(sorted(phase_ct.items()), last_row + 1):
                ws1.cell(row=ri, column=1, value=ph).font = _fn(C_GOLD, False, 9)
                ws1.cell(row=ri, column=1).border = bdr
                
                ws1.merge_cells(start_row=ri, start_column=2, end_row=ri, end_column=4)
                ws1.cell(row=ri, column=2, value=f"{cnt} trades").font = _fn(C_WHT, False, 9)
                
                for c_col in (2, 3, 4):
                    ws1.cell(row=ri, column=c_col).border = bdr
                
                ws1.row_dimensions[ri].height = 18
                last_row = ri

            ws1.column_dimensions["A"].width = 28
            ws1.column_dimensions["B"].width = 30
            ws1.column_dimensions["C"].width = 30
            ws1.column_dimensions["D"].width = 20

            # ═══════ SHEET 2: ALL TRADES ═══════
            ws2=wb.create_sheet("📋 All Trades")
            ws2.sheet_properties.tabColor="00d4ff"
            headers=list(all_trades[0].keys())

            # Colour groups for headers
            HDR_COLS = {
                "Symbol":C_GOLD,"Name":C_GOLD,"Sector":C_DIM,"Investment Type":C_DIM,
                "Tech Regime":C_PURPLE,
                "Analysis Date":C_DIM,"Entry Date":C_CYN,"Entry Price (₹)":C_CYN,
                "Stop Loss (₹)":C_RED,"Target 1 (₹)":C_GRN,"Target 2 (₹)":C_GRN,
                "SL Source":C_RED,"T1 Source":C_GRN,
                "Risk (₹)":C_RED,"Reward 1 (₹)":C_GRN,"Reward 2 (₹)":C_GRN,
                "R:R (T1)":C_GOLD,"R:R (T2)":C_GOLD,
                "Exit Date":C_CYN,"Exit Price (₹)":C_CYN,"Exit Reason":C_DIM,
                "Actual Hold Days":C_WHT,"Trailing SL (₹)":C_ORG,
                "Trailing Active":C_ORG,"T1 Achieved":C_GRN,"T2 Achieved":C_GOLD,
                "P&L (₹)":C_GRN,"P&L (%)":C_GRN,"Actual R:R":C_GOLD,"Outcome":C_GRN,
                "MFE (%)":C_GRN,"MAE (%)":C_RED,"MFE (₹)":C_GRN,"MAE (₹)":C_RED,
                "Capture Efficiency %":C_GOLD,
                "BB Position %":C_PURPLE,"Breakout":C_CYN,
                "Vol Spike":C_CYN,"Range Expansion":C_CYN,
                "Gann Sq9 Near Sup":C_GOLD,
                "Market Regime":C_DIM,"Trend Strength":C_DIM,
                "News Sentiment":C_CYN,"Bulk Deal Signal":C_ORG,"Bulk Net (Cr)":C_ORG,
                "SMA20 at Entry":C_DIM,"SMA50 at Entry":C_DIM,
                "Buy Reasons":"8fd8a0","Sell Reasons":"e08080","Ruling Planet":C_DIM,
                "ML Direction":C_GOLD,"ML Confidence":C_GOLD,"ML Rev Price":C_GOLD,
                "Partial Exit":C_GRN,"Partial PnL (60%)":C_GRN,
                "Validation Regime":C_GOLD,
            }
            for ci,h in enumerate(headers,1):
                c=ws2.cell(row=1,column=ci,value=h)
                c.font=Font(name="Consolas",color=HDR_COLS.get(h,C_GOLD),bold=True,size=9)
                c.fill=_f(C_HDR); c.alignment=_al(wrap=True); c.border=bdr
            ws2.row_dimensions[1].height=36

            for ri,trade in enumerate(all_trades,2):
                is_win=trade["Outcome"] in ("WIN","WIN_T2")
                is_t2=trade["Outcome"]=="WIN_T2"
                rf=_f("072e1a" if is_t2 else (C_WIN if is_win else C_LOSS)) if ri%2==0 \
                   else _f("0a3d20" if is_t2 else (C_ALT if is_win else "1e0808"))
                for ci,h in enumerate(headers,1):
                    v=trade[h]; c=ws2.cell(row=ri,column=ci,value=v)
                    col=C_WHT
                    if h=="Outcome": col=("e6b800" if is_t2 else C_GRN) if is_win else C_RED
                    elif h in ("P&L (%)","P&L (₹)","Actual R:R"): col=C_GRN if (v or 0)>0 else C_RED
                    elif h=="MFE (%)": col=C_GRN
                    elif h=="MAE (%)": col=C_RED
                    elif h=="Capture Efficiency %": col=C_GRN if (v or 0)>=70 else C_GOLD if (v or 0)>=50 else C_RED
                    elif h in ("R:R (T1)","R:R (T2)"): col=C_GRN if (v or 0)>=2 else C_GOLD if (v or 0)>=1.2 else C_RED
                    elif h in ("T1 Achieved","T2 Achieved"): col=C_GRN if v=="Yes" else C_DIM
                    elif h=="Trailing Active": col=C_ORG if v=="Yes" else C_DIM
                    elif h=="BB Position %": col=C_GRN if 30<=(v or 0)<=70 else C_GOLD if (v or 0)<30 else C_RED
                    elif h=="Breakout": col=C_GRN if v=="Yes" else C_DIM
                    elif h=="Gann Sq9 Near Sup": col=C_GOLD if v=="Yes" else C_DIM
                    elif h=="Tech Regime": col=C_PURPLE
                    elif h=="Market Regime": col=C_GRN if "BULL" in str(v) else C_RED if "BEAR" in str(v) else C_DIM
                    elif h=="Buy Reasons": col="8fd8a0"
                    elif h=="Sell Reasons": col="e08080"
                    c.font=Font(name="Consolas",color=col,size=8)
                    c.fill=rf; c.border=bdr
                    c.alignment=_al("left" if h in ("Buy Reasons","Sell Reasons","SL Source","T1 Source","Exit Reason") else "center",
                                    wrap=h in ("Buy Reasons","Sell Reasons","SL Source","T1 Source"))
                ws2.row_dimensions[ri].height=50 if any(len(str(trade.get(k,"")))>50
                    for k in ("Buy Reasons","Sell Reasons")) else 18

            COL_W={
                "Symbol":12,"Name":26,"Sector":14,"Investment Type":13,"Tech Regime":18,"RSI at Entry":12,
                "Analysis Date":12,"Entry Date":12,"Exit Date":12,
                "Entry Price (₹)":14,"Stop Loss (₹)":13,"Target 1 (₹)":13,"Target 2 (₹)":13,
                "SL Source":30,"T1 Source":30,
                "Risk (₹)":11,"Reward 1 (₹)":12,"Reward 2 (₹)":12,
                "R:R (T1)":10,"R:R (T2)":10,
                "Exit Price (₹)":13,"Exit Reason":30,"Actual Hold Days":14,
                "Trailing SL (₹)":13,"Trailing Active":13,"T1 Achieved":11,"T2 Achieved":11,
                "P&L (₹)":11,"P&L (%)":10,"Actual R:R":13,"Outcome":12,
                "MFE (%)":10,"MAE (%)":10,"MFE (₹)":11,"MAE (₹)":11,"Capture Efficiency %":16,
                "BB Position %":12,"Breakout":10,"Vol Spike":10,"Range Expansion":14,
                "Gann Sq9 Near Sup":14,
                "Market Regime":16,"Trend Strength":13,
                "News Sentiment":14,"Bulk Deal Signal":14,"Bulk Net (Cr)":13,
                "SMA20 at Entry":14,"SMA50 at Entry":14,
                "Buy Reasons":55,"Sell Reasons":50,"Ruling Planet":13,
            }
            for ci,h in enumerate(headers,1):
                ws2.column_dimensions[get_column_letter(ci)].width=COL_W.get(h,14)
            ws2.freeze_panes="A2"
            ws2.auto_filter.ref=f"A1:{get_column_letter(len(headers))}1"

            # ═══════ SHEET 3: BY SYMBOL ═══════
            ws3=wb.create_sheet("📈 By Symbol")
            ws3.sheet_properties.tabColor="e6b800"
            from collections import defaultdict
            grp=defaultdict(list)
            for t in all_trades: grp[t["Symbol"]].append(t)
            sh3=["Symbol","Name","Sector","Trades","Wins","T2 Wins","Losses",
                 "Win Rate","Avg P&L%","Total P&L%","Avg R:R",
                 "T1 Rate","T2 Rate","Trailing Rate","Avg Hold Days",
                 "Avg Acc Score","Avg MFE%","Avg MAE%","Avg Capture%",
                 "Best Trade%","Worst Trade%","Avg News Sent","Dominant Phase"]
            for ci,h in enumerate(sh3,1):
                c=ws3.cell(row=1,column=ci,value=h)
                c.font=_fn(C_GOLD,True,9);c.fill=_f(C_HDR);c.alignment=_al(wrap=True);c.border=bdr
            ws3.row_dimensions[1].height=32

            for ri,(sk,tk) in enumerate(sorted(grp.items()),2):
                wk=[t for t in tk if t["Outcome"] in ("WIN","WIN_T2")]
                t2k=[t for t in tk if t["Outcome"]=="WIN_T2"]
                wr=round(len(wk)/len(tk)*100,1)
                ap=round(sum(t["P&L (%)"] for t in tk)/len(tk),2)
                tp=round(sum(t["P&L (%)"] for t in tk),2)
                ar=round(sum(t["R:R (T1)"] for t in tk)/len(tk),2)
                ah=round(sum(t["Actual Hold Days"] for t in tk)/len(tk),1)
                t1r=f"{sum(1 for t in tk if t['T1 Achieved']=='Yes')}/{len(tk)}"
                t2r=f"{len(t2k)}/{len(tk)}"
                trr=f"{sum(1 for t in tk if t['Trailing Active']=='Yes')}/{len(tk)}"
                aa=round(sum(t.get("BB Position %",50) for t in tk)/len(tk),1)
                am=round(sum(t["MFE (%)"] for t in tk)/len(tk),2)
                ama=round(sum(t["MAE (%)"] for t in tk)/len(tk),2)
                ac=round(sum(t["Capture Efficiency %"] for t in tk)/len(tk),1)
                bk=max(t["P&L (%)"] for t in tk); wsk=min(t["P&L (%)"] for t in tk)
                nsv=[float(t["News Sentiment"]) for t in tk if t["News Sentiment"]!="N/A"]
                ans=round(sum(nsv)/len(nsv),3) if nsv else None
                dom_phase=Counter(t.get("Tech Regime","N/A") for t in tk).most_common(1)[0][0]

                row=[sk,tk[0]["Name"],tk[0]["Sector"],len(tk),len(wk),len(t2k),len(tk)-len(wk),
                     f"{wr}%",f"{ap:+.2f}%",f"{tp:+.1f}%",f"{ar}:1",
                     t1r,t2r,trr,ah,aa,f"+{am}%",f"-{ama}%",f"{ac}%",
                     f"+{bk:.1f}%",f"{wsk:.1f}%",
                     f"{ans:+.3f}" if ans is not None else "N/A",dom_phase]
                col=C_GRN if wr>=55 else C_RED if wr<45 else C_WHT
                rf=_f(C_WIN if wr>=55 else C_LOSS if wr<45 else C_ALT)
                for ci,v in enumerate(row,1):
                    c=ws3.cell(row=ri,column=ci,value=v)
                    c.font=Font(name="Consolas",color=col,size=9)
                    c.fill=rf;c.border=bdr;c.alignment=_al()
            for ci,h in enumerate(sh3,1):
                ws3.column_dimensions[get_column_letter(ci)].width=max(len(h)+2,12)
            ws3.freeze_panes="A2"

            buf=_io_bt.BytesIO(); wb.save(buf); buf.seek(0)
            # Return special dict — do_GET intercepts _xlsx_bytes to send raw binary
            fname = f"Backtest_{inv_type_bt}_{start_str}_{end_str}{('_'+sym_filter) if sym_filter else ''}.xlsx"
            return {"_xlsx_bytes": buf.read(), "_filename": fname}



        # ══════════════════════════════════════════════════════════════
        # NOTIFICATION CONFIG + FORWARD SIGNAL DISPATCH
        # ══════════════════════════════════════════════════════════════

        if ep == "notify_config_save":
            # Save notification settings to gann_settings.json
            ok = _notifier_save_cfg({
                "EMAIL_ENABLED":   bool(p.get("email_enabled", False)),
                "EMAIL_TO":        p.get("email_to", ""),
                "EMAIL_FROM":      p.get("email_from", ""),
                "EMAIL_PASS":      p.get("email_pass", ""),
                "WHATSAPP_METHOD": p.get("whatsapp_method", "none"),
                "CALLMEBOT_PHONE": p.get("callmebot_phone", ""),
                "CALLMEBOT_KEY":   p.get("callmebot_key", ""),
                "TWILIO_SID":      p.get("twilio_sid", ""),
                "TWILIO_TOKEN":    p.get("twilio_token", ""),
                "TWILIO_FROM":     p.get("twilio_from", "whatsapp:+14155238886"),
                "TWILIO_TO":       p.get("twilio_to", ""),
            })
            return {"ok": ok}

        if ep == "notify_config_load":
            cfg = _notifier_load_cfg()
            # Never return passwords in plaintext — mask them
            safe_cfg = {k: (("*" * 8) if "PASS" in k or "TOKEN" in k or "KEY" in k else v)
                        for k, v in cfg.items()}
            return {"ok": True, "config": safe_cfg}

        if ep == "notify_test":
            return _notifier_test()

        # ── fetch_deals ── fetch today's bulk+block deals from NSE ──────────
        if ep == "fetch_deals":
            days = int(p.get("days", 5))
            if _fetch_institutional is None:
                return {"ok": False, "error": "fetch_institutional not available"}
            try:
                saved = _fetch_institutional.fetch_and_save_deals_range(days_back=days)
                return {"ok": True, "saved": saved, "days": days,
                        "message": f"{saved} deal rows saved for last {days} trading days"}
            except Exception as _fde:
                return {"ok": False, "error": str(_fde)}

        # ── forward_signal ─────────────────────────────────────────────────────
        # If symbol is provided → generate for that symbol only.
        # If symbol is blank   → scan ALL equities, return those with confidence>=60
        #                        and save them to forward_signals DB table.
        if ep == "forward_signal":
            inv_fs    = p.get("type", "swing")
            send_now  = str(p.get("send", "false")).lower() in ("true", "1", "yes")
            as_of_fs  = p.get("date", today.isoformat())
            sym_fs    = p.get("symbol", "").upper().strip()
            min_conf  = int(p.get("min_confidence", 60))

            def _build_signal(sym, rec, adv_result, inst):
                """Package one advisor recommendation + full master_report analysis."""
                _fc2 = _db()
                _fp2 = _fc2.execute(
                    "SELECT close FROM daily_prices WHERE symbol=? AND trade_date<=? "
                    "AND close IS NOT NULL ORDER BY trade_date DESC LIMIT 1",
                    (sym, as_of_fs)).fetchone()
                price = float(_fp2[0]) if _fp2 else (inst.all_time_high * 0.85)
                _fc2.close()

                pd_d      = adv_result.get("planet_dashboard", {})
                bull_asp  = sum(1 for a in pd_d.get("aspects",[]) if a.get("direction")=="BULLISH")
                bear_asp  = sum(1 for a in pd_d.get("aspects",[]) if a.get("direction")=="BEARISH")

                # Fetch full master_report analysis for rich forward signal
                master = {}
                try:
                    master = self.route("master_report", {
                        "symbol":   sym,
                        "date":     as_of_fs,
                        "price":    str(price),
                        "inv_type": inv_fs,
                    })
                except Exception as _mr_e:
                    print(f"  [FWD ] master_report for {sym} skipped: {_mr_e}", flush=True)

                # ── For long type: override master_report levels with advisor levels ──
                # master_report computes its own entry/sl/t1/t2 independently.
                # This causes contradiction: card shows advisor levels, verdict shows master levels.
                # Fix: inject advisor levels into master trade_setup so verdict is consistent.
                if inv_fs == "long" and rec:
                    _adv_entry = rec.get("entry", price)
                    _adv_sl    = rec.get("stop_loss", 0)
                    _adv_t1    = rec.get("target1", 0)
                    _adv_t2    = rec.get("target2", 0)
                    _adv_rr    = rec.get("rr_ratio", 0)
                    _adv_rr_str = f"1:{_adv_rr:.2f}" if isinstance(_adv_rr, float) else str(_adv_rr)

                    # Override trade_setup in master with advisor levels
                    if master.get("trade_setup"):
                        master["trade_setup"]["entry"]       = _adv_entry
                        master["trade_setup"]["stop_loss"]   = _adv_sl
                        master["trade_setup"]["target1"]     = _adv_t1
                        master["trade_setup"]["target2"]     = _adv_t2
                        master["trade_setup"]["risk_reward"] = _adv_rr_str
                        master["trade_setup"]["bias"]        = "BULLISH"

                    # Rebuild the overall_verdict using the correct advisor levels
                    # Determine cycle position from 52wk range
                    try:
                        _mv_52h = float(rec.get("price_52wk_high", price * 1.2)) if "price_52wk_high" in rec else price * 1.2
                        _mv_52l = float(rec.get("price_52wk_low",  price * 0.7)) if "price_52wk_low"  in rec else price * 0.7
                    except Exception:
                        _mv_52h = price * 1.2; _mv_52l = price * 0.7
                    _mv_rng = _mv_52h - _mv_52l
                    _mv_wvp = (price - _mv_52l) / _mv_rng if _mv_rng > 0 else 0.5
                    _mv_wvp = max(0.0, min(1.0, _mv_wvp))

                    # Choose cycle label
                    if _mv_wvp <= 0.30:
                        _cv_icon, _cv_label, _cv_action = "🟢", "ACCUMULATION ZONE", "ACCUMULATE"
                    elif _mv_wvp <= 0.55:
                        _cv_icon, _cv_label, _cv_action = "🟢", "EARLY MARKUP", "BUY"
                    elif _mv_wvp <= 0.75:
                        _cv_icon, _cv_label, _cv_action = "🟡", "MARKUP PHASE", "HOLD & TRAIL"
                    else:
                        _cv_icon, _cv_label, _cv_action = "🔴", "DISTRIBUTION ZONE", "EXIT / AVOID"

                    _mv_gann_date = rec.get("buy_date", "")
                    _mv_regime    = rec.get("regime", "")

                    if _cv_action in ("ACCUMULATE", "BUY"):
                        _mv_verdict = (
                            f"{_cv_icon} CYCLE PHASE: {_cv_label} — ACTION: {_cv_action} "
                            f"{sym} is in the {_cv_label} of its price-time cycle. "
                            f"The positive cycle has begun — accumulate at support. "
                            f"📋 CYCLE PLAN: "
                            f"Entry ₹{_adv_entry:,.2f} | "
                            f"SL ₹{_adv_sl:,.2f} (below cycle low — invalidation) | "
                            f"T1 ₹{_adv_t1:,.2f} (exit 50%, activate trailing SL) | "
                            f"T2 ₹{_adv_t2:,.2f} (distribution zone — full cycle exit). "
                            f"No time limit — hold until cycle ends. R:R {_adv_rr_str}."
                        )
                    elif _cv_action == "HOLD & TRAIL":
                        _mv_verdict = (
                            f"{_cv_icon} CYCLE PHASE: {_cv_label} — {_cv_action} "
                            f"{sym} is in active markup. "
                            f"If holding: trail SL, target T1 ₹{_adv_t1:,.2f} → trail to T2 ₹{_adv_t2:,.2f}. "
                            f"If not yet in: wait for pullback to ₹{_adv_entry:,.2f} before entering."
                        )
                    else:
                        _mv_verdict = (
                            f"{_cv_icon} CYCLE PHASE: {_cv_label} — {_cv_action} "
                            f"{sym} is near distribution. "
                            f"Begin trimming at T1 ₹{_adv_t1:,.2f}. "
                            f"New long entries not recommended at this stage."
                        )
                    if _mv_gann_date:
                        _mv_verdict += f" 📅 Key Gann date: {_mv_gann_date} — watch for cycle inflection."
                    master["overall_verdict"] = _mv_verdict
                    master["wave_pos_pct"]    = _mv_wvp

                return {
                    "symbol":         sym,
                    "name":           inst.name,
                    "sector":         inst.sector,
                    "ruling_planet":  inst.ruling_planet,
                    "analysis_date":  as_of_fs,
                    "inv_type":       inv_fs,
                    "action":         "BUY",
                    "entry":          rec.get("entry", price),
                    "stop_loss":      rec.get("stop_loss", 0),
                    "target1":        rec.get("target1", 0),
                    "target2":        rec.get("target2", 0),
                    "rr_ratio":       rec.get("rr_ratio", 0),
                    "confidence":     rec.get("confidence", 0),
                    "regime":         rec.get("regime", ""),
                    "tech_momentum":  rec.get("tech_momentum", "N/A"),
                    "wyckoff_phase":  rec.get("tech_momentum", "N/A"),  # DB column compat
                    "news_sentiment": rec.get("news_label", "N/A"),
                    "bulk_signal":    rec.get("bulk_signal", "NEUTRAL"),
                    "hold_days":      rec.get("hold_days", 0),
                    "buy_date":       (rec.get("buy_date","") or "").replace("-","/"),
                    "sell_date":      (rec.get("sell_date","") or "").replace("-","/"),
                    "buy_time":       rec.get("buy_time", "09:15 IST"),
                    "reasons":        (
                        ([f"ML: {rec.get('ml_direction','N/A')} conf {rec.get('ml_confidence',0):.0%} → rev ₹{rec.get('ml_reversal_price',price):,.2f}"]
                         if rec.get("ml_model_trained") and rec.get("ml_confidence",0) > 0.60 else []) +
                        rec.get("buy_reasons", []) + rec.get("sell_reasons", [])
                    ),
                    "engine_scores": {
                        "gann":        f"{rec.get('gann_score',0):.0f}/20",
                        "quant":       f"{rec.get('quant_score',0):.0f}/20",
                        "natal":       f"{rec.get('natal_score',0):.0f}/20",
                        "fundamental": f"{rec.get('fund_score',0):.0f}/25",
                        "sentiment":   rec.get("news_label","N/A"),
                    },
                    "macro_note": f"Sky {bull_asp}B/{bear_asp}b · Retro: {','.join(pd_d.get('retrograde',[])[:3]) or 'none'}",
                    # Full analysis sections from master_report
                    "master_report":     master,
                    "technical_summary": master.get("technical_summary", {}),
                    "gann_summary":      master.get("gann_summary", {}),
                    "natal_summary":     master.get("natal_summary", {}),
                    "simons_summary":    master.get("simons_summary", {}),
                    "fundamental_summary": master.get("fundamental_summary", {}),
                    "sentiment_summary": master.get("sentiment_summary", {}),
                    "trade_setup":       master.get("trade_setup", {}),
                    "overall_verdict":   master.get("overall_verdict", ""),
                    "headline":          master.get("headline", ""),
                    "reversal_dates":    rec.get("reversal_dates", []),
                    "fund_grade":        rec.get("fund_grade", ""),
                    "fund_verdict":      rec.get("fund_verdict", ""),
                    "price":             price,
                    # v3.9 ML augmentation fields
                    "ml_direction":      rec.get("ml_direction", "NEUTRAL"),
                    "ml_confidence":     rec.get("ml_confidence", 0.0),
                    "ml_reversal_price": rec.get("ml_reversal_price", price),
                    "ml_reversal_date":  rec.get("ml_reversal_date", ""),
                    "ml_model_trained":  rec.get("ml_model_trained", False),
                    "ml_expected_move":  rec.get("ml_expected_move", 0.0),
                }

            def _save_signal(sig):
                """Persist signal to forward_signals table. Skip duplicates."""
                try:
                    _sc = _db()
                    _sc.execute("""
                        INSERT OR IGNORE INTO forward_signals
                        (signal_date,symbol,inv_type,entry,stop_loss,target1,target2,
                         rr_ratio,confidence,regime,wyckoff_phase,news_sentiment,bulk_signal,
                         hold_days,buy_date,sell_date,reasons,status,trailing_sl,created_at)
                        VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,'OPEN',?,?)
                    """, (
                        sig["analysis_date"], sig["symbol"], sig["inv_type"],
                        sig["entry"], sig["stop_loss"], sig["target1"], sig["target2"],
                        sig["rr_ratio"], sig["confidence"], sig["regime"],
                        sig.get("tech_momentum", sig.get("wyckoff_phase","N/A")), sig["news_sentiment"], sig["bulk_signal"],
                        sig["hold_days"], sig["buy_date"], sig["sell_date"],
                        json.dumps(sig["reasons"]),
                        sig["stop_loss"],   # trailing_sl starts at SL
                        datetime.now().isoformat()
                    ))
                    _sc.commit()
                    _sc.close()
                except Exception as _se:
                    print(f"  [WARN ] forward_signal save error: {_se}", flush=True)

            # ── Single-symbol mode ─────────────────────────────────────────
            if sym_fs:
                inst_fs = ALL_INSTRUMENTS.get(sym_fs)
                if not inst_fs:
                    raise ValueError(f"Unknown symbol: {sym_fs}")
                adv_result = self.route("advisor", {
                    "symbol": sym_fs, "diversify": "1", "type": inv_fs,
                    "risk": "balanced", "amount": "100000", "date": as_of_fs
                })
                recs = adv_result.get("recommendations", [])
                rec  = recs[0] if recs else {}
                if not rec:
                    return {"ok": False, "error": f"No signal for {sym_fs} on {as_of_fs}"}
                sig = _build_signal(sym_fs, rec, adv_result, inst_fs)
                _save_signal(sig)
                notify_result = {}
                if send_now:
                    notify_result = _notifier_send_signal(sig)
                return {"ok": True, "mode": "single", "signal": sig,
                        "notify": notify_result, "advisor": rec}

            # ── All-symbols scan mode ──────────────────────────────────────
            # Score every equity through the advisor individually (symbols= filter
            # ensures each call only scores that one stock with its real price data).
            all_signals = []
            equity_syms = [s for s, i in ALL_INSTRUMENTS.items()
                           if i.instrument_type == "EQUITY"]
            print(f"  [FWD ] Scanning {len(equity_syms)} equities for forward signals...", flush=True)

            for _sym in equity_syms:
                try:
                    _inst = ALL_INSTRUMENTS[_sym]
                    # Pass symbols=_sym so advisor scores ONLY this stock
                    _adv  = self.route("advisor", {
                        "symbols": _sym, "diversify": "1", "type": inv_fs,
                        "risk": "balanced", "amount": "100000", "date": as_of_fs
                    })
                    _recs = _adv.get("recommendations", [])
                    _rec  = _recs[0] if _recs else {}
                    if not _rec or _rec.get("confidence", 0) < min_conf:
                        continue
                    _sig = _build_signal(_sym, _rec, _adv, _inst)
                    _save_signal(_sig)
                    all_signals.append(_sig)
                    print(f"  [FWD ] {_sym}: conf={_sig['confidence']} entry=₹{_sig['entry']}", flush=True)
                except Exception as _fe:
                    print(f"  [FWD ] {_sym} skip: {_fe}", flush=True)
                    continue

            print(f"  [FWD ] Scan complete: {len(all_signals)} signals generated", flush=True)

            # ── Build Excel report + send ONE combined email ───────────────
            notify_result = {"ok": False, "error": "send=false"}
            if send_now and all_signals:
                try:
                    notify_result = _notifier_send_signal_batch(all_signals, as_of_fs, inv_fs)
                except Exception as _ne:
                    print(f"  [FWD ] Notify error: {_ne}", flush=True)
                    notify_result = {"ok": False, "error": str(_ne)}

            return {
                "ok":             True,
                "mode":           "scan_all",
                "date":           as_of_fs,
                "inv_type":       inv_fs,
                "min_confidence": min_conf,
                "signal_count":   len(all_signals),
                "signals":        all_signals,
                "notify":         notify_result,
            }

        # ── forward_test_update ────────────────────────────────────────────────
        # Called daily (by scheduler or manually) to update open forward signals:
        # checks current price vs SL / T1 / T2 / trailing SL and updates status.
        if ep == "forward_test_update":
            # Helper function to fetch intraday OHLC from yfinance
            def _fetch_intraday_ohlc(yf_sym):
                try:
                    import yfinance as yf
                    ticker = yf.Ticker(yf_sym)
                    df = ticker.history(period="1d")
                    if not df.empty:
                        last_row = df.iloc[-1]
                        return {
                            "close": float(last_row["Close"]),
                            "high":  float(last_row["High"]),
                            "low":   float(last_row["Low"]),
                            "open":  float(last_row["Open"]),
                        }
                except Exception as _yfe:
                    print(f"  [INTRADAY] Failed to fetch {yf_sym}: {_yfe}", flush=True)
                return None

            _ftu = _db()
            open_sigs = _ftu.execute(
                "SELECT id,symbol,inv_type,entry,stop_loss,target1,target2,"
                "       trailing_sl,hold_days,signal_date,buy_date "
                "FROM forward_signals WHERE status='OPEN'"
            ).fetchall()
            _ftu.close()

            updated = 0
            now_str = today.isoformat()

            for row in open_sigs:
                (sig_id, sym, inv_type, entry, sl, t1, t2,
                 trail_sl, hold_days, signal_date, buy_date) = row

                # ── PHASE 3 FIX (T3): Call yfinance for live intraday OHLC ────
                cur_price, cur_high, cur_low = None, None, None
                inst = ALL_INSTRUMENTS.get(sym)
                if inst and getattr(inst, "yfinance_symbol", None):
                    live_data = _fetch_intraday_ohlc(inst.yfinance_symbol)
                    if live_data:
                        cur_price = live_data["close"]
                        cur_high  = live_data["high"]
                        cur_low   = live_data["low"]

                if cur_price is None:
                    # Fallback: EOD daily_prices cache in SQLite
                    _pc = _db()
                    _pr = _pc.execute(
                        "SELECT close, high, low FROM daily_prices WHERE symbol=? "
                        "AND trade_date<=? AND close IS NOT NULL "
                        "ORDER BY trade_date DESC LIMIT 1",
                        (sym, now_str)).fetchone()
                    _pc.close()
                    if _pr:
                        cur_price = float(_pr[0])
                        cur_high  = float(_pr[1] or _pr[0])
                        cur_low   = float(_pr[2] or _pr[0])

                if cur_price is None:
                    continue

                # Update max_high + trailing SL
                _pu = _db()
                _mh_row = _pu.execute(
                    "SELECT max_high FROM forward_signals WHERE id=?", (sig_id,)).fetchone()
                prev_max = float(_mh_row[0] or entry) if _mh_row and _mh_row[0] else entry
                new_max  = max(prev_max, cur_high)

                # Trailing SL: once T1 is hit, trail at 50% of (max_high - entry) above entry
                cur_trail = trail_sl or sl
                if new_max >= t1:
                    trail_gap  = (new_max - entry) * 0.5
                    new_trail  = round(max(sl, entry + trail_gap * 0.3, new_max - trail_gap), 2)
                    cur_trail  = max(cur_trail, new_trail)

                new_status = "OPEN"
                exit_price = None
                exit_reason = None

                # Check exit conditions (in priority order)
                is_intraday_expired = False
                if inv_type == "intraday":
                    from datetime import datetime as dt_now
                    current_ist_time = dt_now.now()
                    is_past_square_off = (current_ist_time.hour > 15) or (current_ist_time.hour == 15 and current_ist_time.minute >= 15)
                    is_past_day = (today > date.fromisoformat(signal_date))
                    if is_past_square_off or is_past_day:
                        is_intraday_expired = True

                if cur_price >= t2:
                    new_status  = "T2_HIT"
                    exit_price  = t2
                    exit_reason = f"Target 2 ₹{t2:.2f} reached"
                elif cur_price >= t1 and new_max >= t1:
                    new_status  = "T1_HIT"
                    exit_price  = t1
                    exit_reason = f"Target 1 ₹{t1:.2f} reached (trailing active)"
                elif cur_price <= cur_trail:
                    if new_max >= t1:
                        new_status  = "TRAILING_SL"
                        exit_price  = cur_trail
                        exit_reason = f"Trailing SL ₹{cur_trail:.2f} hit (locked profit)"
                    else:
                        new_status  = "SL_HIT"
                        exit_price  = sl
                        exit_reason = f"Stop Loss ₹{sl:.2f} hit"
                elif is_intraday_expired:
                    new_status  = "EXPIRED"
                    exit_price  = cur_price
                    exit_reason = "Intraday same-session Auto-Square-Off at 15:15 IST"
                else:
                    # Check hold_days expiry
                    try:
                        start = date.fromisoformat(signal_date)
                        elapsed = (today - start).days
                        if elapsed >= (hold_days or 30):
                            new_status  = "EXPIRED"
                            exit_price  = cur_price
                            exit_reason = f"Hold period {hold_days}d expired"
                    except Exception:
                        pass

                pnl = round((exit_price - entry) / entry * 100, 2) if exit_price else round((cur_price - entry) / entry * 100, 2)

                _pu.execute("""
                    UPDATE forward_signals SET
                        status=?, exit_date=?, exit_price=?, exit_reason=?,
                        max_high=?, trailing_sl=?, pnl_pct=?, updated_at=?
                    WHERE id=?
                """, (
                    new_status,
                    now_str if new_status != "OPEN" else None,
                    exit_price if new_status != "OPEN" else None,
                    exit_reason if new_status != "OPEN" else None,
                    new_max, cur_trail, pnl,
                    datetime.now().isoformat(),
                    sig_id
                ))
                _pu.commit()
                _pu.close()
                if new_status != "OPEN":
                    updated += 1
                    print(f"  [FWD ] {sym} → {new_status} @ ₹{exit_price} ({pnl:+.2f}%)", flush=True)

            return {"ok": True, "open_checked": len(open_sigs), "exited": updated, "date": now_str}

        # ── forward_test_report ───────────────────────────────────────────────
        # Returns all forward signals (open + closed) for display/export
        if ep == "forward_test_report":
            status_filter = p.get("status", "")   # OPEN/SL_HIT/T1_HIT/T2_HIT/TRAILING_SL/EXPIRED/all
            sym_filter    = p.get("symbol", "").upper()
            days_back     = int(p.get("days", 90))
            cutoff        = (today - timedelta(days=days_back)).isoformat()

            _fr = _db()
            q   = "SELECT * FROM forward_signals WHERE signal_date >= ?"
            qp  = [cutoff]
            if status_filter and status_filter.upper() != "ALL":
                q  += " AND status=?"
                qp.append(status_filter.upper())
            if sym_filter:
                q  += " AND symbol=?"
                qp.append(sym_filter)
            q += " ORDER BY signal_date DESC, confidence DESC"

            cur  = _fr.execute(q, qp)
            cols = [d[0] for d in cur.description]
            rows = cur.fetchall()
            _fr.close()

            signals = []
            for r in rows:
                d = dict(zip(cols, r))
                try:
                    d["reasons"] = json.loads(d.get("reasons") or "[]")
                except Exception:
                    d["reasons"] = []
                signals.append(d)

            # Summary stats
            total   = len(signals)
            open_n  = sum(1 for s in signals if s["status"] == "OPEN")
            wins    = sum(1 for s in signals if s["status"] in ("T1_HIT","T2_HIT","TRAILING_SL") and (s.get("pnl_pct") or 0) > 0)
            losses  = sum(1 for s in signals if s["status"] == "SL_HIT")
            avg_pnl = round(sum(s.get("pnl_pct") or 0 for s in signals if s["status"] != "OPEN") / max(1, total - open_n), 2)

            return {
                "ok":       True,
                "signals":  signals,
                "total":    total,
                "open":     open_n,
                "wins":     wins,
                "losses":   losses,
                "avg_pnl_pct": avg_pnl,
                "win_rate": round(wins / max(1, wins + losses) * 100, 1),
                "days":     days_back,
            }
        # ── v3.9: ML model management endpoints ──────────────────────────────
        if ep == "ml_status":
            return {"ok": True, "status": get_model_status()}

        if ep == "ml_train":
            import threading as _thr2
            def _do_train():
                try:
                    meta = train_models(lookback_years=int(p.get("years",3)),
                                        forward_days=int(p.get("forward_days",10)),
                                        verbose=True)
                    print(f"  [ML] Training complete: {meta}", flush=True)
                except Exception as _te:
                    print(f"  [ML] Training error: {_te}", flush=True)
            _thr2.Thread(target=_do_train, daemon=True, name="ML_TRAIN").start()
            return {"ok": True, "message": "Model training started in background. Check logs for progress."}

        # ── market_brain_digest ───────────────────────────────────────────────
        # Returns overall market mood, sector sentiments, and key events from
        # the local news_sentiment database (built by bulk_news_fetch.py).
        if ep == "market_brain_digest":
            if _market_brain_local is None:
                return {
                    "digest_date": today.isoformat(),
                    "mood_state": "NEUTRAL",
                    "mood_score": 0.0,
                    "mood_rationale": "Market Brain module not loaded.",
                    "confidence": 0.0,
                    "narrative": "market_brain_local module failed to import. Check server logs.",
                    "sector_sentiments": [],
                    "key_events": [],
                    "error": "module_not_loaded",
                }
            try:
                return _market_brain_local.get_market_brain_digest(DB_PATH)
            except Exception as _mbd_e:
                return {
                    "digest_date": today.isoformat(),
                    "mood_state": "NEUTRAL",
                    "mood_score": 0.0,
                    "mood_rationale": f"Digest error: {_mbd_e}",
                    "confidence": 0.0,
                    "narrative": "Could not build digest. Ensure news_sentiment table is populated.",
                    "sector_sentiments": [],
                    "key_events": [],
                    "error": str(_mbd_e),
                }

        # ── market_brain_ask ─────────────────────────────────────────────────
        # Answers a natural-language question about the market using local news DB.
        if ep == "market_brain_ask":
            query = p.get("query", "").strip()
            if not query:
                return {"answer": "Please provide a question (query parameter).", "error": "empty_query"}
            if _market_brain_local is None:
                return {"answer": "Market Brain module not available.", "error": "module_not_loaded"}
            try:
                answer = _market_brain_local.local_ask_market_brain(query, DB_PATH)
                return {"answer": answer, "query": query}
            except Exception as _mba_e:
                return {"answer": f"Error: {_mba_e}", "error": str(_mba_e)}

        # ── v3.9: Portfolio Management endpoints ──
        if ep == "portfolio_get":
            try:
                conn = _db()
                user_id = self.current_user["user_id"]
                pf_id = _get_user_portfolio_id(conn, user_id)

                # PHASE 1 FIX (C1): Guardian Risk Engine REMOVED from here.
                # It now runs once at 15:35 IST in scheduler.run_guardian_eod().
                # Running it on every page load caused:
                #   - Stale-price auto-closes (using days-old close prices)
                #   - PnL recorded at close price instead of SL/T2 level
                #   - Potential double-fire from concurrent browser tabs

                sym = p.get("symbol", "").strip()
                if sym:
                    cursor = conn.execute("SELECT * FROM positions WHERE symbol=? AND portfolio_id=? ORDER BY created_at DESC", (sym, pf_id))
                else:
                    cursor = conn.execute("SELECT * FROM positions WHERE portfolio_id=? ORDER BY created_at DESC", (pf_id,))
                
                rows = cursor.fetchall()
                cols = [desc[0] for desc in cursor.description]
                conn.close()
                
                trades = [dict(zip(cols, r)) for r in rows]
                
                # Add 'protected' flag if SL was moved to break even
                for t in trades:
                    if t['status'] == 'OPEN' and t['stop_loss'] >= t['entry_price']:
                        t['protected'] = True

                # Calculate contract-required fields
                open_trades = [t for t in trades if t['status'] == 'OPEN']
                closed_trades = [t for t in trades if t['status'] == 'CLOSED']
                positions = open_trades
                
                total_invested = 0.0
                total_exposure = 0.0

                # PHASE 2 FIX (U1): Replace N+1 DB connection loop with single batch query.
                # Previously opened+closed a new DB connection PER open trade (~5ms each).
                open_syms = list({pos.get("symbol") for pos in open_trades if pos.get("symbol")})
                conn_batch = _db()
                if open_syms:
                    placeholders = ",".join("?" * len(open_syms))
                    # Get latest close for each symbol in ONE query (optimized with INNER JOIN)
                    px_rows = conn_batch.execute(
                        f"SELECT d.symbol, d.close "
                        f"FROM daily_prices d "
                        f"INNER JOIN ( "
                        f"    SELECT symbol, MAX(trade_date) AS max_date "
                        f"    FROM daily_prices "
                        f"    WHERE symbol IN ({placeholders}) "
                        f"    GROUP BY symbol "
                        f") m ON d.symbol = m.symbol AND d.trade_date = m.max_date",
                        open_syms
                    ).fetchall()
                    conn_batch.close()
                    price_map = {r[0]: float(r[1]) for r in px_rows if r[1] is not None}
                else:
                    conn_batch.close()
                    price_map = {}

                for pos in open_trades:
                    shares   = float(pos.get("shares", 0))
                    entry_px = float(pos.get("entry_price", 0))
                    total_invested += entry_px * shares
                    sym = pos.get("symbol", "")
                    cmp = price_map.get(sym, entry_px)
                    total_exposure += cmp * shares
                    pos["cmp"]            = round(cmp, 2)
                    pos["unrealized_pnl"] = round((cmp - entry_px) * shares, 2)
                    pos["unrealized_pct"] = round((cmp - entry_px) / entry_px * 100, 2) if entry_px else 0.0
                    
                realized_pnl = sum(float(t.get("realized_pnl") or 0.0) for t in closed_trades)
                unrealized_pnl = total_exposure - total_invested
                
                return {
                    "ok": True,
                    "positions": positions,
                    "total_invested": round(total_invested, 2),
                    "total_exposure": round(total_exposure, 2),
                    "realized_pnl": round(realized_pnl, 2),
                    "unrealized_pnl": round(unrealized_pnl, 2),
                    "trades": trades
                }
            except Exception as e:
                return {"ok": False, "error": str(e)}

        if ep == "portfolio_add":
            # PHASE 2 FIX (Fix 6): Risk gate validation before insert.
            # PHASE 2 FIX (Fix 10): Round-based position sizing + capital validation.
            try:
                _c = _db()
                user_id = self.current_user["user_id"]
                pf_id = _get_user_portfolio_id(_c, user_id)
                import uuid

                sym         = p.get("symbol", "").strip().upper()
                inv_type    = p.get("inv_type", "swing")
                entry_price = float(p.get("entry_price", 0))
                shares_in   = int(p.get("shares", 0))
                stop_loss   = float(p.get("stop_loss", 0))
                target1     = float(p.get("target1", 0))
                target2     = float(p.get("target2", 0))
                skip_risk   = bool(int(p.get("skip_risk_check", 0)))  # admin override

                if not sym:
                    return {"ok": False, "error": "symbol is required"}
                if entry_price <= 0:
                    return {"ok": False, "error": "entry_price must be > 0"}

                # ── Fetch user risk settings ──────────────────────────────────
                rs_row = _c.execute(
                    "SELECT capital, max_risk_pct, max_positions, daily_loss_limit, "
                    "max_sector_pct, max_position_pct, max_correlation_exposure, kill_switch "
                    "FROM risk_settings WHERE user_id=?", (user_id,)
                ).fetchone()

                if rs_row:
                    risk_settings = {
                        "capital":                  float(rs_row[0]) if rs_row[0] else 100000.0,
                        "max_risk_pct":             float(rs_row[1]) if rs_row[1] else 2.0,
                        "max_positions":            int(rs_row[2])   if rs_row[2] else 5,
                        "daily_loss_limit":         float(rs_row[3]) if rs_row[3] else 50000.0,
                        "max_sector_pct":           float(rs_row[4]) if rs_row[4] else 30.0,
                        "max_position_pct":         float(rs_row[5]) if rs_row[5] else 10.0,
                        "max_correlation_exposure": float(rs_row[6]) if rs_row[6] else 0.7,
                        "kill_switch":              bool(rs_row[7]),
                        "portfolio_id":             pf_id,
                    }
                else:
                    # No saved settings yet — use safe defaults
                    risk_settings = {
                        "capital": 100000.0, "max_risk_pct": 2.0, "max_positions": 5,
                        "daily_loss_limit": 50000.0, "max_sector_pct": 30.0,
                        "max_position_pct": 10.0, "max_correlation_exposure": 0.7,
                        "kill_switch": False, "portfolio_id": pf_id,
                    }

                # ── PHASE 2 FIX (Fix 10): Smart position sizing ───────────────
                # If caller sends shares=0, compute from risk settings automatically.
                if shares_in == 0 and stop_loss > 0 and entry_price > stop_loss:
                    capital      = risk_settings["capital"]
                    price_diff   = abs(entry_price - stop_loss)
                    risk_amount  = capital * (risk_settings["max_risk_pct"] / 100.0)
                    # Use round() not int() to prevent systematic underfunding
                    shares_in    = max(round(risk_amount / price_diff), 1)
                    # Clamp to max_position_pct
                    max_pos_val  = capital * (risk_settings["max_position_pct"] / 100.0)
                    if shares_in * entry_price > max_pos_val:
                        shares_in = max(int(max_pos_val / entry_price), 1)

                shares = max(shares_in, 1)

                # ── PHASE 2 FIX (Fix 10): Capital validation ─────────────────
                from core.portfolio_state import get_available_capital
                avail = get_available_capital(
                    risk_settings=risk_settings, portfolio_id=pf_id, user_id=user_id
                )
                position_cost = shares * entry_price
                if position_cost > avail and not skip_risk:
                    _c.close()
                    return {
                        "ok": False,
                        "error": f"Insufficient capital. Trade costs ₹{position_cost:,.2f} but only ₹{avail:,.2f} available.",
                        "required": round(position_cost, 2),
                        "available": round(avail, 2)
                    }

                # ── PHASE 2 FIX (Fix 6): Risk gate candidate validation ───────
                if not skip_risk:
                    from core.risk_gates import validate_candidate
                    from data.instruments import ALL_INSTRUMENTS as _AI
                    _inst = _AI.get(sym)
                    candidate = {
                        "symbol":     sym,
                        "sector":     _inst.sector if _inst else "Other",
                        "entry":      entry_price,
                        "stop_loss":  stop_loss if stop_loss > 0 else entry_price * 0.95,
                        "portfolio_id": pf_id,
                    }
                    current_alloc = {"total_deployed_capital": (risk_settings["capital"] - avail)}
                    passed, reason = validate_candidate(candidate, current_alloc, risk_settings)
                    if not passed:
                        _c.close()
                        return {"ok": False, "error": f"Risk gate blocked: {reason}", "risk_reason": reason}

                source_signal_id = p.get("source_signal_id")
                if source_signal_id:
                    try:
                        source_signal_id = int(source_signal_id)
                    except ValueError:
                        pass

                # ── Insert position ───────────────────────────────────────────
                trade_id   = str(uuid.uuid4())
                created_at = datetime.utcnow().isoformat()
                entry_date = today.isoformat()

                try:
                    _c.execute("""
                        INSERT INTO positions
                            (id, portfolio_id, symbol, inv_type, entry_date, entry_price,
                             shares, stop_loss, target1, target2, status,
                             lifecycle_state, source_signal_id, created_at, updated_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'OPEN', 'OPEN', ?, ?, ?)""",
                        (trade_id, pf_id, sym, inv_type, entry_date, entry_price,
                         shares, stop_loss, target1, target2, source_signal_id, created_at, created_at))
                    _c.commit()
                except sqlite3.IntegrityError:
                    _c.close()
                    return {
                        "ok": False,
                        "error": f"A trade for {sym} on {entry_date} already exists in the portfolio."
                    }
                _c.close()
                return {
                    "ok": True,
                    "message": "Trade added to portfolio",
                    "trade_id": trade_id,
                    "shares":   shares,
                    "position_cost": round(position_cost, 2),
                    "available_after": round(avail - position_cost, 2)
                }
            except Exception as e:
                import traceback as _tb6
                return {"ok": False, "error": str(e), "trace": _tb6.format_exc()[-400:]}

        if ep == "portfolio_close":
            try:
                _c = _db()
                user_id = self.current_user["user_id"]
                pf_id = _get_user_portfolio_id(_c, user_id)

                trade_id   = p.get("id", "")
                exit_price = float(p.get("exit_price", 0))
                exit_reason = p.get("exit_reason", "Manual Exit")
                exit_date  = today.isoformat()

                trade = _c.execute(
                    "SELECT entry_price, shares FROM positions WHERE id=? AND portfolio_id=?",
                    (trade_id, pf_id)
                ).fetchone()
                if not trade:
                    return {"ok": False, "error": "Trade not found"}

                entry_price, shares = float(trade[0]), int(trade[1])
                from core.indicators import calculate_transaction_costs
                tx_costs = calculate_transaction_costs(entry_price, exit_price, shares)
                realized_pnl = round((exit_price - entry_price) * shares - tx_costs, 2)
                now_iso = datetime.utcnow().isoformat()

                _c.execute("""
                    UPDATE positions
                    SET status='CLOSED', exit_date=?, exit_price=?,
                        realized_pnl=?, exit_reason=?,
                        lifecycle_state='CLOSED', updated_at=?
                    WHERE id=? AND portfolio_id=?""",
                    (exit_date, exit_price, realized_pnl, exit_reason,
                     now_iso, trade_id, pf_id))
                _c.commit()
                _c.close()
                return {"ok": True, "message": "Trade squared off successfully",
                        "realized_pnl": realized_pnl}
            except Exception as e:
                return {"ok": False, "error": str(e)}

        # ── portfolio_modify — edit SL / targets live ────────────────────────
        if ep == "portfolio_modify":
            try:
                _c = _db()
                user_id = self.current_user["user_id"]
                pf_id = _get_user_portfolio_id(_c, user_id)
                
                trade_id  = p.get("id", "")
                new_sl    = float(p.get("stop_loss", 0))
                new_t1    = float(p.get("target1", 0))
                new_t2    = float(p.get("target2", 0))
                reason    = p.get("change_reason", "Manual adjustment")

                if not trade_id:
                    _c.close()
                    return {"ok": False, "error": "id required"}

                # Fetch current values for audit logging
                old_row = _c.execute(
                    "SELECT stop_loss, target1, target2 FROM positions WHERE id=? AND portfolio_id=?",
                    (trade_id, pf_id)
                ).fetchone()

                if not old_row:
                    _c.close()
                    return {"ok": False, "error": "Position not found"}

                old_sl = float(old_row[0] or 0)
                old_t1 = float(old_row[1] or 0)
                old_t2 = float(old_row[2] or 0)
                now_iso = datetime.utcnow().isoformat()

                # Audit individual changes
                if abs(old_sl - new_sl) > 0.01:
                    _c.execute("""
                        INSERT INTO position_audit_log (position_id, field, old_value, new_value, changed_at, change_reason)
                        VALUES (?, 'stop_loss', ?, ?, ?, ?)
                    """, (trade_id, str(old_sl), str(new_sl), now_iso, reason))

                if abs(old_t1 - new_t1) > 0.01:
                    _c.execute("""
                        INSERT INTO position_audit_log (position_id, field, old_value, new_value, changed_at, change_reason)
                        VALUES (?, 'target1', ?, ?, ?, ?)
                    """, (trade_id, str(old_t1), str(new_t1), now_iso, reason))

                if abs(old_t2 - new_t2) > 0.01:
                    _c.execute("""
                        INSERT INTO position_audit_log (position_id, field, old_value, new_value, changed_at, change_reason)
                        VALUES (?, 'target2', ?, ?, ?, ?)
                    """, (trade_id, str(old_t2), str(new_t2), now_iso, reason))

                # Update the position record
                _c.execute("""
                    UPDATE positions
                    SET stop_loss=?, target1=?, target2=?, updated_at=?
                    WHERE id=? AND portfolio_id=?
                """, (new_sl, new_t1, new_t2, now_iso, trade_id, pf_id))

                _c.commit()
                _c.close()
                return {"ok": True, "message": "Order updated successfully"}
            except Exception as e:
                return {"ok": False, "error": str(e)}

        # ── portfolio_partial_exit — close N shares, keep remainder ──────────
        if ep == "portfolio_partial_exit":
            try:
                _c = _db()
                user_id = self.current_user["user_id"]
                pf_id = _get_user_portfolio_id(_c, user_id)
                import uuid
                
                trade_id   = p.get("id", "")
                exit_shares = int(p.get("shares", 0))
                exit_price  = float(p.get("exit_price", 0))
                if not trade_id or not exit_shares:
                    return {"ok": False, "error": "id and shares required"}
                row = _c.execute(
                    "SELECT symbol,inv_type,entry_price,shares,stop_loss,target1,target2,entry_date FROM positions WHERE id=? AND portfolio_id=?",
                    (trade_id, pf_id)).fetchone()
                if not row:
                    return {"ok": False, "error": "Trade not found"}
                sym, inv_type, entry, total_shares, sl, t1, t2, entry_date = row
                exit_shares = min(exit_shares, total_shares)
                remain      = total_shares - exit_shares
                
                from core.indicators import calculate_transaction_costs
                tx_costs = calculate_transaction_costs(entry, exit_price, exit_shares)
                partial_pnl = round((exit_price - entry) * exit_shares - tx_costs, 2)
                
                if remain > 0:
                    # Shrink original trade qty
                    _c.execute("UPDATE positions SET shares=? WHERE id=? AND portfolio_id=?",
                               (remain, trade_id, pf_id))
                    # Create new closed record for the exited portion
                    _c.execute("""INSERT INTO positions
                        (id, portfolio_id, symbol,inv_type,entry_date,entry_price,shares,stop_loss,target1,target2,
                         status,exit_date,exit_price,realized_pnl,created_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'CLOSED', ?, ?, ?, ?)""",
                        (str(uuid.uuid4()), pf_id, sym, inv_type, entry_date, entry, exit_shares, sl, t1, t2,
                         today.isoformat(), exit_price, partial_pnl, datetime.utcnow().isoformat()))
                else:
                    # Full exit
                    _c.execute("""UPDATE positions SET
                        status='CLOSED',exit_date=?,exit_price=?,realized_pnl=?
                        WHERE id=? AND portfolio_id=?""",
                        (today.isoformat(), exit_price, partial_pnl, trade_id, pf_id))
                _c.commit(); _c.close()
                return {"ok": True, "partial_pnl": partial_pnl, "remaining_shares": remain}
            except Exception as e:
                return {"ok": False, "error": str(e)}

        # ── portfolio_csv — export trade history as CSV ──────────────────────
        if ep == "portfolio_csv":
            import io as _io_csv
            try:
                _c = _db()
                user_id = self.current_user["user_id"]
                pf_id = _get_user_portfolio_id(_c, user_id)
                rows = _c.execute("""
                    SELECT symbol,inv_type,entry_date,entry_price,shares,
                           stop_loss,target1,target2,status,exit_date,exit_price,
                           realized_pnl,created_at
                    FROM positions WHERE portfolio_id=? ORDER BY created_at DESC
                """, (pf_id,)).fetchall()
                _c.close()
                buf = _io_csv.StringIO()
                buf.write("Symbol,Type,Entry Date,Entry Price,Shares,SL,T1,T2,Status,Exit Date,Exit Price,Realized P&L,Created At\n")
                for r in rows:
                    buf.write(",".join(str(x or "") for x in r) + "\n")
                csv_bytes = buf.getvalue().encode("utf-8")
                self.send_response(200)
                self.send_header("Content-Type", "text/csv")
                self.send_header("Content-Disposition", 'attachment; filename="trades.csv"')
                self.send_header("Content-Length", str(len(csv_bytes)))
                self.end_headers()
                self.wfile.write(csv_bytes)
                return None  # already sent
            except Exception as e:
                return {"ok": False, "error": str(e)}

        # ── watchlist_get ────────────────────────────────────────────────────
        if ep == "watchlist_get":
            try:
                _c = _db()
                rows = _c.execute(
                    "SELECT id,symbol,added_at,notes FROM watchlist_items ORDER BY added_at DESC"
                ).fetchall()
                _c.close()
                items = [{"id": r[0], "symbol": r[1], "added_at": r[2], "notes": r[3]} for r in rows]
                # Enrich with latest price
                cache = get_cached_prices()
                for item in items:
                    pr = cache.get(item["symbol"], {})
                    item["price"]      = round(float(pr.get("close") or 0), 2) if pr.get("close") else None
                    item["change_pct"] = round(float(pr.get("change_pct") or 0), 2)
                    item["high"]       = round(float(pr.get("high") or 0), 2)
                    item["low"]        = round(float(pr.get("low") or 0), 2)
                    inst = ALL_INSTRUMENTS.get(item["symbol"])
                    item["name"]   = inst.name if inst else item["symbol"]
                    item["sector"] = inst.sector if inst else ""
                symbols = [item["symbol"] for item in items]
                return {"ok": True, "symbols": symbols, "items": items}
            except Exception as e:
                return {"ok": False, "error": str(e)}

        # ── watchlist_add ─────────────────────────────────────────────────────
        if ep == "watchlist_add":
            try:
                sym = p.get("symbol", "").upper().strip()
                if not sym:
                    return {"ok": False, "error": "symbol required"}
                if sym not in ALL_INSTRUMENTS:
                    return {"ok": False, "error": f"Unknown symbol: {sym}"}
                _c = _db()
                _c.execute(
                    "INSERT OR IGNORE INTO watchlist_items(symbol,added_at) VALUES(?,?)",
                    (sym, datetime.now().isoformat()))
                _c.commit(); _c.close()
                return {"ok": True, "symbol": sym}
            except Exception as e:
                return {"ok": False, "error": str(e)}

        # ── watchlist_remove ──────────────────────────────────────────────────
        if ep == "watchlist_remove":
            try:
                sym = p.get("symbol", "").upper().strip()
                _c = _db()
                _c.execute("DELETE FROM watchlist_items WHERE symbol=?", (sym,))
                _c.commit(); _c.close()
                return {"ok": True, "symbol": sym}
            except Exception as e:
                return {"ok": False, "error": str(e)}

        # ── market_depth — simulated 5-level bid/ask from OHLCV ───────────────
        if ep == "market_depth":
            sym = p.get("symbol", "").upper()
            if not sym:
                return {"ok": False, "error": "symbol required"}
            try:
                import math as _mdm
                cache = get_cached_prices()
                pr = cache.get(sym, {})
                cmp  = float(pr.get("close") or 0) if pr.get("close") else 0
                hi   = float(pr.get("high") or cmp * 1.01)
                lo   = float(pr.get("low") or cmp * 0.99)
                if cmp == 0:
                    return {"ok": False, "error": "No price data"}
                # Build synthetic depth from tick size
                tick = max(round(cmp * 0.001, 2), 0.05)
                import random as _rnd
                _rnd.seed(int(cmp * 100))  # deterministic per price
                bids, asks = [], []
                for i in range(1, 6):
                    bpx = round(cmp - tick * i, 2)
                    apx = round(cmp + tick * i, 2)
                    bqty = int(abs(_rnd.gauss(5000, 2000)) * (1 / i))
                    aqty = int(abs(_rnd.gauss(4500, 2000)) * (1 / i))
                    bids.append({"price": bpx, "qty": bqty, "orders": _rnd.randint(3, 20)})
                    asks.append({"price": apx, "qty": aqty, "orders": _rnd.randint(3, 20)})
                total_bid = sum(b["qty"] for b in bids)
                total_ask = sum(a["qty"] for a in asks)
                imbalance = round((total_bid - total_ask) / max(total_bid + total_ask, 1) * 100, 1)
                return {
                    "ok": True, "symbol": sym, "cmp": cmp,
                    "high": hi, "low": lo,
                    "bids": bids, "asks": asks,
                    "total_bid_qty": total_bid, "total_ask_qty": total_ask,
                    "imbalance_pct": imbalance,
                    "note": "Simulated depth (real Level-2 requires broker API)"
                }
            except Exception as e:
                return {"ok": False, "error": str(e)}

        # ── alert_get — list all price alerts ──────────────────────────────────
        if ep == "alert_get":
            try:
                _c = _db()
                rows = _c.execute(
                    "SELECT id,symbol,condition,threshold,notify_browser,notify_whatsapp,status,triggered_at,created_at FROM price_alerts ORDER BY id DESC"
                ).fetchall()
                _c.close()
                alerts = [{"id":r[0],"symbol":r[1],"condition":r[2],"threshold":r[3],
                           "notify_browser":bool(r[4]),"notify_whatsapp":bool(r[5]),
                           "status":r[6],"triggered_at":r[7],"created_at":r[8]} for r in rows]
                return {"ok": True, "alerts": alerts}
            except Exception as e:
                return {"ok": False, "error": str(e)}

        # ── alert_set — create or update a price alert ─────────────────────────
        if ep == "alert_set":
            try:
                sym       = p.get("symbol", "").upper()
                cond      = p.get("condition", "ABOVE").upper()
                threshold = float(p.get("threshold", 0))
                n_browser = int(p.get("notify_browser", 1))
                n_wa      = int(p.get("notify_whatsapp", 0))
                if not sym or not threshold:
                    return {"ok": False, "error": "symbol and threshold required"}
                _c = _db()
                _c.execute("""INSERT INTO price_alerts
                    (symbol,condition,threshold,notify_browser,notify_whatsapp,status,created_at)
                    VALUES(?,?,?,?,?,'ACTIVE',?)""",
                    (sym, cond, threshold, n_browser, n_wa, datetime.now().isoformat()))
                _c.commit(); _c.close()
                return {"ok": True, "symbol": sym, "condition": cond, "threshold": threshold}
            except Exception as e:
                return {"ok": False, "error": str(e)}

        # ── alert_delete — remove a price alert ──────────────────────────────
        if ep == "alert_delete":
            try:
                alert_id = int(p.get("id", 0))
                _c = _db()
                _c.execute("DELETE FROM price_alerts WHERE id=?", (alert_id,))
                _c.commit(); _c.close()
                return {"ok": True, "deleted_id": alert_id}
            except Exception as e:
                return {"ok": False, "error": str(e)}

        # ── alert_check — check alerts against current prices (called by UI) ──
        if ep == "alert_check":
            try:
                cache = get_cached_prices()
                _c = _db()
                active = _c.execute(
                    "SELECT id,symbol,condition,threshold FROM price_alerts WHERE status='ACTIVE'"
                ).fetchall()
                triggered = []
                for row in active:
                    aid, sym, cond, thresh = row
                    pr = cache.get(sym, {})
                    cmp = float(pr.get("close") or 0) if pr.get("close") else 0
                    if cmp == 0: continue
                    hit = False
                    if cond == "ABOVE" and cmp >= thresh: hit = True
                    elif cond == "BELOW" and cmp <= thresh: hit = True
                    elif cond == "PCT_UP" and float(pr.get("change_pct") or 0) >= thresh: hit = True
                    elif cond == "PCT_DOWN" and float(pr.get("change_pct") or 0) <= -thresh: hit = True
                    if hit:
                        _c.execute("UPDATE price_alerts SET status='TRIGGERED', triggered_at=? WHERE id=?",
                                   (datetime.now().isoformat(), aid))
                        triggered.append({"id": aid, "symbol": sym, "condition": cond,
                                          "threshold": thresh, "cmp": cmp})
                _c.commit(); _c.close()
                return {"ok": True, "triggered": triggered}
            except Exception as e:
                return {"ok": False, "error": str(e)}

        # ── risk_dashboard — portfolio risk metrics ────────────────────────────
        if ep == "risk_dashboard":
            try:
                import math as _mrk
                _c = _db()
                user_id = self.current_user["user_id"]
                pf_id = _get_user_portfolio_id(_c, user_id)
                
                open_trades = _c.execute("""
                    SELECT symbol,inv_type,entry_date,entry_price,shares,stop_loss,target1,target2
                    FROM positions WHERE status='OPEN' AND portfolio_id=?
                """, (pf_id,)).fetchall()
                
                closed_trades = _c.execute("""
                    SELECT entry_price,exit_price,shares,realized_pnl,entry_date,exit_date
                    FROM positions WHERE status='CLOSED' AND exit_price IS NOT NULL AND portfolio_id=?
                """, (pf_id,)).fetchall()

                # Get user starting capital
                capital = 100000.0
                profile_row = _c.execute("SELECT starting_capital FROM risk_profiles WHERE user_id=?", (user_id,)).fetchone()
                if profile_row:
                    capital = float(profile_row[0])
                    
                _c.close()

                cache = get_cached_prices()
                positions = []
                total_inv, total_cur, total_exp = 0.0, 0.0, 0.0
                sector_exposure = {}

                for tr in open_trades:
                    sym, inv_type, edate, eprice, shares, sl, t1, t2 = tr
                    pr = cache.get(sym, {})
                    cmp = float(pr.get("close") or eprice) if pr.get("close") else eprice
                    unr_pnl = round((cmp - eprice) * shares, 2)
                    unr_pct = round((cmp - eprice) / eprice * 100, 2)
                    inv_amt = round(eprice * shares, 2)
                    cur_val = round(cmp * shares, 2)
                    total_inv += inv_amt; total_cur += cur_val
                    total_exp += cur_val
                    inst = ALL_INSTRUMENTS.get(sym)
                    sector = inst.sector if inst else "Other"
                    sector_exposure[sector] = sector_exposure.get(sector, 0) + cur_val
                    days = max(1, (date.today() - date.fromisoformat(edate[:10])).days) if edate else 1
                    positions.append({"symbol": sym, "entry": eprice, "cmp": cmp,
                                      "shares": shares, "unrealized_pnl": unr_pnl,
                                      "unrealized_pct": unr_pct, "days_held": days,
                                      "invested": inv_amt, "current": cur_val, "sector": sector})

                # Closed trade stats
                wins = [t for t in closed_trades if (t[3] or 0) > 0]
                losses = [t for t in closed_trades if (t[3] or 0) < 0]
                win_rate = round(len(wins) / max(len(closed_trades), 1) * 100, 1)
                total_realized = sum(t[3] or 0 for t in closed_trades)

                # Sharpe, Profit Factor, Expectancy
                import math as _math
                gross_profit = sum(float(t[3] or 0) for t in wins)
                gross_loss   = sum(abs(float(t[3] or 0)) for t in losses)
                profit_factor = round(gross_profit / max(gross_loss, 1.0), 2) if gross_loss > 0 else (round(gross_profit, 2) if gross_profit > 0 else 1.0)
                expectancy = round(total_realized / max(len(closed_trades), 1), 2)

                pnl_vals = [float(t[3] or 0) for t in closed_trades]
                if len(pnl_vals) >= 2:
                    mean_pnl = sum(pnl_vals) / len(pnl_vals)
                    var_pnl  = sum((x - mean_pnl) ** 2 for x in pnl_vals) / (len(pnl_vals) - 1)
                    std_pnl  = _math.sqrt(var_pnl)
                    sharpe_ratio = round(mean_pnl / std_pnl, 2) if std_pnl > 0 else 0.0
                else:
                    sharpe_ratio = 0.0

                # Max Drawdown from closed trades
                cum_pnl = 0.0; peak = 0.0; max_dd = 0.0
                for t in sorted(closed_trades, key=lambda x: x[5] or ""):  # x[5] is exit_date (index 5)
                    pnl = float(t[3] or 0)
                    cum_pnl += pnl
                    peak = max(peak, cum_pnl)
                    dd = (peak - cum_pnl) / max(abs(peak), 1) * 100
                    max_dd = max(max_dd, dd)

                # VaR (95%) approximate from closed trades
                pnl_list = sorted([float(t[3] or 0) for t in closed_trades])
                var_95 = round(pnl_list[int(len(pnl_list) * 0.05)] if len(pnl_list) >= 20 else 0, 2)

                # Sector % breakdown
                sector_pct = {s: round(v / max(total_exp, 1) * 100, 1) for s, v in sector_exposure.items()}

                # Portfolio health score
                health = 100
                if max_dd > 20: health -= 30
                elif max_dd > 10: health -= 15
                if win_rate < 40: health -= 20
                elif win_rate > 60: health += 10
                if total_exp > 0 and max(sector_pct.values() or [0]) > 40: health -= 15
                health = max(0, min(100, health))

                var_95_pct = round(var_95 / capital * 100, 2) if capital > 0 else 0.0
                sector_list = [{"name": s, "pct": pct} for s, pct in sector_pct.items()]

                return {
                    "ok": True,
                    "open_positions": len(open_trades),
                    "total_invested": round(total_inv, 2),
                    "total_exposure": round(total_cur, 2),
                    "unrealized_pnl": round(total_cur - total_inv, 2),
                    "realized_pnl":   round(total_realized, 2),
                    "max_drawdown_pct": round(max_dd, 2),
                    "var_95":         var_95,
                    "var_95_pct":     var_95_pct,
                    "win_pct":        win_rate,
                    "health_score":   health,
                    "positions":      positions,
                    "sector_exposure": sector_list,
                    "total_trades":   len(closed_trades),
                    "sharpe_ratio":   sharpe_ratio,
                    "profit_factor":  profit_factor,
                    "expectancy":     expectancy,
                }
            except Exception as e:
                import traceback as _tb
                return {"ok": False, "error": str(e), "trace": _tb.format_exc()[-400:]}


        # ── kill_switch_activate ──────────────────────────────────────────────────────────────────
        if ep == "kill_switch_activate":
            # PHASE 2 FIX (Fix 8):
            #   action="new_trades_only"  -> sets kill_switch=1 (blocks portfolio_add)
            #   action="close_all"        -> sets kill_switch=1 AND closes every OPEN position
            try:
                user_id = self.current_user["user_id"]
                action  = p.get("action", "new_trades_only").lower()
                reason  = p.get("reason", "Manual kill switch activation")
                now_iso = datetime.utcnow().isoformat()

                conn = _db()
                conn.execute("""
                    INSERT INTO risk_settings (user_id, kill_switch, updated_at)
                    VALUES (?, 1, ?)
                    ON CONFLICT(user_id) DO UPDATE SET
                        kill_switch=1, updated_at=excluded.updated_at
                """, (user_id, now_iso))
                conn.commit()

                closed_count = 0
                closed_pnl   = 0.0

                if action == "close_all":
                    pf_id = _get_user_portfolio_id(conn, user_id)
                    open_pos = conn.execute(
                        "SELECT id, symbol, entry_price, shares "
                        "FROM positions WHERE status='OPEN' AND portfolio_id=?",
                        (pf_id,)
                    ).fetchall()

                    cache = get_cached_prices()
                    for pos_id, sym, entry_px, shares in open_pos:
                        entry_px = float(entry_px)
                        shares   = int(shares)
                        pr       = cache.get(sym, {})
                        cmp      = float(pr.get("close") or entry_px)
                        exit_px  = round(cmp * (1 - 0.001), 2)
                        pnl      = round((exit_px - entry_px) * shares, 2)
                        closed_pnl += pnl
                        conn.execute("""
                            UPDATE positions
                            SET status='CLOSED', exit_date=?, exit_price=?,
                                realized_pnl=?, exit_reason=?,
                                lifecycle_state='CLOSED', updated_at=?
                            WHERE id=?
                        """, (today.isoformat(), exit_px, pnl,
                              f"Kill Switch \u2014 {reason}", now_iso, pos_id))
                        closed_count += 1
                    conn.commit()

                conn.close()
                return {
                    "ok":           True,
                    "action":       action,
                    "kill_switch":  True,
                    "closed_count": closed_count,
                    "closed_pnl":   round(closed_pnl, 2),
                    "message":      (f"Kill switch activated. {closed_count} positions closed."
                                     if action == "close_all"
                                     else "Kill switch activated. No new trades allowed.")
                }
            except Exception as e:
                return {"ok": False, "error": str(e)}

        # ── daily_pnl_status ──────────────────────────────────────────────────────────────────
        if ep == "daily_pnl_status":
            # PHASE 2 FIX (Fix 7): Today's realized P&L vs daily loss limit.
            try:
                user_id = self.current_user["user_id"]
                conn    = _db()
                pf_id   = _get_user_portfolio_id(conn, user_id)
                today_s = today.isoformat()

                pnl_row = conn.execute(
                    "SELECT SUM(realized_pnl) FROM positions "
                    "WHERE portfolio_id=? AND status='CLOSED' AND exit_date=?",
                    (pf_id, today_s)
                ).fetchone()
                today_pnl = round(float(pnl_row[0]) if pnl_row and pnl_row[0] else 0.0, 2)

                rs_row = conn.execute(
                    "SELECT daily_loss_limit, kill_switch FROM risk_settings WHERE user_id=?",
                    (user_id,)
                ).fetchone()
                daily_limit = float(rs_row[0]) if rs_row and rs_row[0] else 50000.0
                kill_switch = bool(rs_row[1]) if rs_row else False
                conn.close()

                breached = today_pnl <= -daily_limit
                used_pct = round(abs(today_pnl) / daily_limit * 100, 1) if daily_limit > 0 else 0.0

                return {
                    "ok":          True,
                    "date":        today_s,
                    "today_pnl":   today_pnl,
                    "daily_limit": -daily_limit,
                    "used_pct":    used_pct,
                    "breached":    breached,
                    "kill_switch": kill_switch,
                    "status":      ("BREACH" if breached else
                                    "WARNING" if used_pct >= 75 else "NORMAL")
                }
            except Exception as e:
                return {"ok": False, "error": str(e)}

        # ── risk_settings_get ──────────────────────────────────────────────────────────────────────────
        if ep == "risk_settings_get":
            # PHASE 1 FIX (S2): Single source of truth - SQLite per-user row.
            # Previously split across JSON file + risk_profiles, so kill_switch
            # was invisible to risk_gates.validate_candidate().
            try:
                user_id = self.current_user["user_id"]
                conn = _db()
                rs_row = conn.execute(
                    "SELECT capital, max_risk_pct, max_positions, daily_loss_limit, "
                    "max_sector_pct, max_position_pct, max_correlation_exposure, kill_switch "
                    "FROM risk_settings WHERE user_id=?", (user_id,)
                ).fetchone()
                profile = conn.execute(
                    "SELECT starting_capital, max_position_pct, max_sector_pct "
                    "FROM risk_profiles WHERE user_id=?", (user_id,)
                ).fetchone()
                conn.close()
                if rs_row:
                    settings = {
                        "capital":                  float(rs_row[0]) if rs_row[0] else 100000.0,
                        "max_risk_pct":             float(rs_row[1]) if rs_row[1] else 2.0,
                        "max_positions":            int(rs_row[2])   if rs_row[2] else 5,
                        "daily_loss_limit":         float(rs_row[3]) if rs_row[3] else 50000.0,
                        "max_sector_pct":           float(rs_row[4]) if rs_row[4] else 30.0,
                        "max_position_pct":         float(rs_row[5]) if rs_row[5] else 10.0,
                        "max_correlation_exposure": float(rs_row[6]) if rs_row[6] else 0.7,
                        "kill_switch":              bool(rs_row[7]),
                    }
                else:
                    capital     = float(profile[0]) if profile and profile[0] else 100000.0
                    max_pos_pct = float(profile[1]) if profile and profile[1] else 10.0
                    max_sec_pct = float(profile[2]) if profile and profile[2] else 30.0
                    settings = {
                        "capital": capital, "max_risk_pct": 2.0, "max_positions": 5,
                        "daily_loss_limit": 50000.0, "max_sector_pct": max_sec_pct,
                        "max_position_pct": max_pos_pct,
                        "max_correlation_exposure": 0.7, "kill_switch": False,
                    }
                return {"ok": True, "settings": settings}
            except Exception as e:
                return {"ok": False, "error": str(e)}

        # ── risk_settings_save ──────────────────────────────────────────────────────────────────────────
        if ep == "risk_settings_save":
            # PHASE 1 FIX (S2): Write to per-user SQLite row, NOT a JSON file.
            # kill_switch is now stored in DB and visible to risk_gates.
            try:
                user_id     = self.current_user["user_id"]
                capital     = float(p.get("capital", 100000))
                max_risk    = float(p.get("max_risk_pct", 2.0))
                max_pos     = int(p.get("max_positions", 5))
                daily_limit = float(p.get("daily_loss_limit", 50000))
                max_sec     = float(p.get("max_sector_pct", 30.0))
                max_pos_pct = float(p.get("max_position_pct", 10.0))
                max_corr    = float(p.get("max_correlation_exposure", 0.7))
                kill        = int(bool(int(p.get("kill_switch", 0))))
                now_iso     = datetime.utcnow().isoformat()
                conn = _db()
                conn.execute("""
                    INSERT INTO risk_settings
                        (user_id, capital, max_risk_pct, max_positions, daily_loss_limit,
                         max_sector_pct, max_position_pct, max_correlation_exposure,
                         kill_switch, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(user_id) DO UPDATE SET
                        capital=excluded.capital, max_risk_pct=excluded.max_risk_pct,
                        max_positions=excluded.max_positions,
                        daily_loss_limit=excluded.daily_loss_limit,
                        max_sector_pct=excluded.max_sector_pct,
                        max_position_pct=excluded.max_position_pct,
                        max_correlation_exposure=excluded.max_correlation_exposure,
                        kill_switch=excluded.kill_switch, updated_at=excluded.updated_at
                """, (user_id, capital, max_risk, max_pos, daily_limit,
                      max_sec, max_pos_pct, max_corr, kill, now_iso))
                conn.execute("""
                    UPDATE risk_profiles SET
                        starting_capital=?, max_position_pct=?, max_sector_pct=?, updated_at=?
                    WHERE user_id=?
                """, (capital, max_pos_pct, max_sec, now_iso, user_id))
                conn.commit()
                conn.close()
                return {"ok": True, "message": "Risk settings saved"}
            except Exception as e:
                return {"ok": False, "error": str(e)}

        # ── correlation_matrix — price correlation between symbols ─────────────
        if ep == "correlation_matrix":
            import math as _mco
            syms = [s.strip().upper() for s in p.get("symbols", "").split(",") if s.strip()]
            if len(syms) < 2:
                return {"ok": False, "error": "Need at least 2 symbols (comma-separated)"}
            try:
                _c = _db()
                closes_map = {}
                for sym in syms:
                    rows = _c.execute("""
                        SELECT trade_date, close FROM daily_prices
                        WHERE symbol=? AND close IS NOT NULL
                        ORDER BY trade_date DESC LIMIT 120
                    """, (sym,)).fetchall()
                    closes_map[sym] = {r[0]: float(r[1]) for r in rows}
                _c.close()
                # Common dates
                all_dates = sorted(set.intersection(*[set(v.keys()) for v in closes_map.values()]), reverse=True)[:100]
                if len(all_dates) < 10:
                    return {"ok": False, "error": "Not enough common price history"}
                series = {sym: [closes_map[sym][d] for d in all_dates] for sym in syms}
                # Returns
                def _returns(prices):
                    return [(prices[i] - prices[i+1]) / prices[i+1] for i in range(len(prices)-1)]
                ret_map = {sym: _returns(series[sym]) for sym in syms}
                def _pearson(a, b):
                    n = len(a)
                    ma = sum(a)/n; mb = sum(b)/n
                    num = sum((a[i]-ma)*(b[i]-mb) for i in range(n))
                    da  = _mco.sqrt(sum((x-ma)**2 for x in a))
                    db  = _mco.sqrt(sum((x-mb)**2 for x in b))
                    return round(num / max(da * db, 1e-10), 4)
                matrix = [[_pearson(ret_map[s1], ret_map[s2]) for s2 in syms] for s1 in syms]
                # Flag highly correlated pairs
                high_corr = []
                for i in range(len(syms)):
                    for j in range(i+1, len(syms)):
                        if abs(matrix[i][j]) >= 0.75:
                            high_corr.append({"sym1": syms[i], "sym2": syms[j], "corr": matrix[i][j]})
                return {"ok": True, "symbols": syms, "matrix": matrix,
                        "high_corr_pairs": high_corr, "n_dates": len(all_dates)}
            except Exception as e:
                return {"ok": False, "error": str(e)}

        # ── analytics_data — Performance analytics data ────────────────────────
        if ep == "analytics_data":
            dtype = p.get("type", "calendar")
            try:
                _c = _db()
                user_id = self.current_user["user_id"]
                pf_id = _get_user_portfolio_id(_c, user_id)
                closed = _c.execute("""
                    SELECT symbol,entry_date,entry_price,exit_date,exit_price,
                           shares,realized_pnl,inv_type,stop_loss
                    FROM positions
                    WHERE status='CLOSED' AND exit_price IS NOT NULL AND realized_pnl IS NOT NULL AND portfolio_id=?
                    ORDER BY exit_date ASC
                """, (pf_id,)).fetchall()
                _c.close()
                trades = [{"symbol":r[0],"entry_date":r[1],"entry_price":r[2],
                           "exit_date":r[3],"exit_price":r[4],"shares":r[5],
                           "pnl":round(float(r[6] or 0),2),"inv_type":r[7],
                           "sl":r[8]} for r in closed]

                if dtype == "calendar":
                    # Daily P&L map
                    daily = {}
                    for t in trades:
                        d = (t["exit_date"] or "")[:10]
                        if d: daily[d] = daily.get(d, 0) + t["pnl"]
                    best_day  = max(daily.values()) if daily else 0
                    worst_day = min(daily.values()) if daily else 0
                    profitable_days = sum(1 for v in daily.values() if v > 0)
                    return {"ok": True, "type": "calendar", "daily_pnl": daily,
                            "total_pnl": round(sum(daily.values()), 2),
                            "best_day": round(best_day, 2), "worst_day": round(worst_day, 2),
                            "profitable_days": profitable_days, "total_days": len(daily)}

                elif dtype == "equity_curve":
                    capital = 1000000.0
                    nifty_map = {}
                    try:
                        _c = _db()
                        nifty_rows = _c.execute(
                            "SELECT trade_date, close FROM daily_prices WHERE symbol='NIFTY50' ORDER BY trade_date ASC"
                        ).fetchall()
                        _c.close()
                        nifty_map = {r[0][:10]: float(r[1]) for r in nifty_rows if r[0]}
                    except Exception as e:
                        print(f"[ANALYTICS] Error fetching NIFTY50 close: {e}", flush=True)

                    sorted_nifty_dates = sorted(nifty_map.keys())
                    def get_nifty_price(target_date):
                        if not nifty_map: return 1.0
                        if target_date in nifty_map: return nifty_map[target_date]
                        for d in reversed(sorted_nifty_dates):
                            if d <= target_date: return nifty_map[d]
                        return nifty_map[sorted_nifty_dates[0]]

                    start_date = "2026-01-01"
                    if trades:
                        valid_dates = [t["entry_date"][:10] for t in trades if t.get("entry_date")]
                        if valid_dates: start_date = min(valid_dates)

                    start_nifty_val = get_nifty_price(start_date)

                    curve = [{"date": start_date, "value": capital, "nifty": capital}]
                    cum = capital
                    for t in trades:
                        cum += t["pnl"]
                        t_date = (t["exit_date"] or "")[:10]
                        nifty_price = get_nifty_price(t_date)
                        nifty_val = round((nifty_price / start_nifty_val) * capital, 2) if start_nifty_val > 0 else capital
                        curve.append({"date": t_date, "value": round(cum, 2), "nifty": nifty_val})

                    total_ret = round((cum - capital) / capital * 100, 2) if trades else 0.0
                    end_nifty = curve[-1]["nifty"] if curve else capital
                    nifty_ret = round((end_nifty - capital) / capital * 100, 2) if capital > 0 else 0.0

                    peak = capital; max_dd = 0.0
                    for pt in curve:
                        val = pt["value"]
                        peak = max(peak, val)
                        dd = (peak - val) / peak * 100
                        max_dd = max(max_dd, dd)

                    returns = []
                    for i in range(1, len(curve)):
                        prev = curve[i-1]["value"]
                        curr = curve[i]["value"]
                        if prev > 0:
                            returns.append((curr - prev) / prev)
                    if len(returns) > 1:
                        import math
                        mean_ret = sum(returns) / len(returns)
                        var_ret = sum((r - mean_ret)**2 for r in returns) / (len(returns) - 1)
                        std_ret = math.sqrt(var_ret)
                        sharpe = (mean_ret / std_ret * math.sqrt(252)) if std_ret > 0 else 0.0
                    else:
                        sharpe = 0.0

                    return {
                        "ok": True,
                        "type": "equity_curve",
                        "curve": curve,
                        "start_capital": capital,
                        "end_capital": round(cum, 2),
                        "total_return": total_ret,
                        "nifty_return": nifty_ret,
                        "sharpe": sharpe,
                        "max_drawdown": max_dd
                    }

                elif dtype == "statistics":
                    wins = [t for t in trades if t["pnl"] > 0]
                    losses = [t for t in trades if t["pnl"] <= 0]
                    avg_win  = round(sum(t["pnl"] for t in wins) / max(len(wins),1), 2)
                    avg_loss = round(sum(t["pnl"] for t in losses) / max(len(losses),1), 2)
                    win_pct  = round(len(wins) / max(len(trades),1) * 100, 1)
                    loss_pct = round(100.0 - win_pct, 1)
                    pf       = round(abs(sum(t["pnl"] for t in wins)) / max(abs(sum(t["pnl"] for t in losses)),1), 2)
                    exp      = round(win_pct/100 * avg_win + (1-win_pct/100) * avg_loss, 2)
                    # R:R
                    rr_list = [abs((t["exit_price"]-t["entry_price"])/(t["entry_price"]-t["sl"])) 
                               for t in trades if t.get("sl") and t["entry_price"] != t.get("sl",0)]
                    avg_rr   = round(sum(rr_list)/max(len(rr_list),1), 2)
                    best_rr  = round(max(rr_list) if rr_list else 0.0, 2)
                    worst_rr = round(min(rr_list) if rr_list else 0.0, 2)
                    
                    # Consec wins/losses & hold days
                    hold_days_list = []
                    max_consec_wins = 0; max_consec_losses = 0
                    current_wins = 0; current_losses = 0
                    
                    for t in sorted(trades, key=lambda x: x["exit_date"] or ""):
                        try:
                            if t["entry_date"] and t["exit_date"]:
                                hd = (date.fromisoformat(t["exit_date"][:10]) - date.fromisoformat(t["entry_date"][:10])).days
                                hold_days_list.append(hd)
                        except Exception as e:
                            print(f"[STATS] Error calculating hold days: {e}", flush=True)
                            
                        if t["pnl"] > 0:
                            current_wins += 1
                            current_losses = 0
                            max_consec_wins = max(max_consec_wins, current_wins)
                        else:
                            current_losses += 1
                            current_wins = 0
                            max_consec_losses = max(max_consec_losses, current_losses)
                            
                    avg_hold = round(sum(hold_days_list)/max(len(hold_days_list),1), 1)
                    
                    # Recent streak (last 15 trades)
                    recent_trades = sorted(trades, key=lambda x: x["exit_date"] or "")[-15:]
                    recent_streak = ["W" if t["pnl"] > 0 else "L" for t in recent_trades]
                    
                    # Monthly P&L
                    monthly = {}
                    for t in trades:
                        m = (t["exit_date"] or "")[:7]
                        if m: monthly[m] = monthly.get(m, 0) + t["pnl"]
                    monthly_list = [{"month": k, "pnl": round(v, 2)} for k, v in sorted(monthly.items())]
                    
                    # Weekday P&L stats
                    weekday_map = {0: "Mon", 1: "Tue", 2: "Wed", 3: "Thu", 4: "Fri"}
                    weekday_stats = {d: {"wins": 0, "losses": 0} for d in weekday_map.values()}
                    for t in trades:
                        if t["exit_date"]:
                            try:
                                dt_obj = date.fromisoformat(t["exit_date"][:10])
                                day_name = weekday_map.get(dt_obj.weekday())
                                if day_name:
                                    if t["pnl"] > 0:
                                        weekday_stats[day_name]["wins"] += 1
                                    else:
                                        weekday_stats[day_name]["losses"] += 1
                            except Exception: pass
                    weekday_list = [{"day": k, "wins": v["wins"], "losses": v["losses"]} for k, v in weekday_stats.items()]

                    return {"ok": True, "type": "statistics",
                            "total_trades": len(trades), 
                            "win_pct": win_pct,
                            "loss_pct": loss_pct,
                            "avg_win": avg_win, "avg_loss": avg_loss,
                            "profit_factor": pf, "expectancy": exp,
                            "avg_rr": avg_rr, "best_rr": best_rr, "worst_rr": worst_rr,
                            "avg_hold_days": avg_hold,
                            "max_consec_wins": max_consec_wins,
                            "max_consec_losses": max_consec_losses,
                            "recent_streak": recent_streak,
                            "monthly_pnl": monthly_list,
                            "weekday": weekday_list}

                elif dtype == "best_worst":
                    mode = p.get("mode", "best")
                    sorted_trades = sorted(trades, key=lambda x: x["pnl"], reverse=True)
                    if mode == "best":
                        selected_trades = sorted_trades[:10]
                    else:
                        selected_trades = sorted_trades[-10:][::-1]
                    mapped_trades = []
                    for idx, t in enumerate(selected_trades):
                        entry = t["entry_price"]
                        exit_pr = t["exit_price"]
                        pct = round((exit_pr - entry) / entry * 100, 2) if entry else 0.0
                        sl = t.get("sl")
                        rr = round(abs((exit_pr - entry) / (entry - sl)), 2) if sl and entry != sl else 0.0
                        mapped_trades.append({
                            "rank": idx + 1,
                            "symbol": t["symbol"],
                            "entry": round(entry, 2),
                            "exit": round(exit_pr, 2),
                            "entry_date": t["entry_date"][:10] if t["entry_date"] else "",
                            "exit_date": t["exit_date"][:10] if t["exit_date"] else "",
                            "pnl": round(t["pnl"], 2),
                            "pct": pct,
                            "rr": rr,
                            "inv_type": t["inv_type"]
                        })
                    return {"ok": True, "type": "best_worst", "trades": mapped_trades}

                elif dtype == "sector_pnl":
                    sector_map = {}
                    for t in trades:
                        inst = ALL_INSTRUMENTS.get(t["symbol"])
                        sec = inst.sector if inst else "Other"
                        if sec not in sector_map:
                            sector_map[sec] = {"trades": 0, "wins": 0, "pnl": 0}
                        sector_map[sec]["trades"] += 1
                        if t["pnl"] > 0: sector_map[sec]["wins"] += 1
                        sector_map[sec]["pnl"] = round(sector_map[sec]["pnl"] + t["pnl"], 2)
                    sectors = [{"name": k, "trades": v["trades"], "wins": v["wins"],
                                "win_pct": round(v["wins"]/max(v["trades"],1)*100,1),
                                "net_pnl": v["pnl"],
                                "avg_pnl": round(v["pnl"]/max(v["trades"],1),2)}
                               for k, v in sector_map.items()]
                    sectors.sort(key=lambda x: x["net_pnl"], reverse=True)
                    return {"ok": True, "type": "sector_pnl", "sectors": sectors}

                return {"ok": False, "error": f"Unknown analytics type: {dtype}"}
            except Exception as e:
                import traceback as _tb
                return {"ok": False, "error": str(e), "trace": _tb.format_exc()[-400:]}

        # ── analytics_whatsapp_report — send summary via WhatsApp ─────────────
        if ep == "analytics_whatsapp_report":
            try:
                _c = _db()
                user_id = self.current_user["user_id"]
                pf_id = _get_user_portfolio_id(_c, user_id)
                closed = _c.execute("""
                    SELECT realized_pnl FROM positions
                    WHERE status='CLOSED' AND realized_pnl IS NOT NULL AND portfolio_id=?
                """, (pf_id,)).fetchall()
                open_n = _c.execute("SELECT COUNT(*) FROM positions WHERE status='OPEN' AND portfolio_id=?", (pf_id,)).fetchone()[0]
                _c.close()
                wins   = sum(1 for r in closed if (r[0] or 0) > 0)
                total  = len(closed)
                net    = round(sum(r[0] or 0 for r in closed), 2)
                wr     = round(wins/max(total,1)*100,1)
                msg = (f"📊 Vprofitables Portfolio Report\n"
                       f"Total Trades: {total} | Win Rate: {wr}%\n"
                       f"Net Realized P&L: ₹{net:,.2f}\n"
                       f"Open Positions: {open_n}\n"
                       f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}")
                result = _notifier_send_signal({"body": msg, "symbol": "PORTFOLIO", "action": "REPORT"})
                return {"ok": True, "message": msg, "notify_result": result}
            except Exception as e:
                return {"ok": False, "error": str(e)}

        # ── oi_chain — full option chain for a symbol + date ─────────────────
        if ep == "oi_chain":
            sym  = p.get("symbol", "NIFTY").upper()
            dt   = p.get("date", today)
            exp  = p.get("expiry", "")   # optional filter by expiry
            try:
                _c = _db()
                if exp:
                    rows = _c.execute("""
                        SELECT strike, option_type, oi, change_in_oi, volume, iv, ltp, bid, ask, expiry_date
                        FROM option_chain_data
                        WHERE symbol=? AND trade_date=? AND expiry_date=?
                        ORDER BY strike, option_type
                    """, (sym, dt, exp)).fetchall()
                else:
                    # Use nearest available expiry for this date
                    nearest = _c.execute("""
                        SELECT expiry_date FROM option_chain_data
                        WHERE symbol=? AND trade_date=?
                        ORDER BY expiry_date LIMIT 1
                    """, (sym, dt)).fetchone()
                    near_exp = nearest[0] if nearest else ""
                    rows = _c.execute("""
                        SELECT strike, option_type, oi, change_in_oi, volume, iv, ltp, bid, ask, expiry_date
                        FROM option_chain_data
                        WHERE symbol=? AND trade_date=? AND expiry_date=?
                        ORDER BY strike, option_type
                    """, (sym, dt, near_exp)).fetchall()
                _c.close()
                chain = [{"strike": r[0], "option_type": r[1], "oi": r[2],
                          "change_in_oi": r[3], "volume": r[4], "iv": r[5],
                          "ltp": r[6], "bid": r[7], "ask": r[8],
                          "expiry_date": r[9]} for r in rows]
                return {"ok": True, "symbol": sym, "date": dt,
                        "expiry": exp or (rows[0][9] if rows else ""),
                        "chain": chain, "count": len(chain)}
            except Exception as e:
                return {"ok": False, "error": str(e)}

        # ── pcr — PCR history for a symbol (last N days) ─────────────────────
        if ep == "pcr":
            sym   = p.get("symbol", "NIFTY").upper()
            limit = int(p.get("days", 30))
            try:
                _c = _db()
                rows = _c.execute("""
                    SELECT trade_date, expiry_date, total_ce_oi, total_pe_oi,
                           pcr, max_pain, atm_strike, spot_price
                    FROM pcr_summary
                    WHERE symbol=?
                    ORDER BY trade_date DESC LIMIT ?
                """, (sym, limit)).fetchall()
                _c.close()
                history = [{"date": r[0], "expiry": r[1],
                            "ce_oi": r[2], "pe_oi": r[3],
                            "pcr": r[4], "max_pain": r[5],
                            "atm": r[6], "spot": r[7]} for r in rows]
                latest = history[0] if history else {}
                sentiment = (
                    "BEARISH" if latest.get("pcr", 1) > 1.5 else
                    "BULLISH" if latest.get("pcr", 1) < 0.7 else
                    "NEUTRAL"
                )
                return {"ok": True, "symbol": sym, "history": history,
                        "latest_pcr": latest.get("pcr", 0),
                        "sentiment": sentiment,
                        "max_pain": latest.get("max_pain", 0)}
            except Exception as e:
                return {"ok": False, "error": str(e)}

        # ── max_pain — max pain strike for symbol + expiry ───────────────────
        if ep == "max_pain":
            sym    = p.get("symbol", "NIFTY").upper()
            dt     = p.get("date", today)
            expiry = p.get("expiry", "")
            try:
                _c = _db()
                if not expiry:
                    # Get nearest expiry
                    nearest = _c.execute("""
                        SELECT expiry_date FROM pcr_summary
                        WHERE symbol=? AND trade_date=?
                        ORDER BY expiry_date LIMIT 1
                    """, (sym, dt)).fetchone()
                    expiry = nearest[0] if nearest else ""
                row = _c.execute("""
                    SELECT pcr, max_pain, atm_strike, spot_price,
                           total_ce_oi, total_pe_oi
                    FROM pcr_summary
                    WHERE symbol=? AND trade_date=? AND expiry_date=?
                """, (sym, dt, expiry)).fetchone()
                _c.close()
                if not row:
                    return {"ok": False, "error": f"No OI data for {sym} on {dt}"}
                return {"ok": True, "symbol": sym, "date": dt,
                        "expiry": expiry,
                        "pcr": row[0], "max_pain": row[1],
                        "atm_strike": row[2], "spot": row[3],
                        "ce_oi": row[4], "pe_oi": row[5]}
            except Exception as e:
                return {"ok": False, "error": str(e)}

        raise ValueError(f"Unknown endpoint: {ep}")



# ─────────────────────────────────────────────────────────────────────────────
# STARTUP PRICE FETCH — runs every time app.py is started
# Fetches latest EOD prices for all instruments via yfinance batch download
# Compares against yesterday's cached price and prints what changed
# ─────────────────────────────────────────────────────────────────────────────
def fetch_prices_on_startup():
    """
    Pull current/EOD prices from yfinance for every instrument with a yfinance symbol.
    Uses batch download (single HTTP call for up to 100 symbols) — fast.
    Saves to SQLite. /api/price reads from this cache instantly.
    """
    import sqlite3

    print("\n  Fetching current prices from yfinance...")

    try:
        import yfinance as yf
    except ImportError:
        print("  [SKIP] yfinance not installed.")
        print("         Run:  pip install yfinance")
        print("         Prices will use ATH-based fallback until then.\n")
        return 0

    # Collect all yfinance symbols
    sym_map = {}   # { yf_symbol: our_symbol }
    for our_sym, inst in ALL_INSTRUMENTS.items():
        yf_sym = (inst.yfinance_symbol or "").strip()
        if yf_sym:
            sym_map[yf_sym] = our_sym

    if not sym_map:
        print("  [SKIP] No yfinance symbols configured.\n")
        return 0

    yf_syms   = list(sym_map.keys())
    today_str = date.today().isoformat()
    updated   = 0
    errors    = 0

    # Load yesterday's cached prices to detect changes
    init_db()
    prev_cache = get_cached_prices()

    # ── Separate indices (^) from equities (.NS/.BO) — yfinance handles
    #    mixed symbol types poorly in one batch download ──────────────────
    index_syms  = [s for s in yf_syms if s.startswith("^")]
    equity_syms = [s for s in yf_syms if not s.startswith("^")]

    all_data = {}   # { yf_sym: {close, high, low, open, volume, change_pct} }

    def _parse_row(raw, yf_sym, chunk):
        """Extract OHLCV for one symbol from a yf.download() result."""
        import math as _math
        if len(chunk) == 1:
            # Single-ticker download: flat columns (Close, High, Low, Open, Volume)
            if len(raw) == 0:
                return None
            try:
                cl  = float(raw["Close"].iloc[-1])
                pr  = float(raw["Close"].iloc[-2]) if len(raw) > 1 else cl
                hi  = float(raw["High"].iloc[-1])
                lo  = float(raw["Low"].iloc[-1])
                op  = float(raw["Open"].iloc[-1])
                vol = int(raw["Volume"].iloc[-1]) if "Volume" in raw.columns else 0
            except Exception:
                return None
        else:
            # Multi-ticker: MultiIndex (field, ticker) columns
            # yfinance ≥0.2 uses (Price, Ticker) MultiIndex
            try:
                # Try standard MultiIndex access
                close_col = raw.get("Close", raw.get("close", None))
                if close_col is None:
                    return None
                if hasattr(close_col, "columns"):
                    # MultiIndex — access by ticker name
                    if yf_sym not in close_col.columns:
                        return None
                    closes = close_col[yf_sym].dropna()
                else:
                    closes = close_col.dropna()
                if len(closes) == 0:
                    return None
                cl = float(closes.iloc[-1])
                pr = float(closes.iloc[-2]) if len(closes) > 1 else cl

                def _get(field):
                    col = raw.get(field, None)
                    if col is None: return cl
                    if hasattr(col, "columns"):
                        return float(col[yf_sym].dropna().iloc[-1]) if yf_sym in col.columns else cl
                    return float(col.dropna().iloc[-1])

                hi  = _get("High")
                lo  = _get("Low")
                op  = _get("Open")
                try:
                    vol_col = raw.get("Volume", None)
                    if vol_col is not None and hasattr(vol_col, "columns") and yf_sym in vol_col.columns:
                        vol = int(vol_col[yf_sym].dropna().iloc[-1])
                    else:
                        vol = 0
                except Exception:
                    vol = 0
            except Exception:
                return None

        if any(_math.isnan(v) for v in [cl, hi, lo, op] if v is not None):
            return None
        chg = round((cl - pr) / pr * 100, 2) if pr else 0
        return {"close": round(cl, 2), "high": round(hi, 2),
                "low":   round(lo, 2), "open": round(op, 2),
                "volume": vol, "change_pct": chg}

    def _download_chunk(chunk, label=""):
        """Download a chunk and parse all symbols."""
        import io as _io, sys as _sys
        _old_err = _sys.stderr; _sys.stderr = _io.StringIO()
        try:
            raw = yf.download(
                tickers     = " ".join(chunk),
                period      = "5d",        # 5d to ensure we get latest even after holidays
                interval    = "1d",
                auto_adjust = True,
                progress    = False,
                threads     = True,
                group_by    = "column",    # consistent column structure
            )
        finally:
            _sys.stderr = _old_err
        if raw is None or raw.empty:
            return
        for yf_sym in chunk:
            parsed = _parse_row(raw, yf_sym, chunk)
            if parsed:
                all_data[yf_sym] = parsed

    # Download indices (smaller batch — typically 39 symbols)
    CHUNK_IDX = 40
    for i in range(0, len(index_syms), CHUNK_IDX):
        chunk = index_syms[i:i+CHUNK_IDX]
        try:
            _download_chunk(chunk, f"INDEX batch {i//CHUNK_IDX+1}")
        except Exception as e:
            errors += len(chunk)
            print(f"  [ERR] Index batch {i//CHUNK_IDX+1}: {e}")

    # Download equities in chunks of 50 (more conservative for .NS symbols)
    CHUNK_EQ = 50
    eq_ok = eq_err = 0
    for i in range(0, len(equity_syms), CHUNK_EQ):
        chunk = equity_syms[i:i+CHUNK_EQ]
        try:
            before = len(all_data)
            _download_chunk(chunk, f"EQ batch {i//CHUNK_EQ+1}")
            fetched = len(all_data) - before
            eq_ok  += fetched
            eq_err += len(chunk) - fetched
        except Exception as e:
            errors += len(chunk)
            eq_err += len(chunk)
            print(f"  [ERR] Equity batch {i//CHUNK_EQ+1}: {e}")

    print(f"  Indices fetched:  {sum(1 for s in index_syms if s in all_data)}/{len(index_syms)}")
    print(f"  Equities fetched: {sum(1 for s in equity_syms if s in all_data)}/{len(equity_syms)}")

    # ── Write to SQLite & print changes ──
    conn = _db()
    c    = conn.cursor()
    now  = datetime.now().isoformat()

    changed_lines  = []   # price moved vs yesterday
    unchanged_syms = 0

    for yf_sym, row in all_data.items():
        our_sym = sym_map.get(yf_sym)
        if not our_sym:
            continue
        c.execute("""
            INSERT OR REPLACE INTO daily_prices
            (symbol, trade_date, open, high, low, close, volume, change_pct, updated_at)
            VALUES (?,?,?,?,?,?,?,?,?)
        """, (our_sym, today_str,
              row["open"], row["high"], row["low"], row["close"],
              row["volume"], row["change_pct"], now))
        updated += 1

        # Compare to previous cache
        prev = prev_cache.get(our_sym)
        prev_close = float(prev["close"]) if prev and prev.get("close") else None
        new_close  = row["close"]

        if prev_close is None:
            changed_lines.append(f"    NEW  {our_sym:<16} {new_close:>10,.2f}  ({row['change_pct']:+.2f}%)")
        elif abs(new_close - prev_close) / max(prev_close, 1) > 0.0005:  # > 0.05% change
            direction = "▲" if new_close > prev_close else "▼"
            changed_lines.append(
                f"    {direction}    {our_sym:<16} {new_close:>10,.2f}  "
                f"({row['change_pct']:+.2f}%)  prev {prev_close:,.2f}"
            )
        else:
            unchanged_syms += 1

    conn.commit()
    conn.close()

    # ── Print summary ──
    print(f"  ✓  {updated} prices fetched  |  {errors} errors")
    if changed_lines:
        print(f"\n  Price changes since last run  ({len(changed_lines)} instruments):")
        for ln in changed_lines[:30]:   # cap at 30 lines
            print(ln)
        if len(changed_lines) > 30:
            print(f"    ... and {len(changed_lines)-30} more")
    else:
        print(f"  No price changes detected ({unchanged_syms} instruments unchanged).")
    print()
    return updated


def main():
    SEP = "=" * 56
    print("\n" + SEP, flush=True)
    print("  Vprofitables Intelligence System v3.9", flush=True)
    print(f"  Python  : {sys.version.split()[0]}", flush=True)
    print(f"  DB      : {DB_PATH}", flush=True)
    print(f"  Date    : {date.today()}", flush=True)
    print(f"  Port    : {PORT}", flush=True)
    print(SEP, flush=True)

    print("  [MAIN] Step 1/7: Seeding pivots...", flush=True)
    try:
        seed_static_pivots()
        print("  [MAIN] Step 1/7: Pivots OK", flush=True)
    except Exception as e:
        print(f"  [WARN] pivot seed: {e}", flush=True)
    print("  [MAIN] Step 2/7: ATL reconcile...", flush=True)
    # After seeding, reconcile: if instruments.py ATL < DB historical low → update to STATIC_VERIFIED
    try:
        _conn_s = _db()
        _c_s    = _conn_s.cursor()
        updated = 0
        for sym, inst in ALL_INSTRUMENTS.items():
            # 1. Update live Instrument objects with stored DB values (critical for dynamic stocks)
            row_atl = _c_s.execute(
                "SELECT pivot_price FROM pivot_levels WHERE symbol=? AND label='ATL'",
                (sym,)).fetchone()
            if row_atl:
                inst.all_time_low = float(row_atl[0])

            row_ath = _c_s.execute(
                "SELECT pivot_price FROM pivot_levels WHERE symbol=? AND label='ATH'",
                (sym,)).fetchone()
            if row_ath:
                inst.all_time_high = float(row_ath[0])

            # 2. Reconcile hand-researched static ATLs (only for core instruments)
            is_dynamic = getattr(inst, 'atl_date', None) is None
            if is_dynamic:
                continue
                
            static_atl  = getattr(inst, 'all_time_low', None)
            static_date = str(getattr(inst, 'atl_date', None) or getattr(inst, 'inception_date', ''))
            if not static_atl:
                continue
            static_atl = float(static_atl)
            
            # Get DB ATL
            row = _c_s.execute(
                "SELECT pivot_price, source FROM pivot_levels WHERE symbol=? AND label='ATL'",
                (sym,)).fetchone()
            if not row:
                continue
            db_atl, db_src = float(row[0]), row[1]
            if db_src == 'USER':
                continue
            if static_atl < db_atl:
                desc = (f"All-Time Low {static_atl:,.2f} — {static_date} "
                        f"(instruments.py {static_atl:,.2f} < DB auto {db_atl:,.2f}, manually verified)")
                _c_s.execute("""UPDATE pivot_levels
                    SET pivot_price=?, pivot_date=?, source='STATIC_VERIFIED',
                        description=?, updated_at=?
                    WHERE symbol=? AND label='ATL' AND source!='USER'""",
                    (static_atl, static_date, desc, datetime.now().isoformat(), sym))
                updated += 1
        _conn_s.commit(); _conn_s.close()
        if updated: print(f"  [ATL] Reconciled {updated} symbols: instruments.py ATL < DB ATL → STATIC_VERIFIED")
    except Exception as _e_s:
        print(f"  [WARN] ATL reconcile: {_e_s}")
    print(f"  NSE / BSE / MCX — {len(ALL_INSTRUMENTS)} Instruments")
    print(SEP)
    print(f"\n  URL:   http://localhost:{PORT}")
    print(f"  Date:  {date.today().isoformat()}")
    print("\n  Modules loaded:")
    print("    [OK] Ephemeris (VSOP87) + 11 planets")
    print("    [OK] Aspect detection")
    print("    [OK] Gann Math (Sq9, Angles, Time Cycles)")
    print("    [OK] Signal Engine (Confluence scorer)")
    print("    [OK] Quant Engine (Fourier/FFT, Regime, Backtest, S/R)")
    print(f"    [OK] {len(ALL_INSTRUMENTS)} instruments with natal charts")

    # ── Show DB coverage ──────────────────────────────────────────
    print("  [MAIN] Step 3/7: Checking DB coverage...", flush=True)
    try:
        _conn = _db()
        _cnt  = _conn.execute("SELECT COUNT(*) FROM daily_prices").fetchone()[0]
        _syms = _conn.execute("SELECT COUNT(DISTINCT symbol) FROM daily_prices").fetchone()[0]
        _conn.close()
        if _cnt > 0:
            print(f"    [DB] Historical data: {_cnt:,} rows across {_syms} symbols", flush=True)
        else:
            print(f"    [DB] No historical data yet — run: python download_history.py", flush=True)
        print("  [MAIN] Step 3/7: DB OK", flush=True)
    except Exception as _dbe:
        print(f"  [WARN] DB coverage check failed: {_dbe}", flush=True)
    print("  [MAIN] Step 4/7: Starting background init thread...", flush=True)
    def _bg():
        print("  [BG 1/2] Pre-computing VSOP87 planet positions...", flush=True)
        try:
            from core.ephemeris import get_all_planets as _g
            _g(date.today())
            print("  [BG 1/2] Planets ready", flush=True)
        except Exception as _e:
            import traceback as _bgtb
            print(f"  [BG 1/2] Ephemeris error: {_e}", flush=True)
            _bgtb.print_exc()
        print("  [BG 2/2] Fetching live prices from Yahoo Finance...", flush=True)
        try:
            fetch_prices_on_startup()
            print("  [BG 2/2] Prices done -- system fully loaded", flush=True)
        except Exception as _e:
            import traceback as _bgtb2
            print(f"  [BG 2/2] Price error: {_e}", flush=True)
            _bgtb2.print_exc()
    threading.Thread(target=_bg, daemon=True, name="bg-init").start()
    print("  [MAIN] Step 4/7: Background thread started", flush=True)

    # ── Pre-warm RAG embedding model in background ────────────────────────────
    # Loads all-MiniLM-L6-v2 into memory NOW so the first /api/llm_extract
    # request doesn't pay the 20-second cold-start cost.
    def _prewarm_rag():
        try:
            from core.rag_engine import RAG_AVAILABLE, _get_embedder, _get_collection
            from core.llm_extractor import _ensure_tables
            _ensure_tables()
            if RAG_AVAILABLE:
                _get_embedder()      # loads sentence-transformers model into RAM
                _get_collection()    # opens ChromaDB (creates dir if needed)
                print("  [RAG] Pre-warm complete — embedding model in memory", flush=True)
        except Exception as _e:
            print(f"  [RAG] Pre-warm skipped: {_e}", flush=True)
    threading.Thread(target=_prewarm_rag, daemon=True, name="RAG-Prewarm").start()

    # ── Background update-check (non-blocking, non-fatal) ────────────────────
    def _check_for_updates():
        try:
            import urllib.request as _ureq, json as _ujson
            _url = "https://api.github.com/repos/chiragkaura/Vprofitables/releases/latest"
            with _ureq.urlopen(_ureq.Request(_url, headers={"User-Agent": "Vprofitables"}), timeout=5) as _r:
                _data = _ujson.loads(_r.read())
            _latest = _data.get("tag_name", "").lstrip("v")
            _current = "3.9"
            if _latest and _latest != _current:
                print(f"  [UPDATE] ✨ New version v{_latest} available!", flush=True)
                print(f"  [UPDATE]    Download: {_data.get('html_url', '')}", flush=True)
        except Exception:
            pass  # non-fatal — no internet is fine
    threading.Thread(target=_check_for_updates, daemon=True, name="UpdateCheck").start()

    print(f"  [MAIN] Step 5/7: Creating HTTP server on port {PORT}...", flush=True)
    class _THS(socketserver.ThreadingMixIn, http.server.HTTPServer):
        daemon_threads = True
        allow_reuse_address = True
    try:
        # Bind to 0.0.0.0 so the app is reachable from LAN and internet
        # (localhost would refuse all connections except from this machine)
        server = _THS(("0.0.0.0", PORT), Handler)
        print(f"  [MAIN] Step 5/7: HTTP server bound to 0.0.0.0:{PORT} OK", flush=True)
        print(f"  [MAIN]           Local:   http://localhost:{PORT}", flush=True)
        print(f"  [MAIN]           Network: http://<your-ip>:{PORT}", flush=True)
    except OSError as _ose:
        print(f"  [FATAL] Cannot bind to port {PORT}: {_ose}", flush=True)
        print(f"  [FATAL] Port {PORT} may already be in use.", flush=True)
        print(f"  [FATAL] Kill existing python.exe in Task Manager, then retry.", flush=True)
        import traceback as _ostb
        _ostb.print_exc()
        raise

    print(f"  [MAIN] Step 6/7: Opening browser...", flush=True)
    # When running as installed .exe app, always open browser automatically
    # When running as script (dev mode), browser open is suppressed to avoid duplicate tabs
    if _IS_FROZEN:
        threading.Timer(2.5, lambda: webbrowser.open(f"http://localhost:{PORT}")).start()
        print(f"  [MAIN]           Browser will open in 2.5s at http://localhost:{PORT}", flush=True)
    else:
        print(f"  [MAIN]           Dev mode — open manually: http://localhost:{PORT}", flush=True)

    print("  [MAIN] Step 7/7: Server running — press Ctrl+C to stop", flush=True)
    print("  Press Ctrl+C to stop\n", flush=True)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n  Stopped.", flush=True)
    except Exception as _sfe:
        import traceback as _sftb
        print(f"  [FATAL] serve_forever crashed: {_sfe}", flush=True)
        _sftb.print_exc()
        raise


if __name__ == "__main__":
    main()